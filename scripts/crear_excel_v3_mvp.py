#!/usr/bin/env python3
"""
CREAR EXCEL V3.0 - MVP (FASE 1)
================================

PROP√ìSITO: Crear estructura completa de Excel v3.0 desde CERO

APLICA LECCIONES APRENDIDAS:
- Diagn√≥stico PRIMERO (‚úÖ Ya hecho)
- Construcci√≥n con especificaciones completas (‚úÖ Cuestionario + diagn√≥stico)
- Validaciones desde el inicio (‚úÖ Incluidas)
- Manual inline con comentarios (‚úÖ Incluido)

FASE 1 MVP - Due: Nov 19, 2025 (7 d√≠as)
---------------------------------------
‚úÖ Hoja TRANSACCIONES (20 columnas)
‚úÖ Hoja EFECTIVO (9 cuentas bancarias)
‚úÖ Hoja DASHBOARD (KPIs cr√≠ticos)
‚úÖ Hoja ENTIDADES_ALIAS (22 clientes)
‚úÖ Validaciones y protecci√≥n
‚úÖ Manual inline (comentarios)

Ejecutar:
    python scripts/crear_excel_v3_mvp.py

Salida:
    AlvaroVelasco_Finanzas_v3.0.xlsx
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime
import os

# ============================================================================
# CONFIGURACI√ìN GLOBAL
# ============================================================================

EXCEL_FILE = "AlvaroVelasco_Finanzas_v3.0.xlsx"

# Colores corporativos
COLOR_HEADER = "1F4E78"      # Azul oscuro
COLOR_EDITABLE = "FFF2CC"    # Amarillo claro
COLOR_FORMULA = "FFFFFF"     # Blanco
COLOR_WARNING = "FCE4D6"     # Naranja claro
COLOR_SUCCESS = "C6EFCE"     # Verde claro
COLOR_ERROR = "FFC7CE"       # Rojo claro

# Estilos de fuente
FONT_HEADER = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
FONT_NORMAL = Font(name='Calibri', size=10)
FONT_SMALL = Font(name='Calibri', size=9, italic=True)

# Bordes
BORDER_THIN = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ============================================================================
# DATOS DE CONFIGURACI√ìN (Del Cuestionario v3.0)
# ============================================================================

# 22 Clientes REALES (Bloque 2 T3 - Facturaci√≥n Noviembre 2025)
# TOTAL: $9,466.42
CLIENTES = [
    {"nombre": "Grupo Acci√≥n Comercial S.A.", "categoria": "VIP", "monto_nov": 1689.04, "factura": "AR-002"},
    {"nombre": "VWR International Ltda", "categoria": "VIP", "monto_nov": 1400.00, "factura": "AR-001"},
    {"nombre": "Alfipac (Almac√©n Fiscal Pac√≠fico)", "categoria": "VIP", "monto_nov": 761.05, "factura": "AR-003"},
    {"nombre": "3-102-887892 SRL", "categoria": "Regular", "monto_nov": 691.56, "factura": "AR-004"},
    {"nombre": "Waipio S.A.", "categoria": "Regular", "monto_nov": 687.27, "factura": "AR-005"},
    {"nombre": "Centro Integral Oncolog√≠a CIO SRL", "categoria": "Regular", "monto_nov": 687.05, "factura": "AR-006"},
    {"nombre": "Ortodoncia de la Cruz", "categoria": "Regular", "monto_nov": 494.50, "factura": "AR-007"},
    {"nombre": "Global Automotriz GACR S.A.", "categoria": "Regular", "monto_nov": 439.61, "factura": "AR-008"},
    {"nombre": "Solusa Consolidators", "categoria": "Regular", "monto_nov": 378.35, "factura": "AR-009"},
    {"nombre": "Cemso", "categoria": "Regular", "monto_nov": 333.92, "factura": "AR-010"},
    {"nombre": "Acacia (Asoc. CR Agencias Carga)", "categoria": "Regular", "monto_nov": 333.35, "factura": "AR-011"},
    {"nombre": "Rodriguez Rojas Carlos Humberto", "categoria": "Regular", "monto_nov": 282.50, "factura": "AR-012"},
    {"nombre": "Supply Net C.R.W.H S.A.", "categoria": "Regular", "monto_nov": 276.85, "factura": "AR-013"},
    {"nombre": "Operation Managment Tierra Magnifica", "categoria": "Regular", "monto_nov": 209.06, "factura": "AR-014"},
    {"nombre": "Gentra de Costa Rica S.A.", "categoria": "Regular", "monto_nov": 183.63, "factura": "AR-015"},
    {"nombre": "Sevilla Navarro Edgar", "categoria": "Regular", "monto_nov": 169.50, "factura": "AR-016"},
    {"nombre": "Gomez Ajoy Edgar Luis", "categoria": "Regular", "monto_nov": 113.00, "factura": "AR-017"},
    {"nombre": "Melendez Morales Monica", "categoria": "Regular", "monto_nov": 113.00, "factura": "AR-018"},
    {"nombre": "Bandogo Soluciones Tecnol√≥gicas S.A.", "categoria": "Regular", "monto_nov": 67.80, "factura": "AR-019"},
    {"nombre": "CPF Servicios Radiol√≥gicos S.A.", "categoria": "Regular", "monto_nov": 56.50, "factura": "AR-020"},
    {"nombre": "Ortodec S.A.", "categoria": "Regular", "monto_nov": 56.50, "factura": "AR-021"},
    {"nombre": "Perez Morales Francisco", "categoria": "Regular", "monto_nov": 42.38, "factura": "AR-022"},
]

# 9 Cuentas Bancarias REALES (Bloque 3 B1 - Saldos 12 Nov 2025)
# TOTAL EFECTIVO: $3,444.54 (12.9 d√≠as de cobertura)
CUENTAS_BANCARIAS = [
    # BNCR (5 cuentas)
    {"nombre": "BNCR CRC Ahorros (***8618)", "numero": "100-01-000-188618-3", "tipo": "Banco", "moneda": "CRC", "saldo": 211.24, "uso": "NEGOCIO"},
    {"nombre": "BNCR USD Ahorros (***1066)", "numero": "100-02-087-601066-4", "tipo": "Banco", "moneda": "USD", "saldo": 1087.37, "uso": "NEGOCIO"},
    {"nombre": "BNCR CRC Corriente (***2186)", "numero": "200-01-087-042186-9", "tipo": "Banco", "moneda": "CRC", "saldo": 28950.50, "uso": "NEGOCIO/RESERVAS"},
    {"nombre": "BNCR USD Corriente (***9589)", "numero": "200-02-087-009589-4", "tipo": "Banco", "moneda": "USD", "saldo": 0.43, "uso": "PERSONAL"},
    {"nombre": "BNCR USD Corriente (***1112)", "numero": "200-02-087-011112-1", "tipo": "Banco", "moneda": "USD", "saldo": 21.84, "uso": "PERSONAL"},
    # PROMERICA (4 cuentas - A nombre de "ALVARO VELASCONET SRL")
    {"nombre": "Promerica CRC SINPE (***1708)", "numero": "10000003881708", "tipo": "Banco", "moneda": "CRC", "saldo": 1090.00, "uso": "NEGOCIO"},
    {"nombre": "Promerica USD Ahorros (***1691)", "numero": "20000003881691", "tipo": "Banco", "moneda": "USD", "saldo": 0.00, "uso": "NEGOCIO"},
    {"nombre": "Promerica CRC CC Corp (***4229)", "numero": "30000003904229", "tipo": "Banco", "moneda": "CRC", "saldo": 0.00, "uso": "NEGOCIO"},
    {"nombre": "Promerica USD CC Corp (***1774)", "numero": "40000003881774", "tipo": "Banco", "moneda": "USD", "saldo": 2276.44, "uso": "NEGOCIO"},
]

# 5 Tarjetas de Cr√©dito REALES (Bloque 1 C1 - Saldos 12 Nov 2025)
# TOTAL DEUDA TC: $14,867.73 USD + ‚Ç°863,830 CRC (~$16,536 USD equivalente)
TARJETAS_CREDITO = [
    {"nombre": "BNCR VISA ***3519 (Alvaro)", "saldo_usd": 3864.90, "saldo_crc": 0, "uso": "PERSONAL", "estrategia": "Pago total mensual"},
    {"nombre": "BNCR VISA ***9837 (Alvaro)", "saldo_usd": 3299.01, "saldo_crc": 0, "uso": "EMPRESA", "estrategia": "Pago m√≠nimo √ó 1.5"},
    {"nombre": "BNCR VISA ***6386 (Alejandra)", "saldo_usd": 5195.07, "saldo_crc": 0, "uso": "EMPRESA", "estrategia": "Pago m√≠nimo √ó 1.5"},
    {"nombre": "BNCR MC ***8759 (Alvaro)", "saldo_usd": 0, "saldo_crc": 863830, "uso": "EMPRESA", "estrategia": "Pago m√≠nimo √ó 1.5"},
    {"nombre": "BAC VISA ***9550 (Alvaro)", "saldo_usd": 2508.75, "saldo_crc": 0, "uso": "EMPRESA", "estrategia": "Pago m√≠nimo √ó 1.5"},
]

# 5 Proveedores Principales
PROVEEDORES = [
    "Intcomex Costa Rica",
    "Eurocomp S.A.",
    "CompuEcon√≥micos",
    "TD Synex",
    "ICD Soft",
]

# Tipos de Transacci√≥n
TIPOS_TRANSACCION = [
    "INGRESO",
    "GASTO OPERATIVO",
    "GASTO FINANCIERO",
    "COMPRA PARA REVENTA",
    "TRANSFERENCIA",
    "PAGO TARJETA",
    "PAGO PRESTAMO",
    "AJUSTE",
]

# Categor√≠as (Simplificadas para MVP)
CATEGORIAS = [
    "Ventas - Servicios T√©cnicos",
    "Ventas - Hardware",
    "Ventas - Software",
    "Compras - Inventario",
    "Salarios",
    "Seguridad Social (CCSS)",
    "Alquiler",
    "Servicios P√∫blicos",
    "Internet/Tel√©fono",
    "Combustible",
    "Mantenimiento",
    "Papeler√≠a",
    "Intereses - Tarjetas",
    "Intereses - Pr√©stamos",
    "Intereses - Hacienda",
    "IVA Cobrado",
    "IVA Pagado",
    "Transferencia entre cuentas",
    "Otros",
]

# Estados de transacci√≥n
ESTADOS = [
    "COMPLETADA",
    "PENDIENTE",
    "CANCELADA",
]

# ============================================================================
# FUNCIONES DE UTILIDAD
# ============================================================================

def crear_estilo_header(ws, row, col_start, col_end):
    """Aplica estilo de encabezado a un rango"""
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row, col)
        cell.font = FONT_HEADER
        cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER_THIN

def crear_estilo_editable(cell):
    """Marca celda como editable (amarillo)"""
    cell.fill = PatternFill(start_color=COLOR_EDITABLE, end_color=COLOR_EDITABLE, fill_type="solid")
    cell.border = BORDER_THIN
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.font = FONT_NORMAL

def crear_estilo_formula(cell):
    """Marca celda como f√≥rmula (blanco, protegido)"""
    cell.fill = PatternFill(start_color=COLOR_FORMULA, end_color=COLOR_FORMULA, fill_type="solid")
    cell.border = BORDER_THIN
    cell.alignment = Alignment(horizontal='right', vertical='center')
    cell.font = FONT_NORMAL

def agregar_comentario(cell, texto):
    """Agrega comentario de ayuda a una celda"""
    from openpyxl.comments import Comment
    cell.comment = Comment(texto, "Sistema v3.0")
    cell.comment.width = 300
    cell.comment.height = 100

# ============================================================================
# HOJA 1: TRANSACCIONES
# ============================================================================

def crear_hoja_transacciones(wb):
    """
    Crea la hoja TRANSACCIONES con 20 columnas y validaciones.

    Columnas A-T:
    A: Fecha
    B: Tipo
    C: Categor√≠a
    D: Descripci√≥n
    E: Cuenta Origen
    F: Entidad (Cliente/Proveedor)
    G: Factura #
    H: Monto CRC
    I: Monto USD
    J: Tipo Cambio
    K: M√©todo Pago
    L: Estado
    M: IVA Incluido (%)
    N: Referencia Bancaria
    O: Notas
    P: Creado Por
    Q: Fecha Creaci√≥n
    R: Modificado Por
    S: Alerta Duplicados
    T: ID √önico
    """
    print("\nüìä Creando hoja TRANSACCIONES...")

    ws = wb.create_sheet("TRANSACCIONES", 0)

    # Encabezados
    headers = [
        "Fecha", "Tipo", "Categor√≠a", "Descripci√≥n", "Cuenta Origen",
        "Entidad", "Factura #", "Monto CRC", "Monto USD", "Tipo Cambio",
        "M√©todo Pago", "Estado", "IVA %", "Ref. Bancaria", "Notas",
        "Creado Por", "Fecha Creaci√≥n", "Modificado Por", "‚ö†Ô∏è Duplicados", "ID"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col, header)

    crear_estilo_header(ws, 1, 1, len(headers))

    # Anchos de columna
    widths = [12, 18, 22, 30, 28, 30, 12, 12, 12, 10, 15, 12, 8, 15, 25, 12, 12, 12, 15, 10]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Congelar primera fila
    ws.freeze_panes = "A2"

    # ========================================================================
    # VALIDACIONES DE DATOS
    # ========================================================================

    # Validaci√≥n: Tipo (columna B)
    dv_tipo = DataValidation(type="list", formula1=f'"{",".join(TIPOS_TRANSACCION)}"', allow_blank=False)
    dv_tipo.error = "Selecciona un tipo v√°lido de la lista"
    dv_tipo.errorTitle = "Tipo inv√°lido"
    ws.add_data_validation(dv_tipo)
    dv_tipo.add(f"B2:B1000")

    # Validaci√≥n: Categor√≠a (columna C)
    dv_categoria = DataValidation(type="list", formula1=f'"{",".join(CATEGORIAS)}"', allow_blank=False)
    dv_categoria.error = "Selecciona una categor√≠a v√°lida de la lista"
    dv_categoria.errorTitle = "Categor√≠a inv√°lida"
    ws.add_data_validation(dv_categoria)
    dv_categoria.add(f"C2:C1000")

    # Validaci√≥n: Cuenta Origen (columna E) - Todas las cuentas + tarjetas
    todas_cuentas = [c["nombre"] for c in CUENTAS_BANCARIAS] + [t["nombre"] for t in TARJETAS_CREDITO]
    # Crear validaci√≥n (Excel permite hasta ~8000 chars en validaciones modernas)
    cuentas_str = ",".join(todas_cuentas)
    dv_cuenta = DataValidation(type="list", formula1=f'"{cuentas_str}"', allow_blank=False)
    dv_cuenta.error = "Selecciona una cuenta v√°lida de la lista"
    dv_cuenta.errorTitle = "Cuenta inv√°lida"
    ws.add_data_validation(dv_cuenta)
    dv_cuenta.add(f"E2:E1000")

    # Validaci√≥n: M√©todo Pago (columna K)
    metodos = "Transferencia,Efectivo,Cheque,Tarjeta d√©bito,Tarjeta cr√©dito,SINPE M√≥vil,Dep√≥sito"
    dv_metodo = DataValidation(type="list", formula1=f'"{metodos}"', allow_blank=True)
    ws.add_data_validation(dv_metodo)
    dv_metodo.add(f"K2:K1000")

    # Validaci√≥n: Estado (columna L)
    dv_estado = DataValidation(type="list", formula1=f'"{",".join(ESTADOS)}"', allow_blank=False)
    ws.add_data_validation(dv_estado)
    dv_estado.add(f"L2:L1000")

    # Validaci√≥n: IVA % (columna M) - Solo 0, 1, 2, 4, 8, 13
    dv_iva = DataValidation(type="list", formula1='"0,1,2,4,8,13"', allow_blank=True)
    ws.add_data_validation(dv_iva)
    dv_iva.add(f"M2:M1000")

    # ========================================================================
    # F√ìRMULAS EN FILA 2 (para copiar abajo)
    # ========================================================================

    # Tipo Cambio (J2) - Auto-fetch o manual
    ws['J2'] = 540  # Valor por defecto
    agregar_comentario(ws['J2'], "üí° TIPO DE CAMBIO\n\nIngresa el tipo de cambio del d√≠a.\n\nSi dejas vac√≠o, se usar√° 540 por defecto.\n\nFormato: 540 (sin comas)")

    # Alerta Duplicados (S2)
    formula_duplicados = '''=IF(
COUNTIFS(
$A:$A, A2,
$E:$E, E2,
$I:$I, I2
) > 1,
"‚ö†Ô∏è POSIBLE DUPLICADO",
""
)'''
    ws['S2'] = formula_duplicados
    crear_estilo_formula(ws['S2'])

    # ID √önico (T2)
    ws['T2'] = '=ROW()-1'
    crear_estilo_formula(ws['T2'])

    # Fecha Creaci√≥n (Q2) - F√≥rmula NOW()
    ws['Q2'] = '=NOW()'
    crear_estilo_formula(ws['Q2'])
    ws['Q2'].number_format = 'DD/MM/YYYY HH:MM'

    # ========================================================================
    # MANUAL INLINE - Comentarios en columnas editables
    # ========================================================================

    agregar_comentario(ws['A2'], "üí° FECHA DE LA TRANSACCI√ìN\n\nFormato: DD/MM/YYYY\nEjemplo: 15/11/2025\n\n‚ö†Ô∏è Usa la fecha real de la transacci√≥n, no cuando la registras.")
    agregar_comentario(ws['B2'], "üí° TIPO DE TRANSACCI√ìN\n\nOpciones:\n‚Ä¢ INGRESO - Dinero que entra\n‚Ä¢ GASTO OPERATIVO - Gastos del negocio\n‚Ä¢ GASTO FINANCIERO - Intereses, comisiones\n‚Ä¢ COMPRA PARA REVENTA - Inventario\n‚Ä¢ TRANSFERENCIA - Movimiento entre cuentas\n‚Ä¢ PAGO TARJETA - Abono a tarjetas\n‚Ä¢ PAGO PRESTAMO - Abono a pr√©stamos\n‚Ä¢ AJUSTE - Correcciones")
    agregar_comentario(ws['C2'], "üí° CATEGOR√çA\n\nElige la categor√≠a contable.\n\nSi es venta: Especifica qu√© vendiste\nSi es gasto: Especifica en qu√© gastaste\n\n‚ö†Ô∏è Importante para reportes de P&L")
    agregar_comentario(ws['D2'], "üí° DESCRIPCI√ìN\n\nDetalla QU√â fue la transacci√≥n.\n\nEjemplo:\n‚Ä¢ Pago quincenal empleados\n‚Ä¢ Compra inventario laptops HP\n‚Ä¢ Servicio t√©cnico en sitio - Cliente XYZ\n\n‚úÖ S√© espec√≠fico, te ayudar√° despu√©s")
    agregar_comentario(ws['E2'], "üí° CUENTA ORIGEN\n\nElige de d√≥nde sali√≥/entr√≥ el dinero:\n\n‚Ä¢ Cuenta bancaria espec√≠fica\n‚Ä¢ Tarjeta de cr√©dito\n‚Ä¢ Efectivo\n\n‚ö†Ô∏è Debe coincidir exactamente con nombres en hoja EFECTIVO")
    agregar_comentario(ws['F2'], "üí° ENTIDAD (Cliente o Proveedor)\n\nSi es INGRESO: Nombre del cliente\nSi es GASTO: Nombre del proveedor\n\n‚ö†Ô∏è Usa nombres EXACTOS de la hoja ENTIDADES_ALIAS para que sistema los reconozca")
    agregar_comentario(ws['G2'], "üí° FACTURA #\n\nN√∫mero de factura electr√≥nica.\n\nFormato CR:\n‚Ä¢ Clientes: 50601012345678901234567890123456789012345\n‚Ä¢ Proveedores: 50601XXXXXXXXXXXXXXX\n\nSi no hay factura: deja vac√≠o")
    agregar_comentario(ws['H2'], "üí° MONTO EN COLONES (CRC)\n\nSi transacci√≥n fue en colones, ingr√©sala aqu√≠.\n\nFormato: 50000 (sin comas)\n\n‚ö†Ô∏è Ingresa SOLO en una moneda (CRC o USD), no ambas")
    agregar_comentario(ws['I2'], "üí° MONTO EN D√ìLARES (USD)\n\nSi transacci√≥n fue en d√≥lares, ingr√©sala aqu√≠.\n\nFormato: 100.50\n\n‚ö†Ô∏è Ingresa SOLO en una moneda (CRC o USD), no ambas")
    agregar_comentario(ws['K2'], "üí° M√âTODO DE PAGO\n\nC√≥mo se realiz√≥ el pago:\n‚Ä¢ Transferencia\n‚Ä¢ SINPE M√≥vil\n‚Ä¢ Efectivo\n‚Ä¢ Cheque\n‚Ä¢ Tarjeta d√©bito\n‚Ä¢ Tarjeta cr√©dito\n‚Ä¢ Dep√≥sito")
    agregar_comentario(ws['L2'], "üí° ESTADO\n\nCOMPLETADA - Ya se realiz√≥\nPENDIENTE - A√∫n no se ejecuta\nCANCELADA - Se anul√≥")
    agregar_comentario(ws['M2'], "üí° IVA INCLUIDO (%)\n\nSi el monto incluye IVA, indica %:\n‚Ä¢ 13% - Mayor√≠a productos/servicios\n‚Ä¢ 0% - Sin IVA\n‚Ä¢ 1%, 2%, 4% - Casos especiales\n\n‚ö†Ô∏è Esto permite calcular IVA exacto despu√©s")
    agregar_comentario(ws['N2'], "üí° REFERENCIA BANCARIA\n\nN√∫mero de referencia del banco.\n\nEjemplo:\n‚Ä¢ SINPE: 912345678\n‚Ä¢ Transferencia: REF-202511-12345\n\n√ötil para conciliaci√≥n")
    agregar_comentario(ws['O2'], "üí° NOTAS ADICIONALES\n\nCualquier informaci√≥n extra relevante:\n‚Ä¢ Recordatorios\n‚Ä¢ Aclaraciones\n‚Ä¢ Pr√≥ximas acciones\n‚Ä¢ Relaciones con otras transacciones")
    agregar_comentario(ws['P2'], "üí° CREADO POR\n\nIngresa tu nombre o iniciales.\n\nEjemplo:\n‚Ä¢ Alvaro\n‚Ä¢ AV\n‚Ä¢ Contador\n‚Ä¢ Asistente\n\n√ötil para auditor√≠a")

    # Estilos columnas editables (amarillo)
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']:
        crear_estilo_editable(ws[f'{col}2'])

    # Formatos de n√∫mero
    ws['H2'].number_format = '#,##0.00'  # CRC
    ws['I2'].number_format = '#,##0.00'  # USD
    ws['J2'].number_format = '#,##0.00'  # Tipo cambio
    ws['A2'].number_format = 'DD/MM/YYYY'

    print("   ‚úÖ Hoja TRANSACCIONES creada")
    print(f"      - {len(headers)} columnas")
    print(f"      - {len([dv_tipo, dv_categoria, dv_cuenta, dv_metodo, dv_estado, dv_iva])} validaciones")
    print(f"      - Manual inline en 16 columnas")

# ============================================================================
# HOJA 2: EFECTIVO
# ============================================================================

def crear_hoja_efectivo(wb):
    """
    Crea la hoja EFECTIVO con:
    - 9 cuentas bancarias (ACTIVOS)
    - 5 tarjetas de cr√©dito (PASIVOS)
    - C√°lculo de EFECTIVO NETO

    Calcula saldos autom√°ticamente desde TRANSACCIONES.
    """
    print("\nüè¶ Creando hoja EFECTIVO...")

    ws = wb.create_sheet("EFECTIVO")

    # T√≠tulo principal
    ws['A1'] = "CONTROL DE EFECTIVO, BANCOS Y TARJETAS DE CR√âDITO"
    ws['A1'].font = Font(name='Calibri', size=14, bold=True)
    ws.merge_cells('A1:I1')

    # ========================================================================
    # SECCI√ìN 1: CUENTAS BANCARIAS (ACTIVOS)
    # ========================================================================

    ws['A3'] = "üè¶ CUENTAS BANCARIAS (ACTIVOS)"
    ws['A3'].font = Font(size=12, bold=True)
    ws.merge_cells('A3:I3')

    # Encabezados (fila 4)
    headers = ["Cuenta", "N¬∞ Completo", "Tipo", "Moneda", "Saldo 12/Nov", "Ingresos", "Egresos", "Saldo Actual", "Uso"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(4, col, header)
    crear_estilo_header(ws, 4, 1, len(headers))

    # Anchos
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 13
    ws.column_dimensions['F'].width = 13
    ws.column_dimensions['G'].width = 13
    ws.column_dimensions['H'].width = 13
    ws.column_dimensions['I'].width = 15

    # Agregar las 9 cuentas bancarias
    row = 5

    for cuenta in CUENTAS_BANCARIAS:
        ws.cell(row, 1, cuenta["nombre"])
        ws.cell(row, 2, cuenta.get("numero", ""))
        ws.cell(row, 3, cuenta["tipo"])
        ws.cell(row, 4, cuenta["moneda"])

        # Saldo inicial en su moneda ORIGINAL (no convertir)
        saldo_inicial = cuenta["saldo"]
        ws.cell(row, 5, saldo_inicial)
        crear_estilo_editable(ws.cell(row, 5))

        # Formato seg√∫n moneda
        if cuenta["moneda"] == "CRC":
            ws.cell(row, 5).number_format = '‚Ç°#,##0.00'
        else:
            ws.cell(row, 5).number_format = '$#,##0.00'

        # F√≥rmula INGRESOS (columna F) - TODO: Conectar con TRANSACCIONES
        ws.cell(row, 6, 0)
        crear_estilo_formula(ws.cell(row, 6))
        ws.cell(row, 6).number_format = '#,##0.00'

        # F√≥rmula EGRESOS (columna G) - TODO: Conectar con TRANSACCIONES
        ws.cell(row, 7, 0)
        crear_estilo_formula(ws.cell(row, 7))
        ws.cell(row, 7).number_format = '#,##0.00'

        # F√≥rmula SALDO ACTUAL (columna H)
        formula_saldo = f'=E{row}+F{row}+G{row}'
        ws.cell(row, 8, formula_saldo)
        crear_estilo_formula(ws.cell(row, 8))
        ws.cell(row, 8).number_format = '#,##0.00'

        # Uso
        ws.cell(row, 9, cuenta.get("uso", ""))

        row += 1

    # Fila TOTAL BANCOS (en USD equivalente)
    row_total_bancos = row
    ws.cell(row_total_bancos, 1, "TOTAL BANCOS (USD equivalente)")
    ws.cell(row_total_bancos, 1).font = Font(bold=True)

    # Crear f√≥rmula que sume USD + (CRC/517.5)
    # Necesitamos SUMIF por moneda
    tipo_cambio = 517.5

    # Saldo inicial total en USD
    formula_saldo_usd = f'=SUMIF(D5:D{row_total_bancos-1},"USD",E5:E{row_total_bancos-1})+SUMIF(D5:D{row_total_bancos-1},"CRC",E5:E{row_total_bancos-1})/{tipo_cambio}'
    ws.cell(row_total_bancos, 5, formula_saldo_usd)
    ws.cell(row_total_bancos, 5).font = Font(bold=True)
    ws.cell(row_total_bancos, 5).number_format = '$#,##0.00'
    ws.cell(row_total_bancos, 5).fill = PatternFill(start_color=COLOR_SUCCESS, end_color=COLOR_SUCCESS, fill_type="solid")

    # Ingresos, Egresos, Saldo Actual (simple suma, ya deber√≠an estar en USD)
    for col in [6, 7, 8]:  # Columnas F-H
        formula = f'=SUM({get_column_letter(col)}5:{get_column_letter(col)}{row_total_bancos-1})'
        ws.cell(row_total_bancos, col, formula)
        ws.cell(row_total_bancos, col).font = Font(bold=True)
        ws.cell(row_total_bancos, col).number_format = '$#,##0.00'
        ws.cell(row_total_bancos, col).fill = PatternFill(start_color=COLOR_SUCCESS, end_color=COLOR_SUCCESS, fill_type="solid")

    # ========================================================================
    # SECCI√ìN 2: TARJETAS DE CR√âDITO (PASIVOS)
    # ========================================================================

    row += 2  # Espacio
    ws[f'A{row}'] = "üí≥ TARJETAS DE CR√âDITO (PASIVOS - DEUDAS)"
    ws[f'A{row}'].font = Font(size=12, bold=True, color='FF0000')
    ws.merge_cells(f'A{row}:I{row}')

    row += 1
    # Encabezados tarjetas
    headers_tc = ["Tarjeta", "Titular", "Saldo USD", "Saldo CRC", "Equiv. USD Total", "Pagos", "Cargos", "Saldo Actual", "Estrategia"]
    for col, header in enumerate(headers_tc, start=1):
        cell = ws.cell(row, col, header)
    crear_estilo_header(ws, row, 1, len(headers_tc))

    row += 1
    row_inicio_tc = row

    for tc in TARJETAS_CREDITO:
        # Nombre tarjeta
        ws.cell(row, 1, tc["nombre"])

        # Titular (extraer de nombre)
        titular = tc["nombre"].split("(")[1].replace(")", "") if "(" in tc["nombre"] else ""
        ws.cell(row, 2, titular)

        # Saldos
        ws.cell(row, 3, tc["saldo_usd"])
        ws.cell(row, 3).number_format = '$#,##0.00'
        crear_estilo_editable(ws.cell(row, 3))

        ws.cell(row, 4, tc["saldo_crc"])
        ws.cell(row, 4).number_format = '‚Ç°#,##0'
        crear_estilo_editable(ws.cell(row, 4))

        # Equivalente USD Total (USD + CRC convertido)
        formula_equiv = f'=C{row}+(D{row}/517.5)'
        ws.cell(row, 5, formula_equiv)
        ws.cell(row, 5).number_format = '$#,##0.00'
        crear_estilo_formula(ws.cell(row, 5))

        # Pagos (TODO: conectar con TRANSACCIONES)
        ws.cell(row, 6, 0)
        ws.cell(row, 6).number_format = '$#,##0.00'
        crear_estilo_formula(ws.cell(row, 6))

        # Cargos (TODO: conectar con TRANSACCIONES)
        ws.cell(row, 7, 0)
        ws.cell(row, 7).number_format = '$#,##0.00'
        crear_estilo_formula(ws.cell(row, 7))

        # Saldo Actual
        formula_saldo_tc = f'=E{row}+F{row}+G{row}'
        ws.cell(row, 8, formula_saldo_tc)
        ws.cell(row, 8).number_format = '$#,##0.00'
        crear_estilo_formula(ws.cell(row, 8))
        ws.cell(row, 8).fill = PatternFill(start_color=COLOR_ERROR, end_color=COLOR_ERROR, fill_type="solid")

        # Estrategia
        ws.cell(row, 9, tc["estrategia"])
        ws.cell(row, 9).font = FONT_SMALL

        row += 1

    # Fila TOTAL TARJETAS
    row_total_tc = row
    ws.cell(row_total_tc, 1, "TOTAL TARJETAS (PASIVOS)")
    ws.cell(row_total_tc, 1).font = Font(bold=True, color='FF0000')

    for col in [3, 4, 5, 6, 7, 8]:  # Columnas C-H
        formula = f'=SUM({get_column_letter(col)}{row_inicio_tc}:{get_column_letter(col)}{row_total_tc-1})'
        ws.cell(row_total_tc, col, formula)
        ws.cell(row_total_tc, col).font = Font(bold=True)

        if col in [3, 5, 6, 7, 8]:
            ws.cell(row_total_tc, col).number_format = '$#,##0.00'
        else:
            ws.cell(row_total_tc, col).number_format = '‚Ç°#,##0'

        ws.cell(row_total_tc, col).fill = PatternFill(start_color=COLOR_ERROR, end_color=COLOR_ERROR, fill_type="solid")

    # ========================================================================
    # SECCI√ìN 3: EFECTIVO NETO
    # ========================================================================

    row += 2
    ws[f'A{row}'] = "üí∞ EFECTIVO NETO (Bancos - Tarjetas)"
    ws[f'A{row}'].font = Font(size=14, bold=True)
    ws.merge_cells(f'A{row}:E{row}')

    # F√≥rmula efectivo neto
    formula_neto = f'=H{row_total_bancos}-H{row_total_tc}'
    ws[f'F{row}'] = formula_neto
    ws[f'F{row}'].font = Font(size=16, bold=True)
    ws[f'F{row}'].number_format = '$#,##0.00'

    # Color condicional (rojo si negativo, verde si positivo)
    ws[f'F{row}'].fill = PatternFill(start_color=COLOR_WARNING, end_color=COLOR_WARNING, fill_type="solid")

    # Notas explicativas
    row += 2
    ws[f'A{row}'] = "üí° NOTAS:"
    ws[f'A{row}'].font = Font(bold=True)

    row += 1
    ws[f'A{row}'] = "‚Ä¢ Saldos al 12 de Noviembre 2025"
    ws[f'A{row}'].font = FONT_SMALL

    row += 1
    ws[f'A{row}'] = "‚Ä¢ Columnas 'Ingresos', 'Egresos', 'Pagos' y 'Cargos' se conectar√°n autom√°ticamente a TRANSACCIONES"
    ws[f'A{row}'].font = FONT_SMALL

    row += 1
    ws[f'A{row}'] = "‚Ä¢ Tipo de cambio usado: ‚Ç°517.5 por $1 USD"
    ws[f'A{row}'].font = FONT_SMALL

    row += 1
    ws[f'A{row}'] = f"‚Ä¢ EFECTIVO NETO REAL (12/Nov): $3,444.54 en bancos - $16,536 en tarjetas = -$13,091.46 (CRISIS)"
    ws[f'A{row}'].font = Font(size=10, bold=True, color='FF0000')

    print("   ‚úÖ Hoja EFECTIVO creada")
    print(f"      - {len(CUENTAS_BANCARIAS)} cuentas bancarias")
    print(f"      - {len(TARJETAS_CREDITO)} tarjetas de cr√©dito")
    print(f"      - Efectivo neto = Bancos - Tarjetas")

# ============================================================================
# HOJA 3: DASHBOARD
# ============================================================================

def crear_hoja_dashboard(wb):
    """
    Crea el DASHBOARD con KPIs cr√≠ticos.

    M√©tricas prioritarias:
    - Efectivo Total
    - D√≠as de Cobertura
    - Top 5 Clientes
    - Alertas Cr√≠ticas
    """
    print("\nüìä Creando hoja DASHBOARD...")

    ws = wb.create_sheet("DASHBOARD")

    # T√≠tulo principal
    ws['B2'] = "DASHBOARD FINANCIERO"
    ws['B2'].font = Font(name='Calibri', size=16, bold=True)
    ws.merge_cells('B2:F2')
    ws['B2'].alignment = Alignment(horizontal='center')

    # Fecha de actualizaci√≥n
    ws['B3'] = "√öltima actualizaci√≥n:"
    ws['C3'] = "=NOW()"
    ws['C3'].number_format = 'DD/MM/YYYY HH:MM'
    ws['C3'].font = FONT_SMALL

    # ========================================================================
    # SECCI√ìN 1: EFECTIVO
    # ========================================================================

    ws['B5'] = "üí∞ EFECTIVO DISPONIBLE"
    ws['B5'].font = Font(size=12, bold=True)

    ws['B6'] = "Total Efectivo (USD):"
    ws['C6'] = "=EFECTIVO!G13"  # Asumiendo que fila 13 es el total
    ws['C6'].font = Font(size=14, bold=True)
    ws['C6'].number_format = '$#,##0.00'
    ws['C6'].fill = PatternFill(start_color=COLOR_SUCCESS, end_color=COLOR_SUCCESS, fill_type="solid")

    # ========================================================================
    # SECCI√ìN 2: D√çAS DE COBERTURA
    # ========================================================================

    ws['B8'] = "‚è±Ô∏è D√çAS DE COBERTURA"
    ws['B8'].font = Font(size=12, bold=True)

    ws['B9'] = "Gasto Diario Promedio:"
    ws['C9'] = "=SUMIFS(TRANSACCIONES!I:I, TRANSACCIONES!B:B, \"GASTO OPERATIVO\", TRANSACCIONES!L:L, \"COMPLETADA\") / 30"
    ws['C9'].number_format = '$#,##0.00'

    ws['B10'] = "D√≠as de Cobertura:"
    ws['C10'] = "=IF(C9>0, C6/C9, 0)"
    ws['C10'].font = Font(size=14, bold=True)
    ws['C10'].number_format = '0.0'

    # Alerta condicional
    ws['D10'] = '=IF(C10<15, "üö® CR√çTICO", IF(C10<30, "‚ö†Ô∏è PRECAUCI√ìN", "‚úÖ SALUDABLE"))'
    ws['D10'].font = Font(size=11, bold=True)

    # ========================================================================
    # SECCI√ìN 3: TOP 5 CLIENTES (NOVIEMBRE)
    # ========================================================================

    ws['B12'] = "üèÜ TOP 5 CLIENTES (Noviembre)"
    ws['B12'].font = Font(size=12, bold=True)

    # Encabezados
    ws['B13'] = "Cliente"
    ws['C13'] = "Facturado USD"
    crear_estilo_header(ws, 13, 2, 3)

    # Aqu√≠ deber√≠amos usar f√≥rmulas din√°micas, pero para MVP ponemos los datos del cuestionario
    row = 14
    for i, cliente in enumerate(sorted(CLIENTES, key=lambda x: x["monto_nov"], reverse=True)[:5], 1):
        ws.cell(row, 2, cliente["nombre"])
        ws.cell(row, 3, cliente["monto_nov"])
        ws.cell(row, 3).number_format = '$#,##0.00'
        row += 1

    # ========================================================================
    # SECCI√ìN 4: ALERTAS CR√çTICAS
    # ========================================================================

    ws['E5'] = "üö® ALERTAS CR√çTICAS"
    ws['E5'].font = Font(size=12, bold=True, color='FF0000')

    ws['E6'] = "Duplicados:"
    ws['F6'] = '=COUNTIF(TRANSACCIONES!S:S, "‚ö†Ô∏è POSIBLE DUPLICADO")'
    ws['F6'].font = Font(bold=True)
    ws['F6'].fill = PatternFill(start_color=COLOR_ERROR, end_color=COLOR_ERROR, fill_type="solid")

    ws['E7'] = "Pendientes:"
    ws['F7'] = '=COUNTIF(TRANSACCIONES!L:L, "PENDIENTE")'
    ws['F7'].font = Font(bold=True)
    ws['F7'].fill = PatternFill(start_color=COLOR_WARNING, end_color=COLOR_WARNING, fill_type="solid")

    ws['E8'] = "Sin categor√≠a:"
    ws['F8'] = '=COUNTBLANK(TRANSACCIONES!C:C) - 1'
    ws['F8'].font = Font(bold=True)

    # Anchos
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15

    print("   ‚úÖ Hoja DASHBOARD creada")
    print("      - Efectivo total")
    print("      - D√≠as de cobertura con alertas")
    print("      - Top 5 clientes")
    print("      - Alertas de duplicados/pendientes")

# ============================================================================
# HOJA 4: ENTIDADES_ALIAS
# ============================================================================

def crear_hoja_entidades_alias(wb):
    """
    Crea la hoja ENTIDADES_ALIAS con 22 clientes pre-cargados.

    Sistema expandido de normalizaci√≥n para:
    - Clientes
    - Proveedores
    - Bancos
    - Categor√≠as
    """
    print("\nüë• Creando hoja ENTIDADES_ALIAS...")

    ws = wb.create_sheet("ENTIDADES_ALIAS")

    # T√≠tulo
    ws['A1'] = "SISTEMA DE NORMALIZACI√ìN DE ENTIDADES"
    ws['A1'].font = Font(name='Calibri', size=14, bold=True)
    ws.merge_cells('A1:J1')

    # Descripci√≥n
    ws['A2'] = "üí° Esta hoja permite normalizar nombres de clientes, proveedores y bancos que aparecen con variaciones"
    ws['A2'].font = FONT_SMALL
    ws.merge_cells('A2:J2')

    # Encabezados (fila 4)
    headers = ["Tipo", "Nombre Est√°ndar", "Alias 1", "Alias 2", "Alias 3", "Alias 4", "Alias 5", "Categor√≠a", "Notas", "√öltima Actualizaci√≥n"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(4, col, header)
    crear_estilo_header(ws, 4, 1, len(headers))

    # Anchos
    widths = [12, 35, 25, 25, 25, 25, 25, 15, 30, 15]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # ========================================================================
    # CARGAR 22 CLIENTES
    # ========================================================================

    row = 5
    for cliente in sorted(CLIENTES, key=lambda x: x["monto_nov"], reverse=True):
        ws.cell(row, 1, "Cliente")
        ws.cell(row, 2, cliente["nombre"])

        # Generar alias comunes
        nombre = cliente["nombre"]

        # Alias 1: Sin S.A./Ltda
        alias1 = nombre.replace(" S.A.", "").replace(" Ltda", "").replace(" S.A", "").strip()
        ws.cell(row, 3, alias1 if alias1 != nombre else "")

        # Alias 2: Siglas
        palabras = nombre.split()
        if len(palabras) > 2:
            alias2 = "".join([p[0].upper() for p in palabras if len(p) > 3])
            ws.cell(row, 4, alias2 if len(alias2) > 1 else "")

        ws.cell(row, 8, cliente["categoria"])
        ws.cell(row, 9, f"Facturaci√≥n Nov: ${cliente['monto_nov']:.2f}")
        ws.cell(row, 10, datetime.now().strftime("%d/%m/%Y"))

        # Estilos
        for col in range(1, 11):
            if col in [3, 4, 5, 6, 7]:  # Alias editables
                crear_estilo_editable(ws.cell(row, col))

        row += 1

    # ========================================================================
    # CARGAR 5 PROVEEDORES
    # ========================================================================

    for proveedor in PROVEEDORES:
        ws.cell(row, 1, "Proveedor")
        ws.cell(row, 2, proveedor)
        ws.cell(row, 8, "Principal")
        ws.cell(row, 10, datetime.now().strftime("%d/%m/%Y"))

        for col in range(1, 11):
            if col in [3, 4, 5, 6, 7]:
                crear_estilo_editable(ws.cell(row, col))

        row += 1

    # ========================================================================
    # CARGAR 9 BANCOS
    # ========================================================================

    for cuenta in CUENTAS_BANCARIAS:
        ws.cell(row, 1, "Banco")
        ws.cell(row, 2, cuenta["nombre"])
        ws.cell(row, 8, cuenta["tipo"])
        ws.cell(row, 9, f'Moneda: {cuenta["moneda"]}')
        ws.cell(row, 10, datetime.now().strftime("%d/%m/%Y"))

        for col in range(1, 11):
            if col in [3, 4, 5, 6, 7]:
                crear_estilo_editable(ws.cell(row, col))

        row += 1

    # Agregar instrucciones
    ws[f'A{row + 2}'] = "üìã INSTRUCCIONES:"
    ws[f'A{row + 2}'].font = Font(bold=True)

    ws[f'A{row + 3}'] = "1. Cuando aparezca una variaci√≥n de nombre, agr√©gala como 'Alias' en la fila correspondiente"
    ws[f'A{row + 4}'] = "2. Ejecuta: python scripts/normalizar_entidades_universal_v3.py"
    ws[f'A{row + 5}'] = "3. El script unificar√° todos los nombres autom√°ticamente"

    for i in range(3, 6):
        ws[f'A{row + i}'].font = FONT_SMALL
        ws.merge_cells(f'A{row + i}:J{row + i}')

    # Congelar paneles
    ws.freeze_panes = "B5"

    print("   ‚úÖ Hoja ENTIDADES_ALIAS creada")
    print(f"      - {len(CLIENTES)} clientes pre-cargados")
    print(f"      - {len(PROVEEDORES)} proveedores")
    print(f"      - {len(CUENTAS_BANCARIAS)} cuentas bancarias")
    print(f"      - Total: {len(CLIENTES) + len(PROVEEDORES) + len(CUENTAS_BANCARIAS)} entidades")

# ============================================================================
# HOJA 5: CONFIGURACI√ìN
# ============================================================================

def crear_hoja_configuracion(wb):
    """Crea hoja de configuraci√≥n del sistema con TC bi-moneda"""
    print("\n‚öôÔ∏è Creando hoja CONFIGURACI√ìN...")

    ws = wb.create_sheet("CONFIG")

    # ========================================================================
    # T√çTULO PRINCIPAL
    # ========================================================================
    ws['A1'] = "CONFIGURACI√ìN DEL SISTEMA"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:D1')

    # ========================================================================
    # SECCI√ìN 1: INFORMACI√ìN GENERAL
    # ========================================================================
    ws['A3'] = "üìã INFORMACI√ìN GENERAL"
    ws['A3'].font = Font(size=12, bold=True)
    ws.merge_cells('A3:D3')

    configs_generales = [
        ("Versi√≥n", "3.0.0 MVP", False),
        ("Fecha Creaci√≥n", datetime.now().strftime("%d/%m/%Y %H:%M"), False),
        ("Propietario", "Alvaro Velasco - AVN (AlvaroVelascoNet)", False),
        ("Per√≠odo Fiscal", "2025", False),
        ("Mes Activo", "Noviembre 2025", True),  # Editable
    ]

    row = 4
    for key, value, editable in configs_generales:
        ws.cell(row, 1, key)
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2, value)

        if editable:
            crear_estilo_editable(ws.cell(row, 2))
            agregar_comentario(ws.cell(row, 2), "üí° MES ACTIVO\n\nCambia el mes cuando inicies un nuevo per√≠odo.\n\nFormato: Enero 2025, Febrero 2025, etc.")

        row += 1

    # ========================================================================
    # SECCI√ìN 2: TIPO DE CAMBIO (EDITABLE)
    # ========================================================================
    row += 1
    ws[f'A{row}'] = "üí± TIPO DE CAMBIO BI-MONEDA (CRC ‚Üî USD)"
    ws[f'A{row}'].font = Font(size=12, bold=True, color='0000FF')
    ws.merge_cells(f'A{row}:D{row}')

    row += 1
    # Encabezados TC
    ws.cell(row, 1, "Tipo")
    ws.cell(row, 2, "Valor")
    ws.cell(row, 3, "√öltima Actualizaci√≥n")
    ws.cell(row, 4, "Notas")
    crear_estilo_header(ws, row, 1, 4)

    row += 1
    row_tc_compra = row
    # TC COMPRA
    ws.cell(row, 1, "TC Compra")
    ws.cell(row, 1).font = Font(bold=True)
    ws.cell(row, 2, 517.50)
    crear_estilo_editable(ws.cell(row, 2))
    ws.cell(row, 2).number_format = '‚Ç°#,##0.00'

    ws.cell(row, 3, datetime.now().strftime("%d/%m/%Y %H:%M"))
    crear_estilo_editable(ws.cell(row, 3))

    ws.cell(row, 4, "Banco compra d√≥lares (t√∫ vendes USD)")
    agregar_comentario(ws.cell(row, 2), "üí° TIPO DE CAMBIO COMPRA\n\nCu√°ntos colones TE DAN por $1 USD\n\nEjemplo: Si banco te da ‚Ç°517.50 por $1\n\n‚ö†Ô∏è Actualizar 1 vez/semana")

    row += 1
    row_tc_venta = row
    # TC VENTA
    ws.cell(row, 1, "TC Venta")
    ws.cell(row, 1).font = Font(bold=True)
    ws.cell(row, 2, 525.00)
    crear_estilo_editable(ws.cell(row, 2))
    ws.cell(row, 2).number_format = '‚Ç°#,##0.00'

    ws.cell(row, 3, datetime.now().strftime("%d/%m/%Y %H:%M"))
    crear_estilo_editable(ws.cell(row, 3))

    ws.cell(row, 4, "Banco vende d√≥lares (t√∫ compras USD)")
    agregar_comentario(ws.cell(row, 2), "üí° TIPO DE CAMBIO VENTA\n\nCu√°ntos colones PAGAS por $1 USD\n\nEjemplo: Si banco cobra ‚Ç°525.00 por $1\n\n‚ö†Ô∏è Actualizar 1 vez/semana")

    # ========================================================================
    # SECCI√ìN 3: HISTORIAL DE TIPOS DE CAMBIO
    # ========================================================================
    row += 2
    ws[f'A{row}'] = "üìä HISTORIAL DE TIPOS DE CAMBIO (Auditor√≠a)"
    ws[f'A{row}'].font = Font(size=11, bold=True)
    ws.merge_cells(f'A{row}:F{row}')

    row += 1
    # Encabezados historial
    headers_hist = ["Fecha", "TC Compra", "TC Venta", "Promedio", "Variaci√≥n %", "Registrado Por"]
    for col, header in enumerate(headers_hist, start=1):
        ws.cell(row, col, header)
    crear_estilo_header(ws, row, 1, len(headers_hist))

    row += 1
    row_hist_inicio = row
    # Primera entrada del historial (actual)
    ws.cell(row, 1, datetime.now().strftime("%d/%m/%Y"))
    ws.cell(row, 2, 517.50)
    ws.cell(row, 2).number_format = '‚Ç°#,##0.00'
    ws.cell(row, 3, 525.00)
    ws.cell(row, 3).number_format = '‚Ç°#,##0.00'

    # Promedio
    ws.cell(row, 4, f'=(B{row}+C{row})/2')
    ws.cell(row, 4).number_format = '‚Ç°#,##0.00'
    crear_estilo_formula(ws.cell(row, 4))

    # Variaci√≥n % (ser√° 0 para la primera)
    ws.cell(row, 5, "Inicial")
    ws.cell(row, 6, "Sistema")

    # Dejar filas vac√≠as para futuras actualizaciones (formato editable)
    for i in range(1, 6):  # 5 filas adicionales
        row += 1
        for col in [1, 2, 3, 6]:  # Fecha, TC Compra, TC Venta, Usuario
            crear_estilo_editable(ws.cell(row, col))

        # F√≥rmulas para Promedio y Variaci√≥n
        ws.cell(row, 4, f'=IF(B{row}<>"", (B{row}+C{row})/2, "")')
        ws.cell(row, 4).number_format = '‚Ç°#,##0.00'
        crear_estilo_formula(ws.cell(row, 4))

        ws.cell(row, 5, f'=IF(D{row}<>"", IF(D{row-1}<>"", (D{row}-D{row-1})/D{row-1}*100, 0), "")')
        ws.cell(row, 5).number_format = '0.00"%"'
        crear_estilo_formula(ws.cell(row, 5))

    # ========================================================================
    # INSTRUCCIONES
    # ========================================================================
    row += 2
    ws[f'A{row}'] = "üìù INSTRUCCIONES PARA ACTUALIZAR TC:"
    ws[f'A{row}'].font = Font(bold=True, size=10)
    ws.merge_cells(f'A{row}:F{row}')

    row += 1
    instrucciones = [
        "1. Actualiza TC Compra y TC Venta cada semana (o cuando hagas cambio de moneda)",
        "2. Actualiza la fecha en '√öltima Actualizaci√≥n'",
        "3. Registra el cambio en el HISTORIAL para auditor√≠a",
        "4. Usa TC Compra cuando VENDES d√≥lares (USD ‚Üí CRC)",
        "5. Usa TC Venta cuando COMPRAS d√≥lares (CRC ‚Üí USD)",
    ]

    for instruccion in instrucciones:
        ws[f'A{row}'] = instruccion
        ws[f'A{row}'].font = FONT_SMALL
        ws.merge_cells(f'A{row}:F{row}')
        row += 1

    # Anchos de columna
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15

    print("   ‚úÖ Hoja CONFIG creada")
    print(f"      - TC Compra: ‚Ç°517.50 (editable)")
    print(f"      - TC Venta: ‚Ç°525.00 (editable)")
    print(f"      - Historial de TCs para auditor√≠a")

# ============================================================================
# FUNCI√ìN PRINCIPAL
# ============================================================================

def main():
    """Funci√≥n principal de creaci√≥n del Excel v3.0"""

    print("""
‚ïî‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïó
‚ïë                                                               ‚ïë
‚ïë           CREAR EXCEL V3.0 - FASE 1 MVP                     ‚ïë
‚ïë           Alvaro Velasco - AVN (AlvaroVelascoNet)           ‚ïë
‚ïë                                                               ‚ïë
‚ïë           Due: Nov 19, 2025 (7 d√≠as)                        ‚ïë
‚ïë                                                               ‚ïë
‚ïö‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïù
    """)

    # Verificar si archivo ya existe
    if os.path.exists(EXCEL_FILE):
        backup = f"{EXCEL_FILE}.backup.{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        os.rename(EXCEL_FILE, backup)
        print(f"‚ö†Ô∏è Archivo existente respaldado como: {backup}")

    # Crear nuevo workbook
    print("\nüì¶ Creando nuevo archivo Excel...")
    wb = openpyxl.Workbook()

    # Eliminar hoja por defecto
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Crear hojas principales
    crear_hoja_transacciones(wb)
    crear_hoja_efectivo(wb)
    crear_hoja_dashboard(wb)
    crear_hoja_entidades_alias(wb)
    crear_hoja_configuracion(wb)

    # Guardar archivo
    print(f"\nüíæ Guardando archivo: {EXCEL_FILE}")
    wb.save(EXCEL_FILE)

    # Resumen final
    print("\n" + "="*60)
    print("‚úÖ EXCEL V3.0 MVP CREADO EXITOSAMENTE")
    print("="*60)
    print(f"\nüìÑ Archivo: {EXCEL_FILE}")
    print(f"üìä Hojas creadas: {len(wb.sheetnames)}")
    print(f"\nHojas incluidas:")
    for i, sheet in enumerate(wb.sheetnames, 1):
        print(f"   {i}. {sheet}")

    print("\n‚úÖ FASE 1 COMPLETADA:")
    print("   ‚úÖ TRANSACCIONES - 20 columnas con validaciones")
    print("   ‚úÖ EFECTIVO - 9 cuentas bancarias con f√≥rmulas")
    print("   ‚úÖ DASHBOARD - KPIs cr√≠ticos")
    print("   ‚úÖ ENTIDADES_ALIAS - 22 clientes + 5 proveedores + 9 bancos")
    print("   ‚úÖ Manual inline - Comentarios en todas las celdas editables")

    print("\nüìã PR√ìXIMOS PASOS:")
    print("   1. Abre el archivo en Excel/OneDrive")
    print("   2. Verifica validaciones y f√≥rmulas")
    print("   3. Ingresa saldos iniciales en hoja EFECTIVO")
    print("   4. Comienza a registrar transacciones de Noviembre")
    print("   5. Ejecuta: python scripts/importar_datos_noviembre_v2.py (pr√≥ximo)")

    print("\n" + "="*60)
    print(f"üéØ Tiempo de desarrollo: {datetime.now()}")
    print("="*60 + "\n")

if __name__ == "__main__":
    main()
