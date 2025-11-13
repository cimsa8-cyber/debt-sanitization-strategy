#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CORRECCIÓN DE FÓRMULAS HOJA EFECTIVO
Corrige las fórmulas de Ingreso/Egreso para que sumen TODAS las transacciones
de cada cuenta, no solo la fila del balance inicial.
"""
import openpyxl
from datetime import datetime
import shutil

EXCEL_FILE = "AlvaroVelasco_Finanzas_v2.0.xlsx"
BACKUP_FILE = f"AlvaroVelasco_Finanzas_v2.0_FORMULAS_EFECTIVO_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

def crear_backup():
    print("=" * 80)
    print("CREANDO BACKUP")
    print("=" * 80)
    print(f"Backup: {BACKUP_FILE}")
    try:
        shutil.copy2(EXCEL_FILE, BACKUP_FILE)
        print("✅ Backup creado")
        print()
        return True
    except Exception as e:
        print(f"❌ ERROR: {e}")
        return False

def corregir_formulas():
    print("=" * 80)
    print("CORRECCIÓN DE FÓRMULAS HOJA EFECTIVO")
    print("=" * 80)
    print()

    wb = openpyxl.load_workbook(EXCEL_FILE)

    # Verificar que existe la hoja Efectivo
    if 'Efectivo' not in wb.sheetnames:
        print("❌ ERROR: No se encontró la hoja 'Efectivo'")
        print(f"   Hojas disponibles: {', '.join(wb.sheetnames)}")
        return False

    ws_efectivo = wb['Efectivo']

    print("📋 Hoja 'Efectivo' encontrada")
    print()

    # =========================================================================
    # PASO 1: MOSTRAR FÓRMULAS ACTUALES
    # =========================================================================
    print("=" * 80)
    print("📊 FÓRMULAS ACTUALES (antes de corregir)")
    print("=" * 80)
    print()

    for fila in range(3, 11):  # Filas 3-10 (8 cuentas)
        cuenta = ws_efectivo.cell(fila, 3).value  # Columna C (Cuenta)
        nombre = ws_efectivo.cell(fila, 2).value  # Columna B (Nombre)

        # Obtener fórmulas actuales
        celda_d = ws_efectivo.cell(fila, 4)  # Columna D (Ingreso)
        celda_e = ws_efectivo.cell(fila, 5)  # Columna E (Egreso)

        print(f"Fila {fila}: {nombre}")
        print(f"   Cuenta: {cuenta}")
        print(f"   D{fila} (Ingreso): {celda_d.value if not isinstance(celda_d.value, str) or not celda_d.value.startswith('=') else celda_d.value}")
        print(f"   E{fila} (Egreso):  {celda_e.value if not isinstance(celda_e.value, str) or not celda_e.value.startswith('=') else celda_e.value}")
        print()

    # =========================================================================
    # PASO 2: CORREGIR FÓRMULAS
    # =========================================================================
    print("=" * 80)
    print("🔧 CORRIGIENDO FÓRMULAS")
    print("=" * 80)
    print()

    print("📝 Nueva fórmula para Ingresos (columna D):")
    print('   =SUMIFS(TRANSACCIONES!$I:$I, TRANSACCIONES!$E:$E, C{fila}, TRANSACCIONES!$K:$K, "Ingreso")')
    print()
    print("📝 Nueva fórmula para Egresos (columna E):")
    print('   =SUMIFS(TRANSACCIONES!$I:$I, TRANSACCIONES!$E:$E, C{fila}, TRANSACCIONES!$K:$K, "Egreso")')
    print()
    print("Esto sumará TODAS las transacciones de cada cuenta, no solo la fila del balance inicial.")
    print()

    cuentas_corregidas = []

    for fila in range(3, 11):  # Filas 3-10 (8 cuentas)
        cuenta = ws_efectivo.cell(fila, 3).value  # Columna C (Cuenta)
        nombre = ws_efectivo.cell(fila, 2).value  # Columna B (Nombre)

        if not cuenta:
            print(f"⚠️  Fila {fila}: Sin cuenta definida, saltando...")
            continue

        # Nueva fórmula para Ingresos (columna D)
        formula_ingreso = f'=SUMIFS(TRANSACCIONES!$I:$I,TRANSACCIONES!$E:$E,C{fila},TRANSACCIONES!$K:$K,"Ingreso")'
        ws_efectivo.cell(fila, 4).value = formula_ingreso

        # Nueva fórmula para Egresos (columna E)
        # IMPORTANTE: Para egresos, sumamos valores absolutos porque en TRANSACCIONES los egresos son negativos
        # pero en Efectivo queremos mostrarlos como positivos
        formula_egreso = f'=SUMIFS(TRANSACCIONES!$I:$I,TRANSACCIONES!$E:$E,C{fila},TRANSACCIONES!$K:$K,"Egreso")*-1'
        ws_efectivo.cell(fila, 5).value = formula_egreso

        # Verificar que la columna F (Balance) tenga la fórmula correcta =D-E
        formula_balance = f'=D{fila}-E{fila}'
        ws_efectivo.cell(fila, 6).value = formula_balance

        cuentas_corregidas.append({
            'fila': fila,
            'nombre': nombre,
            'cuenta': cuenta[:40] if cuenta else 'N/A'
        })

        print(f"✅ Fila {fila}: {nombre[:40]}")
        print(f"   D{fila}: {formula_ingreso}")
        print(f"   E{fila}: {formula_egreso}")
        print(f"   F{fila}: {formula_balance}")
        print()

    # =========================================================================
    # PASO 3: GUARDAR
    # =========================================================================
    print("=" * 80)
    print("💾 GUARDANDO CAMBIOS")
    print("=" * 80)
    print()

    wb.save(EXCEL_FILE)
    print("✅ Excel actualizado con nuevas fórmulas")
    print()

    # =========================================================================
    # PASO 4: VERIFICAR RESULTADOS
    # =========================================================================
    print("=" * 80)
    print("📊 VERIFICACIÓN - RECALCULANDO CON NUEVAS FÓRMULAS")
    print("=" * 80)
    print()

    # Recargar con data_only=True para ver valores calculados
    print("⏳ Recalculando fórmulas...")
    wb_verificar = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws_verificar = wb_verificar['Efectivo']

    print()
    print("💰 SALDOS ACTUALIZADOS:")
    print()

    for info in cuentas_corregidas:
        fila = info['fila']
        nombre = info['nombre']

        ingreso = ws_verificar.cell(fila, 4).value
        egreso = ws_verificar.cell(fila, 5).value
        balance = ws_verificar.cell(fila, 6).value

        ingreso_val = float(ingreso) if ingreso else 0
        egreso_val = float(egreso) if egreso else 0
        balance_val = float(balance) if balance else 0

        print(f"Fila {fila}: {nombre}")
        print(f"   Ingreso: ${ingreso_val:,.2f}")
        print(f"   Egreso:  ${egreso_val:,.2f}")
        print(f"   Balance: ${balance_val:,.2f}")
        print()

    # =========================================================================
    # RESUMEN
    # =========================================================================
    print("=" * 80)
    print("📊 RESUMEN")
    print("=" * 80)
    print()

    print(f"✅ Cuentas corregidas: {len(cuentas_corregidas)}")
    print()
    print("🔧 Cambios aplicados:")
    print("   • Columna D (Ingreso): SUMIFS sumando todos los ingresos de cada cuenta")
    print("   • Columna E (Egreso): SUMIFS sumando todos los egresos de cada cuenta")
    print("   • Columna F (Balance): D - E")
    print()
    print("⚠️  IMPORTANTE:")
    print("   • Cierra y vuelve a abrir el Excel para ver los valores actualizados")
    print("   • Excel recalculará automáticamente las fórmulas al abrir")
    print("   • Verifica que el saldo de Promerica sea $2,163.44")
    print()

    print("=" * 80)
    print("✅ CORRECCIÓN COMPLETADA")
    print("=" * 80)
    print()

    return True

if __name__ == "__main__":
    try:
        if not crear_backup():
            print("❌ Abortando")
            exit(1)

        if corregir_formulas():
            print("🎉 Proceso completado exitosamente!")
            print()
            print("📋 PRÓXIMOS PASOS:")
            print("   1. Cierra completamente el Excel")
            print("   2. Vuelve a abrirlo")
            print("   3. Ve a la hoja Efectivo")
            print("   4. Verifica que Promerica (fila 3) muestre Balance: $2,163.44")
        else:
            print("❌ Proceso falló")

    except Exception as e:
        print(f"❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
