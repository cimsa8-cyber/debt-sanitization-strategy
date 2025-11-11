#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
NORMALIZACIÓN NOMBRES CUENTA PROMERICA
Unifica todas las variaciones de Promerica USD a un solo nombre estándar
"""
import openpyxl
from datetime import datetime
import shutil

EXCEL_FILE = "AlvaroVelasco_Finanzas_v2.0.xlsx"
BACKUP_FILE = f"AlvaroVelasco_Finanzas_v2.0_NORMALIZAR_PROMERICA_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

# Nombre estándar (según extracto bancario)
NOMBRE_ESTANDAR = "Promerica USD (40000003881774)"

# Variaciones a normalizar
VARIACIONES_A_NORMALIZAR = [
    "Promerica USD",
    "Promerica USD 1774",
    # NO incluimos "Promerica CRC" porque es otra cuenta diferente
]

def crear_backup():
    print("=" * 80)
    print("CREANDO BACKUP")
    print("=" * 80)
    print(f"Backup: {BACKUP_FILE}")
    try:
        shutil.copy2(EXCEL_FILE, BACKUP_FILE)
        print("✅ Backup creado")
        print()
        return True
    except Exception as e:
        print(f"❌ ERROR: {e}")
        return False

def normalizar():
    print("=" * 80)
    print("NORMALIZACIÓN NOMBRES CUENTA PROMERICA USD")
    print("=" * 80)
    print()

    print(f"🎯 Nombre estándar: \"{NOMBRE_ESTANDAR}\"")
    print()
    print("📋 Variaciones que se cambiarán:")
    for var in VARIACIONES_A_NORMALIZAR:
        print(f"   • \"{var}\" → \"{NOMBRE_ESTANDAR}\"")
    print()

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb['TRANSACCIONES']

    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    col_map = {}
    for col in range(1, len(headers) + 1):
        if headers[col-1]:
            col_map[headers[col-1]] = col

    # =========================================================================
    # PASO 1: IDENTIFICAR TRANSACCIONES A CAMBIAR
    # =========================================================================
    print("=" * 80)
    print("🔍 PASO 1: Identificando transacciones a cambiar...")
    print("=" * 80)
    print()

    transacciones_cambiar = []

    for row in range(2, ws.max_row + 1):
        cuenta = ws.cell(row, col_map['Cuenta Bancaria']).value

        if cuenta:
            cuenta_str = str(cuenta).strip()

            if cuenta_str in VARIACIONES_A_NORMALIZAR:
                fecha = ws.cell(row, col_map['Fecha']).value
                concepto = ws.cell(row, col_map['Concepto']).value
                monto = ws.cell(row, col_map['Monto USD']).value

                transacciones_cambiar.append({
                    'fila': row,
                    'nombre_viejo': cuenta_str,
                    'fecha': fecha,
                    'concepto': concepto[:40] if concepto else 'Sin concepto',
                    'monto': float(monto) if monto else 0,
                })

    print(f"📊 Transacciones a cambiar: {len(transacciones_cambiar)}")
    print()

    # Agrupar por variación
    por_variacion = {}
    for trans in transacciones_cambiar:
        nombre = trans['nombre_viejo']
        if nombre not in por_variacion:
            por_variacion[nombre] = []
        por_variacion[nombre].append(trans)

    for nombre, trans_list in sorted(por_variacion.items()):
        print(f"\"{nombre}\": {len(trans_list)} transacciones")

    print()

    # =========================================================================
    # PASO 2: APLICAR CAMBIOS
    # =========================================================================
    print("=" * 80)
    print("✏️  PASO 2: Aplicando cambios...")
    print("=" * 80)
    print()

    contador = 0

    for trans in transacciones_cambiar:
        fila = trans['fila']
        ws.cell(fila, col_map['Cuenta Bancaria']).value = NOMBRE_ESTANDAR
        contador += 1

        if contador <= 10 or contador % 10 == 0:
            fecha_str = trans['fecha'].strftime('%d/%m/%Y') if isinstance(trans['fecha'], datetime) else 'Sin fecha'
            signo = '+' if trans['monto'] > 0 else ''
            print(f"✅ Fila {fila}: {fecha_str} - {signo}${abs(trans['monto']):,.2f} - {trans['concepto']}")

    if len(transacciones_cambiar) > 10:
        print(f"   ... y {len(transacciones_cambiar) - 10} más")

    print()
    print(f"📊 Total transacciones actualizadas: {contador}")
    print()

    # =========================================================================
    # PASO 3: GUARDAR
    # =========================================================================
    print("=" * 80)
    print("💾 PASO 3: Guardando cambios...")
    print("=" * 80)
    print()

    wb.save(EXCEL_FILE)
    print("✅ Excel actualizado")
    print()

    # =========================================================================
    # PASO 4: VERIFICACIÓN
    # =========================================================================
    print("=" * 80)
    print("📊 PASO 4: Verificación final...")
    print("=" * 80)
    print()

    # Recargar y contar
    wb_verificar = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws_verificar = wb_verificar['TRANSACCIONES']

    variaciones_despues = {}

    for row in range(2, ws_verificar.max_row + 1):
        cuenta = ws_verificar.cell(row, col_map['Cuenta Bancaria']).value

        if cuenta and 'Promerica' in str(cuenta):
            cuenta_str = str(cuenta).strip()

            if cuenta_str not in variaciones_despues:
                variaciones_despues[cuenta_str] = 0

            variaciones_despues[cuenta_str] += 1

    print("📋 Variaciones después de normalizar:")
    print()

    for nombre, cantidad in sorted(variaciones_despues.items()):
        print(f"   \"{nombre}\": {cantidad} transacciones")

    print()

    # Calcular balance Promerica USD
    total_ingreso = 0
    total_egreso = 0

    for row in range(2, ws_verificar.max_row + 1):
        cuenta = ws_verificar.cell(row, col_map['Cuenta Bancaria']).value
        monto = ws_verificar.cell(row, col_map['Monto USD']).value
        ing_egr = ws_verificar.cell(row, col_map['Ingreso/Egreso']).value

        if cuenta and str(cuenta).strip() == NOMBRE_ESTANDAR:
            if monto:
                monto_val = float(monto)
                if ing_egr == 'Ingreso':
                    total_ingreso += monto_val
                elif ing_egr == 'Egreso':
                    total_egreso += abs(monto_val)

    balance_calculado = total_ingreso - total_egreso

    print("💰 BALANCE PROMERICA USD (40000003881774):")
    print(f"   Ingreso: ${total_ingreso:,.2f}")
    print(f"   Egreso: ${total_egreso:,.2f}")
    print(f"   Balance: ${balance_calculado:,.2f}")
    print()

    print(f"🏦 Balance extracto bancario: $2,163.44")
    diferencia = abs(balance_calculado - 2163.44)
    print(f"⚖️  Diferencia: ${diferencia:,.2f}")
    print()

    if diferencia < 1.00:
        print("✅ ¡SALDO CORRECTO! Balance coincide con extracto 🎉")
    else:
        print(f"⚠️  Aún hay diferencia de ${diferencia:,.2f}")
        print()
        print("📋 POSIBLES CAUSAS:")
        print("   1. Faltan transacciones por registrar")
        print("   2. Hay transacciones duplicadas")
        print("   3. Saldo inicial incorrecto")
        print("   4. Hay transacciones de otra cuenta mezcladas")

    print()

    # =========================================================================
    # RESUMEN
    # =========================================================================
    print("=" * 80)
    print("📊 RESUMEN")
    print("=" * 80)
    print()

    print(f"✅ Transacciones normalizadas: {len(transacciones_cambiar)}")
    print(f"🎯 Nombre estándar aplicado: \"{NOMBRE_ESTANDAR}\"")
    print(f"📋 Variaciones después: {len(variaciones_despues)}")
    print()

    print("🔧 PRÓXIMOS PASOS:")
    print("   1. Cierra y vuelve a abrir el Excel")
    print("   2. Ve a la hoja Efectivo, fila 3 (Promerica)")
    print("   3. Verifica que el Balance muestre: $2,163.44")
    print()
    print("   Si aún no coincide, habrá que investigar las transacciones faltantes")

    print()
    print("=" * 80)
    print("✅ NORMALIZACIÓN COMPLETADA")
    print("=" * 80)
    print()

if __name__ == "__main__":
    try:
        if not crear_backup():
            print("❌ Abortando")
            exit(1)

        normalizar()
        print("🎉 Proceso completado exitosamente!")

    except Exception as e:
        print(f"❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
