#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Agregar múltiples movimientos de noviembre
- Uber Eats $18.73 (05/11)
- NASA Gasolinera ₡4,500 (08/11)
- Pago parqueos ₡20,000 (07/11)
- Pago SEA Global $58.76 (06/11)
"""
import openpyxl
from datetime import datetime
import sys

# Ruta del archivo
excel_file = "AlvaroVelasco_Finanzas_v1.0_CORREGIDO.xlsx"

try:
    wb = openpyxl.load_workbook(excel_file)
    ws = wb['TRANSACCIONES']

    print("="*80)
    print("AGREGANDO MOVIMIENTOS BATCH - NOVIEMBRE 2025")
    print("="*80)

    # Obtener último ID y última fila
    last_id = 0
    last_row = 1
    for row in range(2, ws.max_row + 1):
        id_val = ws[f'P{row}'].value
        if id_val:
            try:
                last_id = max(last_id, int(id_val))
                last_row = row
            except:
                pass

    next_id = last_id + 1
    next_row = last_row + 1

    print(f"\nÚltimo ID: {last_id}")
    print(f"Próximo ID: {next_id}")
    print(f"Próxima fila: {next_row}")

    # Tasa de cambio aproximada
    tc = 573.0

    # Lista de movimientos a agregar
    movimientos = []

    # 1. Uber Eats - 05/11/2025
    mov1_ref = "530922157846"
    existe = False
    for row in range(2, ws.max_row + 1):
        if ws[f'H{row}'].value == mov1_ref:
            existe = True
            print(f"\n❌ Uber Eats (Ref {mov1_ref}) YA EXISTE en fila {row}")
            break

    if not existe:
        movimientos.append({
            'fecha': '05/11/2025',
            'tipo': 'Gasto',
            'categoria': 'Alimentación',
            'entidad': 'Uber Eats',
            'cuenta': 'Tarjeta BNCR Visa 3519',
            'proveedor': 'Uber Eats',
            'concepto': 'Pedido Uber Eats vía PayPal',
            'referencia': mov1_ref,
            'monto_usd': 18.73,
            'monto_crc': 0,
            'tipo_mov': 'Egreso',
            'estado': 'Completado',
            'prioridad': 'BAJA',
            'notas': 'Gasto operativo - Alimentación'
        })

    # 2. NASA Gasolinera - 08/11/2025
    mov2_ref = "4216832973"
    existe = False
    for row in range(2, ws.max_row + 1):
        if ws[f'H{row}'].value == mov2_ref:
            existe = True
            print(f"\n❌ NASA Gasolinera (Ref {mov2_ref}) YA EXISTE en fila {row}")
            break

    if not existe:
        movimientos.append({
            'fecha': '08/11/2025',
            'tipo': 'Gasto',
            'categoria': 'Combustible',
            'entidad': 'Estación NASA Heredia',
            'cuenta': 'Tarjeta BAC 3873',
            'proveedor': 'NASA',
            'concepto': 'Combustible vehículo',
            'referencia': mov2_ref,
            'monto_usd': round(4500 / tc, 2),
            'monto_crc': 4500,
            'tipo_mov': 'Egreso',
            'estado': 'Completado',
            'prioridad': 'MEDIA',
            'notas': 'Gasto operativo - Combustible'
        })

    # 3. Pago parqueos - 07/11/2025
    mov3_ref = "2025110711631000083110251"
    existe = False
    for row in range(2, ws.max_row + 1):
        if ws[f'H{row}'].value == mov3_ref:
            existe = True
            print(f"\n❌ Pago parqueos (Ref {mov3_ref}) YA EXISTE en fila {row}")
            break

    if not existe:
        movimientos.append({
            'fecha': '07/11/2025',
            'tipo': 'Pago',
            'categoria': 'Servicios',
            'entidad': 'Alejandra Arias Fallas',
            'cuenta': 'Promerica CRC 1774',
            'proveedor': 'Alejandra Arias',
            'concepto': 'Pago peajes y parqueos',
            'referencia': mov3_ref,
            'monto_usd': round(20000 / tc, 2),
            'monto_crc': 20000,
            'tipo_mov': 'Egreso',
            'estado': 'Completado',
            'prioridad': 'MEDIA',
            'notas': 'Gasto operativo - Peajes y parqueos. Comprobante: 2796348'
        })

    # 4. Pago SEA Global - 06/11/2025
    mov4_ref = "2025110611631000083032151"
    existe = False
    for row in range(2, ws.max_row + 1):
        if ws[f'H{row}'].value == mov4_ref:
            existe = True
            print(f"\n❌ SEA Global (Ref {mov4_ref}) YA EXISTE en fila {row}")
            break

    if not existe:
        movimientos.append({
            'fecha': '06/11/2025',
            'tipo': 'Pago Proveedor',
            'categoria': 'Logística',
            'entidad': 'SEA Global Logistics',
            'cuenta': 'Promerica USD 1774',
            'proveedor': 'SEA Global',
            'concepto': 'Pago facturas 3271, 3251, 3234',
            'referencia': mov4_ref,
            'monto_usd': 58.76,
            'monto_crc': 0,
            'tipo_mov': 'Egreso',
            'estado': 'Completado',
            'prioridad': 'ALTA',
            'notas': 'Pago a proveedor - Facturas múltiples. Comprobante: 2795229'
        })

    # Agregar los movimientos
    if not movimientos:
        print("\n⚠️ TODOS LOS MOVIMIENTOS YA EXISTEN EN EL EXCEL")
        print("No hay nada que agregar.")
    else:
        print(f"\n✅ AGREGANDO {len(movimientos)} MOVIMIENTOS NUEVOS:\n")

        for mov in movimientos:
            # Agregar en la siguiente fila
            ws[f'A{next_row}'] = datetime.strptime(mov['fecha'], '%d/%m/%Y')
            ws[f'B{next_row}'] = mov['tipo']
            ws[f'C{next_row}'] = mov['categoria']
            ws[f'D{next_row}'] = mov['entidad']
            ws[f'E{next_row}'] = mov['cuenta']
            ws[f'F{next_row}'] = mov['proveedor']
            ws[f'G{next_row}'] = mov['concepto']
            ws[f'H{next_row}'] = mov['referencia']
            ws[f'I{next_row}'] = mov['monto_usd']
            ws[f'J{next_row}'] = mov['monto_crc']
            ws[f'K{next_row}'] = mov['tipo_mov']
            ws[f'L{next_row}'] = mov['estado']
            ws[f'M{next_row}'] = mov['prioridad']
            ws[f'N{next_row}'] = ''  # Vencimiento vacío
            ws[f'O{next_row}'] = mov['notas']
            ws[f'P{next_row}'] = next_id
            ws[f'Q{next_row}'] = datetime.now()
            ws[f'R{next_row}'] = 'Álvaro Velasco'

            print(f"Fila {next_row} (ID {next_id}): {mov['concepto']} - ${mov['monto_usd']} / ₡{mov['monto_crc']}")

            next_row += 1
            next_id += 1

        # Guardar
        wb.save(excel_file)
        print(f"\n✅ Archivo guardado: {excel_file}")
        print(f"✅ Total movimientos agregados: {len(movimientos)}")

except FileNotFoundError:
    print(f"❌ ERROR: No se encontró el archivo {excel_file}")
    print("Asegúrese de que el archivo esté en el directorio actual.")
except Exception as e:
    print(f"❌ ERROR: {e}")
    import traceback
    traceback.print_exc()
