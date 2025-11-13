#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CORRECCIÓN FINAL PROMERICA
1. Elimina duplicados (filas 211, 212, 213)
2. Agrega movimientos faltantes del extracto
3. Verifica saldo final = $2,163.44
"""
import openpyxl
from datetime import datetime
import shutil

EXCEL_FILE = "AlvaroVelasco_Finanzas_v2.0.xlsx"
BACKUP_FILE = f"AlvaroVelasco_Finanzas_v2.0_CORRECCION_FINAL_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

def crear_backup():
    print("=" * 80)
    print("CREANDO BACKUP")
    print("=" * 80)
    print(f"Backup: {BACKUP_FILE}")
    try:
        shutil.copy2(EXCEL_FILE, BACKUP_FILE)
        print("✅ Backup creado")
        print()
        return True
    except Exception as e:
        print(f"❌ ERROR: {e}")
        return False

def corregir():
    print("=" * 80)
    print("CORRECCIÓN FINAL PROMERICA")
    print("=" * 80)
    print()

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb['TRANSACCIONES']

    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    col_map = {}
    for col in range(1, len(headers) + 1):
        if headers[col-1]:
            col_map[headers[col-1]] = col

    # =========================================================================
    # PASO 1: ELIMINAR DUPLICADOS (filas 211, 212, 213)
    # =========================================================================
    print("📋 PASO 1: Eliminando duplicados...")
    print()

    # Verificar que las filas son duplicados
    duplicados = []

    # Fila 211: Waipio duplicado
    ref_211 = ws.cell(211, col_map['Referencia']).value
    monto_211 = ws.cell(211, col_map['Monto USD']).value
    if ref_211 == '8319443' and abs(float(monto_211) - 1459.96) < 0.01:
        duplicados.append(211)
        print(f"✓ Fila 211: Waipio $1,459.96 duplicado (REF: {ref_211})")

    # Fila 212: Delta duplicado
    ref_212 = ws.cell(212, col_map['Referencia']).value
    monto_212 = ws.cell(212, col_map['Monto USD']).value
    if ref_212 == '204203' and abs(float(monto_212) + 76.28) < 0.01:
        duplicados.append(212)
        print(f"✓ Fila 212: Delta $76.28 duplicado (REF: {ref_212})")

    # Fila 213: Comisión duplicada
    ref_213 = ws.cell(213, col_map['Referencia']).value
    monto_213 = ws.cell(213, col_map['Monto USD']).value
    if ref_213 == '2797944' and abs(float(monto_213) + 0.75) < 0.01:
        duplicados.append(213)
        print(f"✓ Fila 213: Comisión $0.75 duplicada (REF: {ref_213})")

    print()

    if duplicados:
        print(f"🗑️  Eliminando {len(duplicados)} filas duplicadas...")
        # Eliminar de mayor a menor para no cambiar índices
        for fila in sorted(duplicados, reverse=True):
            ws.delete_rows(fila, 1)
            print(f"   ✅ Fila {fila} eliminada")
        print()
    else:
        print("⚠️  No se encontraron duplicados en las filas esperadas")
        print()

    # =========================================================================
    # PASO 2: AGREGAR MOVIMIENTOS FALTANTES DEL EXTRACTO
    # =========================================================================
    print("=" * 80)
    print("📋 PASO 2: Agregando movimientos faltantes del extracto...")
    print("=" * 80)
    print()

    # Movimientos a agregar
    movimientos_faltantes = [
        {
            'fecha': datetime(2025, 11, 3),
            'tipo': 'GASTOS OPERATIVOS',
            'categoria': 'Cargas Sociales',
            'entidad': 'Gobierno',
            'cuenta': 'Promerica USD (40000003881774)',
            'proveedor': 'CCSS',
            'concepto': 'Pago CCSS - Transferencia electrónica',
            'referencia': '67169898',
            'monto_usd': -733.20,
        },
        {
            'fecha': datetime(2025, 11, 4),
            'tipo': 'INGRESOS',
            'categoria': 'Ventas de Productos',
            'entidad': 'Cuentas por Cobrar',
            'cuenta': 'Promerica USD (40000003881774)',
            'proveedor': 'Pérez Benavides Y.',
            'concepto': 'Pago facturas 2535, 2530',
            'referencia': '5490555',
            'monto_usd': 761.06,
        },
        {
            'fecha': datetime(2025, 11, 6),
            'tipo': 'GASTOS FINANCIEROS',
            'categoria': 'Comisiones Bancarias',
            'entidad': 'Banco',
            'cuenta': 'Promerica USD (40000003881774)',
            'proveedor': 'Banco Promerica',
            'concepto': 'Comisión Transferencia SINPE',
            'referencia': '2795229',
            'monto_usd': -0.75,
        },
        {
            'fecha': datetime(2025, 11, 6),
            'tipo': 'GASTOS OPERATIVOS',
            'categoria': 'Logística',
            'entidad': 'Proveedor',
            'cuenta': 'Promerica USD (40000003881774)',
            'proveedor': 'Sea Global Logistics Services',
            'concepto': 'Pago logística - Transferencia',
            'referencia': '32048171',
            'monto_usd': -58.76,
        },
        {
            'fecha': datetime(2025, 11, 7),
            'tipo': 'GASTOS FINANCIEROS',
            'categoria': 'Comisiones Bancarias',
            'entidad': 'Banco',
            'cuenta': 'Promerica USD (40000003881774)',
            'proveedor': 'Banco Promerica',
            'concepto': 'Comisión Transferencia SINPE',
            'referencia': '2796348',
            'monto_usd': -0.75,
        },
        {
            'fecha': datetime(2025, 11, 7),
            'tipo': 'AJUSTES',
            'categoria': 'Ajustes Bancarios',
            'entidad': 'Banco',
            'cuenta': 'Promerica USD (40000003881774)',
            'proveedor': 'Banco Promerica',
            'concepto': 'Reversión de Comisión por transferencia',
            'referencia': '2796348-REV',
            'monto_usd': 0.75,
        },
        {
            'fecha': datetime(2025, 11, 7),
            'tipo': 'GASTOS FINANCIEROS',
            'categoria': 'Comisiones Bancarias',
            'entidad': 'Banco',
            'cuenta': 'Promerica USD (40000003881774)',
            'proveedor': 'Banco Promerica',
            'concepto': 'Comisión Transferencia SINPE',
            'referencia': '2796348-2',
            'monto_usd': -0.75,
        },
        {
            'fecha': datetime(2025, 11, 7),
            'tipo': 'GASTOS OPERATIVOS',
            'categoria': 'Servicios Profesionales',
            'entidad': 'Proveedor',
            'cuenta': 'Promerica USD (40000003881774)',
            'proveedor': 'Alejandra Arias Fallas',
            'concepto': 'Pago servicios profesionales - Transferencia',
            'referencia': '32067344',
            'monto_usd': -40.57,
        },
        {
            'fecha': datetime(2025, 11, 7),
            'tipo': 'INGRESOS',
            'categoria': 'Otros Ingresos',
            'entidad': 'Cuentas por Cobrar',
            'cuenta': 'Promerica USD (40000003881774)',
            'proveedor': 'Asociación Costarricense',
            'concepto': 'Ingreso - PLANESBPPP',
            'referencia': '8313051',
            'monto_usd': 333.35,
        },
        {
            'fecha': datetime(2025, 11, 8),
            'tipo': 'GASTOS OPERATIVOS',
            'categoria': 'Combustible',
            'entidad': 'Vehículo',
            'cuenta': 'Promerica USD (40000003881774)',
            'proveedor': 'Estación Servicios NASA Heredia',
            'concepto': 'Combustible - Estación NASA Heredia',
            'referencia': '141111',
            'monto_usd': -9.13,
        },
        {
            'fecha': datetime(2025, 11, 10),
            'tipo': 'PASIVOS',
            'categoria': 'Pago a Proveedores',
            'entidad': 'Pasivo',
            'cuenta': 'Promerica USD (40000003881774)',
            'proveedor': 'STCR Costa Rica Trust (Intcomex)',
            'concepto': 'Pago Intcomex vía STCR - Facturas pendientes',
            'referencia': '32105690',
            'monto_usd': -3137.26,
        },
    ]

    filas_agregadas = []

    for trans in movimientos_faltantes:
        next_row = ws.max_row + 1

        ws.cell(next_row, col_map['Fecha']).value = trans['fecha']
        ws.cell(next_row, col_map['Fecha']).number_format = 'd/m/yy'
        ws.cell(next_row, col_map['Tipo Transacción']).value = trans['tipo']
        ws.cell(next_row, col_map['Categoría']).value = trans['categoria']
        ws.cell(next_row, col_map['Entidad']).value = trans['entidad']
        ws.cell(next_row, col_map['Cuenta Bancaria']).value = trans['cuenta']
        ws.cell(next_row, col_map['Cliente/Proveedor']).value = trans['proveedor']
        ws.cell(next_row, col_map['Concepto']).value = trans['concepto']
        ws.cell(next_row, col_map['Referencia']).value = trans['referencia']
        ws.cell(next_row, col_map['Monto USD']).value = trans['monto_usd']

        # Ingreso/Egreso se determina por fórmula
        # Estado y Prioridad
        ws.cell(next_row, col_map['Estado']).value = 'Completado' if trans['monto_usd'] > 0 else 'Pagado'
        ws.cell(next_row, col_map['Prioridad']).value = 'Normal'
        ws.cell(next_row, col_map['Notas']).value = f"Según extracto bancario - Doc: {trans['referencia']}"

        filas_agregadas.append({
            'fila': next_row,
            'fecha': trans['fecha'].strftime('%d/%m/%Y'),
            'concepto': trans['concepto'],
            'monto': trans['monto_usd'],
        })

        signo = '+' if trans['monto_usd'] > 0 else ''
        print(f"✅ {trans['fecha'].strftime('%d/%m/%Y')} - {trans['concepto'][:45]}")
        print(f"   Fila: {next_row} | Monto: {signo}${abs(trans['monto_usd']):,.2f} | Ref: {trans['referencia']}")

    print()
    print(f"📊 Total movimientos agregados: {len(filas_agregadas)}")
    print()

    # =========================================================================
    # PASO 3: GUARDAR Y VERIFICAR
    # =========================================================================
    print("=" * 80)
    print("💾 Guardando cambios...")
    print("=" * 80)
    print()

    wb.save(EXCEL_FILE)
    print("✅ Excel actualizado")
    print()

    # =========================================================================
    # PASO 4: VERIFICACIÓN FINAL
    # =========================================================================
    print("=" * 80)
    print("📊 VERIFICACIÓN FINAL")
    print("=" * 80)
    print()

    # Recargar para calcular con fórmulas
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb['TRANSACCIONES']

    # Calcular saldo
    saldo_inicial = 3030.89
    total_movimientos = 0

    for row in range(2, ws.max_row + 1):
        cuenta = ws.cell(row, col_map['Cuenta Bancaria']).value
        fecha = ws.cell(row, col_map['Fecha']).value
        monto = ws.cell(row, col_map['Monto USD']).value
        tipo = ws.cell(row, col_map['Tipo Transacción']).value

        if cuenta and 'Promerica' in str(cuenta) and '40000003881774' in str(cuenta):
            if fecha and isinstance(fecha, datetime):
                if fecha.month == 11 and fecha.year == 2025:
                    if tipo and 'Apertura Inicial' not in str(tipo):
                        if monto:
                            total_movimientos += float(monto)

    saldo_calculado = saldo_inicial + total_movimientos

    print(f"✅ Saldo Inicial: ${saldo_inicial:,.2f}")
    print(f"📊 Total movimientos noviembre: ${total_movimientos:,.2f}")
    print(f"💰 Saldo calculado: ${saldo_calculado:,.2f}")
    print()
    print(f"🏦 Saldo extracto bancario: $2,163.44")
    diferencia = abs(saldo_calculado - 2163.44)
    print(f"⚖️  Diferencia: ${diferencia:,.2f}")
    print()

    if diferencia < 1.00:
        print("✅ ¡SALDO CORRECTO! Reconciliación exitosa 🎉")
    else:
        print(f"⚠️  Aún hay diferencia de ${diferencia:,.2f}")
        print("   Revisar manualmente en el Excel")

    print()

    # =========================================================================
    # RESUMEN FINAL
    # =========================================================================
    print("=" * 80)
    print("📊 RESUMEN DE CAMBIOS")
    print("=" * 80)
    print()

    print(f"🗑️  Duplicados eliminados: {len(duplicados)}")
    for fila in duplicados:
        print(f"   • Fila {fila}")
    print()

    print(f"➕ Movimientos agregados: {len(filas_agregadas)}")
    for mov in filas_agregadas:
        signo = '+' if mov['monto'] > 0 else ''
        print(f"   • Fila {mov['fila']}: {mov['fecha']} - {signo}${abs(mov['monto']):,.2f}")
    print()

    print("=" * 80)
    print("✅ CORRECCIÓN COMPLETADA")
    print("=" * 80)
    print()

if __name__ == "__main__":
    try:
        if not crear_backup():
            print("❌ Abortando")
            exit(1)

        corregir()
        print("🎉 Proceso completado exitosamente!")

    except Exception as e:
        print(f"❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
