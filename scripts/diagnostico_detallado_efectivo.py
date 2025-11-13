#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
DIAGNÓSTICO DETALLADO - HOJA EFECTIVO
Muestra estructura completa para entender dónde están los balances
"""
import openpyxl
from alias_cuentas import es_balance_inicial

EXCEL_FILE = "AlvaroVelasco_Finanzas_v2.0.xlsx"

print("="*80)
print("DIAGNÓSTICO DETALLADO - HOJA EFECTIVO")
print("="*80)

wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
ws = wb['Efectivo']

print(f"\n📄 Hoja: Efectivo")
print(f"   Total filas: {ws.max_row}")
print(f"   Total columnas: {ws.max_column}")

print("\n" + "="*80)
print("PRIMERAS 20 FILAS - TODAS LAS COLUMNAS")
print("="*80)

# Leer headers
headers = []
for col in range(1, ws.max_column + 1):
    header = ws.cell(1, col).value
    headers.append(header if header else f"Col{col}")

print(f"\nHeaders detectados:")
for i, h in enumerate(headers, 1):
    col_letter = openpyxl.utils.get_column_letter(i)
    print(f"   {col_letter}: {h}")

print("\n" + "="*80)
print("CONTENIDO FILA POR FILA")
print("="*80)

for row in range(1, min(21, ws.max_row + 1)):
    print(f"\n{'='*40}")
    print(f"Fila {row}:")
    print(f"{'='*40}")

    for col_idx, header in enumerate(headers, 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        valor = ws.cell(row, col_idx).value

        if valor is not None and str(valor).strip() != "":
            print(f"   {col_letter} ({header}): {valor}")

            # Verificar si es balance inicial
            if header and 'concepto' in str(header).lower():
                if es_balance_inicial(valor):
                    print(f"      ✅ RECONOCIDO como Balance inicial")

print("\n" + "="*80)
print("BÚSQUEDA DE BALANCES INICIALES")
print("="*80)

balances_encontrados = []

for row in range(1, min(30, ws.max_row + 1)):
    # Leer todas las celdas de la fila
    fila_data = {}
    for col_idx, header in enumerate(headers, 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        valor = ws.cell(row, col_idx).value
        fila_data[col_letter] = {
            'header': header,
            'valor': valor
        }

    # Buscar en TODAS las columnas si hay algo que parezca balance inicial
    for col_letter, data in fila_data.items():
        valor = data['valor']
        if valor and es_balance_inicial(valor):
            print(f"\n✅ ENCONTRADO en Fila {row}, Columna {col_letter} ({data['header']})")
            print(f"   Valor: {valor}")
            print(f"   Contenido completo de la fila:")
            for cl, cd in fila_data.items():
                if cd['valor'] is not None and str(cd['valor']).strip() != "":
                    print(f"      {cl} ({cd['header']}): {cd['valor']}")

            balances_encontrados.append({
                'fila': row,
                'col': col_letter,
                'concepto': valor,
                'fila_completa': fila_data
            })

if not balances_encontrados:
    print("\n❌ NO se encontraron balances iniciales en las primeras 30 filas")
    print("\n💡 POSIBLES RAZONES:")
    print("   1. La hoja Efectivo no tiene 'Balance inicial' o 'Apertura Inicial' explícitos")
    print("   2. Los balances están en una columna diferente a la esperada")
    print("   3. La hoja Efectivo solo tiene formulas que apuntan a TRANSACCIONES")

    print("\n🔍 Buscando cualquier texto que contenga 'balance' o 'apertura':")
    for row in range(1, min(30, ws.max_row + 1)):
        for col_idx in range(1, ws.max_column + 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            valor = ws.cell(row, col_idx).value
            if valor:
                valor_str = str(valor).lower()
                if 'balance' in valor_str or 'apertura' in valor_str or 'saldo' in valor_str:
                    print(f"   Fila {row}, {col_letter}: {valor}")
else:
    print(f"\n✅ Total balances iniciales encontrados: {len(balances_encontrados)}")

print("\n" + "="*80)
