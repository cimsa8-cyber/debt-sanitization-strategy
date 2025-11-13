#!/usr/bin/env python3
"""
PLAN DE PAGO CxP - Estrategia de Supervivencia
Prioriza pagos según urgencia, impacto operativo, y liquidez disponible
"""

import openpyxl
from datetime import datetime

V3_FILE = "AlvaroVelasco_Finanzas_v3.0.xlsx"

print("\n" + "="*70)
print("PLAN DE PAGO CxP - ESTRATEGIA DE SUPERVIVENCIA")
print("="*70)

wb = openpyxl.load_workbook(V3_FILE, data_only=True)
ws_cxp = wb['CxP']
ws_efectivo = wb['EFECTIVO']

# ============================================================================
# CALCULAR EFECTIVO NETO
# ============================================================================

total_bancos = 0
total_tarjetas = 0

for row in range(5, 14):
    saldo = ws_efectivo.cell(row, 5).value or 0
    if isinstance(saldo, (int, float)):
        total_bancos += saldo

for row in range(16, 21):
    saldo = ws_efectivo.cell(row, 5).value or 0
    if isinstance(saldo, (int, float)):
        total_tarjetas += abs(saldo)

efectivo_neto = total_bancos - total_tarjetas
efectivo_disponible = total_bancos  # Solo bancos (sin usar tarjetas)

print(f"\n💰 SITUACIÓN ACTUAL:")
print(f"   Efectivo en Bancos: ${total_bancos:,.2f}")
print(f"   Deuda Tarjetas: ${total_tarjetas:,.2f}")
print(f"   Efectivo Neto: ${efectivo_neto:,.2f}")

# ============================================================================
# EXTRAER CxP
# ============================================================================

print(f"\n📊 ANALIZANDO CUENTAS POR PAGAR...")

cxp_data = []

for row in range(3, 25):
    proveedor = ws_cxp.cell(row, 1).value
    factura = ws_cxp.cell(row, 2).value
    fecha_emision = ws_cxp.cell(row, 3).value
    fecha_venc = ws_cxp.cell(row, 4).value
    monto = ws_cxp.cell(row, 5).value or 0
    saldo = ws_cxp.cell(row, 6).value or 0
    dias_vencido = ws_cxp.cell(row, 7).value or 0
    prioridad = ws_cxp.cell(row, 8).value or ""
    estado = ws_cxp.cell(row, 9).value or ""

    if proveedor and isinstance(saldo, (int, float)) and saldo > 0:
        # Calcular días para vencer
        dias_para_vencer = 999
        if fecha_venc and isinstance(fecha_venc, datetime):
            dias_para_vencer = (fecha_venc - datetime.now()).days

        # Calcular días vencido manualmente si es necesario
        if isinstance(dias_vencido, str) or dias_vencido is None:
            if dias_para_vencer < 0:
                dias_vencido = abs(dias_para_vencer)
            else:
                dias_vencido = 0

        cxp_data.append({
            'proveedor': proveedor,
            'factura': factura,
            'saldo': saldo,
            'dias_vencido': dias_vencido,
            'dias_para_vencer': dias_para_vencer,
            'prioridad': prioridad,
            'estado': estado,
            'fecha_venc': fecha_venc
        })

total_cxp = sum(c['saldo'] for c in cxp_data)
print(f"   Total CxP: ${total_cxp:,.2f}")
print(f"   Proveedores: {len(cxp_data)}")

# ============================================================================
# PRIORIZAR PAGOS
# ============================================================================

print(f"\n🎯 PRIORIZANDO PAGOS...")

# Scoring:
# - Prioridad: CRÍTICA=100, ALTA=75, MEDIA=50, BAJA=25
# - Días vencido: +2 puntos por día vencido
# - Días para vencer: -1 punto por día que falta (menos = más urgente)

for c in cxp_data:
    # Score prioridad
    if "CRÍTICA" in str(c['prioridad']).upper() or "CRITICA" in str(c['prioridad']).upper():
        score_prioridad = 100
        c['urgencia_visual'] = "🔴 CRÍTICA"
    elif "ALTA" in str(c['prioridad']).upper():
        score_prioridad = 75
        c['urgencia_visual'] = "🟠 ALTA"
    elif "MEDIA" in str(c['prioridad']).upper():
        score_prioridad = 50
        c['urgencia_visual'] = "🟡 MEDIA"
    else:
        score_prioridad = 25
        c['urgencia_visual'] = "🟢 BAJA"

    # Score días vencido
    score_vencido = c['dias_vencido'] * 2

    # Score días para vencer
    if c['dias_para_vencer'] < 0:  # Ya vencido
        score_dias = 50
    elif c['dias_para_vencer'] < 7:  # Vence en menos de 7 días
        score_dias = 30
    elif c['dias_para_vencer'] < 15:
        score_dias = 15
    else:
        score_dias = max(0, 15 - c['dias_para_vencer'] / 2)

    c['score'] = score_prioridad + score_vencido + score_dias

    # Categoría operativa
    if "ALQUILER" in c['proveedor'].upper():
        c['categoria'] = "🏢 OPERATIVO CRÍTICO"
        c['score'] += 20  # Bonus por criticidad
    elif "NISSAN" in c['proveedor'].upper():
        c['categoria'] = "🚗 FINANCIAMIENTO"
    elif "HACIENDA" in c['proveedor'].upper() or "IVA" in c['proveedor'].upper():
        c['categoria'] = "🏛️ GOBIERNO"
    elif "INTCOMEX" in c['proveedor'].upper():
        c['categoria'] = "📦 PROVEEDOR CLAVE"
        c['score'] += 10  # Bonus por ser proveedor principal
    else:
        c['categoria'] = "🔧 SERVICIOS/OTROS"

# Ordenar por score descendente
cxp_ranking = sorted(cxp_data, key=lambda x: x['score'], reverse=True)

# ============================================================================
# REPORTE PRIORIZADO
# ============================================================================

print("\n" + "="*70)
print("RANKING DE PAGOS (Mayor prioridad primero)")
print("="*70)

print(f"\n{'#':<3} {'PROVEEDOR':<28} {'SALDO':<12} {'VENCE':<8} {'PRIORIDAD':<14} {'CATEGORÍA':<20}")
print("-" * 70)

for i, c in enumerate(cxp_ranking, 1):
    if c['dias_para_vencer'] < 0:
        vence_str = f"{abs(c['dias_para_vencer']):.0f}d atrás"
    else:
        vence_str = f"en {c['dias_para_vencer']:.0f}d"

    print(f"{i:<3} {c['proveedor'][:26]:<28} ${c['saldo']:>10,.2f} {vence_str:<8} {c['urgencia_visual']:<14} {c['categoria']:<20}")

# ============================================================================
# ESCENARIOS DE PAGO
# ============================================================================

print("\n" + "="*70)
print("ESCENARIOS DE PAGO")
print("="*70)

print(f"\n💵 Efectivo disponible: ${efectivo_disponible:,.2f}")
print(f"📊 Total CxP: ${total_cxp:,.2f}")
print(f"❌ Déficit: ${efectivo_disponible - total_cxp:,.2f}")

escenarios = [
    ("Escenario 1: SOLO CRÍTICO (mantener operación)", lambda c: "CRÍTICA" in c['prioridad'].upper() or "CRITICA" in c['prioridad'].upper()),
    ("Escenario 2: CRÍTICO + ALTA (moderado)", lambda c: "CRÍTICA" in c['prioridad'].upper() or "CRITICA" in c['prioridad'].upper() or "ALTA" in c['prioridad'].upper()),
    ("Escenario 3: Todo <15 días (agresivo)", lambda c: c['dias_para_vencer'] < 15),
    ("Escenario 4: PAGAR TODO (no recomendado)", lambda c: True)
]

for nombre, filtro in escenarios:
    pagos = [c for c in cxp_ranking if filtro(c)]
    total_pago = sum(c['saldo'] for c in pagos)
    efectivo_restante = efectivo_disponible - total_pago
    porcentaje = (total_pago / total_cxp * 100) if total_cxp > 0 else 0

    print(f"\n{nombre}")
    print(f"  💵 Pago total: ${total_pago:,.2f} ({porcentaje:.1f}% del CxP)")
    print(f"  💰 Efectivo restante: ${efectivo_restante:,.2f}")
    print(f"  📋 Facturas a pagar: {len(pagos)}")

    if efectivo_restante < 0:
        print(f"  ❌ IMPOSIBLE - Falta ${abs(efectivo_restante):,.2f}")
    elif efectivo_restante < 5000:
        print(f"  ⚠️  ARRIESGADO - Solo queda ${efectivo_restante:,.2f}")
    elif efectivo_restante < 15000:
        print(f"  ✅ VIABLE - Liquidez aceptable")
    else:
        print(f"  ✅ SEGURO - Buena liquidez")

    # Listar facturas
    if len(pagos) <= 5:
        for p in pagos:
            print(f"     • {p['proveedor']}: ${p['saldo']:,.2f}")

# ============================================================================
# RECOMENDACIONES
# ============================================================================

print("\n" + "="*70)
print("💡 RECOMENDACIONES ESTRATÉGICAS")
print("="*70)

# Filtrar críticos
criticos = [c for c in cxp_data if "CRÍTICA" in c['prioridad'].upper() or "CRITICA" in c['prioridad'].upper()]
total_critico = sum(c['saldo'] for c in criticos)

# Filtrar vence <7 días
urgentes = [c for c in cxp_data if c['dias_para_vencer'] < 7 and c['dias_para_vencer'] >= 0]
total_urgente = sum(c['saldo'] for c in urgentes)

print(f"\n1. PRIORIDAD MÁXIMA (Pagar SÍ o SÍ):")
print(f"   🔴 CRÍTICOS: {len(criticos)} facturas = ${total_critico:,.2f}")
for c in criticos:
    print(f"      • {c['proveedor']}: ${c['saldo']:,.2f} ({c['categoria']})")

print(f"\n2. URGENTE (Vence en <7 días):")
print(f"   ⏰ {len(urgentes)} facturas = ${total_urgente:,.2f}")
for c in urgentes:
    print(f"      • {c['proveedor']}: ${c['saldo']:,.2f} (vence en {c['dias_para_vencer']:.0f} días)")

print(f"\n3. ESTRATEGIA RECOMENDADA:")
total_pagar_min = total_critico + total_urgente
efectivo_post_pago = efectivo_disponible - total_pagar_min

print(f"   Pagar: CRÍTICOS + URGENTES = ${total_pagar_min:,.2f}")
print(f"   Efectivo restante: ${efectivo_post_pago:,.2f}")

if efectivo_post_pago > 15000:
    print(f"   ✅ PLAN VIABLE - Buena liquidez post-pago")
elif efectivo_post_pago > 5000:
    print(f"   ⚠️  PLAN AJUSTADO - Liquidez justa")
else:
    print(f"   ❌ PLAN CRÍTICO - Necesitás cobrar CxC primero")

print(f"\n4. NEGOCIACIÓN:")
# Identificar pagos que se pueden negociar (BAJA prioridad y monto grande)
negociables = [c for c in cxp_data if "BAJA" in c['prioridad'].upper() and c['saldo'] > 1000]
if negociables:
    print(f"   💬 Negociar plazo con:")
    for c in negociables[:3]:
        print(f"      • {c['proveedor']}: ${c['saldo']:,.2f} (prioridad {c['prioridad']})")
else:
    print(f"   ℹ️  No hay facturas grandes con baja prioridad para negociar")

print("\n" + "="*70)

wb.close()
