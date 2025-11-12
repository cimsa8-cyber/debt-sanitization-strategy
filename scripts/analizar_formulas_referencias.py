#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ANÁLISIS EXHAUSTIVO DE FÓRMULAS
Identifica referencias sin $$ y fórmulas que se corrieron al agregar filas
"""
import openpyxl
import re

EXCEL_FILE = "AlvaroVelasco_Finanzas_v2.0.xlsx"

def analizar_formula(formula):
    """Analiza una fórmula y detecta referencias sin $$"""
    if not formula or not isinstance(formula, str) or not formula.startswith('='):
        return None

    # Buscar referencias de celdas (ej: A1, $A$1, A$1, $A1)
    patron = r'([A-Z]+\$?\d+|\$[A-Z]+\$?\d+)'
    referencias = re.findall(patron, formula)

    referencias_sin_absoluto = []
    referencias_absolutas = []
    referencias_mixtas = []

    for ref in referencias:
        if '$' not in ref:
            referencias_sin_absoluto.append(ref)
        elif ref.count('$') == 2:
            referencias_absolutas.append(ref)
        else:
            referencias_mixtas.append(ref)

    return {
        'sin_absoluto': referencias_sin_absoluto,
        'absolutas': referencias_absolutas,
        'mixtas': referencias_mixtas,
        'total_refs': len(referencias)
    }

def analizar():
    print("=" * 80)
    print("ANÁLISIS EXHAUSTIVO DE FÓRMULAS")
    print("=" * 80)
    print()

    wb = openpyxl.load_workbook(EXCEL_FILE)

    # =========================================================================
    # PASO 1: ANALIZAR HOJA EFECTIVO
    # =========================================================================
    print("📋 PASO 1: Analizando hoja EFECTIVO...")
    print("=" * 80)
    print()

    ws_efectivo = wb['Efectivo']

    problemas_efectivo = []

    # Analizar fila 3 completa
    print("🔍 FILA 3 (Promerica USD):")
    print()

    for col in range(1, ws_efectivo.max_column + 1):
        celda = ws_efectivo.cell(3, col)
        letra_col = openpyxl.utils.get_column_letter(col)

        if celda.value and isinstance(celda.value, str) and celda.value.startswith('='):
            analisis = analizar_formula(celda.value)

            print(f"{letra_col}3: {celda.value}")

            if analisis['sin_absoluto']:
                print(f"   ⚠️  Referencias SIN $$: {', '.join(analisis['sin_absoluto'])}")
                problemas_efectivo.append({
                    'celda': f"{letra_col}3",
                    'formula': celda.value,
                    'problema': f"Referencias sin $$: {', '.join(analisis['sin_absoluto'])}"
                })

            if analisis['mixtas']:
                print(f"   ⚠️  Referencias MIXTAS: {', '.join(analisis['mixtas'])}")

            if analisis['absolutas']:
                print(f"   ✅ Referencias ABSOLUTAS: {', '.join(analisis['absolutas'])}")

            print()

    # =========================================================================
    # PASO 2: ANALIZAR COLUMNA K EN TRANSACCIONES
    # =========================================================================
    print("=" * 80)
    print("📋 PASO 2: Analizando columna K (Ingreso/Egreso) en TRANSACCIONES...")
    print("=" * 80)
    print()

    ws_trans = wb['TRANSACCIONES']

    # Encontrar columna K
    headers = [ws_trans.cell(1, col).value for col in range(1, ws_trans.max_column + 1)]
    try:
        col_k = headers.index('Ingreso/Egreso') + 1
    except ValueError:
        print("❌ No se encontró columna 'Ingreso/Egreso'")
        return

    print(f"✅ Columna K = columna {col_k}")
    print()

    # Verificar primeras 5 filas y últimas 5 filas
    print("🔍 Primeras 5 filas (después del encabezado):")
    print()

    problemas_transacciones = []

    for row in range(2, min(7, ws_trans.max_row + 1)):
        celda = ws_trans.cell(row, col_k)

        if celda.value and isinstance(celda.value, str) and celda.value.startswith('='):
            analisis = analizar_formula(celda.value)

            print(f"Fila {row}: {celda.value}")

            if analisis['sin_absoluto']:
                print(f"   ⚠️  Referencias SIN $$: {', '.join(analisis['sin_absoluto'])}")
                problemas_transacciones.append({
                    'celda': f"K{row}",
                    'formula': celda.value,
                    'problema': f"Referencias sin $$: {', '.join(analisis['sin_absoluto'])}"
                })
            else:
                print(f"   ✅ OK")

            print()

    print("🔍 Últimas 5 filas (filas nuevas agregadas):")
    print()

    for row in range(max(2, ws_trans.max_row - 4), ws_trans.max_row + 1):
        celda = ws_trans.cell(row, col_k)
        tipo = ws_trans.cell(row, headers.index('Tipo Transacción') + 1).value if 'Tipo Transacción' in headers else None
        cuenta = ws_trans.cell(row, headers.index('Cuenta Bancaria') + 1).value if 'Cuenta Bancaria' in headers else None

        print(f"Fila {row}:")
        print(f"   Cuenta: {cuenta}")
        print(f"   Tipo: {tipo}")

        if celda.value and isinstance(celda.value, str) and celda.value.startswith('='):
            analisis = analizar_formula(celda.value)

            print(f"   Fórmula: {celda.value}")

            if analisis['sin_absoluto']:
                print(f"   ⚠️  Referencias SIN $$: {', '.join(analisis['sin_absoluto'])}")
                problemas_transacciones.append({
                    'celda': f"K{row}",
                    'formula': celda.value,
                    'problema': f"Referencias sin $$: {', '.join(analisis['sin_absoluto'])}"
                })
            else:
                print(f"   ✅ OK")
        else:
            print(f"   Valor: {celda.value}")

        print()

    # =========================================================================
    # PASO 3: VERIFICAR FÓRMULAS EN FILAS 211-221 (LAS AGREGADAS)
    # =========================================================================
    print("=" * 80)
    print("📋 PASO 3: Verificando filas 211-221 (agregadas por scripts)...")
    print("=" * 80)
    print()

    print("🔍 Verificando si tienen fórmulas en columna K:")
    print()

    for row in range(211, 222):
        if row > ws_trans.max_row:
            break

        celda_k = ws_trans.cell(row, col_k)
        cuenta = ws_trans.cell(row, headers.index('Cuenta Bancaria') + 1).value if 'Cuenta Bancaria' in headers else None
        concepto = ws_trans.cell(row, headers.index('Concepto') + 1).value if 'Concepto' in headers else None

        print(f"Fila {row}: {concepto[:40] if concepto else 'N/A'}...")
        print(f"   Cuenta: {cuenta}")

        if celda_k.value:
            if isinstance(celda_k.value, str) and celda_k.value.startswith('='):
                print(f"   ⚠️  TIENE FÓRMULA: {celda_k.value}")
                print(f"   ❌ PROBLEMA: Los scripts deberían haber puesto 'Ingreso' o 'Egreso', no fórmulas")
            else:
                print(f"   ✅ Valor fijo: {celda_k.value}")
        else:
            print(f"   ⚠️  VACÍO - debería tener 'Ingreso' o 'Egreso'")

        print()

    # =========================================================================
    # RESUMEN DE PROBLEMAS
    # =========================================================================
    print("=" * 80)
    print("📊 RESUMEN DE PROBLEMAS DETECTADOS")
    print("=" * 80)
    print()

    total_problemas = len(problemas_efectivo) + len(problemas_transacciones)

    if total_problemas == 0:
        print("✅ No se detectaron problemas con referencias sin $$")
    else:
        print(f"⚠️  Total problemas detectados: {total_problemas}")
        print()

        if problemas_efectivo:
            print(f"🔴 HOJA EFECTIVO: {len(problemas_efectivo)} problemas")
            for p in problemas_efectivo:
                print(f"   • {p['celda']}: {p['problema']}")
            print()

        if problemas_transacciones:
            print(f"🔴 HOJA TRANSACCIONES: {len(problemas_transacciones)} problemas")
            for p in problemas_transacciones:
                print(f"   • {p['celda']}: {p['problema']}")
            print()

    print("=" * 80)
    print("✅ ANÁLISIS COMPLETADO")
    print("=" * 80)
    print()

if __name__ == "__main__":
    try:
        analizar()
    except Exception as e:
        print(f"❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
