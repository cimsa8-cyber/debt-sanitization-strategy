#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CORRECCIÓN PROMERICA USD 1774
Soluciona 3 problemas críticos:
1. Balance inicial duplicado (01/11)
2. Cuentas por cobrar mal categorizadas
3. Facturas vencidas en cuenta incorrecta
"""
import openpyxl
from datetime import datetime
import sys
import os

EXCEL_FILE = "AlvaroVelasco_Finanzas_v2.0.xlsx"

print("="*80)
print("CORRECCIÓN: PROMERICA USD 1774 - PROBLEMAS IDENTIFICADOS")
print("="*80)

if not os.path.exists(EXCEL_FILE):
    print(f"\nERROR: No se encontró {EXCEL_FILE}")
    sys.exit(1)

# Cargar Excel
print("\nCargando Excel...")
wb = openpyxl.load_workbook(EXCEL_FILE)
ws = wb['TRANSACCIONES']

# Contadores
cambios_realizados = []
balance_duplicado_eliminado = False
cuentas_cobrar_movidas = 0
cuentas_pagar_movidas = 0

# Recorrer transacciones
print("\nAnalizando transacciones...")

for row in range(2, ws.max_row + 1):
    cuenta = ws[f'E{row}'].value
    fecha = ws[f'A{row}'].value
    tipo = ws[f'B{row}'].value
    concepto = ws[f'G{row}'].value
    monto_usd = ws[f'I{row}'].value

    if not cuenta:
        continue

    cuenta_str = str(cuenta).strip()

    # Solo procesar cuentas de Promerica
    if 'Promerica USD' not in cuenta_str:
        continue

    # PROBLEMA 1: Balance inicial duplicado del 01/11
    if tipo and 'Balance inicial' in str(tipo):
        if fecha and fecha.month == 11 and fecha.day == 1:
            # Este es el balance duplicado - ELIMINAR FILA
            print(f"\n⚠️ Balance duplicado encontrado (Fila {row}):")
            print(f"   Fecha: {fecha.strftime('%d/%m/%Y')}")
            print(f"   Monto: ${monto_usd}")
            print(f"   Concepto: {concepto}")
            print(f"   → Marcando para eliminación...")

            # Marcar toda la fila para eliminación (limpiar contenido)
            for col in range(1, 21):  # Columnas A-T
                ws.cell(row=row, column=col).value = None

            cambios_realizados.append({
                'fila': row,
                'tipo': 'BALANCE_DUPLICADO_ELIMINADO',
                'monto': monto_usd,
                'fecha': fecha.strftime('%d/%m/%Y') if fecha else 'N/A'
            })
            balance_duplicado_eliminado = True
            continue

    # PROBLEMA 2: Saldos pendientes (Cuentas por Cobrar)
    if concepto and 'Saldo pendiente' in str(concepto):
        print(f"\n⚠️ Cuenta por cobrar mal ubicada (Fila {row}):")
        print(f"   Concepto: {concepto}")
        print(f"   Monto: ${monto_usd}")
        print(f"   → Moviendo a 'Por Cobrar'...")

        # Cambiar cuenta a "Por Cobrar"
        ws[f'E{row}'] = 'Por Cobrar'

        cambios_realizados.append({
            'fila': row,
            'tipo': 'CUENTA_COBRAR_MOVIDA',
            'concepto': concepto,
            'monto': monto_usd
        })
        cuentas_cobrar_movidas += 1
        continue

    # PROBLEMA 3: Facturas vencidas (Cuentas por Pagar)
    if concepto and ('VENCIDA' in str(concepto).upper() or 'Factura pendiente' in str(concepto)):
        print(f"\n⚠️ Cuenta por pagar mal ubicada (Fila {row}):")
        print(f"   Concepto: {concepto}")
        print(f"   Monto: ${monto_usd}")
        print(f"   → Moviendo a 'Por Pagar'...")

        # Cambiar cuenta a "Por Pagar"
        ws[f'E{row}'] = 'Por Pagar'

        cambios_realizados.append({
            'fila': row,
            'tipo': 'CUENTA_PAGAR_MOVIDA',
            'concepto': concepto,
            'monto': monto_usd
        })
        cuentas_pagar_movidas += 1
        continue

print("\n" + "="*80)
print("RESUMEN DE CAMBIOS")
print("="*80)

print(f"\nTotal cambios realizados: {len(cambios_realizados)}")

if balance_duplicado_eliminado:
    print(f"\n✓ Balance inicial duplicado ELIMINADO: 1")

if cuentas_cobrar_movidas > 0:
    print(f"\n✓ Cuentas por cobrar movidas: {cuentas_cobrar_movidas}")
    total_cobrar = sum([c['monto'] for c in cambios_realizados if c['tipo'] == 'CUENTA_COBRAR_MOVIDA'])
    print(f"   Total monto: ${total_cobrar:,.2f}")

if cuentas_pagar_movidas > 0:
    print(f"\n✓ Cuentas por pagar movidas: {cuentas_pagar_movidas}")
    total_pagar = sum([c['monto'] for c in cambios_realizados if c['tipo'] == 'CUENTA_PAGAR_MOVIDA'])
    print(f"   Total monto: ${total_pagar:,.2f}")

# Detalle de cambios
print("\n" + "="*80)
print("DETALLE DE CAMBIOS")
print("="*80)

# Agrupar por tipo
por_tipo = {}
for cambio in cambios_realizados:
    tipo = cambio['tipo']
    if tipo not in por_tipo:
        por_tipo[tipo] = []
    por_tipo[tipo].append(cambio)

for tipo, cambios in por_tipo.items():
    print(f"\n{tipo}:")
    for c in cambios[:10]:  # Primeros 10
        if 'concepto' in c:
            print(f"   Fila {c['fila']} | ${c['monto']:>8.2f} | {c['concepto'][:50]}")
        else:
            print(f"   Fila {c['fila']} | ${c['monto']:>8.2f} | {c.get('fecha', 'N/A')}")

    if len(cambios) > 10:
        print(f"   ... y {len(cambios) - 10} más")

# Guardar cambios
if len(cambios_realizados) > 0:
    print("\n" + "="*80)
    print("GUARDANDO CAMBIOS")
    print("="*80)

    print(f"\nGuardando {EXCEL_FILE}...")
    wb.save(EXCEL_FILE)

    print("\n✅ CAMBIOS GUARDADOS EXITOSAMENTE")

    # Calcular impacto
    print("\n" + "="*80)
    print("IMPACTO EN SALDO PROMERICA USD 1774")
    print("="*80)

    # Balance duplicado
    if balance_duplicado_eliminado:
        balance_eliminado = sum([c['monto'] for c in cambios_realizados if c['tipo'] == 'BALANCE_DUPLICADO_ELIMINADO'])
        print(f"\nBalance duplicado eliminado: -${balance_eliminado:,.2f}")

    # Cuentas por cobrar
    if cuentas_cobrar_movidas > 0:
        total_cobrar = sum([c['monto'] for c in cambios_realizados if c['tipo'] == 'CUENTA_COBRAR_MOVIDA'])
        print(f"Cuentas por cobrar movidas: -${total_cobrar:,.2f}")

    # Cuentas por pagar
    if cuentas_pagar_movidas > 0:
        total_pagar = sum([c['monto'] for c in cambios_realizados if c['tipo'] == 'CUENTA_PAGAR_MOVIDA'])
        print(f"Cuentas por pagar movidas: +${total_pagar:,.2f} (egreso, así que suma)")

    # Total
    total_reduccion = 0
    if balance_duplicado_eliminado:
        total_reduccion += sum([c['monto'] for c in cambios_realizados if c['tipo'] == 'BALANCE_DUPLICADO_ELIMINADO'])
    if cuentas_cobrar_movidas > 0:
        total_reduccion += sum([c['monto'] for c in cambios_realizados if c['tipo'] == 'CUENTA_COBRAR_MOVIDA'])
    if cuentas_pagar_movidas > 0:
        total_reduccion -= sum([c['monto'] for c in cambios_realizados if c['tipo'] == 'CUENTA_PAGAR_MOVIDA'])

    print(f"\n{'='*40}")
    print(f"Reducción total del saldo: ${total_reduccion:,.2f}")
    print(f"{'='*40}")

    print(f"\nSaldo anterior: $13,173.69")
    print(f"Reducción: -${total_reduccion:,.2f}")
    print(f"Saldo esperado después: ${13173.69 - total_reduccion:,.2f}")

    print("\n" + "="*80)
    print("PRÓXIMOS PASOS")
    print("="*80)

    print("""
1. Ejecutar auditoría con alias para verificar cambios
2. El saldo de Promerica debería estar ahora cerca de $3,030.89 (extracto)
3. Las cuentas Por Cobrar y Por Pagar ahora tienen sus saldos correctos
4. Ejecutar: python scripts/auditoria_con_alias.py
""")

else:
    print("\n⚠️ No se encontraron cambios que realizar")
    print("   El Excel ya está correcto o los problemas fueron resueltos")

print("\n" + "="*80)
