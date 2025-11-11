#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
REGISTRO DE TRANSACCIONES - 10 NOVIEMBRE 2025
Registra 3 transacciones pendientes del día 10/11/2025
"""
import openpyxl
from datetime import datetime

EXCEL_FILE = "AlvaroVelasco_Finanzas_v2.0.xlsx"

# Transacciones a registrar
TRANSACCIONES = [
    {
        'num': 1,
        'fecha': datetime(2025, 11, 10, 17, 16, 59),
        'tipo': 'INGRESOS',
        'categoria': 'Ventas de Productos',
        'entidad': 'Cuentas por Cobrar',
        'cuenta': 'Promerica USD (40000003881774)',
        'cliente_proveedor': 'Gentra de Costa Rica',
        'concepto': 'Pago facturas electrónicas #02516, #02518, #02555',
        'referencia': '950456696',
        'monto_usd': 635.63,
        'monto_crc': None,
        'ingreso_egreso': 'Ingreso',
        'estado': 'Completado',
        'prioridad': 'Normal',
        'vencimiento': None,
        'notas': 'Transferencia SINPE - Se aplicará hoy 10pm',
    },
    {
        'num': 2,
        'fecha': datetime(2025, 11, 10, 11, 63, 10),  # Aproximado del número de referencia
        'tipo': 'COMPRAS PARA REVENTA',
        'categoria': 'Productos Tecnológicos',
        'entidad': 'Cuentas por Pagar',
        'cuenta': 'Promerica USD (40000003881774)',
        'cliente_proveedor': 'Intcomex Costa Rica',
        'concepto': 'Pago Factura Intcomex #821720 - AlvaroVelascoNet',
        'referencia': '2025111011631000083186200',
        'monto_usd': -3137.26,  # Negativo porque es egreso
        'monto_crc': None,
        'ingreso_egreso': 'Egreso',
        'estado': 'Pagado',
        'prioridad': 'Normal',
        'vencimiento': None,
        'notas': 'SINPE Programado - Doc: 2797944',
    },
    {
        'num': 3,
        'fecha': datetime(2025, 11, 10, 18, 33, 16),
        'tipo': 'GASTOS OPERATIVOS',
        'categoria': 'Combustible',
        'entidad': 'Vehículo',
        'cuenta': 'Efectivo/Contado',  # Condición venta = 01 (contado)
        'cliente_proveedor': 'Petróleos Delta CR',
        'concepto': 'Diesel 67.759 litros - Estación Tibas',
        'referencia': '50610112500310102878209800193010000008282300073203',
        'monto_usd': None,
        'monto_crc': -37606.00,  # Negativo porque es egreso
        'ingreso_egreso': 'Egreso',
        'estado': 'Pagado',
        'prioridad': 'Normal',
        'vencimiento': None,
        'notas': 'Factura #8282 - Petróleos Delta',
    },
]

def registrar_transacciones():
    """Registra las transacciones pendientes en Excel"""

    print("=" * 80)
    print("REGISTRO DE TRANSACCIONES - 10 NOVIEMBRE 2025")
    print("=" * 80)
    print()

    # Cargar Excel
    print("📂 Cargando Excel...")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb['TRANSACCIONES']
    print(f"✓ Hoja TRANSACCIONES cargada ({ws.max_row} filas actuales)")
    print()

    # Identificar columnas
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]

    # Mapeo de columnas
    col_map = {
        'Fecha': headers.index('Fecha') + 1,
        'Tipo Transacción': headers.index('Tipo Transacción') + 1,
        'Categoría': headers.index('Categoría') + 1,
        'Entidad': headers.index('Entidad') + 1,
        'Cuenta Bancaria': headers.index('Cuenta Bancaria') + 1,
        'Cliente/Proveedor': headers.index('Cliente/Proveedor') + 1,
        'Concepto': headers.index('Concepto') + 1,
        'Referencia': headers.index('Referencia') + 1,
        'Monto USD': headers.index('Monto USD') + 1,
        'Monto CRC': headers.index('Monto CRC') + 1,
        'Ingreso/Egreso': headers.index('Ingreso/Egreso') + 1,
        'Estado': headers.index('Estado') + 1,
        'Prioridad': headers.index('Prioridad') + 1,
        'Vencimiento': headers.index('Vencimiento') + 1,
        'Notas': headers.index('Notas') + 1,
    }

    print("=" * 80)
    print("REGISTRANDO TRANSACCIONES")
    print("=" * 80)
    print()

    filas_agregadas = []

    for trans in TRANSACCIONES:
        next_row = ws.max_row + 1

        # Escribir datos en columnas
        ws.cell(next_row, col_map['Fecha']).value = trans['fecha']
        ws.cell(next_row, col_map['Tipo Transacción']).value = trans['tipo']
        ws.cell(next_row, col_map['Categoría']).value = trans['categoria']
        ws.cell(next_row, col_map['Entidad']).value = trans['entidad']
        ws.cell(next_row, col_map['Cuenta Bancaria']).value = trans['cuenta']
        ws.cell(next_row, col_map['Cliente/Proveedor']).value = trans['cliente_proveedor']
        ws.cell(next_row, col_map['Concepto']).value = trans['concepto']
        ws.cell(next_row, col_map['Referencia']).value = trans['referencia']

        if trans['monto_usd']:
            ws.cell(next_row, col_map['Monto USD']).value = trans['monto_usd']
        if trans['monto_crc']:
            ws.cell(next_row, col_map['Monto CRC']).value = trans['monto_crc']

        ws.cell(next_row, col_map['Ingreso/Egreso']).value = trans['ingreso_egreso']
        ws.cell(next_row, col_map['Estado']).value = trans['estado']
        ws.cell(next_row, col_map['Prioridad']).value = trans['prioridad']

        if trans['vencimiento']:
            ws.cell(next_row, col_map['Vencimiento']).value = trans['vencimiento']

        ws.cell(next_row, col_map['Notas']).value = trans['notas']

        filas_agregadas.append({
            'num': trans['num'],
            'fila': next_row,
            'tipo': trans['tipo'],
            'concepto': trans['concepto'][:50],
            'monto_usd': trans['monto_usd'],
            'monto_crc': trans['monto_crc'],
        })

        print(f"✅ Transacción {trans['num']}: {trans['tipo']}")
        print(f"   Fila: {next_row}")
        print(f"   Concepto: {trans['concepto'][:60]}")
        if trans['monto_usd']:
            print(f"   Monto: ${trans['monto_usd']:,.2f} USD")
        if trans['monto_crc']:
            print(f"   Monto: ₡{trans['monto_crc']:,.2f} CRC")
        print()

    # Guardar
    print("💾 Guardando cambios...")
    wb.save(EXCEL_FILE)
    print("✅ Excel actualizado exitosamente")
    print()

    # Resumen
    print("=" * 80)
    print("📊 RESUMEN DE TRANSACCIONES REGISTRADAS")
    print("=" * 80)
    print()

    print(f"Total transacciones agregadas: {len(filas_agregadas)}")
    print()

    # Por tipo
    tipos_count = {}
    for t in filas_agregadas:
        tipo = t['tipo']
        if tipo not in tipos_count:
            tipos_count[tipo] = {'count': 0, 'usd': 0, 'crc': 0}
        tipos_count[tipo]['count'] += 1
        if t['monto_usd']:
            tipos_count[tipo]['usd'] += t['monto_usd']
        if t['monto_crc']:
            tipos_count[tipo]['crc'] += t['monto_crc']

    print("Por tipo:")
    for tipo, data in tipos_count.items():
        print(f"   • {tipo}: {data['count']} transacciones")
        if data['usd'] != 0:
            print(f"     USD: ${data['usd']:,.2f}")
        if data['crc'] != 0:
            print(f"     CRC: ₡{data['crc']:,.2f}")
    print()

    # Balance del día
    print("💰 BALANCE DEL DÍA (10/11/2025):")
    total_usd = sum(t['monto_usd'] for t in filas_agregadas if t['monto_usd'])
    total_crc = sum(t['monto_crc'] for t in filas_agregadas if t['monto_crc'])

    print(f"   USD: ${total_usd:,.2f}")
    print(f"   CRC: ₡{total_crc:,.2f}")
    print()

    if total_usd > 0:
        print(f"   ✅ Balance USD positivo: +${total_usd:,.2f}")
    elif total_usd < 0:
        print(f"   ⚠️  Balance USD negativo: ${total_usd:,.2f}")
    else:
        print(f"   ➖ Balance USD neutral")
    print()

    print("=" * 80)
    print("✅ REGISTRO COMPLETADO")
    print("=" * 80)
    print()

    print("📋 PRÓXIMOS PASOS:")
    print("   1. Abre el Excel y verifica las nuevas filas")
    print("   2. Revisa que los montos y categorías sean correctos")
    print("   3. Podemos continuar con análisis de utilidades mensuales")
    print()

    return True

if __name__ == "__main__":
    try:
        registrar_transacciones()
        print("🎉 Proceso completado exitosamente!")
    except FileNotFoundError:
        print(f"❌ ERROR: No se encontró el archivo {EXCEL_FILE}")
        print(f"   Asegúrate de ejecutar este script desde la carpeta del proyecto")
    except Exception as e:
        print(f"❌ ERROR INESPERADO: {e}")
        import traceback
        traceback.print_exc()
