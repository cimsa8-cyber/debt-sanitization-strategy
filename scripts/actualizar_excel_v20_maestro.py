#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SCRIPT MAESTRO - Actualizar Excel a V.20
Agrega todos los movimientos conciliados de 6 cuentas BNCR
Fecha: 10 de noviembre 2025
"""
import openpyxl
from datetime import datetime
import sys
import os

# Configuración
EXCEL_ORIGINAL = "AlvaroVelasco_Finanzas_v1.0_CORREGIDO.xlsx"
EXCEL_NUEVO = "AlvaroVelasco_Finanzas_v2.0.xlsx"
TC_APROXIMADA = 506.5  # Tasa de cambio promedio

# ============================================================================
# MOVIMIENTOS A AGREGAR - ORGANIZADOS POR CUENTA
# ============================================================================

movimientos = []

# ----------------------------------------------------------------------------
# 1. BNCR USD 601066 (EMPRESARIAL) - 14 movimientos
# ----------------------------------------------------------------------------
movimientos.extend([
    # 10/11/2025 - Planes de ahorro
    {"fecha": "10/11/2025", "tipo": "Transferencia Interna", "categoria": "Ahorro Personal", "entidad": "BNCR Plan Ahorro", "cuenta": "BNCR USD 601066", "proveedor": "Plan Impuestos Municipales", "concepto": "Aporte plan ahorro 01002273441", "referencia": "76949655", "monto_usd": 75.00, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "MEDIA", "notas": "Aporte automático quincenal"},
    {"fecha": "10/11/2025", "tipo": "Transferencia Interna", "categoria": "Ahorro Personal", "entidad": "BNCR Plan Ahorro", "cuenta": "BNCR USD 601066", "proveedor": "Plan Matrimonio", "concepto": "Aporte plan ahorro 01002335826", "referencia": "76950132", "monto_usd": 50.00, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "MEDIA", "notas": "Aporte automático quincenal"},
    {"fecha": "10/11/2025", "tipo": "Transferencia Interna", "categoria": "Ahorro Personal", "entidad": "BNCR Plan Ahorro", "cuenta": "BNCR USD 601066", "proveedor": "Plan Black Friday", "concepto": "Aporte plan ahorro 01002388223", "referencia": "76952525", "monto_usd": 25.00, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "MEDIA", "notas": "Aporte automático quincenal"},
    {"fecha": "10/11/2025", "tipo": "Transferencia Interna", "categoria": "Ahorro Personal", "entidad": "BNCR Plan Ahorro", "cuenta": "BNCR USD 601066", "proveedor": "Plan Nuevo", "concepto": "Aporte plan ahorro 17000002201", "referencia": "76954232", "monto_usd": 60.00, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "MEDIA", "notas": "Aporte automático quincenal"},

    # 07/11/2025
    {"fecha": "07/11/2025", "tipo": "Transferencia Interna", "categoria": "Cambio de Moneda", "entidad": "BNCR", "cuenta": "BNCR USD 601066", "proveedor": "BNCR", "concepto": "Conversión USD a CRC", "referencia": "98652306", "monto_usd": 60.48, "monto_crc": 30000, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "BAJA", "notas": "Conversión a cuenta CRC 188618. TC: ₡496.03/USD"},
    {"fecha": "07/11/2025", "tipo": "Pago", "categoria": "Servicios", "entidad": "Alejandra Arias", "cuenta": "BNCR USD 601066", "proveedor": "Alejandra Arias Fallas", "concepto": "Servicios de facturación y cobro semanal", "referencia": "34195590", "monto_usd": 25.00, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "ALTA", "notas": "Pago semanal servicios administrativos"},

    # 05/11/2025
    {"fecha": "05/11/2025", "tipo": "Gasto", "categoria": "Gastos Bancarios", "entidad": "BNCR", "cuenta": "BNCR USD 601066", "proveedor": "Banco Nacional", "concepto": "Comisión transferencia SINPE", "referencia": "16951587", "monto_usd": 3.02, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "BAJA", "notas": "Comisión TFT - Pago Siman"},
    {"fecha": "05/11/2025", "tipo": "Pago Proveedor", "categoria": "Pagos", "entidad": "Siman", "cuenta": "BNCR USD 601066", "proveedor": "Tarjeta Siman", "concepto": "Pago tarjeta crédito Siman", "referencia": "16951584", "monto_usd": 681.42, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "ALTA", "notas": "Pago TC Siman"},

    # 03/11/2025
    {"fecha": "03/11/2025", "tipo": "Pago", "categoria": "Tarjetas de Crédito", "entidad": "BNCR", "cuenta": "BNCR USD 601066", "proveedor": "TC BNCR MC 8759", "concepto": "Pago tarjeta Mastercard 8759", "referencia": "12961759", "monto_usd": 607.29, "monto_crc": 300000, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "ALTA", "notas": "Pago TC. Equivalente: ₡300,000. TC: ₡494.10/USD"},
    {"fecha": "03/11/2025", "tipo": "Pago", "categoria": "Tarjetas de Crédito", "entidad": "BNCR", "cuenta": "BNCR USD 601066", "proveedor": "TC BNCR Visa 6386", "concepto": "Pago tarjeta Visa 6386", "referencia": "12956019", "monto_usd": 321.86, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "ALTA", "notas": "Pago TC Visa 6386"},

    # 03/11/2025 - Planes de ahorro quincenales
    {"fecha": "03/11/2025", "tipo": "Transferencia Interna", "categoria": "Ahorro Personal", "entidad": "BNCR Plan Ahorro", "cuenta": "BNCR USD 601066", "proveedor": "Plan Impuestos Municipales", "concepto": "Aporte plan ahorro 01002273441", "referencia": "74976051", "monto_usd": 75.00, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "MEDIA", "notas": "Aporte automático quincenal"},
    {"fecha": "03/11/2025", "tipo": "Transferencia Interna", "categoria": "Ahorro Personal", "entidad": "BNCR Plan Ahorro", "cuenta": "BNCR USD 601066", "proveedor": "Plan Matrimonio", "concepto": "Aporte plan ahorro 01002335826", "referencia": "74976604", "monto_usd": 50.00, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "MEDIA", "notas": "Aporte automático quincenal"},
    {"fecha": "03/11/2025", "tipo": "Transferencia Interna", "categoria": "Ahorro Personal", "entidad": "BNCR Plan Ahorro", "cuenta": "BNCR USD 601066", "proveedor": "Plan Black Friday", "concepto": "Aporte plan ahorro 01002388223", "referencia": "74977457", "monto_usd": 25.00, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "MEDIA", "notas": "Aporte automático quincenal"},
    {"fecha": "03/11/2025", "tipo": "Transferencia Interna", "categoria": "Ahorro Personal", "entidad": "BNCR Plan Ahorro", "cuenta": "BNCR USD 601066", "proveedor": "Plan Nuevo", "concepto": "Aporte plan ahorro 17000002201", "referencia": "74978531", "monto_usd": 60.00, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "MEDIA", "notas": "Aporte automático quincenal"},
])

# ----------------------------------------------------------------------------
# 2. BNCR USD 11121 (PERSONAL) - 7 movimientos
# ----------------------------------------------------------------------------
movimientos.extend([
    # 03/11/2025
    {"fecha": "03/11/2025", "tipo": "Gasto", "categoria": "Alimentación", "entidad": "Comunidad PAS", "cuenta": "BNCR USD 11121", "proveedor": "Comunidad PAS", "concepto": "Cena PAS", "referencia": "90535048", "monto_usd": 50.00, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "BAJA", "notas": "Gasto personal - Cena comunidad"},

    # 31/10/2025
    {"fecha": "31/10/2025", "tipo": "Gasto", "categoria": "Entretenimiento", "entidad": "Feria", "cuenta": "BNCR USD 11121", "proveedor": "José Alejandro Alfaro", "concepto": "Feria 31 octubre Halloween", "referencia": "17250058", "monto_usd": 101.42, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "BAJA", "notas": "Gasto personal - Feria Halloween"},

    # 30/10/2025
    {"fecha": "30/10/2025", "tipo": "Transferencia Interna", "categoria": "Cambio de Moneda", "entidad": "BNCR", "cuenta": "BNCR USD 11121", "proveedor": "BNCR", "concepto": "Cambio de moneda", "referencia": "96755715", "monto_usd": 100.00, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "BAJA", "notas": "Cambio USD a otra moneda"},
    {"fecha": "30/10/2025", "tipo": "Ingreso", "categoria": "Reintegros", "entidad": "Auto Mercado", "cuenta": "BNCR USD 11121", "proveedor": "Auto Mercado", "concepto": "Reintegro compras oficina Auto Mercado", "referencia": "15151419", "monto_usd": 109.66, "monto_crc": 0, "tipo_mov": "Ingreso", "estado": "Completado", "prioridad": "MEDIA", "notas": "Reembolso gastos empresa"},
    {"fecha": "30/10/2025", "tipo": "Pago", "categoria": "Tarjetas de Crédito", "entidad": "BNCR", "cuenta": "BNCR USD 11121", "proveedor": "TC BNCR Visa 3519", "concepto": "Pago tarjeta Visa 3519", "referencia": "15137123", "monto_usd": 202.84, "monto_crc": 100000, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "ALTA", "notas": "Pago TC. Equivalente: ₡100,000. TC: ₡493.00/USD"},
    {"fecha": "30/10/2025", "tipo": "Pago", "categoria": "Tarjetas de Crédito", "entidad": "BNCR", "cuenta": "BNCR USD 11121", "proveedor": "TC BNCR Visa 3519", "concepto": "Pago tarjeta Visa 3519", "referencia": "15135145", "monto_usd": 139.72, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "ALTA", "notas": "Pago TC"},
    {"fecha": "30/10/2025", "tipo": "Ingreso", "categoria": "Salario", "entidad": "Empleador", "cuenta": "BNCR USD 11121", "proveedor": "Empleador", "concepto": "Salario 2da quincena octubre", "referencia": "15133894", "monto_usd": 500.00, "monto_crc": 0, "tipo_mov": "Ingreso", "estado": "Completado", "prioridad": "ALTA", "notas": "Ingreso quincenal"},
])

# ----------------------------------------------------------------------------
# 3. BNCR CRC 188618 - Movimientos clave
# ----------------------------------------------------------------------------
movimientos.extend([
    # 10/11/2025
    {"fecha": "10/11/2025", "tipo": "Gasto", "categoria": "Alimentación", "entidad": "McDonald's", "cuenta": "BNCR CRC 188618", "proveedor": "José Alejandro Alfaro", "concepto": "McDonald's", "referencia": "91351363", "monto_usd": round(12126/TC_APROXIMADA, 2), "monto_crc": 12126, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "BAJA", "notas": "Gasto personal"},
    {"fecha": "10/11/2025", "tipo": "Pago", "categoria": "Servicios", "entidad": "Alejandra Arias", "cuenta": "BNCR CRC 188618", "proveedor": "Alejandra Arias Fallas", "concepto": "Pagos Alejandra", "referencia": "93662715", "monto_usd": round(3000/TC_APROXIMADA, 2), "monto_crc": 3000, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "MEDIA", "notas": "Pago servicios"},

    # 07/11/2025
    {"fecha": "07/11/2025", "tipo": "Ingreso", "categoria": "Ingresos Varios", "entidad": "BNCR", "cuenta": "BNCR CRC 188618", "proveedor": "BNCR", "concepto": "Conversión USD a CRC", "referencia": "98652306", "monto_usd": 60.48, "monto_crc": 30000, "tipo_mov": "Ingreso", "estado": "Completado", "prioridad": "BAJA", "notas": "Recepción conversión desde USD 601066. TC: ₡496.03/USD"},
    {"fecha": "07/11/2025", "tipo": "Ingreso", "categoria": "Ingresos Varios", "entidad": "Thomas Davidovich", "cuenta": "BNCR CRC 188618", "proveedor": "Thomas Davidovich", "concepto": "Ingreso Thomas", "referencia": "58419485", "monto_usd": round(20000/TC_APROXIMADA, 2), "monto_crc": 20000, "tipo_mov": "Ingreso", "estado": "Completado", "prioridad": "BAJA", "notas": "Ingreso"},
    {"fecha": "07/11/2025", "tipo": "Gasto", "categoria": "Educación", "entidad": "Andrés Velasco", "cuenta": "BNCR CRC 188618", "proveedor": "Andrés Velasco Arana", "concepto": "Beca Cisco Andrés", "referencia": "54261209", "monto_usd": round(25000/TC_APROXIMADA, 2), "monto_crc": 25000, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "ALTA", "notas": "Pago beca universitaria"},
    {"fecha": "07/11/2025", "tipo": "Gasto", "categoria": "Vivienda", "entidad": "Pagos Casa", "cuenta": "BNCR CRC 188618", "proveedor": "Casa", "concepto": "Pago Casa", "referencia": "54545292", "monto_usd": round(10000/TC_APROXIMADA, 2), "monto_crc": 10000, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "ALTA", "notas": "Gasto vivienda"},

    # 05/11/2025
    {"fecha": "05/11/2025", "tipo": "Gasto", "categoria": "Personal", "entidad": "Adelanto Salario", "cuenta": "BNCR CRC 188618", "proveedor": "Álvaro Velasco", "concepto": "Adelanto de salario", "referencia": "96297548", "monto_usd": round(70000/TC_APROXIMADA, 2), "monto_crc": 70000, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "ALTA", "notas": "Adelanto salarial"},

    # 03/11/2025
    {"fecha": "03/11/2025", "tipo": "Pago", "categoria": "Tarjetas de Crédito", "entidad": "BNCR", "cuenta": "BNCR CRC 188618", "proveedor": "TC BNCR Visa 9837", "concepto": "Pago tarjeta Visa 9837", "referencia": "74750901", "monto_usd": 62.00, "monto_crc": 31434, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "ALTA", "notas": "Pago débito automático. Equivalente: $62 USD. TC: ₡506.97/USD"},
    {"fecha": "03/11/2025", "tipo": "Pago", "categoria": "Tarjetas de Crédito", "entidad": "BNCR", "cuenta": "BNCR CRC 188618", "proveedor": "TC BNCR Visa 9837", "concepto": "Pago tarjeta Visa 9837", "referencia": "74750898", "monto_usd": round(271600/TC_APROXIMADA, 2), "monto_crc": 271600, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "ALTA", "notas": "Pago débito automático"},
    {"fecha": "03/11/2025", "tipo": "Pago", "categoria": "Tarjetas de Crédito", "entidad": "BNCR", "cuenta": "BNCR CRC 188618", "proveedor": "TC BNCR MC 8759", "concepto": "Pago tarjeta MC 8759", "referencia": "74750907", "monto_usd": round(5070/TC_APROXIMADA, 2), "monto_crc": 5070, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "ALTA", "notas": "Pago TC"},
    {"fecha": "03/11/2025", "tipo": "Gasto", "categoria": "Personal", "entidad": "Andrés Velasco", "cuenta": "BNCR CRC 188618", "proveedor": "Andrés Velasco", "concepto": "Mesada semanal Andrés", "referencia": "9048950", "monto_usd": round(5000/TC_APROXIMADA, 2), "monto_crc": 5000, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "MEDIA", "notas": "Mesada semanal"},

    # 31/10/2025
    {"fecha": "31/10/2025", "tipo": "Ingreso", "categoria": "Ingresos Varios", "entidad": "BNCR", "cuenta": "BNCR CRC 188618", "proveedor": "BNCR", "concepto": "Conversión USD a CRC", "referencia": "98652306", "monto_usd": 60.85, "monto_crc": 30000, "tipo_mov": "Ingreso", "estado": "Completado", "prioridad": "BAJA", "notas": "Recepción conversión desde USD 601066. TC: ₡493.09/USD"},
    {"fecha": "31/10/2025", "tipo": "Ingreso", "categoria": "Ahorro Personal", "entidad": "BNCR", "cuenta": "BNCR CRC 188618", "proveedor": "BNCR", "concepto": "Liquidación Plan Ahorro 01002061574", "referencia": "312742", "monto_usd": round(386080.45/TC_APROXIMADA, 2), "monto_crc": 386080.45, "tipo_mov": "Ingreso", "estado": "Completado", "prioridad": "ALTA", "notas": "Liquidación automática plan ahorro"},
    {"fecha": "31/10/2025", "tipo": "Pago", "categoria": "Personal", "entidad": "Rosdeyli Salome", "cuenta": "BNCR CRC 188618", "proveedor": "Rosdeyli Salome Lopez", "concepto": "Pago Rosdeyli", "referencia": "54748812", "monto_usd": round(36000/TC_APROXIMADA, 2), "monto_crc": 36000, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "MEDIA", "notas": "Pago personal"},
    {"fecha": "31/10/2025", "tipo": "Ingreso", "categoria": "Ingresos Varios", "entidad": "José Alejandro Alfaro", "cuenta": "BNCR CRC 188618", "proveedor": "José Alejandro Alfaro", "concepto": "Ingreso Gas", "referencia": "97756190", "monto_usd": round(2000/TC_APROXIMADA, 2), "monto_crc": 2000, "tipo_mov": "Ingreso", "estado": "Completado", "prioridad": "BAJA", "notas": "Ingreso varios"},
])

# ----------------------------------------------------------------------------
# 4. TC BNCR VISA 3519 - Movimientos adicionales
# ----------------------------------------------------------------------------
movimientos.extend([
    # 05/11/2025
    {"fecha": "05/11/2025", "tipo": "Gasto", "categoria": "Alimentación", "entidad": "Uber Eats", "cuenta": "Tarjeta BNCR Visa 3519", "proveedor": "Uber Eats", "concepto": "Uber Eats vía PayPal", "referencia": "742570", "monto_usd": 21.16, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "BAJA", "notas": "Pedido Uber Eats $18.73 + IVA $2.43"},

    # 03/11/2025
    {"fecha": "03/11/2025", "tipo": "Gasto", "categoria": "Supermercado", "entidad": "Auto Mercado", "cuenta": "Tarjeta BNCR Visa 3519", "proveedor": "Auto Mercado Online", "concepto": "Compra Auto Mercado", "referencia": "492231", "monto_usd": round(6774.35/TC_APROXIMADA, 2), "monto_crc": 6774.35, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "BAJA", "notas": "Compra supermercado"},

    # 02/11/2025
    {"fecha": "02/11/2025", "tipo": "Gasto", "categoria": "Alimentación", "entidad": "Uber Eats", "cuenta": "Tarjeta BNCR Visa 3519", "proveedor": "Uber Eats", "concepto": "Uber Eats vía PayPal", "referencia": "454169", "monto_usd": 30.06, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "BAJA", "notas": "Pedido Uber Eats $26.60 + IVA $3.46"},

    # 01/11/2025
    {"fecha": "01/11/2025", "tipo": "Gasto", "categoria": "Vivienda", "entidad": "Comunidad PAS", "cuenta": "Tarjeta BNCR Visa 3519", "proveedor": "Comunidad PAS", "concepto": "Cuota mantenimiento comunidad", "referencia": "367716", "monto_usd": round(5700/TC_APROXIMADA, 2), "monto_crc": 5700, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "ALTA", "notas": "Cuota comunidad"},
    {"fecha": "01/11/2025", "tipo": "Gasto", "categoria": "Alimentación", "entidad": "La Frikitona", "cuenta": "Tarjeta BNCR Visa 3519", "proveedor": "La Frikitona", "concepto": "Restaurante La Frikitona", "referencia": "388577", "monto_usd": round(22800/TC_APROXIMADA, 2), "monto_crc": 22800, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "BAJA", "notas": "Gasto alimentación"},
    {"fecha": "01/11/2025", "tipo": "Gasto", "categoria": "Alimentación", "entidad": "La Frikitona", "cuenta": "Tarjeta BNCR Visa 3519", "proveedor": "La Frikitona", "concepto": "Restaurante La Frikitona", "referencia": "392646", "monto_usd": round(3750/TC_APROXIMADA, 2), "monto_crc": 3750, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "BAJA", "notas": "Gasto alimentación"},

    # 30/10/2025
    {"fecha": "30/10/2025", "tipo": "Gasto", "categoria": "Supermercado", "entidad": "Tienda Pronto", "cuenta": "Tarjeta BNCR Visa 3519", "proveedor": "Tienda Pronto Barreal", "concepto": "Compra Tienda Pronto", "referencia": "197982", "monto_usd": round(6550/TC_APROXIMADA, 2), "monto_crc": 6550, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "BAJA", "notas": "Compra supermercado"},
    {"fecha": "30/10/2025", "tipo": "Gasto", "categoria": "Supermercado", "entidad": "Auto Mercado", "cuenta": "Tarjeta BNCR Visa 3519", "proveedor": "Auto Mercado Online", "concepto": "Compra Auto Mercado", "referencia": "115950", "monto_usd": round(54010/TC_APROXIMADA, 2), "monto_crc": 54010, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "BAJA", "notas": "Compra supermercado"},
    {"fecha": "30/10/2025", "tipo": "Gasto", "categoria": "Supermercado", "entidad": "Auto Mercado", "cuenta": "Tarjeta BNCR Visa 3519", "proveedor": "Auto Mercado Online", "concepto": "Compra Auto Mercado", "referencia": "117529", "monto_usd": round(122605/TC_APROXIMADA, 2), "monto_crc": 122605, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "BAJA", "notas": "Compra supermercado grande"},
    {"fecha": "30/10/2025", "tipo": "Gasto", "categoria": "Supermercado", "entidad": "Auto Mercado", "cuenta": "Tarjeta BNCR Visa 3519", "proveedor": "Auto Mercado Online", "concepto": "Compra Auto Mercado", "referencia": "117336", "monto_usd": round(50/TC_APROXIMADA, 2), "monto_crc": 50, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "BAJA", "notas": "Compra pequeña"},
    {"fecha": "30/10/2025", "tipo": "Devolución", "categoria": "Supermercado", "entidad": "Auto Mercado", "cuenta": "Tarjeta BNCR Visa 3519", "proveedor": "Auto Mercado Online", "concepto": "Devolución Auto Mercado", "referencia": "119128", "monto_usd": round(-3350/TC_APROXIMADA, 2), "monto_crc": -3350, "tipo_mov": "Ingreso", "estado": "Completado", "prioridad": "BAJA", "notas": "Devolución productos"},

    # 27/10/2025
    {"fecha": "27/10/2025", "tipo": "Gasto", "categoria": "Tecnología", "entidad": "Amazon", "cuenta": "Tarjeta BNCR Visa 3519", "proveedor": "Amazon", "concepto": "Compra Amazon", "referencia": "615169", "monto_usd": 89.71, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "MEDIA", "notas": "Compra online Amazon"},

    # 25/10/2025
    {"fecha": "25/10/2025", "tipo": "Gasto", "categoria": "Supermercado", "entidad": "AM PM", "cuenta": "Tarjeta BNCR Visa 3519", "proveedor": "Multimercado AM PM", "concepto": "Compra AM PM", "referencia": "635394", "monto_usd": round(3200/TC_APROXIMADA, 2), "monto_crc": 3200, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "BAJA", "notas": "Compra supermercado"},
    {"fecha": "25/10/2025", "tipo": "Gasto", "categoria": "Supermercado", "entidad": "Mini Super", "cuenta": "Tarjeta BNCR Visa 3519", "proveedor": "Mini Super San Agustín", "concepto": "Compra Mini Super", "referencia": "670264", "monto_usd": round(600/TC_APROXIMADA, 2), "monto_crc": 600, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "BAJA", "notas": "Compra supermercado"},
    {"fecha": "25/10/2025", "tipo": "Gasto", "categoria": "Supermercado", "entidad": "Abastecedor La Esquina", "cuenta": "Tarjeta BNCR Visa 3519", "proveedor": "Abastecedor La Esquina", "concepto": "Compra Abastecedor", "referencia": "659870", "monto_usd": round(2400/TC_APROXIMADA, 2), "monto_crc": 2400, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "BAJA", "notas": "Compra supermercado"},

    # 22/10/2025
    {"fecha": "22/10/2025", "tipo": "Gasto", "categoria": "Tecnología", "entidad": "Amazon", "cuenta": "Tarjeta BNCR Visa 3519", "proveedor": "Amazon", "concepto": "Compra Amazon", "referencia": "123612", "monto_usd": 38.51, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "MEDIA", "notas": "Compra online Amazon"},
    {"fecha": "22/10/2025", "tipo": "Gasto", "categoria": "Transporte", "entidad": "Uber", "cuenta": "Tarjeta BNCR Visa 3519", "proveedor": "Uber", "concepto": "Transporte Uber vía PayPal", "referencia": "376751", "monto_usd": 7.33, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "BAJA", "notas": "Uber $6.49 + IVA $0.84"},
])

# ----------------------------------------------------------------------------
# 5. TC BNCR MC 8759 - Movimientos adicionales
# ----------------------------------------------------------------------------
movimientos.extend([
    # 01/11/2025
    {"fecha": "01/11/2025", "tipo": "Gasto", "categoria": "Tecnología", "entidad": "Anthropic", "cuenta": "Tarjeta BNCR MC 8759", "proveedor": "Anthropic", "concepto": "Suscripción Claude Pro", "referencia": "356611", "monto_usd": 10.00, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "MEDIA", "notas": "Suscripción mensual gasto operativo"},

    # 25/10/2025
    {"fecha": "25/10/2025", "tipo": "Gasto", "categoria": "Tecnología", "entidad": "Apple", "cuenta": "Tarjeta BNCR MC 8759", "proveedor": "Apple Inc", "concepto": "Apple.com compra/suscripción", "referencia": "556829", "monto_usd": 39.99, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "MEDIA", "notas": "Compra Apple (iCloud/apps)"},

    # 18/10/2025
    {"fecha": "18/10/2025", "tipo": "Gasto", "categoria": "Entretenimiento", "entidad": "Netflix", "cuenta": "Tarjeta BNCR MC 8759", "proveedor": "Netflix", "concepto": "Suscripción Netflix", "referencia": "845031", "monto_usd": 21.45, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "BAJA", "notas": "Netflix $18.98 + IVA $2.47"},
])

# ----------------------------------------------------------------------------
# 6. TC BNCR VISA 9837 - Movimientos adicionales
# ----------------------------------------------------------------------------
movimientos.extend([
    # 22/10/2025
    {"fecha": "22/10/2025", "tipo": "Gasto", "categoria": "Tecnología", "entidad": "Cleverbridge", "cuenta": "Tarjeta BNCR Visa 9837", "proveedor": "PayPal Cleverbridge", "concepto": "Software empresarial vía PayPal", "referencia": "174348", "monto_usd": 275.00, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "ALTA", "notas": "Compra software/licencia"},
])

print("="*80)
print("SCRIPT MAESTRO - ACTUALIZACIÓN EXCEL V.20")
print("="*80)
print(f"\nTotal movimientos preparados: {len(movimientos)}")
print(f"Archivo origen: {EXCEL_ORIGINAL}")
print(f"Archivo destino: {EXCEL_NUEVO}")

# Verificar que el archivo existe
if not os.path.exists(EXCEL_ORIGINAL):
    print(f"\n❌ ERROR: No se encontró el archivo {EXCEL_ORIGINAL}")
    print("   Asegúrese de que el archivo esté en el directorio actual.")
    sys.exit(1)

# Cargar el Excel
print("\n📂 Cargando Excel...")
wb = openpyxl.load_workbook(EXCEL_ORIGINAL)
ws = wb['TRANSACCIONES']

# Obtener último ID y última fila
last_id = 0
last_row = 1
print("\n🔍 Analizando Excel existente...")

for row in range(2, ws.max_row + 1):
    id_val = ws[f'P{row}'].value
    if id_val:
        try:
            last_id = max(last_id, int(id_val))
            last_row = row
        except:
            pass

print(f"   Último ID encontrado: {last_id}")
print(f"   Última fila con datos: {last_row}")

# Crear conjunto de referencias existentes para evitar duplicados
referencias_existentes = set()
for row in range(2, ws.max_row + 1):
    ref = ws[f'H{row}'].value
    fecha = ws[f'A{row}'].value
    if ref and fecha:
        try:
            fecha_str = fecha.strftime('%Y-%m-%d') if hasattr(fecha, 'strftime') else str(fecha)
            referencias_existentes.add(f"{fecha_str}_{ref}")
        except:
            pass

print(f"   Referencias únicas existentes: {len(referencias_existentes)}")

# Filtrar movimientos que no existen
movimientos_nuevos = []
movimientos_duplicados = []

print("\n🔍 Verificando duplicados...")
for mov in movimientos:
    try:
        fecha_obj = datetime.strptime(mov['fecha'], '%d/%m/%Y')
        fecha_str = fecha_obj.strftime('%Y-%m-%d')
        clave = f"{fecha_str}_{mov['referencia']}"

        if clave in referencias_existentes:
            movimientos_duplicados.append(mov)
        else:
            movimientos_nuevos.append(mov)
    except:
        # Si hay error, agregar de todos modos
        movimientos_nuevos.append(mov)

print(f"   Movimientos nuevos a agregar: {len(movimientos_nuevos)}")
print(f"   Movimientos duplicados (omitidos): {len(movimientos_duplicados)}")

if len(movimientos_nuevos) == 0:
    print("\n✅ No hay movimientos nuevos que agregar.")
    print("   Todos los movimientos ya existen en el Excel.")
    sys.exit(0)

# Agregar movimientos nuevos
print(f"\n📝 Agregando {len(movimientos_nuevos)} movimientos nuevos...")

next_id = last_id + 1
next_row = last_row + 1

movimientos_agregados = []

for mov in movimientos_nuevos:
    try:
        # Convertir fecha
        fecha_obj = datetime.strptime(mov['fecha'], '%d/%m/%Y')

        # Agregar en la siguiente fila
        ws[f'A{next_row}'] = fecha_obj
        ws[f'B{next_row}'] = mov['tipo']
        ws[f'C{next_row}'] = mov['categoria']
        ws[f'D{next_row}'] = mov['entidad']
        ws[f'E{next_row}'] = mov['cuenta']
        ws[f'F{next_row}'] = mov['proveedor']
        ws[f'G{next_row}'] = mov['concepto']
        ws[f'H{next_row}'] = mov['referencia']
        ws[f'I{next_row}'] = mov['monto_usd']
        ws[f'J{next_row}'] = mov['monto_crc']
        ws[f'K{next_row}'] = mov['tipo_mov']
        ws[f'L{next_row}'] = mov['estado']
        ws[f'M{next_row}'] = mov.get('prioridad', 'MEDIA')
        ws[f'N{next_row}'] = ''  # Vencimiento vacío
        ws[f'O{next_row}'] = mov.get('notas', '')
        ws[f'P{next_row}'] = next_id
        ws[f'Q{next_row}'] = datetime.now()
        ws[f'R{next_row}'] = 'Álvaro Velasco'
        ws[f'S{next_row}'] = False  # Duplicado
        ws[f'T{next_row}'] = 'OK'  # Validación

        movimientos_agregados.append({
            'id': next_id,
            'fila': next_row,
            'fecha': mov['fecha'],
            'concepto': mov['concepto'],
            'monto_usd': mov['monto_usd'],
            'monto_crc': mov['monto_crc']
        })

        next_row += 1
        next_id += 1

    except Exception as e:
        print(f"   ⚠️ Error al agregar movimiento: {mov.get('concepto', 'desconocido')} - {e}")

# Guardar archivo
print(f"\n💾 Guardando archivo como {EXCEL_NUEVO}...")
wb.save(EXCEL_NUEVO)

print("\n" + "="*80)
print("✅ ACTUALIZACIÓN COMPLETADA")
print("="*80)

print(f"\n📊 RESUMEN:")
print(f"   Total movimientos procesados: {len(movimientos)}")
print(f"   Movimientos nuevos agregados: {len(movimientos_agregados)}")
print(f"   Movimientos duplicados omitidos: {len(movimientos_duplicados)}")
print(f"   Archivo guardado: {EXCEL_NUEVO}")

print(f"\n📋 MOVIMIENTOS AGREGADOS POR CUENTA:")
cuentas_resumen = {}
for mov in movimientos_agregados:
    # Buscar cuenta original
    mov_original = next((m for m in movimientos_nuevos if m['concepto'] == mov['concepto'] and m['fecha'] == mov['fecha']), None)
    if mov_original:
        cuenta = mov_original['cuenta']
        if cuenta not in cuentas_resumen:
            cuentas_resumen[cuenta] = {'count': 0, 'total_usd': 0, 'total_crc': 0}
        cuentas_resumen[cuenta]['count'] += 1
        cuentas_resumen[cuenta]['total_usd'] += mov['monto_usd']
        cuentas_resumen[cuenta]['total_crc'] += mov['monto_crc']

for cuenta, datos in sorted(cuentas_resumen.items()):
    print(f"\n   {cuenta}:")
    print(f"      Movimientos: {datos['count']}")
    if datos['total_usd'] != 0:
        print(f"      Total USD: ${datos['total_usd']:,.2f}")
    if datos['total_crc'] != 0:
        print(f"      Total CRC: ₡{datos['total_crc']:,.2f}")

print(f"\n📝 DETALLE DE MOVIMIENTOS AGREGADOS:")
print("-" * 80)
for mov in movimientos_agregados[:20]:  # Primeros 20
    monto_str = f"${mov['monto_usd']:.2f}"
    if mov['monto_crc'] != 0:
        monto_str += f" / ₡{mov['monto_crc']:,.2f}"
    print(f"   Fila {mov['fila']} (ID {mov['id']}): {mov['fecha']} | {monto_str:20} | {mov['concepto'][:50]}")

if len(movimientos_agregados) > 20:
    print(f"   ... y {len(movimientos_agregados) - 20} movimientos más")

print("\n" + "="*80)
print("🎉 EXCEL ACTUALIZADO A VERSIÓN 2.0")
print("="*80)
print(f"\n✅ Siguiente paso: Abrir {EXCEL_NUEVO} y verificar los datos")
print("✅ Las hojas Dashboard, Efectivo, A_P, etc. se actualizarán automáticamente")
print("\n" + "="*80)
