#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Agregar movimientos del 10 de noviembre 2025
1. Uber/PayPal - Tarjeta 3519 - $2.21
2. Pago INTCOMEX - Promerica 1774 - $3,137.26
3. Comisión diferida - Promerica 1774 - $0.75
4. Actualizar factura INTCOMEX fila 98 a "Pagada"
"""
import openpyxl
from datetime import datetime
import sys

if sys.platform == 'win32':
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')

EXCEL_PATH = r"C:\Users\Alvaro Velasco\Desktop\debt-sanitization-strategy\AlvaroVelasco_Finanzas_v1.0_CORREGIDO.xlsx"

print("=" * 80)
print("AGREGANDO MOVIMIENTOS DEL 10 DE NOVIEMBRE 2025")
print("=" * 80)

wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb['TRANSACCIONES']

# Encontrar último ID
ultimo_id = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[15]:  # Columna P = ID
        try:
            id_val = int(row[15])
            if id_val > ultimo_id:
                ultimo_id = id_val
        except:
            pass

print(f"\n[1] Último ID encontrado: {ultimo_id}")
print(f"    Siguiente ID a usar: {ultimo_id + 1}")
print(f"    Última fila: {ws.max_row}")

# Nueva fila para agregar
nueva_fila = ws.max_row + 1

# MOVIMIENTO 1: Uber/PayPal - Tarjeta 3519 - $2.21
print(f"\n[2] Agregando MOVIMIENTO 1 en fila {nueva_fila}:")
print(f"    Uber/PayPal - Tarjeta 3519 - $2.21")

ws.cell(row=nueva_fila, column=1).value = datetime(2025, 11, 10)  # A: Fecha
ws.cell(row=nueva_fila, column=1).number_format = 'DD/MM/YYYY'
ws.cell(row=nueva_fila, column=2).value = "Gasto"  # B: Tipo
ws.cell(row=nueva_fila, column=3).value = "Transporte"  # C: Categoría
ws.cell(row=nueva_fila, column=4).value = "PayPal"  # D: Entidad
ws.cell(row=nueva_fila, column=5).value = "Tarjeta BNCR Visa 3519"  # E: Cuenta Bancaria
ws.cell(row=nueva_fila, column=6).value = "Uber"  # F: Cliente/Proveedor
ws.cell(row=nueva_fila, column=7).value = "Transporte Uber vía PayPal"  # G: Concepto
ws.cell(row=nueva_fila, column=8).value = "531413364318"  # H: Referencia
ws.cell(row=nueva_fila, column=9).value = 2.21  # I: Monto USD
ws.cell(row=nueva_fila, column=9).number_format = '#,##0.00'
ws.cell(row=nueva_fila, column=10).value = 0  # J: Monto CRC
ws.cell(row=nueva_fila, column=11).value = "Egreso"  # K: Ingreso/Egreso
ws.cell(row=nueva_fila, column=12).value = "Completado"  # L: Estado
ws.cell(row=nueva_fila, column=13).value = "BAJA"  # M: Prioridad
ws.cell(row=nueva_fila, column=14).value = None  # N: Vencimiento
ws.cell(row=nueva_fila, column=15).value = "Pago PayPal cargado a Visa 3519"  # O: Notas
ws.cell(row=nueva_fila, column=16).value = ultimo_id + 1  # P: ID Transacción
ws.cell(row=nueva_fila, column=17).value = datetime(2025, 11, 10)  # Q: Fecha Creación
ws.cell(row=nueva_fila, column=17).number_format = 'DD/MM/YYYY'
ws.cell(row=nueva_fila, column=18).value = "Álvaro Velasco"  # R: Usuario

nueva_fila += 1
ultimo_id += 1

# MOVIMIENTO 2: Pago INTCOMEX - Promerica 1774 - $3,137.26
print(f"\n[3] Agregando MOVIMIENTO 2 en fila {nueva_fila}:")
print(f"    Pago INTCOMEX - Promerica 1774 - $3,137.26")

ws.cell(row=nueva_fila, column=1).value = datetime(2025, 11, 10)  # A: Fecha
ws.cell(row=nueva_fila, column=1).number_format = 'DD/MM/YYYY'
ws.cell(row=nueva_fila, column=2).value = "Pago Proveedor"  # B: Tipo
ws.cell(row=nueva_fila, column=3).value = "Compras"  # C: Categoría
ws.cell(row=nueva_fila, column=4).value = "STCR Costa Rica Trust"  # D: Entidad
ws.cell(row=nueva_fila, column=5).value = "Promerica USD 1774"  # E: Cuenta Bancaria
ws.cell(row=nueva_fila, column=6).value = "INTCOMEX Costa Rica"  # F: Cliente/Proveedor
ws.cell(row=nueva_fila, column=7).value = "Pago Factura #2509011 (programado)"  # G: Concepto
ws.cell(row=nueva_fila, column=8).value = "5588600"  # H: Referencia
ws.cell(row=nueva_fila, column=9).value = 3137.26  # I: Monto USD
ws.cell(row=nueva_fila, column=9).number_format = '#,##0.00'
ws.cell(row=nueva_fila, column=10).value = 0  # J: Monto CRC
ws.cell(row=nueva_fila, column=11).value = "Egreso"  # K: Ingreso/Egreso
ws.cell(row=nueva_fila, column=12).value = "Completado"  # L: Estado
ws.cell(row=nueva_fila, column=13).value = "CRÍTICA"  # M: Prioridad
ws.cell(row=nueva_fila, column=14).value = None  # N: Vencimiento
ws.cell(row=nueva_fila, column=15).value = "Pago programado 10/11/2025 22:00 via STCR Scotiabank"  # O: Notas
ws.cell(row=nueva_fila, column=16).value = ultimo_id + 1  # P: ID Transacción
ws.cell(row=nueva_fila, column=17).value = datetime(2025, 11, 10)  # Q: Fecha Creación
ws.cell(row=nueva_fila, column=17).number_format = 'DD/MM/YYYY'
ws.cell(row=nueva_fila, column=18).value = "Álvaro Velasco"  # R: Usuario

nueva_fila += 1
ultimo_id += 1

# MOVIMIENTO 3: Comisión transacción diferida - $0.75
print(f"\n[4] Agregando MOVIMIENTO 3 en fila {nueva_fila}:")
print(f"    Comisión transacción diferida - $0.75")

ws.cell(row=nueva_fila, column=1).value = datetime(2025, 11, 10)  # A: Fecha
ws.cell(row=nueva_fila, column=1).number_format = 'DD/MM/YYYY'
ws.cell(row=nueva_fila, column=2).value = "Gasto"  # B: Tipo
ws.cell(row=nueva_fila, column=3).value = "Gastos Bancarios"  # C: Categoría
ws.cell(row=nueva_fila, column=4).value = "Banco Promerica"  # D: Entidad
ws.cell(row=nueva_fila, column=5).value = "Promerica USD 1774"  # E: Cuenta Bancaria
ws.cell(row=nueva_fila, column=6).value = "Banco Promerica"  # F: Cliente/Proveedor
ws.cell(row=nueva_fila, column=7).value = "Comisión transacción programada/diferida"  # G: Concepto
ws.cell(row=nueva_fila, column=8).value = "5588600-COM"  # H: Referencia
ws.cell(row=nueva_fila, column=9).value = 0.75  # I: Monto USD
ws.cell(row=nueva_fila, column=9).number_format = '#,##0.00'
ws.cell(row=nueva_fila, column=10).value = 0  # J: Monto CRC
ws.cell(row=nueva_fila, column=11).value = "Egreso"  # K: Ingreso/Egreso
ws.cell(row=nueva_fila, column=12).value = "Completado"  # L: Estado
ws.cell(row=nueva_fila, column=13).value = "BAJA"  # M: Prioridad
ws.cell(row=nueva_fila, column=14).value = None  # N: Vencimiento
ws.cell(row=nueva_fila, column=15).value = "Comisión por pago programado INTCOMEX"  # O: Notas
ws.cell(row=nueva_fila, column=16).value = ultimo_id + 1  # P: ID Transacción
ws.cell(row=nueva_fila, column=17).value = datetime(2025, 11, 10)  # Q: Fecha Creación
ws.cell(row=nueva_fila, column=17).number_format = 'DD/MM/YYYY'
ws.cell(row=nueva_fila, column=18).value = "Álvaro Velasco"  # R: Usuario

# ACTUALIZACIÓN: Cambiar estado de factura INTCOMEX fila 98 a "Pagada"
print(f"\n[5] Actualizando estado de factura INTCOMEX en fila 98:")
print(f"    Estado: Pendiente → Pagada")

ws.cell(row=98, column=12).value = "Pagada"  # L: Estado
ws.cell(row=98, column=15).value = "Pagada el 10/11/2025 - Promerica 1774 - Ref: 5588600"  # O: Notas

print(f"\n[6] Guardando Excel...")
wb.save(EXCEL_PATH)
wb.close()

print(f"\n✓ COMPLETADO")
print("=" * 80)
print("RESUMEN DE CAMBIOS:")
print("=" * 80)
print(f"  ✓ Fila 101: Uber/PayPal $2.21 (ID 64)")
print(f"  ✓ Fila 102: Pago INTCOMEX $3,137.26 (ID 65)")
print(f"  ✓ Fila 103: Comisión bancaria $0.75 (ID 66)")
print(f"  ✓ Fila 98: Factura INTCOMEX marcada como 'Pagada'")
print(f"\nTotal agregado: 3 nuevas transacciones")
print(f"Total egresos: $3,140.22")
print("\n" + "=" * 80)
