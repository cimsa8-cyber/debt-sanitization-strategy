#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CORREGIR SIGNOS DE EGRESOS
Cambia el signo de egresos positivos a negativo
"""
import openpyxl
from datetime import datetime
import shutil

EXCEL_FILE = "AlvaroVelasco_Finanzas_v2.0.xlsx"
BACKUP_FILE = f"AlvaroVelasco_Finanzas_v2.0_CORREGIR_SIGNOS_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

def crear_backup():
    print("=" * 80)
    print("CREANDO BACKUP")
    print("=" * 80)
    print(f"Backup: {BACKUP_FILE}")
    try:
        shutil.copy2(EXCEL_FILE, BACKUP_FILE)
        print("✅ Backup creado")
        print()
        return True
    except Exception as e:
        print(f"❌ ERROR: {e}")
        return False

def corregir_signos():
    print("=" * 80)
    print("CORRECCIÓN DE SIGNOS EN EGRESOS")
    print("=" * 80)
    print()

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb['TRANSACCIONES']

    # Encontrar columnas
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    col_map = {}
    for col in range(1, len(headers) + 1):
        if headers[col-1]:
            col_map[headers[col-1]] = col

    print("🔍 Identificando egresos con signo positivo...")
    print()

    # Identificar egresos con signo positivo
    egresos_a_corregir = []

    for row in range(2, ws.max_row + 1):
        cuenta = ws.cell(row, col_map['Cuenta Bancaria']).value
        monto = ws.cell(row, col_map['Monto USD']).value
        ing_egr = ws.cell(row, col_map['Ingreso/Egreso']).value
        concepto = ws.cell(row, col_map['Concepto']).value

        # Solo Promerica USD (o cualquier cuenta donde el problema exista)
        if cuenta and 'Promerica USD (40000003881774)' in str(cuenta):
            if ing_egr == 'Egreso' and monto:
                monto_val = float(monto)

                if monto_val > 0:
                    egresos_a_corregir.append({
                        'fila': row,
                        'monto_actual': monto_val,
                        'monto_correcto': -monto_val,
                        'concepto': concepto
                    })

    print(f"⚠️  Egresos a corregir: {len(egresos_a_corregir)}")
    print()

    if len(egresos_a_corregir) == 0:
        print("✅ No hay egresos que corregir - todos tienen signo correcto")
        return False

    # Mostrar primeros 10
    print("📋 Ejemplos (primeros 10):")
    print()

    for egreso in egresos_a_corregir[:10]:
        print(f"Fila {egreso['fila']}: ${egreso['monto_actual']:,.2f} → -${egreso['monto_actual']:,.2f}")
        print(f"   {egreso['concepto'][:50] if egreso['concepto'] else 'N/A'}...")

    if len(egresos_a_corregir) > 10:
        print(f"   ... y {len(egresos_a_corregir) - 10} más")

    print()

    # =========================================================================
    # APLICAR CORRECCIONES
    # =========================================================================
    print("=" * 80)
    print("✏️  Aplicando correcciones...")
    print("=" * 80)
    print()

    total_corregido = 0
    suma_antes = sum(e['monto_actual'] for e in egresos_a_corregir)

    for egreso in egresos_a_corregir:
        fila = egreso['fila']
        ws.cell(fila, col_map['Monto USD']).value = egreso['monto_correcto']
        total_corregido += 1

    print(f"✅ {total_corregido} egresos corregidos")
    print(f"   Total antes: +${suma_antes:,.2f}")
    print(f"   Total después: -${suma_antes:,.2f}")
    print()

    # =========================================================================
    # GUARDAR
    # =========================================================================
    print("=" * 80)
    print("💾 Guardando cambios...")
    print("=" * 80)
    print()

    wb.save(EXCEL_FILE)
    print("✅ Excel actualizado")
    print()

    # =========================================================================
    # VERIFICACIÓN
    # =========================================================================
    print("=" * 80)
    print("📊 VERIFICACIÓN")
    print("=" * 80)
    print()

    # Recargar para verificar
    wb_verif = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws_verif = wb_verif['TRANSACCIONES']

    # Verificar primeras 5 filas corregidas
    print("🔍 Verificando primeras 5 filas corregidas:")
    print()

    for egreso in egresos_a_corregir[:5]:
        fila = egreso['fila']
        monto_nuevo = ws_verif.cell(fila, col_map['Monto USD']).value

        print(f"Fila {fila}: {egreso['concepto'][:40] if egreso['concepto'] else 'N/A'}...")
        print(f"   Antes: +${egreso['monto_actual']:,.2f}")
        print(f"   Ahora: ${float(monto_nuevo):,.2f}" if monto_nuevo else "   Ahora: N/A")

        if monto_nuevo and float(monto_nuevo) < 0:
            print(f"   ✅ CORRECTO")
        else:
            print(f"   ⚠️  Aún no está negativo (Excel necesita recalcular)")

        print()

    # =========================================================================
    # RESUMEN
    # =========================================================================
    print("=" * 80)
    print("📊 RESUMEN")
    print("=" * 80)
    print()

    print(f"✅ Signos corregidos: {total_corregido} egresos")
    print(f"   Total monto afectado: ${suma_antes:,.2f}")
    print()
    print("🔧 PRÓXIMOS PASOS:")
    print("   1. Cierra Excel (si está abierto)")
    print("   2. Vuelve a abrirlo")
    print("   3. Ve a hoja Efectivo, fila 3")
    print("   4. Verifica los nuevos valores:")
    print()
    print("💡 Valores esperados en Efectivo (fila 3):")
    print(f"   Ingreso (D3): ~$11,951.71")
    print(f"   Egreso (E3): ~$10,170.96 (POSITIVO)")
    print(f"   Balance (F3): ~$1,780.75")
    print()
    print("   (Nota: Si el balance no es $2,163.44 del extracto,")
    print("    tendremos que ajustar el saldo inicial)")

    print()
    print("=" * 80)
    print("✅ CORRECCIÓN COMPLETADA")
    print("=" * 80)
    print()

    return True

if __name__ == "__main__":
    try:
        if not crear_backup():
            print("❌ Abortando")
            exit(1)

        if corregir_signos():
            print("🎉 Signos corregidos exitosamente!")
            print()
            print("👉 Ahora cierra y vuelve a abrir el Excel")
        else:
            print("✅ No se requirieron cambios")

    except Exception as e:
        print(f"❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
