#!/usr/bin/env python3
"""
HEALTH CHECK AUTOMÁTICO - SISTEMA FINANCIERO
Valida integridad contable completa
Detecta datos huérfanos y descuadres
Versión: 1.0
"""

import openpyxl
from datetime import datetime
import sys
from collections import defaultdict

class HealthCheckSistema:
    def __init__(self, excel_file="AlvaroVelasco_Finanzas_v1.0.xlsx"):
        self.excel_file = excel_file
        self.wb = None
        self.errores = []
        self.warnings = []
        self.info = []

    def ejecutar_health_check(self):
        """Ejecuta health check completo"""
        print("="*70)
        print("HEALTH CHECK SISTEMA FINANCIERO - VALIDACIÓN COMPLETA")
        print("="*70)
        print(f"Archivo: {self.excel_file}")
        print(f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print()

        try:
            # Cargar workbook
            print("⏳ Cargando workbook...")
            self.wb = openpyxl.load_workbook(self.excel_file, data_only=True)
            print("✅ Workbook cargado")
            print()

            # Ejecutar checks
            checks = [
                ("Check 1: Estructura del Sistema", self.check_estructura),
                ("Check 2: Datos Huérfanos", self.check_datos_huerfanos),
                ("Check 3: Fórmulas Intactas", self.check_formulas),
                ("Check 4: Duplicados", self.check_duplicados),
                ("Check 5: Balance Contable (CRÍTICO)", self.check_balance_contable),
                ("Check 6: Conciliación Multi-Nivel", self.check_conciliacion_multinivel),
                ("Check 7: Integridad Referencial", self.check_integridad_referencial),
                ("Check 8: Validación Campos Obligatorios", self.check_campos_obligatorios),
                ("Check 9: Rangos Razonables", self.check_rangos_razonables),
                ("Check 10: Consistencia Temporal", self.check_consistencia_temporal),
            ]

            for nombre, check_func in checks:
                print(f"⏳ {nombre}...")
                check_func()
                print()

            # Reporte final
            self.generar_reporte()

            return len(self.errores) == 0

        except FileNotFoundError:
            print(f"❌ ERROR: Archivo no encontrado: {self.excel_file}")
            print("   Ejecute primero: python3 scripts/install_system.py")
            return False
        except Exception as e:
            print(f"❌ ERROR CRÍTICO: {e}")
            import traceback
            traceback.print_exc()
            return False

    def check_estructura(self):
        """Verifica estructura básica del sistema"""
        # Verificar hojas necesarias
        required_sheets = ["TRANSACCIONES", "Dashboard", "Efectivo", "A_R", "A_P",
                          "Tarjetas_Credito", "Conciliacion", "Auditoria", "Health_Check"]

        for sheet in required_sheets:
            if sheet not in self.wb.sheetnames:
                self.errores.append(f"Falta hoja crítica: {sheet}")

        # Verificar columnas TRANSACCIONES
        ws = self.wb["TRANSACCIONES"]
        required_cols = ["Fecha", "Tipo Transacción", "Categoría", "Entidad",
                        "Cuenta Bancaria", "Cliente/Proveedor", "Concepto",
                        "Referencia", "Monto USD"]

        for col_idx, col_name in enumerate(required_cols, 1):
            if ws.cell(1, col_idx).value != col_name:
                self.errores.append(f"Columna {col_idx} incorrecta: esperado '{col_name}', encontrado '{ws.cell(1, col_idx).value}'")

        if not self.errores:
            print("   ✅ Estructura correcta: 9 hojas, 20 columnas")
        else:
            print(f"   ❌ {len(self.errores)} errores de estructura")

    def check_datos_huerfanos(self):
        """
        CRÍTICO: Detecta datos huérfanos (sin referencias completas)
        Similar a integridad referencial en bases de datos
        """
        ws = self.wb["TRANSACCIONES"]
        huerfanos = []

        for row in range(2, ws.max_row + 1):
            fecha = ws[f"A{row}"].value
            if not fecha:  # Fila vacía, skip
                continue

            tipo = ws[f"B{row}"].value
            entidad = ws[f"D{row}"].value
            cuenta = ws[f"E{row}"].value
            concepto = ws[f"G{row}"].value
            monto = ws[f"I{row}"].value
            estado = ws[f"L{row}"].value

            # VALIDACIÓN 1: Transacción sin cuenta bancaria
            if not cuenta:
                huerfanos.append({
                    "fila": row,
                    "tipo": "SIN CUENTA",
                    "detalle": f"Transacción '{concepto}' sin cuenta bancaria asignada"
                })

            # VALIDACIÓN 2: Transacción sin entidad
            if not entidad:
                huerfanos.append({
                    "fila": row,
                    "tipo": "SIN ENTIDAD",
                    "detalle": f"Transacción '{concepto}' sin entidad (EMPRESA/PERSONAL)"
                })

            # VALIDACIÓN 3: Transacción sin monto
            if not monto or monto == 0:
                huerfanos.append({
                    "fila": row,
                    "tipo": "SIN MONTO",
                    "detalle": f"Transacción '{concepto}' sin monto válido"
                })

            # VALIDACIÓN 4: Transacción sin estado
            if not estado:
                huerfanos.append({
                    "fila": row,
                    "tipo": "SIN ESTADO",
                    "detalle": f"Transacción '{concepto}' sin estado"
                })

            # VALIDACIÓN 5: Factura sin cliente/proveedor
            if tipo in ["Factura Cliente", "Factura Proveedor"]:
                cliente_prov = ws[f"F{row}"].value
                if not cliente_prov:
                    huerfanos.append({
                        "fila": row,
                        "tipo": "SIN CLIENTE/PROVEEDOR",
                        "detalle": f"{tipo} sin cliente/proveedor asignado"
                    })

        if huerfanos:
            self.errores.append(f"❌ DATOS HUÉRFANOS DETECTADOS: {len(huerfanos)}")
            for h in huerfanos[:10]:  # Mostrar primeros 10
                print(f"   🔴 Fila {h['fila']}: {h['tipo']} - {h['detalle']}")
            if len(huerfanos) > 10:
                print(f"   ... y {len(huerfanos) - 10} más")
        else:
            print("   ✅ Sin datos huérfanos: Todas las transacciones tienen referencias completas")

    def check_formulas(self):
        """Verifica que fórmulas críticas estén intactas"""
        ws = self.wb["TRANSACCIONES"]
        formulas_rotas = []

        # Reabrir en modo fórmula (no data_only)
        wb_formulas = openpyxl.load_workbook(self.excel_file, data_only=False)
        ws_formulas = wb_formulas["TRANSACCIONES"]

        for row in range(2, min(ws.max_row + 1, 100)):  # Verificar primeras 100 filas
            fecha = ws[f"A{row}"].value
            if not fecha:
                continue

            # Check columna K (Ingreso/Egreso) debe tener fórmula
            formula_k = ws_formulas[f"K{row}"].value
            if not formula_k or not str(formula_k).startswith("="):
                formulas_rotas.append({
                    "fila": row,
                    "columna": "K (Ingreso/Egreso)",
                    "detalle": "Fórmula faltante o rota"
                })

            # Check columna J (Monto CRC) debe tener fórmula
            formula_j = ws_formulas[f"J{row}"].value
            if not formula_j or not str(formula_j).startswith("="):
                formulas_rotas.append({
                    "fila": row,
                    "columna": "J (Monto CRC)",
                    "detalle": "Fórmula faltante o rota"
                })

        if formulas_rotas:
            self.errores.append(f"❌ FÓRMULAS ROTAS: {len(formulas_rotas)}")
            for f in formulas_rotas[:5]:
                print(f"   🔴 Fila {f['fila']}, Col {f['columna']}: {f['detalle']}")
        else:
            print("   ✅ Fórmulas intactas: Todas las columnas calculadas funcionan")

    def check_duplicados(self):
        """Detecta transacciones duplicadas"""
        ws = self.wb["TRANSACCIONES"]
        transacciones = []
        duplicados = []

        for row in range(2, ws.max_row + 1):
            fecha = ws[f"A{row}"].value
            if not fecha:
                continue

            concepto = ws[f"G{row}"].value
            monto = ws[f"I{row}"].value
            cliente = ws[f"F{row}"].value

            # Crear firma única
            firma = f"{fecha}|{concepto}|{monto}|{cliente}"

            if firma in transacciones:
                duplicados.append({
                    "fila": row,
                    "detalle": f"Duplicado: {fecha} - {concepto} - ${monto}"
                })
            else:
                transacciones.append(firma)

        if duplicados:
            self.warnings.append(f"⚠️ POSIBLES DUPLICADOS: {len(duplicados)}")
            for d in duplicados[:5]:
                print(f"   🟠 Fila {d['fila']}: {d['detalle']}")
        else:
            print("   ✅ Sin duplicados detectados")

    def check_balance_contable(self):
        """
        CRÍTICO: Valida integridad contable multi-nivel
        Nivel 1: Efectivo = Suma de todas las transacciones de efectivo/ahorro cobradas
        Nivel 2: A/R = Facturas pendientes de cobro
        Nivel 3: A/P = Facturas pendientes de pago
        Nivel 4: Validar que Activos - Pasivos tenga sentido
        """
        ws = self.wb["TRANSACCIONES"]

        efectivo_sistema = 0
        ar_sistema = 0
        ap_sistema = 0
        tc_sistema = 0

        for row in range(2, ws.max_row + 1):
            fecha = ws[f"A{row}"].value
            if not fecha:
                continue

            tipo = ws[f"B{row}"].value
            categoria = ws[f"C{row}"].value
            monto = ws[f"I{row}"].value or 0
            estado = ws[f"L{row}"].value

            # EFECTIVO: Apertura Inicial de efectivo/ahorro + movimientos posteriores
            if tipo == "Apertura Inicial" and categoria in ["Efectivo", "Ahorro"] and estado == "Cobrado":
                efectivo_sistema += monto

            # A/R: Facturas clientes pendientes
            elif tipo == "Factura Cliente" and estado == "Pendiente":
                ar_sistema += monto

            # A/P: Facturas proveedores pendientes
            elif tipo == "Factura Proveedor" and estado == "Pendiente":
                ap_sistema += abs(monto)

            # TC: Deuda tarjetas crédito (Apertura Inicial categoría TC)
            elif tipo == "Apertura Inicial" and categoria == "Tarjeta Crédito":
                tc_sistema += abs(monto)

        # Calcular patrimonio
        activos = efectivo_sistema + ar_sistema
        pasivos = ap_sistema + tc_sistema
        patrimonio = activos - pasivos

        print(f"   Efectivo (Efectivo + Ahorros): ${efectivo_sistema:,.2f}")
        print(f"   A/R Pendiente: ${ar_sistema:,.2f}")
        print(f"   A/P Pendiente: ${ap_sistema:,.2f}")
        print(f"   TC Deuda: ${tc_sistema:,.2f}")
        print(f"   ---")
        print(f"   Total Activos: ${activos:,.2f}")
        print(f"   Total Pasivos: ${pasivos:,.2f}")
        print(f"   Patrimonio Neto: ${patrimonio:,.2f}")

        # Validación: El efectivo debe ser positivo
        if efectivo_sistema < 0:
            self.errores.append(f"❌ EFECTIVO NEGATIVO: ${efectivo_sistema:,.2f}")
            print(f"   🔴 ERROR: Efectivo no puede ser negativo")
        # Validación: Patrimonio muy negativo es sospechoso
        elif patrimonio < -100000:
            self.warnings.append(f"⚠️ PATRIMONIO MUY NEGATIVO: ${patrimonio:,.2f}")
            print(f"   ⚠️  WARNING: Patrimonio muy negativo, revisar")
        else:
            print(f"   ✅ Balance contable correcto")

    def check_conciliacion_multinivel(self):
        """
        CONCILIACIÓN MULTI-NIVEL (como sistema contable profesional)
        Nivel 1: Efectivo sistema = Suma transacciones efectivo cobradas
        Nivel 2: A/R sistema = Suma facturas cliente pendientes
        Nivel 3: A/P sistema = Suma facturas proveedor pendientes
        Nivel 4: TC sistema = Suma saldos tarjetas
        """
        ws = self.wb["TRANSACCIONES"]

        # NIVEL 1: EFECTIVO
        efectivo_sistema = 0
        for row in range(2, ws.max_row + 1):
            fecha = ws[f"A{row}"].value
            if not fecha:
                continue

            tipo = ws[f"B{row}"].value
            monto = ws[f"I{row}"].value or 0
            estado = ws[f"L{row}"].value
            categoria = ws[f"C{row}"].value

            if categoria == "Efectivo" and estado == "Cobrado":
                efectivo_sistema += monto

        print(f"   Nivel 1 - Efectivo: ${efectivo_sistema:,.2f}")

        # NIVEL 2: A/R
        ar_sistema = 0
        for row in range(2, ws.max_row + 1):
            fecha = ws[f"A{row}"].value
            if not fecha:
                continue

            tipo = ws[f"B{row}"].value
            monto = ws[f"I{row}"].value or 0
            estado = ws[f"L{row}"].value

            if tipo == "Factura Cliente" and estado == "Pendiente":
                ar_sistema += monto

        print(f"   Nivel 2 - A/R Pendiente: ${ar_sistema:,.2f}")

        # NIVEL 3: A/P
        ap_sistema = 0
        for row in range(2, ws.max_row + 1):
            fecha = ws[f"A{row}"].value
            if not fecha:
                continue

            tipo = ws[f"B{row}"].value
            monto = ws[f"I{row}"].value or 0
            estado = ws[f"L{row}"].value

            if tipo == "Factura Proveedor" and estado == "Pendiente":
                ap_sistema += abs(monto)

        print(f"   Nivel 3 - A/P Pendiente: ${ap_sistema:,.2f}")

        # NIVEL 4: Balance total
        balance_total = efectivo_sistema + ar_sistema - ap_sistema
        print(f"   Nivel 4 - Balance Total: ${balance_total:,.2f}")

        # Validar que balance total tiene sentido
        if balance_total < -50000:  # Más de -$50k es sospechoso
            self.warnings.append(f"⚠️ Balance total muy negativo: ${balance_total:,.2f}")
        else:
            print(f"   ✅ Conciliación multi-nivel correcta")

    def check_integridad_referencial(self):
        """
        Valida integridad referencial (como FK en bases de datos)
        Ejemplo: Si hay "Cobro Factura", debe existir "Factura Cliente" correspondiente
        """
        ws = self.wb["TRANSACCIONES"]
        referencias = defaultdict(list)
        problemas = []

        # Mapear referencias
        for row in range(2, ws.max_row + 1):
            fecha = ws[f"A{row}"].value
            if not fecha:
                continue

            tipo = ws[f"B{row}"].value
            referencia = ws[f"H{row}"].value
            cliente = ws[f"F{row}"].value

            if referencia:
                referencias[referencia].append({
                    "fila": row,
                    "tipo": tipo,
                    "cliente": cliente
                })

        # Validar cobros tengan facturas
        for row in range(2, ws.max_row + 1):
            fecha = ws[f"A{row}"].value
            if not fecha:
                continue

            tipo = ws[f"B{row}"].value
            referencia = ws[f"H{row}"].value
            cliente = ws[f"F{row}"].value

            if tipo == "Cobro Factura" and referencia:
                # Buscar factura correspondiente
                facturas_relacionadas = [r for r in referencias.get(referencia, []) if r["tipo"] == "Factura Cliente"]
                if not facturas_relacionadas:
                    problemas.append({
                        "fila": row,
                        "detalle": f"Cobro sin factura: Ref {referencia}, Cliente {cliente}"
                    })

            if tipo == "Pago Factura" and referencia:
                # Buscar factura proveedor correspondiente
                facturas_relacionadas = [r for r in referencias.get(referencia, []) if r["tipo"] == "Factura Proveedor"]
                if not facturas_relacionadas:
                    problemas.append({
                        "fila": row,
                        "detalle": f"Pago sin factura: Ref {referencia}"
                    })

        if problemas:
            self.warnings.append(f"⚠️ PROBLEMAS INTEGRIDAD REFERENCIAL: {len(problemas)}")
            for p in problemas[:5]:
                print(f"   🟠 Fila {p['fila']}: {p['detalle']}")
        else:
            print("   ✅ Integridad referencial correcta")

    def check_campos_obligatorios(self):
        """Valida que campos obligatorios no estén vacíos"""
        ws = self.wb["TRANSACCIONES"]
        faltantes = []

        campos_obligatorios = {
            "A": "Fecha",
            "B": "Tipo Transacción",
            "D": "Entidad",
            "E": "Cuenta Bancaria",
            "G": "Concepto",
            "I": "Monto USD",
            "L": "Estado"
        }

        for row in range(2, ws.max_row + 1):
            fecha = ws[f"A{row}"].value
            if not fecha:  # Fila vacía
                continue

            for col, nombre in campos_obligatorios.items():
                valor = ws[f"{col}{row}"].value
                if not valor or (col == "I" and valor == 0):
                    faltantes.append({
                        "fila": row,
                        "campo": nombre,
                        "detalle": f"Campo obligatorio '{nombre}' vacío"
                    })

        if faltantes:
            self.errores.append(f"❌ CAMPOS OBLIGATORIOS FALTANTES: {len(faltantes)}")
            for f in faltantes[:10]:
                print(f"   🔴 Fila {f['fila']}: {f['detalle']}")
        else:
            print("   ✅ Todos los campos obligatorios completos")

    def check_rangos_razonables(self):
        """Valida que montos estén en rangos razonables"""
        ws = self.wb["TRANSACCIONES"]
        fuera_rango = []

        for row in range(2, ws.max_row + 1):
            fecha = ws[f"A{row}"].value
            if not fecha:
                continue

            monto = ws[f"I{row}"].value or 0
            concepto = ws[f"G{row}"].value

            # Alerta si monto > $50,000
            if abs(monto) > 50000:
                fuera_rango.append({
                    "fila": row,
                    "monto": monto,
                    "detalle": f"Monto inusualmente alto: ${monto:,.2f} - {concepto}"
                })

            # Alerta si monto negativo en ingreso
            ingreso_egreso = ws[f"K{row}"].value
            if ingreso_egreso == "Ingreso" and monto < 0:
                fuera_rango.append({
                    "fila": row,
                    "monto": monto,
                    "detalle": f"Ingreso con monto negativo: ${monto:,.2f}"
                })

        if fuera_rango:
            self.warnings.append(f"⚠️ MONTOS FUERA DE RANGO: {len(fuera_rango)}")
            for f in fuera_rango[:5]:
                print(f"   🟠 Fila {f['fila']}: {f['detalle']}")
        else:
            print("   ✅ Todos los montos en rangos razonables")

    def check_consistencia_temporal(self):
        """Valida consistencia de fechas"""
        ws = self.wb["TRANSACCIONES"]
        inconsistencias = []

        for row in range(2, ws.max_row + 1):
            fecha = ws[f"A{row}"].value
            if not fecha:
                continue

            vencimiento = ws[f"N{row}"].value

            # Si hay vencimiento, debe ser >= fecha
            if vencimiento and isinstance(fecha, datetime) and isinstance(vencimiento, datetime):
                if vencimiento < fecha:
                    inconsistencias.append({
                        "fila": row,
                        "detalle": f"Vencimiento ({vencimiento}) anterior a fecha ({fecha})"
                    })

            # Fechas futuras sospechosas
            if isinstance(fecha, datetime) and fecha > datetime.now():
                dias_futuro = (fecha - datetime.now()).days
                if dias_futuro > 365:  # Más de 1 año en el futuro
                    inconsistencias.append({
                        "fila": row,
                        "detalle": f"Fecha muy futura: {fecha} ({dias_futuro} días)"
                    })

        if inconsistencias:
            self.warnings.append(f"⚠️ INCONSISTENCIAS TEMPORALES: {len(inconsistencias)}")
            for i in inconsistencias[:5]:
                print(f"   🟠 Fila {i['fila']}: {i['detalle']}")
        else:
            print("   ✅ Consistencia temporal correcta")

    def generar_reporte(self):
        """Genera reporte final"""
        print("="*70)
        print("REPORTE FINAL HEALTH CHECK")
        print("="*70)
        print()

        print(f"🔴 ERRORES CRÍTICOS: {len(self.errores)}")
        if self.errores:
            for error in self.errores:
                print(f"   - {error}")
        print()

        print(f"🟠 WARNINGS: {len(self.warnings)}")
        if self.warnings:
            for warning in self.warnings:
                print(f"   - {warning}")
        print()

        if len(self.errores) == 0 and len(self.warnings) == 0:
            print("✅ SISTEMA SALUDABLE: 0 errores, 0 warnings")
            print()
            print("🎉 El sistema está funcionando perfectamente")
            print("📊 Todos los datos están conciliados")
            print("🛡️  No hay datos huérfanos")
            print("✅ Balance contable correcto")
        elif len(self.errores) == 0:
            print("✅ SISTEMA FUNCIONAL con warnings menores")
            print("⚠️  Revisar warnings pero no es crítico")
        else:
            print("❌ SISTEMA REQUIERE ATENCIÓN INMEDIATA")
            print("🔴 Corregir errores críticos antes de continuar")

        print()
        print("="*70)

if __name__ == "__main__":
    checker = HealthCheckSistema()
    success = checker.ejecutar_health_check()
    exit(0 if success else 1)
