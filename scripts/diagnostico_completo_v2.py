#!/usr/bin/env python3
"""
DIAGNÓSTICO COMPLETO - Excel v2.0
==================================

PROPÓSITO: Entender estructura REAL de v2.0 ANTES de crear v3.0

LECCIÓN APRENDIDA DEL PROYECTO ANTERIOR:
"Crear script de diagnóstico PRIMERO antes de implementar"

Este script analiza:
1. Qué hojas existen
2. Qué columnas tiene cada hoja
3. Cuántos datos hay
4. Tipos de datos
5. Valores únicos importantes
6. Fórmulas vs valores hardcoded

Ejecutar:
    python scripts/diagnostico_completo_v2.py

Salida:
    diagnostico_v2_YYYYMMDD_HHMMSS.txt
"""

import openpyxl
from datetime import datetime
import os

# Posibles rutas del Excel v2.0
EXCEL_PATHS = [
    "AlvaroVelasco_Finanzas_v2.0.xlsx",
    "../AlvaroVelasco_Finanzas_v2.0.xlsx",
    "C:\\Users\\Alvaro Velasco\\Desktop\\debt-sanitization-strategy\\AlvaroVelasco_Finanzas_v2.0.xlsx",
]

def encontrar_excel():
    """Busca el archivo Excel en rutas posibles"""
    for path in EXCEL_PATHS:
        if os.path.exists(path):
            print(f"✅ Excel encontrado: {path}")
            return path

    print("❌ Excel v2.0 NO encontrado en ninguna ruta esperada")
    print("\nRutas buscadas:")
    for path in EXCEL_PATHS:
        print(f"  - {path}")

    return None

def diagnosticar_excel(excel_path):
    """Diagnóstico completo del Excel v2.0"""

    print("\n" + "="*80)
    print("DIAGNÓSTICO COMPLETO - EXCEL V2.0")
    print("="*80)
    print(f"Archivo: {excel_path}")
    print(f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("="*80)

    # Abrir Excel
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=False)
    except Exception as e:
        print(f"\n❌ ERROR al abrir Excel: {e}")
        return

    output_lines = []
    output_lines.append("="*80)
    output_lines.append(f"DIAGNÓSTICO COMPLETO - EXCEL V2.0")
    output_lines.append(f"Archivo: {excel_path}")
    output_lines.append(f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    output_lines.append("="*80)

    # 1. RESUMEN GENERAL
    print(f"\n📊 RESUMEN GENERAL:")
    output_lines.append(f"\n📊 RESUMEN GENERAL:")

    total_hojas = len(wb.sheetnames)
    print(f"   Total de hojas: {total_hojas}")
    output_lines.append(f"   Total de hojas: {total_hojas}")

    print(f"   Nombres de hojas:")
    output_lines.append(f"   Nombres de hojas:")
    for i, name in enumerate(wb.sheetnames, 1):
        print(f"      {i}. {name}")
        output_lines.append(f"      {i}. {name}")

    # 2. ANÁLISIS POR HOJA
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        print(f"\n{'='*80}")
        print(f"HOJA: {sheet_name}")
        print(f"{'='*80}")

        output_lines.append(f"\n{'='*80}")
        output_lines.append(f"HOJA: {sheet_name}")
        output_lines.append(f"{'='*80}")

        # Dimensiones
        max_row = ws.max_row
        max_col = ws.max_column

        print(f"\n📐 DIMENSIONES:")
        print(f"   Filas: {max_row}")
        print(f"   Columnas: {max_col}")

        output_lines.append(f"\n📐 DIMENSIONES:")
        output_lines.append(f"   Filas: {max_row}")
        output_lines.append(f"   Columnas: {max_col}")

        # Encabezados (fila 1)
        print(f"\n📋 ENCABEZADOS (Fila 1):")
        output_lines.append(f"\n📋 ENCABEZADOS (Fila 1):")

        encabezados = []
        for col in range(1, min(max_col + 1, 30)):  # Max 30 columnas
            cell = ws.cell(1, col)
            header = cell.value
            if header:
                encabezados.append(header)
                print(f"   {openpyxl.utils.get_column_letter(col)}: {header}")
                output_lines.append(f"   {openpyxl.utils.get_column_letter(col)}: {header}")

        # Primeras 3 filas de datos (para ver formato)
        print(f"\n📝 PRIMERAS 3 FILAS DE DATOS:")
        output_lines.append(f"\n📝 PRIMERAS 3 FILAS DE DATOS:")

        for row_num in range(2, min(5, max_row + 1)):
            print(f"\n   Fila {row_num}:")
            output_lines.append(f"\n   Fila {row_num}:")

            for col_num in range(1, min(max_col + 1, 15)):  # Max 15 cols
                cell = ws.cell(row_num, col_num)
                col_letter = openpyxl.utils.get_column_letter(col_num)

                # Ver si es fórmula o valor
                if cell.value is not None:
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        tipo = "FÓRMULA"
                        valor = cell.value[:50]  # Primeros 50 chars
                    else:
                        tipo = "VALOR"
                        valor = str(cell.value)[:50]

                    header = encabezados[col_num - 1] if col_num <= len(encabezados) else col_letter
                    print(f"      {col_letter} ({header}): [{tipo}] {valor}")
                    output_lines.append(f"      {col_letter} ({header}): [{tipo}] {valor}")

        # Análisis de columnas clave (según nombres estándar)
        if sheet_name.upper() == "TRANSACCIONES" or "TRANS" in sheet_name.upper():
            print(f"\n🔍 ANÁLISIS TRANSACCIONES:")
            output_lines.append(f"\n🔍 ANÁLISIS TRANSACCIONES:")

            # Buscar columnas importantes
            col_fecha = None
            col_tipo = None
            col_cuenta = None
            col_entidad = None
            col_monto = None

            for col in range(1, max_col + 1):
                header = ws.cell(1, col).value
                if header:
                    header_upper = str(header).upper()

                    if "FECHA" in header_upper:
                        col_fecha = col
                    elif "TIPO" in header_upper:
                        col_tipo = col
                    elif "CUENTA" in header_upper:
                        col_cuenta = col
                    elif "ENTIDAD" in header_upper or "CLIENTE" in header_upper:
                        col_entidad = col
                    elif "MONTO" in header_upper or "USD" in header_upper:
                        col_monto = col

            print(f"   Columnas detectadas:")
            output_lines.append(f"   Columnas detectadas:")
            if col_fecha:
                print(f"      Fecha: Columna {openpyxl.utils.get_column_letter(col_fecha)}")
                output_lines.append(f"      Fecha: Columna {openpyxl.utils.get_column_letter(col_fecha)}")
            if col_tipo:
                print(f"      Tipo: Columna {openpyxl.utils.get_column_letter(col_tipo)}")
                output_lines.append(f"      Tipo: Columna {openpyxl.utils.get_column_letter(col_tipo)}")
            if col_cuenta:
                print(f"      Cuenta: Columna {openpyxl.utils.get_column_letter(col_cuenta)}")
                output_lines.append(f"      Cuenta: Columna {openpyxl.utils.get_column_letter(col_cuenta)}")
            if col_entidad:
                print(f"      Entidad: Columna {openpyxl.utils.get_column_letter(col_entidad)}")
                output_lines.append(f"      Entidad: Columna {openpyxl.utils.get_column_letter(col_entidad)}")
            if col_monto:
                print(f"      Monto: Columna {openpyxl.utils.get_column_letter(col_monto)}")
                output_lines.append(f"      Monto: Columna {openpyxl.utils.get_column_letter(col_monto)}")

            # Contar transacciones por tipo
            if col_tipo:
                tipos = {}
                for row in range(2, max_row + 1):
                    tipo = ws.cell(row, col_tipo).value
                    if tipo:
                        tipos[tipo] = tipos.get(tipo, 0) + 1

                print(f"\n   Distribución por Tipo:")
                output_lines.append(f"\n   Distribución por Tipo:")
                for tipo, count in sorted(tipos.items(), key=lambda x: x[1], reverse=True):
                    print(f"      {tipo}: {count} transacciones")
                    output_lines.append(f"      {tipo}: {count} transacciones")

            # Cuentas únicas
            if col_cuenta:
                cuentas = set()
                for row in range(2, max_row + 1):
                    cuenta = ws.cell(row, col_cuenta).value
                    if cuenta:
                        cuentas.add(str(cuenta))

                print(f"\n   Cuentas Únicas ({len(cuentas)}):")
                output_lines.append(f"\n   Cuentas Únicas ({len(cuentas)}):")
                for cuenta in sorted(cuentas)[:20]:  # Primeras 20
                    print(f"      - {cuenta}")
                    output_lines.append(f"      - {cuenta}")

                if len(cuentas) > 20:
                    print(f"      ... y {len(cuentas) - 20} más")
                    output_lines.append(f"      ... y {len(cuentas) - 20} más")

            # Entidades únicas
            if col_entidad:
                entidades = set()
                for row in range(2, max_row + 1):
                    entidad = ws.cell(row, col_entidad).value
                    if entidad:
                        entidades.add(str(entidad))

                print(f"\n   Entidades Únicas ({len(entidades)}):")
                output_lines.append(f"\n   Entidades Únicas ({len(entidades)}):")
                for entidad in sorted(entidades)[:20]:  # Primeras 20
                    print(f"      - {entidad}")
                    output_lines.append(f"      - {entidad}")

                if len(entidades) > 20:
                    print(f"      ... y {len(entidades) - 20} más")
                    output_lines.append(f"      ... y {len(entidades) - 20} más")

        # Análisis hoja EFECTIVO
        if "EFECTIVO" in sheet_name.upper() or "BANCO" in sheet_name.upper():
            print(f"\n🏦 ANÁLISIS EFECTIVO/BANCOS:")
            output_lines.append(f"\n🏦 ANÁLISIS EFECTIVO/BANCOS:")

            # Contar fórmulas vs valores
            formulas = 0
            valores = 0

            for row in range(1, min(max_row + 1, 50)):  # Primeras 50 filas
                for col in range(1, max_col + 1):
                    cell = ws.cell(row, col)
                    if cell.value:
                        if isinstance(cell.value, str) and cell.value.startswith('='):
                            formulas += 1
                        else:
                            valores += 1

            print(f"   Fórmulas detectadas: {formulas}")
            print(f"   Valores estáticos: {valores}")
            output_lines.append(f"   Fórmulas detectadas: {formulas}")
            output_lines.append(f"   Valores estáticos: {valores}")

    # 3. CONCLUSIÓN
    print(f"\n{'='*80}")
    print(f"DIAGNÓSTICO COMPLETADO")
    print(f"{'='*80}")

    output_lines.append(f"\n{'='*80}")
    output_lines.append(f"DIAGNÓSTICO COMPLETADO")
    output_lines.append(f"{'='*80}")

    # Guardar a archivo
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"diagnostico_v2_{timestamp}.txt"

    with open(output_file, 'w', encoding='utf-8') as f:
        f.write("\n".join(output_lines))

    print(f"\n📄 Reporte guardado en: {output_file}")

    return output_file

def main():
    """Función principal"""

    # Buscar Excel
    excel_path = encontrar_excel()

    if not excel_path:
        print("\n⚠️ INSTRUCCIONES:")
        print("   1. Copia el archivo Excel v2.0 a la carpeta del proyecto")
        print("   2. O actualiza EXCEL_PATHS en este script con la ruta correcta")
        return

    # Diagnosticar
    output_file = diagnosticar_excel(excel_path)

    print(f"\n✅ DIAGNÓSTICO COMPLETADO")
    print(f"   Revisa: {output_file}")
    print(f"\n💡 PRÓXIMO PASO: Diseñar v3.0 basado en este análisis")

if __name__ == "__main__":
    main()
