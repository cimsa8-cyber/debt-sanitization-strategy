#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ELIMINAR TRANSACCIONES DUPLICADAS
Elimina duplicados identificados en diagnóstico - mantiene primera ocurrencia
"""
import openpyxl
from datetime import datetime
import shutil

EXCEL_FILE = "AlvaroVelasco_Finanzas_v2.0.xlsx"
BACKUP_FILE = f"AlvaroVelasco_Finanzas_v2.0_ELIMINAR_DUPLICADOS_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

def crear_backup():
    print("=" * 80)
    print("CREANDO BACKUP")
    print("=" * 80)
    print(f"Backup: {BACKUP_FILE}")
    try:
        shutil.copy2(EXCEL_FILE, BACKUP_FILE)
        print("✅ Backup creado")
        print()
        return True
    except Exception as e:
        print(f"❌ ERROR: {e}")
        return False

def eliminar_duplicados():
    print("=" * 80)
    print("ELIMINACIÓN DE TRANSACCIONES DUPLICADAS")
    print("=" * 80)
    print()

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb['TRANSACCIONES']

    # Encontrar columnas
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    col_map = {}
    for col in range(1, len(headers) + 1):
        if headers[col-1]:
            col_map[headers[col-1]] = col

    print("🔍 Identificando duplicados en Promerica USD...")
    print()

    cuenta_buscar = "Promerica USD (40000003881774)"

    # Estructura: {key: [lista de filas]}
    transacciones = {}

    for row in range(2, ws.max_row + 1):
        cuenta = ws.cell(row, col_map['Cuenta Bancaria']).value
        monto = ws.cell(row, col_map['Monto USD']).value
        ing_egr = ws.cell(row, col_map['Ingreso/Egreso']).value
        fecha = ws.cell(row, col_map['Fecha']).value
        concepto = ws.cell(row, col_map['Concepto']).value

        if cuenta and str(cuenta).strip() == cuenta_buscar:
            if monto and ing_egr and fecha:
                monto_val = float(monto)
                fecha_str = fecha.strftime('%d/%m/%Y') if isinstance(fecha, datetime) else str(fecha)

                # Key: fecha + monto absoluto + tipo
                key = f"{fecha_str}_{abs(monto_val)}_{ing_egr}"

                if key not in transacciones:
                    transacciones[key] = []

                transacciones[key].append({
                    'fila': row,
                    'fecha': fecha_str,
                    'tipo': ing_egr,
                    'monto': monto_val,
                    'concepto': concepto[:50] if concepto else 'N/A'
                })

    # Identificar duplicados
    duplicados_a_eliminar = []

    for key, trans_list in transacciones.items():
        if len(trans_list) > 1:
            # Mantener la primera, eliminar las demás
            primera = trans_list[0]
            duplicados = trans_list[1:]

            for dup in duplicados:
                duplicados_a_eliminar.append(dup)

    print(f"📊 DUPLICADOS DETECTADOS: {len(duplicados_a_eliminar)} transacciones")
    print()

    if len(duplicados_a_eliminar) == 0:
        print("✅ No hay duplicados que eliminar")
        return False

    # Agrupar por tipo para mostrar resumen
    ingresos_dup = [d for d in duplicados_a_eliminar if d['tipo'] == 'Ingreso']
    egresos_dup = [d for d in duplicados_a_eliminar if d['tipo'] == 'Egreso']

    print(f"📈 Ingresos duplicados: {len(ingresos_dup)}")
    if ingresos_dup:
        total_ing_dup = sum(d['monto'] for d in ingresos_dup)
        print(f"   Total a eliminar: ${total_ing_dup:,.2f}")
        print()
        for d in ingresos_dup:
            print(f"   • Fila {d['fila']}: {d['fecha']} - ${d['monto']:,.2f}")
            print(f"     {d['concepto']}")
        print()

    print(f"📉 Egresos duplicados: {len(egresos_dup)}")
    if egresos_dup:
        total_egr_dup = sum(abs(d['monto']) for d in egresos_dup)
        print(f"   Total a eliminar: ${total_egr_dup:,.2f}")
        print()

        # Mostrar primeros 15 para no saturar
        for d in egresos_dup[:15]:
            print(f"   • Fila {d['fila']}: {d['fecha']} - ${abs(d['monto']):,.2f}")
            print(f"     {d['concepto']}")

        if len(egresos_dup) > 15:
            print(f"   ... y {len(egresos_dup) - 15} más")
        print()

    # =========================================================================
    # ELIMINAR FILAS (DE ABAJO HACIA ARRIBA PARA NO PERDER ÍNDICES)
    # =========================================================================
    print("=" * 80)
    print("🗑️  ELIMINANDO DUPLICADOS...")
    print("=" * 80)
    print()

    # Ordenar por fila descendente para eliminar de abajo hacia arriba
    duplicados_a_eliminar_sorted = sorted(duplicados_a_eliminar, key=lambda x: x['fila'], reverse=True)

    total_eliminado = 0
    for dup in duplicados_a_eliminar_sorted:
        fila = dup['fila']
        ws.delete_rows(fila, 1)
        total_eliminado += 1

        if total_eliminado <= 10:  # Mostrar primeras 10
            print(f"✅ Eliminada fila {fila}: {dup['fecha']} - ${abs(dup['monto']):,.2f}")
            print(f"   {dup['concepto']}")

    if total_eliminado > 10:
        print(f"   ... y {total_eliminado - 10} filas más eliminadas")

    print()
    print(f"📊 Total filas eliminadas: {total_eliminado}")
    print()

    # =========================================================================
    # GUARDAR
    # =========================================================================
    print("=" * 80)
    print("💾 GUARDANDO CAMBIOS...")
    print("=" * 80)
    print()

    wb.save(EXCEL_FILE)
    print("✅ Excel actualizado")
    print()

    # =========================================================================
    # VERIFICACIÓN
    # =========================================================================
    print("=" * 80)
    print("📊 VERIFICACIÓN POST-ELIMINACIÓN")
    print("=" * 80)
    print()

    # Recargar y recalcular
    wb_verif = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws_verif = wb_verif['TRANSACCIONES']

    # Recalcular ingresos y egresos
    ingresos_final = []
    egresos_final = []

    for row in range(2, ws_verif.max_row + 1):
        cuenta = ws_verif.cell(row, col_map['Cuenta Bancaria']).value
        monto = ws_verif.cell(row, col_map['Monto USD']).value
        ing_egr = ws_verif.cell(row, col_map['Ingreso/Egreso']).value

        if cuenta and str(cuenta).strip() == cuenta_buscar:
            if monto and ing_egr:
                monto_val = float(monto)

                if ing_egr == 'Ingreso':
                    ingresos_final.append(monto_val)
                elif ing_egr == 'Egreso':
                    egresos_final.append(monto_val)

    total_ingresos_final = sum(ingresos_final)
    total_egresos_final = sum(abs(e) for e in egresos_final)
    balance_final = total_ingresos_final - total_egresos_final

    print(f"📈 INGRESOS: {len(ingresos_final)} transacciones")
    print(f"   Total: ${total_ingresos_final:,.2f}")
    print()

    print(f"📉 EGRESOS: {len(egresos_final)} transacciones")
    print(f"   Total: ${total_egresos_final:,.2f}")
    print()

    print(f"💰 BALANCE CALCULADO: ${balance_final:,.2f}")
    print()

    # Comparación
    print("=" * 80)
    print("📊 COMPARACIÓN")
    print("=" * 80)
    print()

    print("ANTES:")
    print(f"   Ingresos: $14,983.00")
    print(f"   Egresos: $13,057.00")
    print(f"   Balance: $1,925.63")
    print()

    print("AHORA:")
    print(f"   Ingresos: ${total_ingresos_final:,.2f}")
    print(f"   Egresos: ${total_egresos_final:,.2f}")
    print(f"   Balance: ${balance_final:,.2f}")
    print()

    print("EXTRACTO BANCARIO:")
    print(f"   Balance objetivo: $2,163.44")
    print()

    diferencia = 2163.44 - balance_final
    print(f"⚖️  Diferencia vs extracto: ${abs(diferencia):,.2f}")
    print()

    if abs(diferencia) < 500:
        print("💡 PRÓXIMO PASO:")
        print(f"   Ajustar saldo inicial de $3,030.89")
        print(f"   Incremento necesario: ${diferencia:,.2f}")
        print(f"   Nuevo saldo inicial: ${3030.89 + diferencia:,.2f}")
    else:
        print("⚠️  La diferencia es mayor a $500")
        print("   Puede haber transacciones faltantes o incorrectas")

    print()
    print("=" * 80)
    print("✅ ELIMINACIÓN COMPLETADA")
    print("=" * 80)
    print()

    return True

if __name__ == "__main__":
    try:
        if not crear_backup():
            print("❌ Abortando")
            exit(1)

        if eliminar_duplicados():
            print("🎉 Duplicados eliminados exitosamente!")
            print()
            print("👉 Ahora cierra y vuelve a abrir el Excel")
            print("   Verifica los nuevos valores en hoja Efectivo, fila 3")
        else:
            print("✅ No se requirieron cambios")

    except Exception as e:
        print(f"❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
