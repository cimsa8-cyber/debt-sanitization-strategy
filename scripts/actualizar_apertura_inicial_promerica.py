#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ACTUALIZAR APERTURA INICIAL PROMERICA - TRANSACCIONES FILA 2
Actualiza el balance de "Apertura Inicial" del 01/11/2025
que es lo que muestra la hoja Efectivo
"""
import openpyxl
import sys
import os

EXCEL_FILE = "AlvaroVelasco_Finanzas_v2.0.xlsx"

print("="*80)
print("ACTUALIZAR APERTURA INICIAL PROMERICA - TRANSACCIONES FILA 2")
print("="*80)

if not os.path.exists(EXCEL_FILE):
    print(f"\n❌ ERROR: No se encontró {EXCEL_FILE}")
    sys.exit(1)

# Cargar Excel
print(f"\nCargando {EXCEL_FILE}...")
wb = openpyxl.load_workbook(EXCEL_FILE)

if 'TRANSACCIONES' not in wb.sheetnames:
    print("\n❌ ERROR: No se encontró la hoja 'TRANSACCIONES'")
    sys.exit(1)

ws = wb['TRANSACCIONES']

# Fila 2 (Apertura Inicial Promerica)
fila = 2

print("\n" + "="*80)
print(f"VERIFICACIÓN FILA {fila}")
print("="*80)

print(f"\nContenido actual:")
print(f"   A (Fecha): {ws[f'A{fila}'].value}")
print(f"   B (Tipo): {ws[f'B{fila}'].value}")
print(f"   E (Cuenta): {ws[f'E{fila}'].value}")
print(f"   G (Concepto): {ws[f'G{fila}'].value}")
print(f"   I (Monto USD): {ws[f'I{fila}'].value}")
print(f"   K (Ingreso/Egreso): {ws[f'K{fila}'].value}")

# Verificar que es la fila correcta
tipo = ws[f'B{fila}'].value
cuenta = ws[f'E{fila}'].value
fecha = ws[f'A{fila}'].value

if tipo and 'Apertura Inicial' in str(tipo):
    if cuenta and 'Promerica USD' in str(cuenta) and '40000003881774' in str(cuenta):
        if fecha and fecha.month == 11 and fecha.day == 1:
            print(f"\n✅ Confirmado: Fila {fila} es Apertura Inicial Promerica USD (01/11/2025)")

            # Nuevo balance según extracto bancario 31/10
            nuevo_balance = 3030.89

            print(f"\n📝 Actualizando balance:")
            print(f"   Balance anterior: ${ws[f'I{fila}'].value}")
            print(f"   Balance nuevo: ${nuevo_balance}")

            # Actualizar monto USD (columna I)
            ws[f'I{fila}'] = nuevo_balance

            # Actualizar notas para documentar el cambio (columna O)
            ws[f'O{fila}'] = f"Balance actualizado a ${nuevo_balance:,.2f} según extracto bancario 31/10/2025"

            # Guardar
            print(f"\n💾 Guardando {EXCEL_FILE}...")
            wb.save(EXCEL_FILE)

            print("\n✅ BALANCE ACTUALIZADO EXITOSAMENTE")

            print("\n" + "="*80)
            print("RESUMEN")
            print("="*80)

            print(f"""
✅ TRANSACCIONES Fila 2 actualizada:
   Tipo: Apertura Inicial
   Cuenta: Promerica USD (40000003881774)
   Fecha: 01/11/2025
   Balance anterior: $2,999.24
   Balance nuevo: ${nuevo_balance:,.2f}

📊 IMPACTO:
   La hoja Efectivo ahora mostrará: ${nuevo_balance:,.2f}
   (Efectivo tiene fórmulas que apuntan a TRANSACCIONES!I2)

✅ ESTE ES EL SALDO CORRECTO:
   Según extracto bancario al 31/10/2025: ${nuevo_balance:,.2f}

📋 PRÓXIMOS PASOS:
   1. Abrir Excel y verificar:
      - Hoja TRANSACCIONES Fila 2: ${nuevo_balance:,.2f}
      - Hoja Efectivo Fila 3: ${nuevo_balance:,.2f}
   2. Ambos deben mostrar el mismo valor ahora
   3. Este es el saldo REAL de Promerica al 31/10

⚠️ NOTA IMPORTANTE:
   El saldo calculado desde movimientos es $2,761.43
   El saldo del extracto (31/10) es ${nuevo_balance:,.2f}
   Diferencia: ${nuevo_balance - 2761.43:.2f}

   Esta diferencia es NORMAL porque:
   - La "Apertura Inicial" del 01/11 es un saldo de CORTE
   - Los movimientos de octubre YA están incluidos en este saldo
   - No deben sumarse dos veces
""")

            print("\n" + "="*80)

        else:
            print(f"\n⚠️ Fecha incorrecta: {fecha}")
            print("   Se esperaba 01/11/2025")
    else:
        print(f"\n⚠️ Cuenta incorrecta: {cuenta}")
        print("   Se esperaba Promerica USD (40000003881774)")
else:
    print(f"\n⚠️ Tipo incorrecto: {tipo}")
    print("   Se esperaba 'Apertura Inicial'")

print("\n" + "="*80)
