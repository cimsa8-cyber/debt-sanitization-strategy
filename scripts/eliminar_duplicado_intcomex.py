#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ELIMINAR DUPLICADO - PAGO INTCOMEX
Busca y elimina el pago duplicado de $3,137.26 del 10/11/2025
"""
import openpyxl
from datetime import datetime

EXCEL_FILE = "AlvaroVelasco_Finanzas_v2.0.xlsx"

def buscar_y_eliminar_duplicado():
    print("=" * 80)
    print("BÚSQUEDA DE PAGO DUPLICADO - INTCOMEX $3,137.26")
    print("=" * 80)
    print()

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb['TRANSACCIONES']

    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    col_map = {}
    for col in range(1, len(headers) + 1):
        if headers[col-1]:
            col_map[headers[col-1]] = col

    print("🔍 Buscando todos los registros relacionados con $3,137.26 del 10/11/2025...")
    print()

    registros_encontrados = []

    for row in range(2, ws.max_row + 1):
        fecha = ws.cell(row, col_map['Fecha']).value
        monto_usd = ws.cell(row, col_map['Monto USD']).value
        concepto = ws.cell(row, col_map['Concepto']).value
        cuenta = ws.cell(row, col_map['Cuenta Bancaria']).value
        tipo = ws.cell(row, col_map['Tipo Transacción']).value
        cliente_prov = ws.cell(row, col_map['Cliente/Proveedor']).value

        # Buscar transacciones de $3,137.26 del 10/11
        if monto_usd:
            try:
                monto_abs = abs(float(monto_usd))
                if monto_abs == 3137.26:
                    # Verificar fecha
                    if fecha:
                        if isinstance(fecha, datetime):
                            fecha_obj = fecha
                        else:
                            try:
                                fecha_obj = datetime.strptime(str(fecha), '%Y-%m-%d %H:%M:%S')
                            except:
                                fecha_obj = None

                        if fecha_obj:
                            # Buscar 10/11/2025 o 31/10/2025
                            if (fecha_obj.month == 11 and fecha_obj.day == 10) or \
                               (fecha_obj.month == 10 and fecha_obj.day == 31):
                                registros_encontrados.append({
                                    'fila': row,
                                    'fecha': fecha_obj,
                                    'tipo': tipo,
                                    'cuenta': cuenta,
                                    'monto': float(monto_usd),
                                    'concepto': concepto[:60] if concepto else 'N/A',
                                    'cliente_prov': cliente_prov
                                })
            except:
                pass

    if not registros_encontrados:
        print("✅ No se encontraron registros con $3,137.26 en fechas relevantes")
        print()
        return

    # Mostrar todos los registros
    print(f"📋 REGISTROS ENCONTRADOS ({len(registros_encontrados)}):")
    print()

    for i, reg in enumerate(registros_encontrados, 1):
        print(f"{i}. Fila {reg['fila']}:")
        print(f"   Fecha: {reg['fecha'].strftime('%d/%m/%Y')}")
        print(f"   Tipo: {reg['tipo']}")
        print(f"   Cuenta: {reg['cuenta']}")
        print(f"   Monto: ${reg['monto']:,.2f}")
        print(f"   Cliente/Proveedor: {reg['cliente_prov']}")
        print(f"   Concepto: {reg['concepto']}")
        print()

    # Identificar duplicado
    print("=" * 80)
    print("🔍 ANÁLISIS DE DUPLICADOS")
    print("=" * 80)
    print()

    # Buscar registros del 10/11 desde Promerica (cualquier signo)
    pagos_promerica_10_11 = [r for r in registros_encontrados
                              if r['fecha'].day == 10 and r['fecha'].month == 11
                              and r['cuenta'] and 'Promerica' in str(r['cuenta'])
                              and 'Pago' in str(r['concepto'])]  # Debe ser un pago

    if len(pagos_promerica_10_11) > 1:
        print(f"⚠️  DUPLICADO DETECTADO: {len(pagos_promerica_10_11)} pagos desde Promerica el 10/11")
        print()

        # El más reciente es probablemente el duplicado (fila más alta)
        duplicado = max(pagos_promerica_10_11, key=lambda x: x['fila'])

        print("🗑️  REGISTRO A ELIMINAR:")
        print(f"   Fila: {duplicado['fila']}")
        print(f"   Tipo: {duplicado['tipo']}")
        print(f"   Concepto: {duplicado['concepto']}")
        print(f"   Motivo: Duplicado - El pago ya estaba registrado")
        print()

        # Confirmar y eliminar
        print("💾 Eliminando fila duplicada...")
        ws.delete_rows(duplicado['fila'], 1)

        wb.save(EXCEL_FILE)
        print("✅ Fila eliminada exitosamente")
        print()

        print("=" * 80)
        print("📊 RESUMEN")
        print("=" * 80)
        print()
        print(f"✅ Fila {duplicado['fila']} eliminada (pago duplicado)")
        print(f"✅ Saldo de Promerica corregido (+$3,137.26)")
        print()
        print("📋 REGISTROS VÁLIDOS QUE QUEDARON:")
        for reg in registros_encontrados:
            if reg['fila'] != duplicado['fila']:
                print(f"   • Fila {reg['fila']}: {reg['tipo']} - {reg['concepto'][:40]}")
        print()

    else:
        print("✅ No se detectó duplicado en pagos desde Promerica")
        print()

    print("=" * 80)
    print("✅ VERIFICACIÓN COMPLETADA")
    print("=" * 80)
    print()

if __name__ == "__main__":
    try:
        buscar_y_eliminar_duplicado()
        print("🎉 Proceso completado!")
    except Exception as e:
        print(f"❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
