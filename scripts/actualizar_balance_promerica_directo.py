#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ACTUALIZAR BALANCE PROMERICA - DIRECTO
Actualiza directamente la Fila 3 de la hoja Efectivo
"""
import openpyxl
import sys
import os

EXCEL_FILE = "AlvaroVelasco_Finanzas_v2.0.xlsx"

print("="*80)
print("ACTUALIZAR BALANCE PROMERICA EN HOJA EFECTIVO")
print("="*80)

if not os.path.exists(EXCEL_FILE):
    print(f"\n❌ ERROR: No se encontró {EXCEL_FILE}")
    sys.exit(1)

# Cargar Excel
print(f"\nCargando {EXCEL_FILE}...")
wb = openpyxl.load_workbook(EXCEL_FILE)

if 'Efectivo' not in wb.sheetnames:
    print("\n❌ ERROR: No se encontró la hoja 'Efectivo'")
    sys.exit(1)

ws = wb['Efectivo']

# Verificar Fila 3 (Promerica USD)
print("\n" + "="*80)
print("VERIFICACIÓN FILA 3 - PROMERICA USD")
print("="*80)

fila = 3

print(f"\nContenido actual de Fila {fila}:")
print(f"   A (Fecha): {ws[f'A{fila}'].value}")
print(f"   B (Concepto): {ws[f'B{fila}'].value}")
print(f"   C (Cuenta): {ws[f'C{fila}'].value}")
print(f"   D (Ingreso): {ws[f'D{fila}'].value}")
print(f"   E (Egreso): {ws[f'E{fila}'].value}")
print(f"   F (Balance): {ws[f'F{fila}'].value}")

# Verificar que es la fila correcta
concepto = ws[f'B{fila}'].value
cuenta = ws[f'C{fila}'].value

if concepto and 'Balance inicial Promerica USD' in str(concepto):
    if cuenta and 'Promerica USD' in str(cuenta) and '40000003881774' in str(cuenta):
        print(f"\n✅ Confirmado: Fila {fila} es Promerica USD (40000003881774)")

        # Nuevo balance según extracto bancario 31/10
        nuevo_balance = 3030.89

        print(f"\n📝 Actualizando balance:")
        print(f"   Balance anterior: ${ws[f'F{fila}'].value}")
        print(f"   Balance nuevo: ${nuevo_balance}")

        # Actualizar columnas D (Ingreso) y F (Balance)
        ws[f'D{fila}'] = nuevo_balance
        ws[f'F{fila}'] = nuevo_balance

        # Guardar
        print(f"\n💾 Guardando {EXCEL_FILE}...")
        wb.save(EXCEL_FILE)

        print("\n✅ BALANCE ACTUALIZADO EXITOSAMENTE")

        print("\n" + "="*80)
        print("RESUMEN")
        print("="*80)

        print(f"""
✅ Hoja Efectivo - Fila 3 actualizada:
   Cuenta: Promerica USD (40000003881774)
   Balance anterior: $2,999.24
   Balance nuevo: ${nuevo_balance:,.2f}

📊 PRÓXIMOS PASOS:
   1. Abrir Excel y verificar hoja Efectivo
   2. El balance de Promerica ahora debe mostrar ${nuevo_balance:,.2f}
   3. Este es el saldo real según extracto bancario al 31/10/2025

⚠️ NOTA:
   - El saldo calculado desde TRANSACCIONES es $2,761.43
   - El saldo del extracto bancario (31/10) es ${nuevo_balance:,.2f}
   - Diferencia de ${nuevo_balance - 2761.43:.2f} probablemente son:
     * Movimientos de noviembre ya registrados (-$269)
     * Pequeñas comisiones o ajustes
""")

        print("\n" + "="*80)

    else:
        print(f"\n⚠️ ADVERTENCIA: Fila {fila} no parece ser Promerica USD 1774")
        print(f"   Cuenta encontrada: {cuenta}")
        print("   No se realizarán cambios por seguridad")
else:
    print(f"\n⚠️ ADVERTENCIA: Fila {fila} no tiene el concepto esperado")
    print(f"   Concepto encontrado: {concepto}")
    print("   No se realizarán cambios por seguridad")

print("\n" + "="*80)
