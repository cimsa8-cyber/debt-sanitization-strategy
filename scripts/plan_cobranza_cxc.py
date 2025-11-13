#!/usr/bin/env python3
"""
PLAN DE COBRANZA CxC - Estrategia de Supervivencia
Prioriza clientes según días vencido, monto, y necesidad de liquidez
"""

import openpyxl
from datetime import datetime

V3_FILE = "AlvaroVelasco_Finanzas_v3.0.xlsx"

print("\n" + "="*70)
print("PLAN DE COBRANZA CxC - ESTRATEGIA DE SUPERVIVENCIA")
print("="*70)

wb = openpyxl.load_workbook(V3_FILE, data_only=True)
ws_cxc = wb['CxC']
ws_efectivo = wb['EFECTIVO']

# ============================================================================
# CALCULAR EFECTIVO NETO
# ============================================================================

total_bancos = 0
total_tarjetas = 0

for row in range(5, 14):
    saldo = ws_efectivo.cell(row, 5).value or 0
    if isinstance(saldo, (int, float)):
        total_bancos += saldo

for row in range(16, 21):
    saldo = ws_efectivo.cell(row, 5).value or 0
    if isinstance(saldo, (int, float)):
        total_tarjetas += abs(saldo)

efectivo_neto = total_bancos - total_tarjetas

# Gasto diario promedio (calculado anteriormente: $23,231/30)
gasto_diario = 855.77  # Aproximado
dias_cobertura_actual = efectivo_neto / gasto_diario if gasto_diario > 0 else 0

print(f"\n💰 SITUACIÓN ACTUAL:")
print(f"   Efectivo Neto: ${efectivo_neto:,.2f}")
print(f"   Gasto Diario: ${gasto_diario:,.2f}")
print(f"   Días Cobertura: {dias_cobertura_actual:.1f} días")

# ============================================================================
# EXTRAER CxC
# ============================================================================

print(f"\n📊 ANALIZANDO CUENTAS POR COBRAR...")

cxc_data = []

for row in range(3, 25):
    cliente = ws_cxc.cell(row, 1).value
    factura = ws_cxc.cell(row, 2).value
    fecha_emision = ws_cxc.cell(row, 3).value
    fecha_venc = ws_cxc.cell(row, 4).value
    monto = ws_cxc.cell(row, 5).value or 0
    saldo = ws_cxc.cell(row, 6).value or 0
    dias_credito = ws_cxc.cell(row, 7).value or 0
    dias_vencido = ws_cxc.cell(row, 8).value or 0
    estado = ws_cxc.cell(row, 9).value or ""

    if cliente and isinstance(saldo, (int, float)) and saldo > 0:
        # Calcular días vencido manualmente si es necesario
        if isinstance(dias_vencido, str) or dias_vencido is None:
            if fecha_venc and isinstance(fecha_venc, datetime):
                dias_vencido = (datetime.now() - fecha_venc).days
                if dias_vencido < 0:
                    dias_vencido = 0
            else:
                dias_vencido = 0

        cxc_data.append({
            'cliente': cliente,
            'factura': factura,
            'saldo': saldo,
            'dias_vencido': dias_vencido,
            'dias_credito': dias_credito,
            'estado': estado,
            'impacto_dias': saldo / gasto_diario if gasto_diario > 0 else 0
        })

total_cxc = sum(c['saldo'] for c in cxc_data)
print(f"   Total CxC: ${total_cxc:,.2f}")
print(f"   Clientes con saldo: {len(cxc_data)}")

# ============================================================================
# PRIORIZAR COBRANZA
# ============================================================================

print(f"\n🎯 PRIORIZANDO COBRANZA...")

# Scoring:
# - Días vencido: peso 50% (más vencido = más puntos)
# - Monto: peso 30% (más grande = más puntos)
# - Impacto días: peso 20%

max_dias = max([c['dias_vencido'] for c in cxc_data]) if cxc_data else 1
max_monto = max([c['saldo'] for c in cxc_data]) if cxc_data else 1

for c in cxc_data:
    score_dias = (c['dias_vencido'] / max_dias * 50) if max_dias > 0 else 0
    score_monto = (c['saldo'] / max_monto * 30) if max_monto > 0 else 0
    score_impacto = (c['impacto_dias'] / 10 * 20)  # Normalizado a 10 días

    c['score'] = score_dias + score_monto + score_impacto

    # Urgencia
    if c['dias_vencido'] > 60:
        c['urgencia'] = "🔴 CRÍTICA"
    elif c['dias_vencido'] > 30:
        c['urgencia'] = "🟠 ALTA"
    elif c['dias_vencido'] > 0:
        c['urgencia'] = "🟡 MEDIA"
    else:
        c['urgencia'] = "🟢 BAJA"

# Ordenar por score descendente
cxc_ranking = sorted(cxc_data, key=lambda x: x['score'], reverse=True)

# ============================================================================
# REPORTE PRIORIZADO
# ============================================================================

print("\n" + "="*70)
print("RANKING DE COBRANZA (Mayor prioridad primero)")
print("="*70)

cobrado_acumulado = 0
dias_ganados_acumulado = 0

print(f"\n{'#':<3} {'CLIENTE':<30} {'SALDO':<12} {'DÍAS':<6} {'URGENCIA':<12} {'IMPACTO':<10}")
print("-" * 70)

for i, c in enumerate(cxc_ranking[:15], 1):  # Top 15
    cobrado_acumulado += c['saldo']
    dias_ganados_acumulado += c['impacto_dias']

    print(f"{i:<3} {c['cliente'][:28]:<30} ${c['saldo']:>10,.2f} {c['dias_vencido']:>4.0f}d {c['urgencia']:<12} +{c['impacto_dias']:.1f}d")

# ============================================================================
# ESCENARIOS DE COBRANZA
# ============================================================================

print("\n" + "="*70)
print("ESCENARIOS DE COBRANZA")
print("="*70)

escenarios = [
    ("Escenario 1: TOP 3 (más críticos)", 3),
    ("Escenario 2: TOP 5 (moderado)", 5),
    ("Escenario 3: TOP 10 (agresivo)", 10),
    ("Escenario 4: TODO (100%)", len(cxc_ranking))
]

for nombre, top_n in escenarios:
    cobro_total = sum(c['saldo'] for c in cxc_ranking[:top_n])
    dias_ganados = sum(c['impacto_dias'] for c in cxc_ranking[:top_n])
    nuevo_efectivo = efectivo_neto + cobro_total
    nuevos_dias = nuevo_efectivo / gasto_diario

    print(f"\n{nombre}")
    print(f"  💵 Cobro: ${cobro_total:,.2f} ({cobro_total/total_cxc*100:.1f}% del total)")
    print(f"  ⏰ Días ganados: +{dias_ganados:.1f} días")
    print(f"  📊 Nueva cobertura: {nuevos_dias:.1f} días")

    if nuevos_dias >= 30:
        print(f"  ✅ OBJETIVO ALCANZADO (≥30 días)")
    elif nuevos_dias >= 20:
        print(f"  ⚠️  CERCA DEL OBJETIVO")
    else:
        print(f"  🔴 INSUFICIENTE - Necesitás más")

# ============================================================================
# RECOMENDACIONES
# ============================================================================

print("\n" + "="*70)
print("💡 RECOMENDACIONES ESTRATÉGICAS")
print("="*70)

# Calcular cuánto necesitás cobrar para llegar a 30 días
dias_objetivo = 30
efectivo_necesario = dias_objetivo * gasto_diario
faltante = efectivo_necesario - efectivo_neto

print(f"\n1. OBJETIVO: {dias_objetivo} días de cobertura")
print(f"   Necesitás: ${efectivo_necesario:,.2f}")
print(f"   Tenés: ${efectivo_neto:,.2f}")
print(f"   Faltante: ${faltante:,.2f}")

# Cuántos clientes necesitás cobrar
cobro_acumulado = 0
clientes_necesarios = 0
for c in cxc_ranking:
    if cobro_acumulado < faltante:
        clientes_necesarios += 1
        cobro_acumulado += c['saldo']

print(f"\n2. PLAN MÍNIMO:")
print(f"   Cobrar TOP {clientes_necesarios} clientes = ${cobro_acumulado:,.2f}")
print(f"   Esto te da {(efectivo_neto + cobro_acumulado) / gasto_diario:.1f} días de cobertura")

print(f"\n3. ACCIONES INMEDIATAS:")
print(f"   🔴 CRÍTICO: Cobrar {sum(1 for c in cxc_data if c['dias_vencido'] > 60)} facturas vencidas >60 días")
print(f"   🟠 ALTA: Cobrar {sum(1 for c in cxc_data if 30 < c['dias_vencido'] <= 60)} facturas vencidas 30-60 días")

print(f"\n4. CLIENTES PRIORITARIOS (TOP 5):")
for i, c in enumerate(cxc_ranking[:5], 1):
    print(f"   {i}. {c['cliente']}: ${c['saldo']:,.2f} ({c['dias_vencido']:.0f} días vencido)")

print("\n" + "="*70)

wb.close()
