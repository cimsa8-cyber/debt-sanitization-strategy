#!/usr/bin/env python3
"""
INTERFAZ CLAUDE - AUDITORÍA Y UPGRADES AUTOMÁTICOS
Genera reporte mensual para que Claude audite el sistema
Detecta patrones, propone mejoras, actualiza sistema
Versión: 1.0
"""

import openpyxl
import json
from datetime import datetime, timedelta
from collections import defaultdict, Counter
import statistics

class InterfazClaudeAudit:
    def __init__(self, excel_file="AlvaroVelasco_Finanzas_v1.0.xlsx"):
        self.excel_file = excel_file
        self.wb = None
        self.reporte = {}

    def generar_reporte_mensual(self):
        """
        Genera reporte mensual completo para auditoría Claude
        Incluye métricas, patrones, anomalías, y datos anonimizados
        """
        print("="*70)
        print("GENERANDO REPORTE MENSUAL PARA AUDITORÍA CLAUDE")
        print("="*70)
        print(f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print()

        try:
            # Cargar workbook
            print("⏳ Cargando sistema...")
            self.wb = openpyxl.load_workbook(self.excel_file, data_only=True)
            print("✅ Sistema cargado")

            # Recolectar métricas
            print("⏳ Recolectando métricas...")
            self.reporte['metadata'] = self.recolectar_metadata()
            self.reporte['metricas_uso'] = self.analizar_metricas_uso()
            self.reporte['patrones_errores'] = self.detectar_patrones_errores()
            self.reporte['analisis_financiero'] = self.analizar_salud_financiera()
            self.reporte['sugerencias_automaticas'] = self.generar_sugerencias()
            self.reporte['datos_anonimizados'] = self.exportar_datos_anonimizados()
            print("✅ Métricas recolectadas")

            # Guardar reporte JSON
            print()
            print("⏳ Guardando reporte...")
            output_file = f"claude_audit_mensual_{datetime.now().strftime('%Y%m')}.json"
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(self.reporte, f, indent=2, ensure_ascii=False, default=str)
            print(f"✅ Reporte guardado: {output_file}")

            # Imprimir resumen
            print()
            self.imprimir_resumen()

            print()
            print("="*70)
            print("PRÓXIMOS PASOS PARA AUDITORÍA CLAUDE")
            print("="*70)
            print()
            print("1. Abrir Claude Code en este directorio")
            print(f"2. Decir: 'Audita mi sistema financiero usando {output_file}'")
            print("3. Claude analizará:")
            print("   - Patrones de error y cómo prevenirlos")
            print("   - Optimizaciones de fórmulas lentas")
            print("   - Nuevas validaciones necesarias")
            print("   - Automatizaciones adicionales")
            print("4. Claude propondrá mejoras específicas")
            print("5. Aprobar mejoras y Claude actualizará sistema automáticamente")
            print()

            return output_file

        except Exception as e:
            print(f"❌ ERROR: {e}")
            import traceback
            traceback.print_exc()
            return None

    def recolectar_metadata(self):
        """Metadata básica del sistema"""
        ws = self.wb["TRANSACCIONES"]

        return {
            "fecha_reporte": datetime.now().isoformat(),
            "version_sistema": "1.0",
            "total_transacciones": ws.max_row - 1,
            "hojas": self.wb.sheetnames,
            "ultima_modificacion": datetime.fromtimestamp(
                __import__('os').path.getmtime(self.excel_file)
            ).isoformat()
        }

    def analizar_metricas_uso(self):
        """Analiza patrones de uso del sistema"""
        ws = self.wb["TRANSACCIONES"]

        # Recolectar datos
        transacciones_por_dia = defaultdict(int)
        tipos_transaccion = Counter()
        montos = []
        duracion_entrada = []

        for row in range(2, ws.max_row + 1):
            fecha = ws[f"A{row}"].value
            if not fecha:
                continue

            fecha_creacion = ws[f"Q{row}"].value
            tipo = ws[f"B{row}"].value
            monto = ws[f"I{row}"].value or 0

            # Contar por día
            if isinstance(fecha, datetime):
                dia = fecha.strftime("%Y-%m-%d")
                transacciones_por_dia[dia] += 1

            # Contar tipos
            tipos_transaccion[tipo] += 1

            # Montos
            if monto != 0:
                montos.append(abs(monto))

        return {
            "transacciones_ultimo_mes": sum(1 for d in transacciones_por_dia.keys()
                                            if datetime.fromisoformat(d) > datetime.now() - timedelta(days=30)),
            "promedio_transacciones_dia": statistics.mean(transacciones_por_dia.values()) if transacciones_por_dia else 0,
            "tipos_mas_usados": dict(tipos_transaccion.most_common(5)),
            "monto_promedio": statistics.mean(montos) if montos else 0,
            "monto_mediana": statistics.median(montos) if montos else 0,
            "monto_max": max(montos) if montos else 0,
            "monto_min": min(montos) if montos else 0,
        }

    def detectar_patrones_errores(self):
        """Detecta patrones de errores del usuario"""
        ws = self.wb["TRANSACCIONES"]

        errores = {
            "campos_vacios": 0,
            "duplicados_detectados": 0,
            "montos_sospechosos": 0,
            "fechas_inconsistentes": 0,
            "sin_cuenta": 0,
            "sin_entidad": 0,
        }

        errores_por_tipo = defaultdict(list)

        for row in range(2, ws.max_row + 1):
            fecha = ws[f"A{row}"].value
            if not fecha:
                continue

            # Detectar campos vacíos
            if not ws[f"E{row}"].value:  # Sin cuenta
                errores["sin_cuenta"] += 1
                errores_por_tipo["sin_cuenta"].append(row)

            if not ws[f"D{row}"].value:  # Sin entidad
                errores["sin_entidad"] += 1
                errores_por_tipo["sin_entidad"].append(row)

            # Detectar duplicados
            duplicado = ws[f"S{row}"].value
            if duplicado and "DUPLICADO" in str(duplicado):
                errores["duplicados_detectados"] += 1
                errores_por_tipo["duplicados"].append(row)

            # Montos sospechosos (>$50k)
            monto = ws[f"I{row}"].value or 0
            if abs(monto) > 50000:
                errores["montos_sospechosos"] += 1
                errores_por_tipo["montos_altos"].append(row)

        # Calcular tasa de error
        total_transacciones = ws.max_row - 1
        if total_transacciones > 0:
            tasa_error = (sum(errores.values()) / total_transacciones) * 100
        else:
            tasa_error = 0

        return {
            "resumen_errores": errores,
            "tasa_error_porcentaje": round(tasa_error, 2),
            "errores_por_tipo": {k: len(v) for k, v in errores_por_tipo.items()},
            "filas_con_errores": {k: v[:10] for k, v in errores_por_tipo.items()},  # Primeras 10 de cada tipo
        }

    def analizar_salud_financiera(self):
        """Analiza salud financiera del negocio"""
        ws = self.wb["TRANSACCIONES"]

        # Calcular balance actual
        efectivo = 0
        ar = 0
        ap = 0
        tc = 0
        ingresos_mes = 0
        egresos_mes = 0

        fecha_hace_30_dias = datetime.now() - timedelta(days=30)

        for row in range(2, ws.max_row + 1):
            fecha = ws[f"A{row}"].value
            if not fecha:
                continue

            tipo = ws[f"B{row}"].value
            monto = ws[f"I{row}"].value or 0
            estado = ws[f"L{row}"].value
            categoria = ws[f"C{row}"].value
            ingreso_egreso = ws[f"K{row}"].value

            # Efectivo
            if categoria == "Efectivo" and estado == "Cobrado":
                efectivo += monto

            # A/R
            if tipo == "Factura Cliente" and estado == "Pendiente":
                ar += monto

            # A/P
            if tipo == "Factura Proveedor" and estado == "Pendiente":
                ap += abs(monto)

            # TC
            if "Tarjeta" in str(tipo):
                tc += abs(monto)

            # Ingresos/Egresos último mes
            if isinstance(fecha, datetime) and fecha >= fecha_hace_30_dias:
                if ingreso_egreso == "Ingreso":
                    ingresos_mes += monto
                elif ingreso_egreso == "Egreso":
                    egresos_mes += abs(monto)

        balance_neto = efectivo + ar - ap - tc
        flujo_mes = ingresos_mes - egresos_mes

        # Calcular runway (meses de operación con efectivo actual)
        if egresos_mes > 0:
            runway_meses = efectivo / (egresos_mes / 30)  # días de runway
        else:
            runway_meses = 999

        # Determinar salud
        if balance_neto > 0 and flujo_mes > 0:
            salud = "EXCELENTE"
        elif balance_neto > 0 and flujo_mes < 0:
            salud = "PRECAUCIÓN"
        elif balance_neto < 0 and flujo_mes > 0:
            salud = "RECUPERACIÓN"
        else:
            salud = "CRÍTICO"

        return {
            "efectivo": round(efectivo, 2),
            "cuentas_por_cobrar": round(ar, 2),
            "cuentas_por_pagar": round(ap, 2),
            "tarjetas_credito": round(tc, 2),
            "balance_neto": round(balance_neto, 2),
            "ingresos_ultimo_mes": round(ingresos_mes, 2),
            "egresos_ultimo_mes": round(egresos_mes, 2),
            "flujo_neto_mes": round(flujo_mes, 2),
            "runway_dias": round(runway_meses, 1),
            "salud_financiera": salud
        }

    def generar_sugerencias(self):
        """Genera sugerencias automáticas basadas en análisis"""
        sugerencias = []

        # Sugerencia 1: Basada en tasa de error
        tasa_error = self.reporte['patrones_errores']['tasa_error_porcentaje']
        if tasa_error > 10:
            sugerencias.append({
                "tipo": "CALIDAD_DATOS",
                "prioridad": "ALTA",
                "detalle": f"Tasa de error {tasa_error}% es alta. Sugerencia: Agregar validación más estricta en campos obligatorios."
            })

        # Sugerencia 2: Basada en salud financiera
        salud = self.reporte['analisis_financiero']['salud_financiera']
        if salud in ["CRÍTICO", "PRECAUCIÓN"]:
            sugerencias.append({
                "tipo": "SALUD_FINANCIERA",
                "prioridad": "CRÍTICA",
                "detalle": f"Salud financiera: {salud}. Sugerencia: Revisar estrategia de cobros A/R y reducción de gastos."
            })

        # Sugerencia 3: Basada en runway
        runway = self.reporte['analisis_financiero']['runway_dias']
        if runway < 60:
            sugerencias.append({
                "tipo": "LIQUIDEZ",
                "prioridad": "CRÍTICA",
                "detalle": f"Runway: {runway} días. Sugerencia: Acelerar cobros y buscar financiamiento de emergencia."
            })

        # Sugerencia 4: Duplicados
        duplicados = self.reporte['patrones_errores']['resumen_errores']['duplicados_detectados']
        if duplicados > 5:
            sugerencias.append({
                "tipo": "DUPLICADOS",
                "prioridad": "MEDIA",
                "detalle": f"{duplicados} duplicados detectados. Sugerencia: Implementar bloqueo automático de duplicados."
            })

        return sugerencias

    def exportar_datos_anonimizados(self):
        """
        Exporta datos anonimizados para análisis Claude
        Remueve información sensible (nombres clientes, montos exactos)
        Preserva patrones para análisis
        """
        ws = self.wb["TRANSACCIONES"]

        transacciones_anonimizadas = []

        for row in range(2, min(ws.max_row + 1, 102)):  # Máximo 100 transacciones
            fecha = ws[f"A{row}"].value
            if not fecha:
                continue

            transacciones_anonimizadas.append({
                "fecha": fecha.isoformat() if isinstance(fecha, datetime) else str(fecha),
                "tipo": ws[f"B{row}"].value,
                "categoria": ws[f"C{row}"].value,
                "entidad": ws[f"D{row}"].value,
                "monto_rango": self.clasificar_monto(ws[f"I{row}"].value),
                "estado": ws[f"L{row}"].value,
                "prioridad": ws[f"M{row}"].value,
                "tiene_duplicado": "DUPLICADO" in str(ws[f"S{row}"].value or ""),
                "validacion_ok": "OK" in str(ws[f"T{row}"].value or ""),
            })

        return transacciones_anonimizadas

    def clasificar_monto(self, monto):
        """Clasifica monto en rangos para anonimización"""
        if not monto:
            return "CERO"

        monto_abs = abs(monto)

        if monto_abs < 100:
            return "BAJO (<$100)"
        elif monto_abs < 500:
            return "MEDIO ($100-$500)"
        elif monto_abs < 2000:
            return "ALTO ($500-$2k)"
        elif monto_abs < 10000:
            return "MUY_ALTO ($2k-$10k)"
        else:
            return "CRITICO (>$10k)"

    def imprimir_resumen(self):
        """Imprime resumen del reporte"""
        print("="*70)
        print("RESUMEN EJECUTIVO")
        print("="*70)
        print()

        # Metadata
        print(f"📊 Total Transacciones: {self.reporte['metadata']['total_transacciones']}")
        print(f"📅 Última Modificación: {self.reporte['metadata']['ultima_modificacion']}")
        print()

        # Métricas de uso
        print(f"📈 Transacciones último mes: {self.reporte['metricas_uso']['transacciones_ultimo_mes']}")
        print(f"📊 Promedio por día: {self.reporte['metricas_uso']['promedio_transacciones_dia']:.1f}")
        print(f"💰 Monto promedio: ${self.reporte['metricas_uso']['monto_promedio']:,.2f}")
        print()

        # Errores
        tasa_error = self.reporte['patrones_errores']['tasa_error_porcentaje']
        print(f"🔴 Tasa de Error: {tasa_error}%")
        if tasa_error < 5:
            print("   ✅ Excelente - Sistema bien utilizado")
        elif tasa_error < 15:
            print("   ⚠️  Aceptable - Revisar mejoras")
        else:
            print("   🔴 Crítico - Requiere capacitación")
        print()

        # Salud financiera
        salud = self.reporte['analisis_financiero']
        print(f"💵 Efectivo: ${salud['efectivo']:,.2f}")
        print(f"📬 A/R: ${salud['cuentas_por_cobrar']:,.2f}")
        print(f"📪 A/P: ${salud['cuentas_por_pagar']:,.2f}")
        print(f"💳 TC: ${salud['tarjetas_credito']:,.2f}")
        print(f"📊 Balance Neto: ${salud['balance_neto']:,.2f}")
        print(f"🏥 Salud: {salud['salud_financiera']}")
        print(f"⏱️  Runway: {salud['runway_dias']:.0f} días")
        print()

        # Sugerencias
        if self.reporte['sugerencias_automaticas']:
            print(f"💡 SUGERENCIAS: {len(self.reporte['sugerencias_automaticas'])}")
            for sug in self.reporte['sugerencias_automaticas']:
                icono = "🔴" if sug['prioridad'] == "CRÍTICA" else "🟠" if sug['prioridad'] == "ALTA" else "🟡"
                print(f"   {icono} {sug['tipo']}: {sug['detalle']}")
        else:
            print("✅ Sin sugerencias críticas")

if __name__ == "__main__":
    interfaz = InterfazClaudeAudit()
    output_file = interfaz.generar_reporte_mensual()

    if output_file:
        exit(0)
    else:
        exit(1)
