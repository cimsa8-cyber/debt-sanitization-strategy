#!/usr/bin/env python3
"""
PROYECCIONES DE FLUJO - Estrategia de Supervivencia
Escenarios combinados: Cobranza CxC + Pagos CxP + Gastos proyectados
"""

import openpyxl
from datetime import datetime, timedelta

V3_FILE = "AlvaroVelasco_Finanzas_v3.0.xlsx"

print("\n" + "="*70)
print("PROYECCIONES DE FLUJO - ESTRATEGIA DE SUPERVIVENCIA")
print("="*70)

wb = openpyxl.load_workbook(V3_FILE, data_only=True)
ws_efectivo = wb['EFECTIVO']
ws_cxc = wb['CxC']
ws_cxp = wb['CxP']
ws_trans = wb['TRANSACCIONES']

# ============================================================================
# CALCULAR SITUACIÓN ACTUAL
# ============================================================================

# Efectivo
total_bancos = 0
total_tarjetas = 0

for row in range(5, 14):
    saldo = ws_efectivo.cell(row, 5).value or 0
    if isinstance(saldo, (int, float)):
        total_bancos += saldo

for row in range(16, 21):
    saldo = ws_efectivo.cell(row, 5).value or 0
    if isinstance(saldo, (int, float)):
        total_tarjetas += abs(saldo)

efectivo_actual = total_bancos - total_tarjetas

# CxC total
total_cxc = 0
for row in range(3, 25):
    saldo = ws_cxc.cell(row, 6).value or 0
    if isinstance(saldo, (int, float)):
        total_cxc += saldo

# CxP por categoría
cxp_critico = 0
cxp_alta = 0
cxp_total = 0

for row in range(3, 25):
    saldo = ws_cxp.cell(row, 6).value or 0
    prioridad = ws_cxp.cell(row, 8).value or ""

    if isinstance(saldo, (int, float)) and saldo > 0:
        cxp_total += saldo
        if "CRÍTICA" in str(prioridad).upper() or "CRITICA" in str(prioridad).upper():
            cxp_critico += saldo
        elif "ALTA" in str(prioridad).upper():
            cxp_alta += saldo

# Gasto mensual promedio
gastos_totales = 0
for row in range(2, ws_trans.max_row + 1):
    tipo = ws_trans.cell(row, 2).value
    monto_usd = ws_trans.cell(row, 9).value or 0
    if tipo and ("GASTO" in str(tipo).upper() or "COMPRA" in str(tipo).upper()):
        gastos_totales += monto_usd

gasto_mensual = gastos_totales  # Ya es del mes actual
gasto_diario = gastos_totales / 30

# Ingreso mensual promedio
ingresos_totales = 0
for row in range(2, ws_trans.max_row + 1):
    tipo = ws_trans.cell(row, 2).value
    monto_usd = ws_trans.cell(row, 9).value or 0
    if tipo and "INGRESO" in str(tipo).upper():
        ingresos_totales += monto_usd

ingreso_mensual = ingresos_totales

print(f"\n📊 SITUACIÓN ACTUAL ({datetime.now().strftime('%d/%b/%Y')})")
print("-" * 70)
print(f"   💰 Efectivo: ${efectivo_actual:,.2f}")
print(f"   📥 CxC Total: ${total_cxc:,.2f}")
print(f"   📤 CxP Total: ${cxp_total:,.2f}")
print(f"      • Crítico: ${cxp_critico:,.2f}")
print(f"      • Alta: ${cxp_alta:,.2f}")
print(f"   📊 Gasto Mensual: ${gasto_mensual:,.2f} (${gasto_diario:,.2f}/día)")
print(f"   📈 Ingreso Mensual: ${ingreso_mensual:,.2f}")

# ============================================================================
# ESCENARIOS DE PROYECCIÓN
# ============================================================================

print("\n" + "="*70)
print("ESCENARIOS DE PROYECCIÓN (30, 60, 90 DÍAS)")
print("="*70)

escenarios = [
    {
        "nombre": "🟢 CONSERVADOR",
        "descripcion": "Cobro 50% CxC | Pago solo CRÍTICO | Ingreso 80%",
        "cobro_cxc": total_cxc * 0.50,
        "pago_cxp": cxp_critico,
        "factor_ingreso": 0.80,  # 80% de los ingresos habituales
    },
    {
        "nombre": "🟡 MODERADO",
        "descripcion": "Cobro 75% CxC | Pago CRÍTICO+ALTA | Ingreso 100%",
        "cobro_cxc": total_cxc * 0.75,
        "pago_cxp": cxp_critico + cxp_alta,
        "factor_ingreso": 1.00,
    },
    {
        "nombre": "🔴 AGRESIVO",
        "descripcion": "Cobro 100% CxC | Pago TODO urgente | Ingreso 120%",
        "cobro_cxc": total_cxc,
        "pago_cxp": cxp_critico + cxp_alta * 1.5,  # Más pagos urgentes
        "factor_ingreso": 1.20,  # Crecimiento optimista
    },
]

for esc in escenarios:
    print(f"\n{esc['nombre']}: {esc['descripcion']}")
    print("-" * 70)

    # Punto de partida
    efectivo = efectivo_actual

    # Operación inicial (mes 1)
    cobro = esc['cobro_cxc']
    pago = esc['pago_cxp']
    ingreso_proyectado = ingreso_mensual * esc['factor_ingreso']
    gasto_proyectado = gasto_mensual

    # Mes 1
    efectivo_mes1 = efectivo + cobro - pago + ingreso_proyectado - gasto_proyectado
    dias_cobertura_mes1 = efectivo_mes1 / gasto_diario if gasto_diario > 0 else 0

    # Mes 2 (sin cobranza CxC extraordinaria, solo flujo normal)
    efectivo_mes2 = efectivo_mes1 + ingreso_proyectado - gasto_proyectado
    dias_cobertura_mes2 = efectivo_mes2 / gasto_diario if gasto_diario > 0 else 0

    # Mes 3
    efectivo_mes3 = efectivo_mes2 + ingreso_proyectado - gasto_proyectado
    dias_cobertura_mes3 = efectivo_mes3 / gasto_diario if gasto_diario > 0 else 0

    print(f"\n   OPERACIÓN INICIAL:")
    print(f"   + Cobro CxC: ${cobro:,.2f}")
    print(f"   - Pago CxP: ${pago:,.2f}")
    print(f"   = Neto operación: ${cobro - pago:,.2f}")

    print(f"\n   MES 1 (0-30 días):")
    print(f"   Efectivo inicial: ${efectivo:,.2f}")
    print(f"   + Cobro CxC: ${cobro:,.2f}")
    print(f"   - Pago CxP: ${pago:,.2f}")
    print(f"   + Ingresos: ${ingreso_proyectado:,.2f}")
    print(f"   - Gastos: ${gasto_proyectado:,.2f}")
    print(f"   = Efectivo final: ${efectivo_mes1:,.2f}")
    print(f"   📅 Días cobertura: {dias_cobertura_mes1:.1f} días")

    if dias_cobertura_mes1 < 15:
        print(f"   ❌ CRÍTICO - Menos de 15 días")
    elif dias_cobertura_mes1 < 30:
        print(f"   ⚠️  AJUSTADO - Entre 15-30 días")
    else:
        print(f"   ✅ SALUDABLE - Más de 30 días")

    print(f"\n   MES 2 (31-60 días):")
    print(f"   + Ingresos: ${ingreso_proyectado:,.2f}")
    print(f"   - Gastos: ${gasto_proyectado:,.2f}")
    print(f"   = Efectivo final: ${efectivo_mes2:,.2f}")
    print(f"   📅 Días cobertura: {dias_cobertura_mes2:.1f} días")

    if dias_cobertura_mes2 < 15:
        print(f"   ❌ CRÍTICO")
    elif dias_cobertura_mes2 < 30:
        print(f"   ⚠️  AJUSTADO")
    else:
        print(f"   ✅ SALUDABLE")

    print(f"\n   MES 3 (61-90 días):")
    print(f"   + Ingresos: ${ingreso_proyectado:,.2f}")
    print(f"   - Gastos: ${gasto_proyectado:,.2f}")
    print(f"   = Efectivo final: ${efectivo_mes3:,.2f}")
    print(f"   📅 Días cobertura: {dias_cobertura_mes3:.1f} días")

    if dias_cobertura_mes3 < 15:
        print(f"   ❌ CRÍTICO")
    elif dias_cobertura_mes3 < 30:
        print(f"   ⚠️  AJUSTADO")
    else:
        print(f"   ✅ SALUDABLE")

    # Resumen del escenario
    print(f"\n   📊 RESUMEN 90 DÍAS:")
    variacion = efectivo_mes3 - efectivo
    variacion_pct = (variacion / efectivo * 100) if efectivo > 0 else 0

    print(f"   Efectivo inicial: ${efectivo:,.2f}")
    print(f"   Efectivo final (90d): ${efectivo_mes3:,.2f}")
    print(f"   Variación: ${variacion:,.2f} ({variacion_pct:+.1f}%)")

    if efectivo_mes3 > efectivo * 1.1:
        print(f"   ✅ CRECIMIENTO SOSTENIDO")
    elif efectivo_mes3 > efectivo * 0.9:
        print(f"   ⚠️  ESTABLE (±10%)")
    else:
        print(f"   ❌ DECRECIMIENTO - Ajustar gastos")

# ============================================================================
# RECOMENDACIONES
# ============================================================================

print("\n" + "="*70)
print("💡 RECOMENDACIONES FINALES")
print("="*70)

# Calcular escenario recomendado
if efectivo_actual > 30000:
    esc_recomendado = "🟡 MODERADO"
    razon = "Tenés buena liquidez actual, podés ser proactivo"
elif total_cxc > cxp_total:
    esc_recomendado = "🟢 CONSERVADOR"
    razon = "Cobrás más de lo que debés, pero mantené prudencia"
else:
    esc_recomendado = "🔴 AGRESIVO"
    razon = "Debés más de lo que te deben, necesitás cobranza agresiva"

print(f"\n1. ESCENARIO RECOMENDADO: {esc_recomendado}")
print(f"   {razon}")

print(f"\n2. ACCIONES PRIORITARIAS:")
print(f"   🎯 Cobrar TOP 5 clientes CxC: ${total_cxc * 0.75:,.2f}")
print(f"   💵 Pagar facturas CRÍTICAS: ${cxp_critico:,.2f}")
print(f"   📊 Mantener gastos bajo control: ${gasto_mensual:,.2f}/mes")

print(f"\n3. INDICADORES A MONITOREAR:")
print(f"   • Días de cobertura: mantener >30 días")
print(f"   • Ratio CxC/CxP: actual {(total_cxc/cxp_total) if cxp_total > 0 else 0:.2f} (ideal >1.0)")
print(f"   • Gasto vs Ingreso: actual {(gasto_mensual/ingreso_mensual) if ingreso_mensual > 0 else 0:.2f} (ideal <0.8)")

print(f"\n4. PLAN DE CONTINGENCIA:")
print(f"   Si efectivo <$15,000:")
print(f"      • Negociar plazo con Hacienda IVA/Renta ($10,000)")
print(f"      • Cobrar VWR International ($3,567) - cliente grande")
print(f"      • Diferir pagos prioridad BAJA")

print("\n" + "="*70)

wb.close()
