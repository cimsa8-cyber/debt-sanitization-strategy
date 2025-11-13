#!/usr/bin/env python3
"""
Poblar IVA_CONTROL con datos de TRANSACCIONES
Extrae INGRESOS (ventas) y GASTOS (compras) de noviembre 2025
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

V3_FILE = "AlvaroVelasco_Finanzas_v3.0.xlsx"

COLOR_EDITABLE = "FFF2CC"
COLOR_CALCULATED = "E7E6E6"

# Clientes zona franca (no pagan IVA)
ZONA_FRANCA = ["VWR International", "RSHughes", "VWR", "RS Hughes"]

print("\n" + "="*60)
print("POBLAR IVA_CONTROL desde TRANSACCIONES")
print("="*60)

wb = openpyxl.load_workbook(V3_FILE)
ws_trans = wb['TRANSACCIONES']
ws_iva = wb['IVA_CONTROL']

# ============================================================================
# EXTRAER INGRESOS (VENTAS)
# ============================================================================

print("\n📊 Extrayendo INGRESOS (Ventas)...")
ventas = []

for row in range(2, ws_trans.max_row + 1):
    tipo = ws_trans.cell(row, 2).value
    if tipo == "INGRESOS":
        fecha = ws_trans.cell(row, 1).value
        desc = ws_trans.cell(row, 4).value or ""
        entidad = ws_trans.cell(row, 6).value or desc
        factura = ws_trans.cell(row, 7).value or f"T-{row-1}"
        monto_usd = ws_trans.cell(row, 9).value or 0

        # Detectar zona franca
        es_zona_franca = any(zf.lower() in str(entidad).lower() for zf in ZONA_FRANCA)

        if monto_usd > 0:
            ventas.append({
                'fecha': fecha,
                'factura': factura,
                'cliente': entidad,
                'monto': monto_usd,
                'zona_franca': es_zona_franca,
                'ref_trans': row - 1
            })

print(f"   ✅ {len(ventas)} ventas encontradas")

# Poblar sección VENTAS (filas 6-20)
row_iva = 6
for venta in ventas[:15]:  # Máximo 15 ventas (filas 6-20)
    # Fecha
    ws_iva.cell(row_iva, 1, venta['fecha'])
    ws_iva.cell(row_iva, 1).number_format = 'DD/MM/YY'

    # Factura
    ws_iva.cell(row_iva, 2, venta['factura'])

    # Cliente
    ws_iva.cell(row_iva, 3, venta['cliente'])

    # Base Gravable USD (monto / 1.13 si incluía IVA, o monto directo)
    # Asumimos que los montos en USD NO incluyen IVA (se factura aparte)
    base_gravable = venta['monto']
    ws_iva.cell(row_iva, 4, base_gravable)
    ws_iva.cell(row_iva, 4).number_format = '$#,##0.00'
    ws_iva.cell(row_iva, 4).fill = PatternFill(start_color=COLOR_EDITABLE, end_color=COLOR_EDITABLE, fill_type="solid")

    # IVA 13% (fórmula ya existe)
    ws_iva.cell(row_iva, 5, f'=D{row_iva}*0.13')
    ws_iva.cell(row_iva, 5).number_format = '$#,##0.00'
    ws_iva.cell(row_iva, 5).fill = PatternFill(start_color=COLOR_CALCULATED, end_color=COLOR_CALCULATED, fill_type="solid")

    # Total USD (fórmula ya existe)
    ws_iva.cell(row_iva, 6, f'=D{row_iva}+E{row_iva}')
    ws_iva.cell(row_iva, 6).number_format = '$#,##0.00'
    ws_iva.cell(row_iva, 6).fill = PatternFill(start_color=COLOR_CALCULATED, end_color=COLOR_CALCULATED, fill_type="solid")

    # Zona Franca
    ws_iva.cell(row_iva, 7, "SI" if venta['zona_franca'] else "NO")
    ws_iva.cell(row_iva, 7).fill = PatternFill(start_color=COLOR_EDITABLE, end_color=COLOR_EDITABLE, fill_type="solid")

    # Retención 2% (fórmula ya existe)
    ws_iva.cell(row_iva, 8, f'=IF(G{row_iva}="SI",0,D{row_iva}*0.02)')
    ws_iva.cell(row_iva, 8).number_format = '$#,##0.00'
    ws_iva.cell(row_iva, 8).fill = PatternFill(start_color=COLOR_CALCULATED, end_color=COLOR_CALCULATED, fill_type="solid")

    # Neto Cobrado (fórmula ya existe)
    ws_iva.cell(row_iva, 9, f'=F{row_iva}-H{row_iva}')
    ws_iva.cell(row_iva, 9).number_format = '$#,##0.00'
    ws_iva.cell(row_iva, 9).fill = PatternFill(start_color=COLOR_CALCULATED, end_color=COLOR_CALCULATED, fill_type="solid")

    # Estado
    ws_iva.cell(row_iva, 10, "EMITIDA")

    # Ref Trans
    ws_iva.cell(row_iva, 11, venta['ref_trans'])

    # Notas zona franca
    if venta['zona_franca']:
        ws_iva.cell(row_iva, 12, "ZONA FRANCA - No aplica IVA")

    row_iva += 1

if len(ventas) > 15:
    print(f"   ⚠️  Solo se poblaron las primeras 15 ventas (hay {len(ventas)} total)")

# ============================================================================
# EXTRAER GASTOS (COMPRAS)
# ============================================================================

print("\n📊 Extrayendo GASTOS OPERATIVOS y COMPRAS...")
compras = []

for row in range(2, ws_trans.max_row + 1):
    tipo = ws_trans.cell(row, 2).value
    if tipo in ["GASTOS OPERATIVOS", "COMPRAS PARA REVENTA"]:
        fecha = ws_trans.cell(row, 1).value
        desc = ws_trans.cell(row, 4).value or ""
        entidad = ws_trans.cell(row, 6).value or desc
        factura = ws_trans.cell(row, 7).value or f"C-{row-1}"
        monto_usd = ws_trans.cell(row, 9).value or 0
        metodo = ws_trans.cell(row, 11).value or "N/D"

        # Gastos operativos y compras son deducibles
        deducible = True

        if monto_usd > 0:
            compras.append({
                'fecha': fecha,
                'factura': factura,
                'proveedor': entidad,
                'monto': monto_usd,
                'deducible': deducible,
                'metodo': metodo,
                'ref_trans': row - 1
            })

print(f"   ✅ {len(compras)} compras encontradas")

# Poblar sección COMPRAS (filas 25-40)
row_iva = 25
for compra in compras[:16]:  # Máximo 16 compras (filas 25-40)
    # Fecha
    ws_iva.cell(row_iva, 1, compra['fecha'])
    ws_iva.cell(row_iva, 1).number_format = 'DD/MM/YY'

    # Factura
    ws_iva.cell(row_iva, 2, compra['factura'])

    # Proveedor
    ws_iva.cell(row_iva, 3, compra['proveedor'])

    # Base Gravable USD
    # Asumimos que el monto incluye IVA, entonces base = monto / 1.13
    base_gravable = compra['monto'] / 1.13
    ws_iva.cell(row_iva, 4, base_gravable)
    ws_iva.cell(row_iva, 4).number_format = '$#,##0.00'
    ws_iva.cell(row_iva, 4).fill = PatternFill(start_color=COLOR_EDITABLE, end_color=COLOR_EDITABLE, fill_type="solid")

    # IVA 13% (fórmula ya existe)
    ws_iva.cell(row_iva, 5, f'=D{row_iva}*0.13')
    ws_iva.cell(row_iva, 5).number_format = '$#,##0.00'
    ws_iva.cell(row_iva, 5).fill = PatternFill(start_color=COLOR_CALCULATED, end_color=COLOR_CALCULATED, fill_type="solid")

    # Total USD (fórmula ya existe)
    ws_iva.cell(row_iva, 6, f'=D{row_iva}+E{row_iva}')
    ws_iva.cell(row_iva, 6).number_format = '$#,##0.00'
    ws_iva.cell(row_iva, 6).fill = PatternFill(start_color=COLOR_CALCULATED, end_color=COLOR_CALCULATED, fill_type="solid")

    # Deducible
    ws_iva.cell(row_iva, 7, "SI" if compra['deducible'] else "NO")
    ws_iva.cell(row_iva, 7).fill = PatternFill(start_color=COLOR_EDITABLE, end_color=COLOR_EDITABLE, fill_type="solid")

    # IVA Acreditable (fórmula ya existe)
    ws_iva.cell(row_iva, 8, f'=IF(G{row_iva}="SI",E{row_iva},0)')
    ws_iva.cell(row_iva, 8).number_format = '$#,##0.00'
    ws_iva.cell(row_iva, 8).fill = PatternFill(start_color=COLOR_CALCULATED, end_color=COLOR_CALCULATED, fill_type="solid")

    # Método Pago
    ws_iva.cell(row_iva, 9, compra['metodo'])

    # Estado
    ws_iva.cell(row_iva, 10, "PAGADA")

    # Ref Trans
    ws_iva.cell(row_iva, 11, compra['ref_trans'])

    row_iva += 1

if len(compras) > 16:
    print(f"   ⚠️  Solo se poblaron las primeras 16 compras (hay {len(compras)} total)")

# ============================================================================
# GUARDAR
# ============================================================================

print(f"\n💾 Guardando {V3_FILE}...")
wb.save(V3_FILE)
wb.close()

# ============================================================================
# RESUMEN
# ============================================================================

print("\n" + "="*60)
print("✅ IVA_CONTROL POBLADO")
print("="*60)
print(f"   📈 Ventas: {min(len(ventas), 15)} registros")
print(f"   📉 Compras: {min(len(compras), 16)} registros")
print(f"   🎯 Zona Franca detectada: {sum(1 for v in ventas if v['zona_franca'])} ventas")
print("\n   📝 SIGUIENTE PASO: Revisar IVA_CONTROL y ajustar si necesario")
print("="*60)
