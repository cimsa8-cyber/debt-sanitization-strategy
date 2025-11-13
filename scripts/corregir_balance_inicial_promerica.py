#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CORRECCI√ìN BALANCE INICIAL PROMERICA
Diagnostica y corrige el balance inicial que aparece como egreso
"""
import openpyxl

EXCEL_FILE = "AlvaroVelasco_Finanzas_v2.0.xlsx"

def corregir_balance_inicial():
    print("=" * 80)
    print("DIAGN√ìSTICO Y CORRECCI√ìN - BALANCE INICIAL PROMERICA")
    print("=" * 80)
    print()

    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=False)
    ws_trans = wb['TRANSACCIONES']

    headers = [ws_trans.cell(1, col).value for col in range(1, ws_trans.max_column + 1)]
    col_map = {}
    for col in range(1, len(headers) + 1):
        if headers[col-1]:
            col_map[headers[col-1]] = col

    # =========================================================================
    # PASO 1: BUSCAR BALANCE INICIAL PROMERICA
    # =========================================================================
    print("üìã PASO 1: Buscando Balance Inicial Promerica...")
    print()

    balance_inicial_fila = None
    for row in range(2, 20):  # Buscar en las primeras filas
        tipo = ws_trans.cell(row, col_map['Tipo Transacci√≥n']).value
        cuenta = ws_trans.cell(row, col_map['Cuenta Bancaria']).value
        concepto = ws_trans.cell(row, col_map['Concepto']).value

        # Buscar balance inicial de Promerica
        if cuenta and 'Promerica' in str(cuenta) and '40000003881774' in str(cuenta):
            if tipo and ('Balance' in str(tipo) or 'TRANSFERENCIAS' in str(tipo)):
                if concepto and 'Balance inicial' in str(concepto):
                    balance_inicial_fila = row
                    break

        # Tambi√©n buscar por concepto solo
        if concepto and 'Balance inicial Promerica' in str(concepto):
            balance_inicial_fila = row
            break

    if not balance_inicial_fila:
        print("‚ö†Ô∏è  No se encontr√≥ Balance Inicial Promerica en las primeras 20 filas")
        print("   Mostrando primeras 5 filas para diagn√≥stico:")
        print()
        for row in range(2, 7):
            fecha = ws_trans.cell(row, col_map['Fecha']).value
            tipo = ws_trans.cell(row, col_map['Tipo Transacci√≥n']).value
            concepto = ws_trans.cell(row, col_map['Concepto']).value
            cuenta = ws_trans.cell(row, col_map['Cuenta Bancaria']).value
            monto = ws_trans.cell(row, col_map['Monto USD']).value
            print(f"   Fila {row}:")
            print(f"      Fecha: {fecha}")
            print(f"      Tipo: {tipo}")
            print(f"      Cuenta: {cuenta}")
            print(f"      Concepto: {concepto[:50] if concepto else 'N/A'}")
            print(f"      Monto USD: ${monto:,.2f}" if monto else "      Monto USD: N/A")
            print()
        return

    # Leer datos del balance inicial
    fecha = ws_trans.cell(balance_inicial_fila, col_map['Fecha']).value
    tipo_actual = ws_trans.cell(balance_inicial_fila, col_map['Tipo Transacci√≥n']).value
    categoria = ws_trans.cell(balance_inicial_fila, col_map['Categor√≠a']).value
    cuenta = ws_trans.cell(balance_inicial_fila, col_map['Cuenta Bancaria']).value
    concepto = ws_trans.cell(balance_inicial_fila, col_map['Concepto']).value
    monto = ws_trans.cell(balance_inicial_fila, col_map['Monto USD']).value
    ing_egr = ws_trans.cell(balance_inicial_fila, col_map['Ingreso/Egreso']).value

    print(f"‚úÖ Balance Inicial encontrado en fila {balance_inicial_fila}:")
    print(f"   Fecha: {fecha}")
    print(f"   Tipo: {tipo_actual}")
    print(f"   Categor√≠a: {categoria}")
    print(f"   Cuenta: {cuenta}")
    print(f"   Concepto: {concepto}")
    print(f"   Monto: ${monto:,.2f}" if monto else "   Monto: N/A")
    print(f"   Ingreso/Egreso: {ing_egr}")
    print()

    # =========================================================================
    # PASO 2: DIAGNOSTICAR PROBLEMA
    # =========================================================================
    print("=" * 80)
    print("üìã PASO 2: Diagnosticando problema...")
    print()

    problema_detectado = False

    # Verificar si Ingreso/Egreso es una f√≥rmula
    if isinstance(ing_egr, str) and ing_egr.startswith('='):
        print("üìã Columna Ingreso/Egreso contiene una F√ìRMULA:")
        print(f"   {ing_egr[:80]}...")
        print()

        # Verificar si el Tipo actual est√° en la lista de la f√≥rmula
        if tipo_actual == "TRANSFERENCIAS":
            if "TRANSFERENCIAS" not in ing_egr and "Apertura Inicial" in ing_egr:
                print("‚ö†Ô∏è  PROBLEMA DETECTADO:")
                print(f"   Tipo actual: '{tipo_actual}'")
                print("   La f√≥rmula NO incluye 'TRANSFERENCIAS' como Ingreso")
                print("   Por lo tanto eval√∫a a 'Egreso'")
                print("   Esto hace que aparezca como egreso en hoja Efectivo")
                print()
                problema_detectado = True

    elif ing_egr and ing_egr == "Egreso":
        print("‚ö†Ô∏è  PROBLEMA DETECTADO:")
        print("   Balance Inicial est√° marcado como 'Egreso'")
        print("   Esto hace que aparezca en columna Egresos de hoja Efectivo")
        print()
        problema_detectado = True

    # =========================================================================
    # PASO 3: CORREGIR
    # =========================================================================
    if problema_detectado:
        print("=" * 80)
        print("üìã PASO 3: Aplicando correcci√≥n...")
        print()

        # Cambiar Tipo a "Apertura Inicial" para que la f√≥rmula eval√∫e a "Ingreso"
        ws_trans.cell(balance_inicial_fila, col_map['Tipo Transacci√≥n']).value = 'Apertura Inicial'

        # Categor√≠a
        ws_trans.cell(balance_inicial_fila, col_map['Categor√≠a']).value = 'Saldos Iniciales'

        print("‚úÖ Correcciones aplicadas:")
        print(f"   Tipo: '{tipo_actual}' ‚Üí 'Apertura Inicial'")
        print(f"   Categor√≠a: '{categoria}' ‚Üí 'Saldos Iniciales'")
        print()
        print("üìã Efecto de la correcci√≥n:")
        print("   La f√≥rmula en Ingreso/Egreso ahora evaluar√° a 'Ingreso'")
        print("   El balance inicial NO aparecer√° como egreso en Efectivo")
        print()

        # Guardar
        print("üíæ Guardando cambios...")
        wb.save(EXCEL_FILE)
        print("‚úÖ Excel actualizado")
        print()

        print("=" * 80)
        print("üìä RESULTADO ESPERADO")
        print("=" * 80)
        print()
        print("En hoja Efectivo, Promerica ahora deber√≠a mostrar:")
        print(f"   Saldo Inicial: ${monto:,.2f}" if monto else "   Saldo Inicial: (monto)")
        print("   Egresos: $0.00 (o suma correcta de egresos reales)")
        print(f"   Balance: Positivo (saldo inicial + ingresos - egresos)")
        print()

    else:
        print("‚úÖ No se detectaron problemas obvios")
        print("   El balance inicial ya est√° configurado correctamente")
        print()

    print("=" * 80)
    print("‚úÖ DIAGN√ìSTICO COMPLETADO")
    print("=" * 80)
    print()

if __name__ == "__main__":
    try:
        corregir_balance_inicial()
        print("üéâ Proceso completado!")
    except Exception as e:
        print(f"‚ùå ERROR: {e}")
        import traceback
        traceback.print_exc()
