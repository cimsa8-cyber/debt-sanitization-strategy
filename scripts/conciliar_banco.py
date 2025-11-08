#!/usr/bin/env python3
"""
CONCILIACIÓN BANCARIA AUTOMÁTICA
Match automático de extractos bancarios con sistema
Detección de diferencias y transacciones faltantes
Versión: 1.0
"""

import pandas as pd
import openpyxl
from datetime import datetime, timedelta
import sys

class ConciliacionBancaria:
    def __init__(self, excel_file="AlvaroVelasco_Finanzas_v1.0.xlsx"):
        self.excel_file = excel_file
        self.wb = None

    def conciliar_extracto(self, extracto_csv, cuenta_bancaria):
        """
        Concilia extracto bancario CSV con sistema

        Args:
            extracto_csv: Ruta al archivo CSV del banco
            cuenta_bancaria: Nombre de cuenta (ej: "Promerica USD")
        """
        print("="*70)
        print("CONCILIACIÓN BANCARIA AUTOMÁTICA")
        print("="*70)
        print(f"Extracto: {extracto_csv}")
        print(f"Cuenta: {cuenta_bancaria}")
        print(f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print()

        try:
            # Cargar extracto bancario
            print("⏳ Cargando extracto bancario...")
            extracto = self.cargar_extracto_csv(extracto_csv)
            print(f"✅ Extracto cargado: {len(extracto)} transacciones")

            # Cargar transacciones del sistema
            print("⏳ Cargando transacciones del sistema...")
            sistema = self.cargar_transacciones_sistema(cuenta_bancaria)
            print(f"✅ Sistema cargado: {len(sistema)} transacciones")

            # Realizar conciliación
            print()
            print("⏳ Ejecutando conciliación automática...")
            resultados = self.ejecutar_conciliacion(extracto, sistema)

            # Generar reporte
            self.generar_reporte_conciliacion(resultados, cuenta_bancaria)

            # Actualizar hoja Conciliación en Excel
            print()
            print("⏳ Actualizando hoja Conciliación...")
            self.actualizar_excel_conciliacion(resultados, cuenta_bancaria)
            print("✅ Hoja Conciliación actualizada")

            return resultados

        except FileNotFoundError as e:
            print(f"❌ ERROR: Archivo no encontrado: {e}")
            return None
        except Exception as e:
            print(f"❌ ERROR CRÍTICO: {e}")
            import traceback
            traceback.print_exc()
            return None

    def cargar_extracto_csv(self, csv_path):
        """
        Carga extracto bancario CSV
        Formato esperado: Fecha,Descripcion,Monto,Referencia
        """
        try:
            # Intentar cargar con diferentes encodings
            for encoding in ['utf-8', 'latin-1', 'iso-8859-1']:
                try:
                    df = pd.read_csv(csv_path, encoding=encoding)
                    break
                except UnicodeDecodeError:
                    continue

            # Normalizar nombres de columnas
            df.columns = df.columns.str.strip().str.lower()

            # Mapear columnas comunes
            columna_mapping = {
                'fecha': ['fecha', 'date', 'fecha_transaccion'],
                'descripcion': ['descripcion', 'description', 'concepto', 'detalle'],
                'monto': ['monto', 'amount', 'valor', 'importe'],
                'referencia': ['referencia', 'reference', 'ref', 'numero']
            }

            df_normalizado = pd.DataFrame()
            for col_standard, posibles in columna_mapping.items():
                for posible in posibles:
                    if posible in df.columns:
                        df_normalizado[col_standard] = df[posible]
                        break

            # Convertir fecha
            df_normalizado['fecha'] = pd.to_datetime(df_normalizado['fecha'], errors='coerce')

            # Convertir monto a float
            df_normalizado['monto'] = pd.to_numeric(df_normalizado['monto'], errors='coerce')

            return df_normalizado

        except Exception as e:
            print(f"❌ Error cargando extracto: {e}")
            raise

    def cargar_transacciones_sistema(self, cuenta_bancaria):
        """Carga transacciones del sistema para cuenta específica"""
        self.wb = openpyxl.load_workbook(self.excel_file, data_only=True)
        ws = self.wb["TRANSACCIONES"]

        transacciones = []
        for row in range(2, ws.max_row + 1):
            fecha = ws[f"A{row}"].value
            if not fecha:
                continue

            cuenta = ws[f"E{row}"].value
            if cuenta_bancaria not in str(cuenta):
                continue

            transacciones.append({
                'fecha': fecha if isinstance(fecha, datetime) else pd.to_datetime(fecha),
                'descripcion': ws[f"G{row}"].value,
                'monto': ws[f"I{row}"].value or 0,
                'referencia': ws[f"H{row}"].value,
                'fila': row
            })

        return pd.DataFrame(transacciones)

    def ejecutar_conciliacion(self, extracto, sistema):
        """
        Ejecuta conciliación automática con 3 niveles:
        1. Match exacto: Fecha + Monto exacto
        2. Match parcial: Fecha ±3 días + Monto ±$1
        3. Sin match: Transacciones solo en extracto o solo en sistema
        """
        resultados = {
            'conciliados': [],
            'no_en_sistema': [],
            'no_en_banco': [],
            'diferencias': []
        }

        extracto_procesado = set()
        sistema_procesado = set()

        # NIVEL 1: Match exacto (Fecha + Monto)
        for idx_ext, row_ext in extracto.iterrows():
            for idx_sis, row_sis in sistema.iterrows():
                if idx_sis in sistema_procesado:
                    continue

                # Match exacto
                if (row_ext['fecha'].date() == row_sis['fecha'].date() and
                    abs(row_ext['monto'] - row_sis['monto']) < 0.01):

                    resultados['conciliados'].append({
                        'fecha_banco': row_ext['fecha'],
                        'fecha_sistema': row_sis['fecha'],
                        'monto': row_ext['monto'],
                        'descripcion_banco': row_ext.get('descripcion', ''),
                        'descripcion_sistema': row_sis['descripcion'],
                        'tipo_match': 'EXACTO',
                        'fila_sistema': row_sis['fila']
                    })

                    extracto_procesado.add(idx_ext)
                    sistema_procesado.add(idx_sis)
                    break

        # NIVEL 2: Match parcial (Fecha ±3 días, Monto ±$1)
        for idx_ext, row_ext in extracto.iterrows():
            if idx_ext in extracto_procesado:
                continue

            for idx_sis, row_sis in sistema.iterrows():
                if idx_sis in sistema_procesado:
                    continue

                # Match parcial
                dias_diferencia = abs((row_ext['fecha'] - row_sis['fecha']).days)
                monto_diferencia = abs(row_ext['monto'] - row_sis['monto'])

                if dias_diferencia <= 3 and monto_diferencia <= 1.0:
                    resultados['conciliados'].append({
                        'fecha_banco': row_ext['fecha'],
                        'fecha_sistema': row_sis['fecha'],
                        'monto': row_ext['monto'],
                        'descripcion_banco': row_ext.get('descripcion', ''),
                        'descripcion_sistema': row_sis['descripcion'],
                        'tipo_match': 'PARCIAL',
                        'fila_sistema': row_sis['fila'],
                        'diferencia_dias': dias_diferencia,
                        'diferencia_monto': monto_diferencia
                    })

                    extracto_procesado.add(idx_ext)
                    sistema_procesado.add(idx_sis)
                    break

        # NIVEL 3: No conciliados
        for idx_ext, row_ext in extracto.iterrows():
            if idx_ext not in extracto_procesado:
                resultados['no_en_sistema'].append({
                    'fecha': row_ext['fecha'],
                    'monto': row_ext['monto'],
                    'descripcion': row_ext.get('descripcion', ''),
                    'referencia': row_ext.get('referencia', ''),
                    'status': '🟠 EN BANCO, NO EN SISTEMA'
                })

        for idx_sis, row_sis in sistema.iterrows():
            if idx_sis not in sistema_procesado:
                resultados['no_en_banco'].append({
                    'fecha': row_sis['fecha'],
                    'monto': row_sis['monto'],
                    'descripcion': row_sis['descripcion'],
                    'referencia': row_sis['referencia'],
                    'fila': row_sis['fila'],
                    'status': '🟡 EN SISTEMA, NO EN BANCO'
                })

        return resultados

    def generar_reporte_conciliacion(self, resultados, cuenta):
        """Genera reporte de conciliación"""
        print()
        print("="*70)
        print("REPORTE DE CONCILIACIÓN")
        print("="*70)
        print(f"Cuenta: {cuenta}")
        print()

        total_conciliados = len(resultados['conciliados'])
        total_no_sistema = len(resultados['no_en_sistema'])
        total_no_banco = len(resultados['no_en_banco'])

        print(f"✅ CONCILIADOS: {total_conciliados}")
        print(f"   - Exactos: {len([c for c in resultados['conciliados'] if c['tipo_match'] == 'EXACTO'])}")
        print(f"   - Parciales: {len([c for c in resultados['conciliados'] if c['tipo_match'] == 'PARCIAL'])}")
        print()

        if total_no_sistema > 0:
            print(f"🟠 EN BANCO, NO EN SISTEMA: {total_no_sistema}")
            print("   Acción requerida: Ingresar estas transacciones al sistema")
            for item in resultados['no_en_sistema'][:5]:
                print(f"   - {item['fecha'].strftime('%Y-%m-%d')}: ${item['monto']:,.2f} - {item['descripcion']}")
            if total_no_sistema > 5:
                print(f"   ... y {total_no_sistema - 5} más")
            print()

        if total_no_banco > 0:
            print(f"🟡 EN SISTEMA, NO EN BANCO: {total_no_banco}")
            print("   Posible razón: Transacciones aún no procesadas por banco")
            for item in resultados['no_en_banco'][:5]:
                print(f"   - {item['fecha'].strftime('%Y-%m-%d')}: ${item['monto']:,.2f} - {item['descripcion']}")
            if total_no_banco > 5:
                print(f"   ... y {total_no_banco - 5} más")
            print()

        # Calcular tasa de conciliación
        total = total_conciliados + total_no_sistema + total_no_banco
        if total > 0:
            tasa = (total_conciliados / total) * 100
            print(f"📊 TASA DE CONCILIACIÓN: {tasa:.1f}%")

            if tasa >= 90:
                print("   ✅ Excelente - Sistema bien conciliado")
            elif tasa >= 70:
                print("   ⚠️  Aceptable - Revisar diferencias")
            else:
                print("   🔴 Crítico - Requiere atención inmediata")

        print()
        print("="*70)

    def actualizar_excel_conciliacion(self, resultados, cuenta):
        """Actualiza hoja Conciliación en Excel"""
        ws = self.wb["Conciliacion"]

        # Limpiar hoja
        ws.delete_rows(2, ws.max_row)

        # Headers
        ws["A1"] = f"CONCILIACIÓN - {cuenta}"
        ws["A2"] = "Fecha Banco"
        ws["B2"] = "Fecha Sistema"
        ws["C2"] = "Monto"
        ws["D2"] = "Descripción"
        ws["E2"] = "Estado"
        ws["F2"] = "Notas"

        row = 3

        # Conciliados
        for item in resultados['conciliados']:
            ws[f"A{row}"] = item['fecha_banco']
            ws[f"B{row}"] = item['fecha_sistema']
            ws[f"C{row}"] = item['monto']
            ws[f"D{row}"] = item['descripcion_sistema']
            ws[f"E{row}"] = f"✅ {item['tipo_match']}"
            if item['tipo_match'] == 'PARCIAL':
                ws[f"F{row}"] = f"Dif: {item.get('diferencia_dias', 0)} días, ${item.get('diferencia_monto', 0):.2f}"
            row += 1

        # No en sistema
        for item in resultados['no_en_sistema']:
            ws[f"A{row}"] = item['fecha']
            ws[f"C{row}"] = item['monto']
            ws[f"D{row}"] = item['descripcion']
            ws[f"E{row}"] = item['status']
            ws[f"F{row}"] = "ACCIÓN: Ingresar al sistema"
            row += 1

        # No en banco
        for item in resultados['no_en_banco']:
            ws[f"B{row}"] = item['fecha']
            ws[f"C{row}"] = item['monto']
            ws[f"D{row}"] = item['descripcion']
            ws[f"E{row}"] = item['status']
            ws[f"F{row}"] = "Pendiente procesamiento banco"
            row += 1

        # Guardar
        self.wb.save(self.excel_file)

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Uso: python3 conciliar_banco.py <extracto.csv> <cuenta>")
        print("Ejemplo: python3 conciliar_banco.py extracto_promerica.csv 'Promerica USD'")
        sys.exit(1)

    extracto_csv = sys.argv[1]
    cuenta = sys.argv[2]

    conciliador = ConciliacionBancaria()
    resultados = conciliador.conciliar_extracto(extracto_csv, cuenta)

    if resultados:
        sys.exit(0)
    else:
        sys.exit(1)
