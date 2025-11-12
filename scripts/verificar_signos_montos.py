#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
VERIFICAR SIGNOS DE MONTOS EN TRANSACCIONES
Revisa si los egresos tienen signo correcto (negativo)
"""
import openpyxl

EXCEL_FILE = "AlvaroVelasco_Finanzas_v2.0.xlsx"

def verificar():
    print("=" * 80)
    print("VERIFICACIÓN DE SIGNOS EN MONTOS USD")
    print("=" * 80)
    print()

    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb['TRANSACCIONES']

    # Encontrar columnas
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    col_map = {}
    for col in range(1, len(headers) + 1):
        if headers[col-1]:
            col_map[headers[col-1]] = col

    print("🔍 Analizando signos de Monto USD para transacciones de Promerica...")
    print()

    # Analizar transacciones de Promerica
    ingresos_positivos = []
    ingresos_negativos = []
    egresos_positivos = []
    egresos_negativos = []

    for row in range(2, ws.max_row + 1):
        cuenta = ws.cell(row, col_map['Cuenta Bancaria']).value
        monto = ws.cell(row, col_map['Monto USD']).value
        ing_egr = ws.cell(row, col_map['Ingreso/Egreso']).value
        concepto = ws.cell(row, col_map['Concepto']).value

        # Solo Promerica USD
        if cuenta and 'Promerica USD (40000003881774)' in str(cuenta):
            if monto:
                monto_val = float(monto)

                if ing_egr == 'Ingreso':
                    if monto_val > 0:
                        ingresos_positivos.append({'fila': row, 'monto': monto_val, 'concepto': concepto})
                    else:
                        ingresos_negativos.append({'fila': row, 'monto': monto_val, 'concepto': concepto})
                elif ing_egr == 'Egreso':
                    if monto_val > 0:
                        egresos_positivos.append({'fila': row, 'monto': monto_val, 'concepto': concepto})
                    else:
                        egresos_negativos.append({'fila': row, 'monto': monto_val, 'concepto': concepto})

    # Mostrar resultados
    print("📊 INGRESOS:")
    print(f"   ✅ Positivos (correcto): {len(ingresos_positivos)}")
    if len(ingresos_positivos) > 0:
        total_ing_pos = sum(t['monto'] for t in ingresos_positivos)
        print(f"      Total: ${total_ing_pos:,.2f}")

    print(f"   ⚠️  Negativos (INCORRECTO): {len(ingresos_negativos)}")
    if len(ingresos_negativos) > 0:
        total_ing_neg = sum(t['monto'] for t in ingresos_negativos)
        print(f"      Total: ${total_ing_neg:,.2f}")
        print()
        print("      Ejemplos:")
        for t in ingresos_negativos[:5]:
            print(f"      • Fila {t['fila']}: ${t['monto']:,.2f} - {t['concepto'][:40] if t['concepto'] else 'N/A'}")

    print()

    print("📊 EGRESOS:")
    print(f"   ⚠️  Positivos (INCORRECTO): {len(egresos_positivos)}")
    if len(egresos_positivos) > 0:
        total_egr_pos = sum(t['monto'] for t in egresos_positivos)
        print(f"      Total: ${total_egr_pos:,.2f}")
        print()
        print("      Ejemplos (primeros 10):")
        for t in egresos_positivos[:10]:
            print(f"      • Fila {t['fila']}: ${t['monto']:,.2f} - {t['concepto'][:40] if t['concepto'] else 'N/A'}")

    print(f"   ✅ Negativos (correcto): {len(egresos_negativos)}")
    if len(egresos_negativos) > 0:
        total_egr_neg = sum(abs(t['monto']) for t in egresos_negativos)
        print(f"      Total: ${total_egr_neg:,.2f}")

    print()

    # Análisis
    print("=" * 80)
    print("📋 ANÁLISIS")
    print("=" * 80)
    print()

    if len(egresos_positivos) > 0:
        print("🚨 PROBLEMA DETECTADO:")
        print()
        print(f"   Hay {len(egresos_positivos)} egresos con signo POSITIVO")
        print(f"   Total: ${sum(t['monto'] for t in egresos_positivos):,.2f}")
        print()
        print("   Los egresos deberían tener signo NEGATIVO en columna I (Monto USD)")
        print("   para que la fórmula de Efectivo funcione correctamente.")
        print()
        print("✅ SOLUCIÓN:")
        print("   Cambiar el signo de estos egresos a negativo")
        print(f"   Multiplicar por -1: {len(egresos_positivos)} filas")
    else:
        print("✅ Todos los egresos tienen signo correcto (negativo)")

    print()

    if len(ingresos_negativos) > 0:
        print("🚨 PROBLEMA DETECTADO:")
        print()
        print(f"   Hay {len(ingresos_negativos)} ingresos con signo NEGATIVO")
        print(f"   Total: ${sum(t['monto'] for t in ingresos_negativos):,.2f}")
        print()
        print("   Los ingresos deberían tener signo POSITIVO en columna I (Monto USD)")
        print()
        print("✅ SOLUCIÓN:")
        print("   Cambiar el signo de estos ingresos a positivo")
        print(f"   Multiplicar por -1: {len(ingresos_negativos)} filas")
    else:
        print("✅ Todos los ingresos tienen signo correcto (positivo)")

    print()

    # Verificar fórmula Efectivo
    print("=" * 80)
    print("🔍 VERIFICACIÓN FÓRMULA EFECTIVO")
    print("=" * 80)
    print()

    if len(egresos_positivos) > 0:
        total_egr_pos = sum(t['monto'] for t in egresos_positivos)

        print("💡 EXPLICACIÓN DEL PROBLEMA:")
        print()
        print(f"   1. Fórmula E3: =SUMIFS(...,\"Egreso\")*-1")
        print(f"   2. SUMIFS suma egresos POSITIVOS: +${total_egr_pos:,.2f}")
        print(f"   3. Multiplica por -1: ${total_egr_pos:,.2f} * -1 = -${total_egr_pos:,.2f}")
        print(f"   4. Resultado en E3: NEGATIVO ❌")
        print()
        print("   Si los egresos fueran NEGATIVOS:")
        print(f"   1. SUMIFS sumaría: -${total_egr_pos:,.2f}")
        print(f"   2. Multiplica por -1: -${total_egr_pos:,.2f} * -1 = +${total_egr_pos:,.2f}")
        print(f"   3. Resultado en E3: POSITIVO ✅")

    print()
    print("=" * 80)
    print("✅ VERIFICACIÓN COMPLETADA")
    print("=" * 80)
    print()

if __name__ == "__main__":
    try:
        verificar()
    except Exception as e:
        print(f"❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
