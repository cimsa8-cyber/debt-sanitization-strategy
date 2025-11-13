#!/usr/bin/env python3
"""
FASE 3: Agregar hoja IVA_CONTROL a v3.0
Control de IVA para compliance Costa Rica
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

V3_FILE = "AlvaroVelasco_Finanzas_v3.0.xlsx"

COLOR_HEADER = "1F4E78"
COLOR_EDITABLE = "FFF2CC"
COLOR_CALCULATED = "E7E6E6"
COLOR_WARNING = "FCE4D6"

BORDER_THIN = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

print("\n" + "="*60)
print("FASE 3: Agregar IVA_CONTROL")
print("="*60)

wb = openpyxl.load_workbook(V3_FILE)

# ============================================================================
# HOJA IVA_CONTROL
# ============================================================================

print("\n📊 Creando hoja IVA_CONTROL...")
ws = wb.create_sheet("IVA_CONTROL", 5)  # Insertar después de CONFIG

ws['A1'] = "CONTROL DE IVA - COSTA RICA"
ws['A1'].font = Font(size=14, bold=True)
ws.merge_cells('A1:L1')

ws['A2'] = f"Período: NOVIEMBRE 2025 | Vencimiento D-104: 26/Nov/2025"
ws['A2'].font = Font(size=10, italic=True)
ws.merge_cells('A2:L2')

# ============================================================================
# SECCIÓN 1: VENTAS (IVA COBRADO)
# ============================================================================

ws['A4'] = "VENTAS - IVA COBRADO (13%)"
ws['A4'].font = Font(size=12, bold=True, color='FFFFFF')
ws['A4'].fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
ws.merge_cells('A4:L4')

# Headers ventas
headers_ventas = ["Fecha", "Factura", "Cliente", "Base Gravable USD", "IVA 13% USD", "Total USD", "Zona Franca", "Retención 2%", "Neto Cobrado", "Estado", "Ref Trans", "Notas"]
for col, header in enumerate(headers_ventas, start=1):
    cell = ws.cell(5, col, header)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = BORDER_THIN

# Anchos columnas
ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 30
ws.column_dimensions['D'].width = 14
ws.column_dimensions['E'].width = 14
ws.column_dimensions['F'].width = 14
ws.column_dimensions['G'].width = 10
ws.column_dimensions['H'].width = 12
ws.column_dimensions['I'].width = 14
ws.column_dimensions['J'].width = 10
ws.column_dimensions['K'].width = 10
ws.column_dimensions['L'].width = 20

# Ejemplo fila (será llenada con fórmulas dinámicas)
row = 6
ws.cell(row, 1, "01/11/25")
ws.cell(row, 1).number_format = 'DD/MM/YY'
ws.cell(row, 2, "FAC-001")
ws.cell(row, 3, "Cliente Ejemplo")
ws.cell(row, 4, 1000.00)
ws.cell(row, 4).number_format = '$#,##0.00'
ws.cell(row, 4).fill = PatternFill(start_color=COLOR_EDITABLE, end_color=COLOR_EDITABLE, fill_type="solid")
ws.cell(row, 5, '=D6*0.13')  # IVA 13%
ws.cell(row, 5).number_format = '$#,##0.00'
ws.cell(row, 5).fill = PatternFill(start_color=COLOR_CALCULATED, end_color=COLOR_CALCULATED, fill_type="solid")
ws.cell(row, 6, '=D6+E6')  # Total
ws.cell(row, 6).number_format = '$#,##0.00'
ws.cell(row, 6).fill = PatternFill(start_color=COLOR_CALCULATED, end_color=COLOR_CALCULATED, fill_type="solid")
ws.cell(row, 7, "NO")  # Zona Franca
ws.cell(row, 7).fill = PatternFill(start_color=COLOR_EDITABLE, end_color=COLOR_EDITABLE, fill_type="solid")
ws.cell(row, 8, '=IF(G6="SI",0,D6*0.02)')  # Retención 2%
ws.cell(row, 8).number_format = '$#,##0.00'
ws.cell(row, 8).fill = PatternFill(start_color=COLOR_CALCULATED, end_color=COLOR_CALCULATED, fill_type="solid")
ws.cell(row, 9, '=F6-H6')  # Neto Cobrado
ws.cell(row, 9).number_format = '$#,##0.00'
ws.cell(row, 9).fill = PatternFill(start_color=COLOR_CALCULATED, end_color=COLOR_CALCULATED, fill_type="solid")
ws.cell(row, 10, "EMITIDA")

# Filas adicionales para datos (7-20)
for r in range(7, 21):
    ws.cell(r, 1).number_format = 'DD/MM/YY'
    ws.cell(r, 4).number_format = '$#,##0.00'
    ws.cell(r, 4).fill = PatternFill(start_color=COLOR_EDITABLE, end_color=COLOR_EDITABLE, fill_type="solid")
    ws.cell(r, 5, f'=D{r}*0.13')
    ws.cell(r, 5).number_format = '$#,##0.00'
    ws.cell(r, 5).fill = PatternFill(start_color=COLOR_CALCULATED, end_color=COLOR_CALCULATED, fill_type="solid")
    ws.cell(r, 6, f'=D{r}+E{r}')
    ws.cell(r, 6).number_format = '$#,##0.00'
    ws.cell(r, 6).fill = PatternFill(start_color=COLOR_CALCULATED, end_color=COLOR_CALCULATED, fill_type="solid")
    ws.cell(r, 7).fill = PatternFill(start_color=COLOR_EDITABLE, end_color=COLOR_EDITABLE, fill_type="solid")
    ws.cell(r, 8, f'=IF(G{r}="SI",0,D{r}*0.02)')
    ws.cell(r, 8).number_format = '$#,##0.00'
    ws.cell(r, 8).fill = PatternFill(start_color=COLOR_CALCULATED, end_color=COLOR_CALCULATED, fill_type="solid")
    ws.cell(r, 9, f'=F{r}-H{r}')
    ws.cell(r, 9).number_format = '$#,##0.00'
    ws.cell(r, 9).fill = PatternFill(start_color=COLOR_CALCULATED, end_color=COLOR_CALCULATED, fill_type="solid")

# Total ventas
row = 21
ws.cell(row, 1, "TOTAL VENTAS")
ws.cell(row, 1).font = Font(bold=True)
ws.cell(row, 4, '=SUM(D6:D20)')
ws.cell(row, 4).number_format = '$#,##0.00'
ws.cell(row, 4).font = Font(bold=True)
ws.cell(row, 5, '=SUM(E6:E20)')
ws.cell(row, 5).number_format = '$#,##0.00'
ws.cell(row, 5).font = Font(bold=True)
ws.cell(row, 5).fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type="solid")
ws.cell(row, 6, '=SUM(F6:F20)')
ws.cell(row, 6).number_format = '$#,##0.00'
ws.cell(row, 6).font = Font(bold=True)

# ============================================================================
# SECCIÓN 2: COMPRAS (IVA PAGADO / CRÉDITO)
# ============================================================================

ws['A23'] = "COMPRAS - IVA PAGADO (CRÉDITO FISCAL)"
ws['A23'].font = Font(size=12, bold=True, color='FFFFFF')
ws['A23'].fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
ws.merge_cells('A23:L23')

# Headers compras
headers_compras = ["Fecha", "Factura", "Proveedor", "Base Gravable USD", "IVA 13% USD", "Total USD", "Deducible", "IVA Acreditable", "Método Pago", "Estado", "Ref Trans", "Notas"]
for col, header in enumerate(headers_compras, start=1):
    cell = ws.cell(24, col, header)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = BORDER_THIN

# Ejemplo fila
row = 25
ws.cell(row, 1, "05/11/25")
ws.cell(row, 1).number_format = 'DD/MM/YY'
ws.cell(row, 2, "PROV-001")
ws.cell(row, 3, "Proveedor Ejemplo")
ws.cell(row, 4, 500.00)
ws.cell(row, 4).number_format = '$#,##0.00'
ws.cell(row, 4).fill = PatternFill(start_color=COLOR_EDITABLE, end_color=COLOR_EDITABLE, fill_type="solid")
ws.cell(row, 5, '=D25*0.13')
ws.cell(row, 5).number_format = '$#,##0.00'
ws.cell(row, 5).fill = PatternFill(start_color=COLOR_CALCULATED, end_color=COLOR_CALCULATED, fill_type="solid")
ws.cell(row, 6, '=D25+E25')
ws.cell(row, 6).number_format = '$#,##0.00'
ws.cell(row, 6).fill = PatternFill(start_color=COLOR_CALCULATED, end_color=COLOR_CALCULATED, fill_type="solid")
ws.cell(row, 7, "SI")  # Deducible
ws.cell(row, 7).fill = PatternFill(start_color=COLOR_EDITABLE, end_color=COLOR_EDITABLE, fill_type="solid")
ws.cell(row, 8, '=IF(G25="SI",E25,0)')  # IVA Acreditable
ws.cell(row, 8).number_format = '$#,##0.00'
ws.cell(row, 8).fill = PatternFill(start_color=COLOR_CALCULATED, end_color=COLOR_CALCULATED, fill_type="solid")
ws.cell(row, 9, "Transferencia")
ws.cell(row, 10, "PAGADA")

# Filas adicionales (26-40)
for r in range(26, 41):
    ws.cell(r, 1).number_format = 'DD/MM/YY'
    ws.cell(r, 4).number_format = '$#,##0.00'
    ws.cell(r, 4).fill = PatternFill(start_color=COLOR_EDITABLE, end_color=COLOR_EDITABLE, fill_type="solid")
    ws.cell(r, 5, f'=D{r}*0.13')
    ws.cell(r, 5).number_format = '$#,##0.00'
    ws.cell(r, 5).fill = PatternFill(start_color=COLOR_CALCULATED, end_color=COLOR_CALCULATED, fill_type="solid")
    ws.cell(r, 6, f'=D{r}+E{r}')
    ws.cell(r, 6).number_format = '$#,##0.00'
    ws.cell(r, 6).fill = PatternFill(start_color=COLOR_CALCULATED, end_color=COLOR_CALCULATED, fill_type="solid")
    ws.cell(r, 7).fill = PatternFill(start_color=COLOR_EDITABLE, end_color=COLOR_EDITABLE, fill_type="solid")
    ws.cell(r, 8, f'=IF(G{r}="SI",E{r},0)')
    ws.cell(r, 8).number_format = '$#,##0.00'
    ws.cell(r, 8).fill = PatternFill(start_color=COLOR_CALCULATED, end_color=COLOR_CALCULATED, fill_type="solid")

# Total compras
row = 41
ws.cell(row, 1, "TOTAL COMPRAS")
ws.cell(row, 1).font = Font(bold=True)
ws.cell(row, 4, '=SUM(D25:D40)')
ws.cell(row, 4).number_format = '$#,##0.00'
ws.cell(row, 4).font = Font(bold=True)
ws.cell(row, 5, '=SUM(E25:E40)')
ws.cell(row, 5).number_format = '$#,##0.00'
ws.cell(row, 5).font = Font(bold=True)
ws.cell(row, 8, '=SUM(H25:H40)')
ws.cell(row, 8).number_format = '$#,##0.00'
ws.cell(row, 8).font = Font(bold=True)
ws.cell(row, 8).fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type="solid")

# ============================================================================
# SECCIÓN 3: RESUMEN IVA
# ============================================================================

ws['A43'] = "RESUMEN DECLARACIÓN D-104"
ws['A43'].font = Font(size=12, bold=True, color='FFFFFF')
ws['A43'].fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
ws.merge_cells('A43:E43')

row = 44
ws.cell(row, 1, "IVA Cobrado (Ventas)")
ws.cell(row, 1).font = Font(bold=True)
ws.cell(row, 2, '=E21')
ws.cell(row, 2).number_format = '$#,##0.00'
ws.cell(row, 2).fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type="solid")

row = 45
ws.cell(row, 1, "IVA Acreditable (Compras)")
ws.cell(row, 1).font = Font(bold=True)
ws.cell(row, 2, '=H41')
ws.cell(row, 2).number_format = '$#,##0.00'
ws.cell(row, 2).fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type="solid")

row = 46
ws.cell(row, 1, "IVA A PAGAR")
ws.cell(row, 1).font = Font(size=12, bold=True, color='FF0000')
ws.cell(row, 2, '=B44-B45')
ws.cell(row, 2).number_format = '$#,##0.00'
ws.cell(row, 2).font = Font(size=12, bold=True, color='FF0000')
ws.cell(row, 2).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type="solid")

row = 47
ws.cell(row, 1, "Vencimiento D-104")
ws.cell(row, 2, "26/Nov/2025")
ws.cell(row, 2).font = Font(bold=True)
ws.cell(row, 2).fill = PatternFill(start_color=COLOR_WARNING, end_color=COLOR_WARNING, fill_type="solid")

# Notas
row = 49
ws.cell(row, 1, "NOTAS:")
ws.cell(row, 1).font = Font(bold=True)
row = 50
ws.cell(row, 1, "• VWR International y RSHughes: ZONA FRANCA (marcar Zona Franca='SI')")
row = 51
ws.cell(row, 1, "• Retención 2%: Aplicable a servicios profesionales")
row = 52
ws.cell(row, 1, "• IVA Acreditable: Solo gastos deducibles (marcar Deducible='SI')")
row = 53
ws.cell(row, 1, "• Ref Trans: Vincular a fila en hoja TRANSACCIONES")

print("   ✅ Hoja IVA_CONTROL creada")
print("   📋 Secciones: Ventas, Compras, Resumen D-104")
print("   🧮 Fórmulas: IVA 13%, Retención 2%, IVA Neto")

# Guardar
print(f"\n💾 Guardando {V3_FILE}...")
wb.save(V3_FILE)
wb.close()

print("\n" + "="*60)
print("✅ FASE 3 COMPLETADA")
print("="*60)
print("   📊 Hoja IVA_CONTROL agregada")
print("   ⚠️  Vencimiento D-104: 26/Nov/2025")
print("   📝 PRÓXIMO PASO: Llenar datos de ventas/compras Nov 2025")
print("="*60)
