#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CORRECCIÓN FINAL COMPLETA
1. Elimina filas vacías (211-214)
2. Corrige formato de fechas (dd/mm/yy)
3. Investiga saldo Promerica en Efectivo
"""
import openpyxl
from datetime import datetime

EXCEL_FILE = "AlvaroVelasco_Finanzas_v2.0.xlsx"

def corregir_todo():
    print("=" * 80)
    print("CORRECCIÓN FINAL DE INCONSISTENCIAS")
    print("=" * 80)
    print()

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb['TRANSACCIONES']

    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    col_map = {}
    for col in range(1, len(headers) + 1):
        if headers[col-1]:
            col_map[headers[col-1]] = col

    # =========================================================================
    # PASO 1: ELIMINAR FILAS VACÍAS
    # =========================================================================
    print("📋 PASO 1: Eliminando filas vacías...")
    print()

    filas_eliminadas = []

    # Revisar filas 211-214
    for row in range(211, 215):
        if row > ws.max_row:
            break

        # Verificar si la fila está vacía
        fila_vacia = True
        for col in range(1, ws.max_column + 1):
            if ws.cell(row, col).value:
                fila_vacia = False
                break

        if fila_vacia:
            filas_eliminadas.append(row)

    if filas_eliminadas:
        # Eliminar de abajo hacia arriba para no desajustar índices
        for row in sorted(filas_eliminadas, reverse=True):
            ws.delete_rows(row, 1)
            print(f"   ✓ Fila {row} eliminada (vacía)")
        print()
        print(f"✅ {len(filas_eliminadas)} filas vacías eliminadas")
        print()
    else:
        print("✅ No se encontraron filas vacías")
        print()

    # =========================================================================
    # PASO 2: CORREGIR FORMATO DE FECHAS
    # =========================================================================
    print("=" * 80)
    print("📋 PASO 2: Corrigiendo formato de fechas...")
    print()

    fechas_corregidas = 0

    # Revisar todas las filas recientes (últimas 20)
    inicio_revision = max(2, ws.max_row - 20)

    for row in range(inicio_revision, ws.max_row + 1):
        fecha_celda = ws.cell(row, col_map['Fecha'])
        fecha_valor = fecha_celda.value

        if fecha_valor and isinstance(fecha_valor, datetime):
            # Aplicar formato correcto: dd/mm/yy
            fecha_celda.number_format = 'd/m/yy'
            fechas_corregidas += 1

    if fechas_corregidas > 0:
        print(f"✅ {fechas_corregidas} fechas formateadas a dd/mm/yy")
    else:
        print("✅ Todas las fechas ya tienen formato correcto")
    print()

    # =========================================================================
    # PASO 3: INVESTIGAR SALDO PROMERICA EN EFECTIVO
    # =========================================================================
    print("=" * 80)
    print("📋 PASO 3: Investigando saldo Promerica en Efectivo...")
    print()

    if 'Efectivo' in wb.sheetnames:
        ws_efectivo = wb['Efectivo']

        # Buscar fila de Promerica
        promerica_fila = None
        for row in range(2, ws_efectivo.max_row + 1):
            cuenta = ws_efectivo.cell(row, 1).value
            if cuenta and 'Promerica' in str(cuenta) and 'USD' in str(cuenta):
                promerica_fila = row
                break

        if promerica_fila:
            # Leer valores actuales
            cuenta = ws_efectivo.cell(promerica_fila, 1).value
            saldo_inicial = ws_efectivo.cell(promerica_fila, 2).value
            ingresos = ws_efectivo.cell(promerica_fila, 3).value
            egresos = ws_efectivo.cell(promerica_fila, 4).value
            balance = ws_efectivo.cell(promerica_fila, 5).value

            print(f"📊 Estado actual de Promerica (Efectivo, fila {promerica_fila}):")
            print(f"   Cuenta: {cuenta}")
            print(f"   Saldo Inicial: ${saldo_inicial:,.2f}" if saldo_inicial else "   Saldo Inicial: N/A")
            print(f"   Ingresos: ${ingresos:,.2f}" if ingresos else "   Ingresos: N/A")
            print(f"   Egresos: ${egresos:,.2f}" if egresos else "   Egresos: N/A")
            print(f"   Balance: ${balance:,.2f}" if balance else "   Balance: N/A")
            print()

            # Verificar fórmulas
            print("🔍 Verificando fórmulas:")
            for col in range(2, 6):
                celda = ws_efectivo.cell(promerica_fila, col)
                if celda.value and isinstance(celda.value, str) and celda.value.startswith('='):
                    col_letter = openpyxl.utils.get_column_letter(col)
                    print(f"   Columna {col_letter}: {celda.value}")
            print()

            if balance and balance < 0:
                print("⚠️  PROBLEMA DETECTADO: Balance negativo")
                print("   Causa probable: Saldo inicial incorrecto o fórmulas erradas")
                print()
        else:
            print("⚠️  No se encontró Promerica USD en hoja Efectivo")
            print()
    else:
        print("⚠️  Hoja 'Efectivo' no encontrada")
        print()

    # Guardar
    print("=" * 80)
    print("💾 Guardando cambios...")
    wb.save(EXCEL_FILE)
    print("✅ Excel actualizado")
    print()

    # Resumen
    print("=" * 80)
    print("📊 RESUMEN DE CORRECCIONES")
    print("=" * 80)
    print()

    if filas_eliminadas:
        print(f"✅ Filas vacías eliminadas: {len(filas_eliminadas)}")
    print(f"✅ Fechas formateadas: {fechas_corregidas}")
    print(f"⚠️  Saldo Promerica en Efectivo: Requiere revisión manual")
    print()

    print("=" * 80)
    print("✅ CORRECCIÓN COMPLETADA")
    print("=" * 80)
    print()

    print("📋 PRÓXIMOS PASOS:")
    print("   1. Verifica que las fechas se vean como dd/mm/yy")
    print("   2. Revisa hoja Efectivo - fórmulas de Promerica")
    print("   3. Confirma saldo Promerica")
    print()

if __name__ == "__main__":
    try:
        corregir_todo()
        print("🎉 Correcciones aplicadas!")
    except Exception as e:
        print(f"❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
