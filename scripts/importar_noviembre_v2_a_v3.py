#!/usr/bin/env python3
"""
IMPORTAR DATOS NOVIEMBRE 2025 - v2.0 → v3.0
============================================

Lee transacciones de Noviembre 2025 desde v2.0
Las valida y las inserta en v3.0
Detecta y reporta duplicados potenciales

IMPORTANTE: Ejecutar en el directorio donde están ambos archivos Excel

Uso:
    python scripts/importar_noviembre_v2_a_v3.py
"""

import openpyxl
from datetime import datetime
import sys

# Archivos
V2_FILE = "AlvaroVelasco_Finanzas_v2.0.xlsx"
V3_FILE = "AlvaroVelasco_Finanzas_v3.0.xlsx"

def analizar_v2():
    """Analiza estructura de v2.0 y muestra primeras filas"""
    print("\n🔍 ANALIZANDO v2.0...")

    try:
        wb = openpyxl.load_workbook(V2_FILE, read_only=True, data_only=True)
    except FileNotFoundError:
        print(f"❌ ERROR: No se encuentra {V2_FILE}")
        print(f"   Asegúrate de estar en el directorio correcto")
        sys.exit(1)

    print(f"   Hojas encontradas: {', '.join(wb.sheetnames)}")

    # Buscar hoja de transacciones (puede tener varios nombres)
    ws = None
    for nombre in ['TRANSACCIONES', 'Transacciones', 'DATOS', 'Sheet1', wb.sheetnames[0]]:
        if nombre in wb.sheetnames:
            ws = wb[nombre]
            break

    if ws is None:
        ws = wb.active

    print(f"   Analizando hoja: {ws.title}")
    print(f"   Total filas: {ws.max_row}")

    # Leer encabezados
    print("\n   Columnas encontradas:")
    headers = []
    for col in range(1, 30):  # Revisar primeras 30 columnas
        header = ws.cell(1, col).value
        if header:
            headers.append((col, header))
            print(f"      {col}. {header}")

    # Mostrar primeras 3 filas de datos
    print("\n   Primeras 3 transacciones:")
    for row in range(2, min(5, ws.max_row + 1)):
        print(f"\n   Fila {row}:")
        for col_idx, header in headers[:10]:  # Primeras 10 columnas
            value = ws.cell(row, col_idx).value
            if value:
                print(f"      {header}: {value}")

    wb.close()
    return headers, ws.title

def detectar_columnas(headers):
    """Detecta índices de columnas importantes basado en nombres"""
    columnas = {
        'fecha': None,
        'tipo': None,
        'categoria': None,
        'descripcion': None,
        'cuenta': None,
        'entidad': None,
        'factura': None,
        'monto_crc': None,
        'monto_usd': None,
        'tipo_cambio': None,
        'metodo': None,
        'estado': None,
    }

    # Mapeo flexible de nombres
    mapeo = {
        'fecha': ['fecha', 'date'],
        'tipo': ['tipo', 'type', 'tipo transaccion'],
        'categoria': ['categoria', 'category'],
        'descripcion': ['descripcion', 'description', 'detalle'],
        'cuenta': ['cuenta', 'cuenta origen', 'banco'],
        'entidad': ['entidad', 'cliente', 'proveedor'],
        'factura': ['factura', 'invoice', 'factura #'],
        'monto_crc': ['monto crc', 'colones', 'crc'],
        'monto_usd': ['monto usd', 'dolares', 'usd'],
        'tipo_cambio': ['tipo cambio', 'tc', 't.c.'],
        'metodo': ['metodo', 'metodo pago', 'forma pago'],
        'estado': ['estado', 'status'],
    }

    for col_idx, header in headers:
        header_lower = str(header).lower().strip()

        for key, variantes in mapeo.items():
            if any(var in header_lower for var in variantes):
                columnas[key] = col_idx
                break

    return columnas

def importar_noviembre():
    """Importa transacciones de Noviembre 2025"""

    print("\n" + "="*60)
    print("IMPORTAR NOVIEMBRE 2025: v2.0 → v3.0")
    print("="*60)

    # Analizar v2.0
    headers, hoja_nombre = analizar_v2()
    columnas = detectar_columnas(headers)

    print("\n📋 Mapeo de columnas detectado:")
    for key, col_idx in columnas.items():
        if col_idx:
            print(f"   {key}: columna {col_idx}")

    # Preguntar confirmación
    print("\n⚠️  ADVERTENCIA: Este script insertará datos en v3.0")
    print(f"   Se leerán transacciones de NOVIEMBRE 2025 desde: {V2_FILE}")
    print(f"   Se insertarán en: {V3_FILE}")

    respuesta = input("\n¿Continuar? (SI/NO): ").strip().upper()
    if respuesta != "SI":
        print("❌ Importación cancelada")
        sys.exit(0)

    # Leer v2.0
    print(f"\n📖 Leyendo {V2_FILE}...")
    wb_v2 = openpyxl.load_workbook(V2_FILE, read_only=True, data_only=True)
    ws_v2 = wb_v2[hoja_nombre]

    # Filtrar transacciones de noviembre
    transacciones_nov = []
    duplicados_detectados = []

    print("🔍 Filtrando transacciones de Noviembre 2025...")

    for row in range(2, ws_v2.max_row + 1):
        fecha = ws_v2.cell(row, columnas['fecha']).value if columnas['fecha'] else None

        # Validar que sea noviembre 2025
        if fecha:
            if isinstance(fecha, datetime):
                if fecha.month == 11 and fecha.year == 2025:
                    # Extraer datos
                    transaccion = {
                        'fecha': fecha,
                        'tipo': ws_v2.cell(row, columnas['tipo']).value if columnas['tipo'] else "",
                        'categoria': ws_v2.cell(row, columnas['categoria']).value if columnas['categoria'] else "",
                        'descripcion': ws_v2.cell(row, columnas['descripcion']).value if columnas['descripcion'] else "",
                        'cuenta': ws_v2.cell(row, columnas['cuenta']).value if columnas['cuenta'] else "",
                        'entidad': ws_v2.cell(row, columnas['entidad']).value if columnas['entidad'] else "",
                        'factura': ws_v2.cell(row, columnas['factura']).value if columnas['factura'] else "",
                        'monto_crc': ws_v2.cell(row, columnas['monto_crc']).value if columnas['monto_crc'] else 0,
                        'monto_usd': ws_v2.cell(row, columnas['monto_usd']).value if columnas['monto_usd'] else 0,
                        'tipo_cambio': ws_v2.cell(row, columnas['tipo_cambio']).value if columnas['tipo_cambio'] else 508,
                        'metodo': ws_v2.cell(row, columnas['metodo']).value if columnas['metodo'] else "",
                        'estado': ws_v2.cell(row, columnas['estado']).value if columnas['estado'] else "COMPLETADA",
                        'fila_origen': row,
                    }

                    transacciones_nov.append(transaccion)

    wb_v2.close()

    print(f"   ✅ {len(transacciones_nov)} transacciones encontradas")

    if len(transacciones_nov) == 0:
        print("\n❌ No se encontraron transacciones de Noviembre 2025")
        print("   Verifica que v2.0 tenga datos de ese mes")
        sys.exit(1)

    # Cargar v3.0
    print(f"\n📝 Insertando en {V3_FILE}...")
    wb_v3 = openpyxl.load_workbook(V3_FILE)
    ws_v3 = wb_v3["TRANSACCIONES"]

    # Encontrar primera fila vacía
    fila_inicio = 2
    while ws_v3.cell(fila_inicio, 1).value is not None:
        fila_inicio += 1

    print(f"   Insertando desde fila {fila_inicio}...")

    # Insertar transacciones
    insertadas = 0
    for idx, trans in enumerate(transacciones_nov, start=fila_inicio):
        ws_v3.cell(idx, 1, trans['fecha'])  # A: Fecha
        ws_v3.cell(idx, 2, trans['tipo'])  # B: Tipo
        ws_v3.cell(idx, 3, trans['categoria'])  # C: Categoría
        ws_v3.cell(idx, 4, trans['descripcion'])  # D: Descripción
        ws_v3.cell(idx, 5, trans['cuenta'])  # E: Cuenta Origen
        ws_v3.cell(idx, 6, trans['entidad'])  # F: Entidad
        ws_v3.cell(idx, 7, trans['factura'])  # G: Factura #
        ws_v3.cell(idx, 8, trans['monto_crc'])  # H: Monto CRC
        ws_v3.cell(idx, 9, trans['monto_usd'])  # I: Monto USD
        ws_v3.cell(idx, 10, trans['tipo_cambio'])  # J: Tipo Cambio
        ws_v3.cell(idx, 11, trans['metodo'])  # K: Método Pago
        ws_v3.cell(idx, 12, trans['estado'])  # L: Estado
        ws_v3.cell(idx, 16, "Importado v2.0")  # P: Creado Por
        ws_v3.cell(idx, 17, datetime.now())  # Q: Fecha Creación

        insertadas += 1

    # Guardar
    backup = f"{V3_FILE}.backup.pre_import.{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    print(f"\n💾 Creando backup: {backup}")
    wb_v3.save(backup)

    print(f"💾 Guardando {V3_FILE}...")
    wb_v3.save(V3_FILE)
    wb_v3.close()

    # Resumen
    print("\n" + "="*60)
    print("✅ IMPORTACIÓN COMPLETADA")
    print("="*60)
    print(f"   📊 Transacciones importadas: {insertadas}")
    print(f"   📄 Archivo actualizado: {V3_FILE}")
    print(f"   💾 Backup creado: {backup}")

    print("\n📋 PRÓXIMOS PASOS:")
    print("   1. Abre v3.0 en Excel")
    print("   2. Revisa columna S (⚠️ Duplicados)")
    print("   3. Elimina duplicados manualmente si hay")
    print("   4. Verifica que los datos se vean correctos")
    print("   5. Ajusta categorías/cuentas si es necesario")

    print("\n" + "="*60)

if __name__ == "__main__":
    try:
        importar_noviembre()
    except Exception as e:
        print(f"\n❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
