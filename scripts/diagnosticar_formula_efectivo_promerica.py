#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
DIAGNÓSTICO FÓRMULA EFECTIVO - PROMERICA
Muestra EXACTAMENTE qué transacciones está sumando la fórmula SUMIFS de Efectivo fila 3
"""
import openpyxl
from datetime import datetime

EXCEL_FILE = "AlvaroVelasco_Finanzas_v2.0.xlsx"

def diagnosticar():
    print("=" * 80)
    print("DIAGNÓSTICO FÓRMULA EFECTIVO - PROMERICA (FILA 3)")
    print("=" * 80)
    print()

    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws_efectivo = wb['Efectivo']
    ws_trans = wb['TRANSACCIONES']

    # Leer valores de Efectivo fila 3
    print("📊 VALORES ACTUALES EN HOJA EFECTIVO (Fila 3):")
    print()

    cuenta_efectivo = ws_efectivo.cell(3, 3).value  # C3
    ingreso_efectivo = ws_efectivo.cell(3, 4).value  # D3
    egreso_efectivo = ws_efectivo.cell(3, 5).value  # E3
    balance_efectivo = ws_efectivo.cell(3, 6).value  # F3

    print(f"   Cuenta (C3): {cuenta_efectivo}")
    print(f"   Ingreso (D3): ${float(ingreso_efectivo):,.2f}" if ingreso_efectivo else "   Ingreso (D3): N/A")
    print(f"   Egreso (E3): ${float(egreso_efectivo):,.2f}" if egreso_efectivo else "   Egreso (E3): N/A")
    print(f"   Balance (F3): ${float(balance_efectivo):,.2f}" if balance_efectivo else "   Balance (F3): N/A")
    print()

    # Leer fórmulas
    wb_formulas = openpyxl.load_workbook(EXCEL_FILE)
    ws_efectivo_formulas = wb_formulas['Efectivo']

    formula_d3 = ws_efectivo_formulas.cell(3, 4).value  # D3
    formula_e3 = ws_efectivo_formulas.cell(3, 5).value  # E3

    print("📋 FÓRMULAS EN EFECTIVO FILA 3:")
    print()
    print(f"   D3: {formula_d3}")
    print(f"   E3: {formula_e3}")
    print()

    # Buscar columnas en TRANSACCIONES
    headers = [ws_trans.cell(1, col).value for col in range(1, ws_trans.max_column + 1)]
    col_map = {}
    for col in range(1, len(headers) + 1):
        if headers[col-1]:
            col_map[headers[col-1]] = col

    # Simular el SUMIFS manualmente
    print("=" * 80)
    print("🔍 SIMULANDO SUMIFS - Qué transacciones captura la fórmula")
    print("=" * 80)
    print()

    print(f"📋 Criterio de búsqueda: Cuenta Bancaria = \"{cuenta_efectivo}\"")
    print()

    # Buscar transacciones que coincidan
    transacciones_ingresos = []
    transacciones_egresos = []

    for row in range(2, ws_trans.max_row + 1):
        cuenta_trans = ws_trans.cell(row, col_map['Cuenta Bancaria']).value
        monto = ws_trans.cell(row, col_map['Monto USD']).value
        ing_egr = ws_trans.cell(row, col_map['Ingreso/Egreso']).value
        fecha = ws_trans.cell(row, col_map['Fecha']).value
        concepto = ws_trans.cell(row, col_map['Concepto']).value

        # Comparar cuenta (exacta)
        if cuenta_trans and str(cuenta_trans).strip() == str(cuenta_efectivo).strip():
            if monto:
                monto_val = float(monto)

                trans_info = {
                    'fila': row,
                    'fecha': fecha,
                    'concepto': concepto[:50] if concepto else 'Sin concepto',
                    'monto': monto_val,
                    'ing_egr': ing_egr
                }

                if ing_egr == 'Ingreso':
                    transacciones_ingresos.append(trans_info)
                elif ing_egr == 'Egreso':
                    transacciones_egresos.append(trans_info)

    # Mostrar transacciones de Ingreso
    print(f"📈 INGRESOS: {len(transacciones_ingresos)} transacciones")
    print()

    total_ingresos = 0
    for i, trans in enumerate(transacciones_ingresos, 1):
        fecha_str = trans['fecha'].strftime('%d/%m/%Y') if isinstance(trans['fecha'], datetime) else 'Sin fecha'
        print(f"{i}. Fila {trans['fila']}: {fecha_str} - +${trans['monto']:,.2f}")
        print(f"   {trans['concepto']}")
        total_ingresos += trans['monto']

    print()
    print(f"💰 TOTAL INGRESOS: ${total_ingresos:,.2f}")
    print(f"📊 VALOR EN D3: ${float(ingreso_efectivo):,.2f}" if ingreso_efectivo else "📊 VALOR EN D3: N/A")

    if ingreso_efectivo and abs(total_ingresos - float(ingreso_efectivo)) < 0.01:
        print("✅ COINCIDE con el valor en D3")
    else:
        print("⚠️  NO COINCIDE con el valor en D3")

    print()

    # Mostrar transacciones de Egreso
    print("=" * 80)
    print(f"📉 EGRESOS: {len(transacciones_egresos)} transacciones")
    print()

    total_egresos = 0
    for i, trans in enumerate(transacciones_egresos, 1):
        fecha_str = trans['fecha'].strftime('%d/%m/%Y') if isinstance(trans['fecha'], datetime) else 'Sin fecha'
        print(f"{i}. Fila {trans['fila']}: {fecha_str} - ${abs(trans['monto']):,.2f}")
        print(f"   {trans['concepto']}")
        total_egresos += abs(trans['monto'])

    print()
    print(f"💰 TOTAL EGRESOS: ${total_egresos:,.2f}")
    print(f"📊 VALOR EN E3: ${float(egreso_efectivo):,.2f}" if egreso_efectivo else "📊 VALOR EN E3: N/A")

    if egreso_efectivo and abs(total_egresos - float(egreso_efectivo)) < 0.01:
        print("✅ COINCIDE con el valor en E3")
    else:
        print("⚠️  NO COINCIDE con el valor en E3")

    print()

    # Calcular balance
    balance_calculado = total_ingresos - total_egresos

    print("=" * 80)
    print("💰 BALANCE CALCULADO")
    print("=" * 80)
    print()

    print(f"   Ingresos: ${total_ingresos:,.2f}")
    print(f"   Egresos: ${total_egresos:,.2f}")
    print(f"   Balance: ${balance_calculado:,.2f}")
    print()

    print(f"📊 Balance en F3: ${float(balance_efectivo):,.2f}" if balance_efectivo else "📊 Balance en F3: N/A")
    print(f"🏦 Balance esperado (extracto): $2,163.44")
    print()

    diferencia_f3 = abs(balance_calculado - float(balance_efectivo)) if balance_efectivo else 0
    diferencia_extracto = abs(balance_calculado - 2163.44)

    if diferencia_f3 < 0.01:
        print("✅ Balance calculado COINCIDE con F3")
    else:
        print(f"⚠️  Balance calculado NO COINCIDE con F3 (dif: ${diferencia_f3:,.2f})")

    print()

    if diferencia_extracto < 1.00:
        print("✅ ¡Balance CORRECTO según extracto!")
    else:
        print(f"⚠️  Diferencia con extracto: ${diferencia_extracto:,.2f}")

    print()

    # Análisis de causa
    print("=" * 80)
    print("🔍 ANÁLISIS")
    print("=" * 80)
    print()

    if balance_calculado > 10000:
        print("⚠️  PROBLEMA DETECTADO: Balance demasiado alto")
        print()
        print("📋 POSIBLES CAUSAS:")
        print("   1. Celda C3 contiene un valor incorrecto o fórmula")
        print("   2. Hay transacciones de otra cuenta con nombre similar")
        print("   3. Hay transacciones duplicadas")
        print()
        print("🔧 RECOMENDACIÓN:")
        print("   1. Verificar que C3 contenga EXACTAMENTE: 'Promerica USD (40000003881774)'")
        print("   2. Verificar que no sea una fórmula (=TRANSACCIONES!E2)")
        print("   3. Si C3 es fórmula, cambiarla por texto fijo")

    elif diferencia_extracto > 1:
        print(f"⚠️  Diferencia con extracto: ${diferencia_extracto:,.2f}")
        print()
        print("📋 POSIBLES CAUSAS:")
        print("   1. Faltan transacciones por registrar")
        print("   2. Saldo inicial incorrecto")
        print("   3. Hay transacciones incorrectas")

    else:
        print("✅ Los datos parecen correctos")

    print()
    print("=" * 80)
    print("✅ DIAGNÓSTICO COMPLETADO")
    print("=" * 80)
    print()

if __name__ == "__main__":
    try:
        diagnosticar()
    except Exception as e:
        print(f"❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
