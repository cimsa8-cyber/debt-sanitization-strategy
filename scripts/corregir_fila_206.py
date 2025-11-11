#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CORRECCIÓN FILA 206 - Cliente/Proveedor
Corrige columna F (Cliente/Proveedor) de "USD" a "Intcomex Costa Rica"
"""
import openpyxl

EXCEL_FILE = "AlvaroVelasco_Finanzas_v2.0.xlsx"

def corregir_fila_206():
    """Corrige el cliente/proveedor en fila 206"""

    print("=" * 80)
    print("CORRECCIÓN FILA 206 - INTCOMEX")
    print("=" * 80)
    print()

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb['TRANSACCIONES']

    # Identificar columnas
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    idx_cliente = headers.index('Cliente/Proveedor') + 1

    # Leer valor actual
    valor_actual = ws.cell(206, idx_cliente).value

    print(f"📋 FILA 206:")
    print(f"   Columna: F (Cliente/Proveedor)")
    print(f"   Valor actual: '{valor_actual}'")
    print(f"   Valor correcto: 'Intcomex Costa Rica'")
    print()

    # Corregir
    ws.cell(206, idx_cliente).value = "Intcomex Costa Rica"

    # Guardar
    wb.save(EXCEL_FILE)

    print("✅ Fila 206 corregida exitosamente")
    print()
    print("=" * 80)

    return True

if __name__ == "__main__":
    try:
        corregir_fila_206()
        print("🎉 Corrección completada!")
    except Exception as e:
        print(f"❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
