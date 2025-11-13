#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
DIAGNÓSTICO: Leer hoja Efectivo tal como está
Muestra toda la estructura para identificar el formato exacto
"""
import openpyxl
import sys
import os

EXCEL_FILE = "AlvaroVelasco_Finanzas_v2.0.xlsx"

print("="*80)
print("DIAGNÓSTICO: LECTURA HOJA EFECTIVO")
print("="*80)

if not os.path.exists(EXCEL_FILE):
    print(f"\n❌ ERROR: No se encontró {EXCEL_FILE}")
    sys.exit(1)

# Cargar Excel
print(f"\nCargando {EXCEL_FILE}...")
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

if 'Efectivo' not in wb.sheetnames:
    print("\n❌ ERROR: No se encontró la hoja 'Efectivo'")
    sys.exit(1)

ws = wb['Efectivo']

print(f"\n✓ Hoja Efectivo encontrada")
print(f"   Filas con datos: {ws.max_row}")
print(f"   Columnas con datos: {ws.max_column}")

# Leer primeras 20 filas completas
print("\n" + "="*80)
print("CONTENIDO DE LA HOJA EFECTIVO (Primeras 20 filas)")
print("="*80)

for row in range(1, min(21, ws.max_row + 1)):
    print(f"\nFila {row}:")

    # Leer todas las columnas (A-H probablemente)
    cols_data = []
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        valor = ws[f'{col}{row}'].value
        if valor is not None:
            # Formatear fechas
            if hasattr(valor, 'strftime'):
                valor_str = valor.strftime('%Y-%m-%d')
            else:
                valor_str = str(valor)
            cols_data.append(f"{col}: {valor_str}")

    if cols_data:
        for dato in cols_data:
            print(f"   {dato}")
    else:
        print(f"   (fila vacía)")

# Buscar específicamente líneas con "Promerica"
print("\n" + "="*80)
print("BÚSQUEDA: Filas que contienen 'Promerica'")
print("="*80)

encontradas = []

for row in range(1, ws.max_row + 1):
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        valor = ws[f'{col}{row}'].value
        if valor and 'Promerica' in str(valor):
            encontradas.append({
                'fila': row,
                'columna': col,
                'valor': valor
            })
            break

if encontradas:
    print(f"\n✓ Encontradas {len(encontradas)} filas con 'Promerica':")
    for item in encontradas:
        print(f"\n   Fila {item['fila']}, Columna {item['columna']}:")
        print(f"      Valor: {item['valor']}")

        # Mostrar toda la fila
        print(f"      Contenido completo de la fila:")
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            val = ws[f'{col}{item["fila"]}'].value
            if val:
                if hasattr(val, 'strftime'):
                    val = val.strftime('%Y-%m-%d')
                print(f"         {col}: {val}")
else:
    print("\n⚠️ No se encontraron filas con 'Promerica'")

# Estructura esperada
print("\n" + "="*80)
print("ESTRUCTURA ESPERADA POR EL USUARIO")
print("="*80)

print("""
Según el usuario, la hoja Efectivo debería tener:

Columnas:
  A: Fecha
  B: Concepto
  C: Cuenta
  D: Ingreso
  E: Egreso
  F: Balance

Ejemplo de fila:
  A: 2025-11-01
  B: Balance inicial Promerica USD
  C: Promerica USD (40000003881774)
  D: 2999.24
  E: 0
  F: 2999.24
""")

print("\n" + "="*80)
