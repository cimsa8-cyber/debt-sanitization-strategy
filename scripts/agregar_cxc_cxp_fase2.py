#!/usr/bin/env python3
"""
FASE 2: Agregar hojas CxC y CxP a v3.0
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime, timedelta

V3_FILE = "AlvaroVelasco_Finanzas_v3.0.xlsx"

COLOR_HEADER = "1F4E78"
COLOR_EDITABLE = "FFF2CC"
COLOR_WARNING = "FCE4D6"
COLOR_ERROR = "FFC7CE"

BORDER_THIN = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# CxC Data (del cuestionario)
CXC_DATA = [
    {"cliente": "Grupo Acción Comercial S.A.", "factura": "AR-002", "monto": 1689.04, "dias_credito": 30},
    {"cliente": "VWR International Ltda", "factura": "AR-001", "monto": 3567.08, "dias_credito": 30},
    {"cliente": "3-102-887892 SRL", "factura": "AR-004", "monto": 691.56, "dias_credito": 30},
    {"cliente": "Centro Integral Oncología CIO SRL", "factura": "AR-006", "monto": 687.05, "dias_credito": 30},
    {"cliente": "Ortodoncia de la Cruz", "factura": "AR-007", "monto": 494.50, "dias_credito": 30},
    {"cliente": "Global Automotriz GACR S.A.", "factura": "AR-008", "monto": 439.61, "dias_credito": 30},
    {"cliente": "Solusa Consolidators", "factura": "AR-009", "monto": 378.35, "dias_credito": 30},
    {"cliente": "Cemso", "factura": "AR-010", "monto": 333.92, "dias_credito": 30},
    {"cliente": "Rodriguez Rojas Carlos Humberto", "factura": "AR-012", "monto": 282.50, "dias_credito": 30},
    {"cliente": "Supply Net C.R.W.H S.A.", "factura": "AR-013", "monto": 276.85, "dias_credito": 30},
    {"cliente": "Operation Managment Tierra Magnifica", "factura": "AR-014", "monto": 209.06, "dias_credito": 30},
    {"cliente": "Sevilla Navarro Edgar", "factura": "AR-016", "monto": 169.50, "dias_credito": 30},
    {"cliente": "Gomez Ajoy Edgar Luis", "factura": "AR-017", "monto": 113.00, "dias_credito": 30},
    {"cliente": "Melendez Morales Monica", "factura": "AR-018", "monto": 113.00, "dias_credito": 30},
    {"cliente": "Bandogo Soluciones Tecnológicas S.A.", "factura": "AR-019", "monto": 67.80, "dias_credito": 30},
    {"cliente": "CPF Servicios Radiológicos S.A.", "factura": "AR-020", "monto": 56.50, "dias_credito": 30},
    {"cliente": "Ortodec S.A.", "factura": "AR-021", "monto": 56.50, "dias_credito": 30},
    {"cliente": "Perez Morales Francisco", "factura": "AR-022", "monto": 42.38, "dias_credito": 30},
]

# CxP Data
CXP_DATA = [
    {"proveedor": "Intcomex Costa Rica", "factura": "2518439", "monto": 2317.09, "vencimiento": "2025-12-03", "prioridad": "ALTA"},
    {"proveedor": "Intcomex Costa Rica", "factura": "2518765", "monto": 679.12, "vencimiento": "2025-12-04", "prioridad": "ALTA"},
    {"proveedor": "Intcomex Costa Rica", "factura": "2520652", "monto": 565.34, "vencimiento": "2025-12-10", "prioridad": "ALTA"},
    {"proveedor": "Alquiler Oficina", "factura": "Mensual", "monto": 775.00, "vencimiento": "2025-12-15", "prioridad": "CRÍTICA"},
    {"proveedor": "Nissan - Financiamiento", "factura": "Cuota mensual", "monto": 800.00, "vencimiento": "2025-12-15", "prioridad": "ALTA"},
    {"proveedor": "Hacienda - IVA/Renta", "factura": "Arreglo de pago", "monto": 10000.00, "vencimiento": "2025-12-31", "prioridad": "BAJA"},
    {"proveedor": "Sea Global Logistics (SGL)", "factura": "Importación 1", "monto": 14.69, "vencimiento": "2025-11-13", "prioridad": "MEDIA"},
    {"proveedor": "Sea Global Logistics (SGL)", "factura": "Importación 2", "monto": 14.69, "vencimiento": "2025-11-13", "prioridad": "MEDIA"},
    {"proveedor": "Corporacion Geoalerta TTI, S.A.", "factura": "Seguridad Sept", "monto": 33.90, "vencimiento": "2025-11-13", "prioridad": "MEDIA"},
    {"proveedor": "Computadores Economicos, S.A.", "factura": "LCD/Bateria/Pantalla", "monto": 284.90, "vencimiento": "2025-11-23", "prioridad": "MEDIA"},
    {"proveedor": "Eurocomp de Costa Rica, S.A.", "factura": "Teclado/Mouse", "monto": 16.92, "vencimiento": "2025-11-19", "prioridad": "MEDIA"},
    {"proveedor": "Eurocomp de Costa Rica, S.A.", "factura": "PC/Laptop/UPS/Monitor", "monto": 2007.68, "vencimiento": "2025-11-16", "prioridad": "ALTA"},
]

print("\n" + "="*60)
print("FASE 2: Agregar CxC y CxP")
print("="*60)

wb = openpyxl.load_workbook(V3_FILE)

# ============================================================================
# HOJA CxC (CUENTAS POR COBRAR)
# ============================================================================

print("\n📊 Creando hoja CxC...")
ws = wb.create_sheet("CxC")

ws['A1'] = "CUENTAS POR COBRAR (CxC)"
ws['A1'].font = Font(size=14, bold=True)
ws.merge_cells('A1:J1')

# Headers
headers = ["Cliente", "Factura", "Fecha Emisión", "Fecha Vencimiento", "Días Crédito", "Monto USD", "Saldo USD", "Días Vencido", "Estado", "Notas"]
for col, header in enumerate(headers, start=1):
    cell = ws.cell(2, col, header)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = BORDER_THIN

# Anchos
ws.column_dimensions['A'].width = 35
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 10
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 12
ws.column_dimensions['H'].width = 12
ws.column_dimensions['I'].width = 12
ws.column_dimensions['J'].width = 30

# Data
row = 3
for cxc in CXC_DATA:
    fecha_emision = datetime(2025, 11, 1)  # Asumiendo emisión Nov 1
    fecha_venc = fecha_emision + timedelta(days=cxc["dias_credito"])

    ws.cell(row, 1, cxc["cliente"])
    ws.cell(row, 2, cxc["factura"])
    ws.cell(row, 3, fecha_emision)
    ws.cell(row, 3).number_format = 'DD/MM/YY'
    ws.cell(row, 4, fecha_venc)
    ws.cell(row, 4).number_format = 'DD/MM/YY'
    ws.cell(row, 5, cxc["dias_credito"])
    ws.cell(row, 6, cxc["monto"])
    ws.cell(row, 6).number_format = '$#,##0.00'
    ws.cell(row, 7, cxc["monto"])  # Saldo = monto (no hay pagos aún)
    ws.cell(row, 7).number_format = '$#,##0.00'

    # Días vencido (fórmula)
    ws.cell(row, 8, f'=IF(D{row}<TODAY(), TODAY()-D{row}, 0)')
    ws.cell(row, 8).number_format = '0'

    # Estado (fórmula)
    ws.cell(row, 9, f'=IF(H{row}=0, "AL DÍA", IF(H{row}<=30, "VENCIDA", "CRÍTICA"))')

    row += 1

# Total
ws.cell(row, 1, "TOTAL CxC")
ws.cell(row, 1).font = Font(bold=True)
ws.cell(row, 6, f'=SUM(F3:F{row-1})')
ws.cell(row, 6).number_format = '$#,##0.00'
ws.cell(row, 6).font = Font(bold=True)
ws.cell(row, 7, f'=SUM(G3:G{row-1})')
ws.cell(row, 7).number_format = '$#,##0.00'
ws.cell(row, 7).font = Font(bold=True)

print(f"   ✅ {len(CXC_DATA)} cuentas por cobrar")
print(f"   💰 Total: ${sum([c['monto'] for c in CXC_DATA]):,.2f}")

# ============================================================================
# HOJA CxP (CUENTAS POR PAGAR)
# ============================================================================

print("\n📊 Creando hoja CxP...")
ws = wb.create_sheet("CxP")

ws['A1'] = "CUENTAS POR PAGAR (CxP)"
ws['A1'].font = Font(size=14, bold=True, color='FF0000')
ws.merge_cells('A1:J1')

# Headers
headers = ["Proveedor", "Factura", "Fecha Emisión", "Fecha Vencimiento", "Monto USD", "Saldo USD", "Días Vencido", "Prioridad", "Estado", "Notas"]
for col, header in enumerate(headers, start=1):
    cell = ws.cell(2, col, header)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = BORDER_THIN

# Anchos
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 12
ws.column_dimensions['H'].width = 12
ws.column_dimensions['I'].width = 12
ws.column_dimensions['J'].width = 30

# Data
row = 3
for cxp in CXP_DATA:
    fecha_venc = datetime.strptime(cxp["vencimiento"], "%Y-%m-%d")
    fecha_emision = fecha_venc - timedelta(days=30)

    ws.cell(row, 1, cxp["proveedor"])
    ws.cell(row, 2, cxp["factura"])
    ws.cell(row, 3, fecha_emision)
    ws.cell(row, 3).number_format = 'DD/MM/YY'
    ws.cell(row, 4, fecha_venc)
    ws.cell(row, 4).number_format = 'DD/MM/YY'
    ws.cell(row, 5, cxp["monto"])
    ws.cell(row, 5).number_format = '$#,##0.00'
    ws.cell(row, 6, cxp["monto"])
    ws.cell(row, 6).number_format = '$#,##0.00'

    # Días vencido
    ws.cell(row, 7, f'=IF(D{row}<TODAY(), TODAY()-D{row}, 0)')
    ws.cell(row, 7).number_format = '0'

    ws.cell(row, 8, cxp["prioridad"])

    # Estado
    ws.cell(row, 9, f'=IF(G{row}=0, "AL DÍA", IF(G{row}<=15, "POR VENCER", "VENCIDA"))')

    row += 1

# Total
ws.cell(row, 1, "TOTAL CxP")
ws.cell(row, 1).font = Font(bold=True, color='FF0000')
ws.cell(row, 5, f'=SUM(E3:E{row-1})')
ws.cell(row, 5).number_format = '$#,##0.00'
ws.cell(row, 5).font = Font(bold=True)
ws.cell(row, 6, f'=SUM(F3:F{row-1})')
ws.cell(row, 6).number_format = '$#,##0.00'
ws.cell(row, 6).font = Font(bold=True)

print(f"   ✅ {len(CXP_DATA)} cuentas por pagar")
print(f"   💰 Total: ${sum([c['monto'] for c in CXP_DATA]):,.2f}")

# Guardar
print(f"\n💾 Guardando {V3_FILE}...")
wb.save(V3_FILE)
wb.close()

print("\n" + "="*60)
print("✅ FASE 2 COMPLETADA")
print("="*60)
print(f"   CxC: ${sum([c['monto'] for c in CXC_DATA]):,.2f}")
print(f"   CxP: ${sum([c['monto'] for c in CXP_DATA]):,.2f}")
print(f"   Neto: ${sum([c['monto'] for c in CXC_DATA]) - sum([c['monto'] for c in CXP_DATA]):,.2f}")
print("="*60)
