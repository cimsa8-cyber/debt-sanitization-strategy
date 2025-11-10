#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
VER CONTENIDO DE TRANSACCIONES FILA 2
La hoja Efectivo tiene fórmulas que apuntan a TRANSACCIONES!A2, B2, etc.
Necesitamos ver qué hay ahí
"""
import openpyxl
import sys
import os

EXCEL_FILE = "AlvaroVelasco_Finanzas_v2.0.xlsx"

print("="*80)
print("INVESTIGACIÓN: TRANSACCIONES FILA 2")
print("="*80)

if not os.path.exists(EXCEL_FILE):
    print(f"\n❌ ERROR: No se encontró {EXCEL_FILE}")
    sys.exit(1)

# Cargar Excel
print(f"\nCargando {EXCEL_FILE}...")
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

if 'TRANSACCIONES' not in wb.sheetnames:
    print("\n❌ ERROR: No se encontró la hoja 'TRANSACCIONES'")
    sys.exit(1)

ws = wb['TRANSACCIONES']

# Ver fila 2 (la que usa Efectivo)
print("\n" + "="*80)
print("FILA 2 DE TRANSACCIONES (usada por hoja Efectivo)")
print("="*80)

fila = 2

# Columnas relevantes
columnas = {
    'A': 'Fecha',
    'B': 'Tipo',
    'C': 'Categoría',
    'D': 'Entidad',
    'E': 'Cuenta',
    'F': 'Proveedor',
    'G': 'Concepto',
    'H': 'Referencia',
    'I': 'Monto USD',
    'J': 'Monto CRC',
    'K': 'Ingreso/Egreso',
    'L': 'Estado',
    'M': 'Prioridad',
    'N': 'Vencimiento',
    'O': 'Notas',
    'P': 'ID',
    'Q': 'Fecha Creación',
    'R': 'Usuario',
    'S': 'Duplicado',
    'T': 'Validación'
}

print(f"\nContenido de Fila {fila}:")

for col, nombre in columnas.items():
    valor = ws[f'{col}{fila}'].value
    if valor is not None:
        if hasattr(valor, 'strftime'):
            valor_str = valor.strftime('%Y-%m-%d %H:%M:%S')
        else:
            valor_str = str(valor)
        print(f"   {col} ({nombre}): {valor_str}")

# Ver primeras 10 filas para contexto
print("\n" + "="*80)
print("PRIMERAS 10 FILAS DE TRANSACCIONES (para contexto)")
print("="*80)

for fila in range(1, 11):
    # Solo mostrar columnas clave: Fecha, Tipo, Cuenta, Concepto, Monto
    fecha = ws[f'A{fila}'].value
    tipo = ws[f'B{fila}'].value
    cuenta = ws[f'E{fila}'].value
    concepto = ws[f'G{fila}'].value
    monto = ws[f'I{fila}'].value
    tipo_mov = ws[f'K{fila}'].value

    if fecha or tipo or cuenta:
        print(f"\nFila {fila}:")
        if fecha:
            fecha_str = fecha.strftime('%d/%m/%Y') if hasattr(fecha, 'strftime') else str(fecha)
            print(f"   Fecha: {fecha_str}")
        if tipo:
            print(f"   Tipo: {tipo}")
        if cuenta:
            print(f"   Cuenta: {cuenta}")
        if concepto:
            print(f"   Concepto: {concepto[:60] if len(str(concepto)) > 60 else concepto}")
        if monto:
            print(f"   Monto: ${monto}")
        if tipo_mov:
            print(f"   Tipo Mov: {tipo_mov}")

# Buscar balance inicial en primeras 50 filas
print("\n" + "="*80)
print("BÚSQUEDA: Balances iniciales Promerica (primeras 50 filas)")
print("="*80)

encontrados = []

for fila in range(1, 51):
    tipo = ws[f'B{fila}'].value
    cuenta = ws[f'E{fila}'].value
    concepto = ws[f'G{fila}'].value
    monto = ws[f'I{fila}'].value

    if cuenta and 'Promerica USD' in str(cuenta):
        if tipo and ('Balance inicial' in str(tipo) or 'Saldo Inicial' in str(tipo)):
            fecha = ws[f'A{fila}'].value
            fecha_str = fecha.strftime('%d/%m/%Y') if hasattr(fecha, 'strftime') else 'SIN FECHA'

            encontrados.append({
                'fila': fila,
                'fecha': fecha_str,
                'tipo': tipo,
                'cuenta': cuenta,
                'concepto': concepto,
                'monto': monto
            })

if encontrados:
    print(f"\n✓ Encontrados {len(encontrados)} balance(s) inicial(es) de Promerica:")
    for item in encontrados:
        print(f"\n   Fila {item['fila']}:")
        print(f"      Fecha: {item['fecha']}")
        print(f"      Tipo: {item['tipo']}")
        print(f"      Cuenta: {item['cuenta']}")
        print(f"      Concepto: {item['concepto']}")
        print(f"      Monto: ${item['monto']}")

    if len(encontrados) > 1:
        print(f"\n⚠️ PROBLEMA: Hay {len(encontrados)} balances iniciales de Promerica")
        print("   Debería haber solo 1 (del 15/10/2025 con $3,121.51)")
else:
    print("\n⚠️ No se encontraron balances iniciales de Promerica")

print("\n" + "="*80)
print("DIAGNÓSTICO")
print("="*80)

print("""
La hoja EFECTIVO tiene FÓRMULAS que apuntan a TRANSACCIONES.

Por ejemplo:
  Efectivo!F3 = =D3-E3
  Efectivo!D3 = =IF(TRANSACCIONES!K2="Ingreso",TRANSACCIONES!I2,"")

Esto significa que:
  - Efectivo muestra lo que está en TRANSACCIONES fila 2
  - Para cambiar Efectivo, debo actualizar TRANSACCIONES
  - NO puedo modificar Efectivo directamente (son fórmulas)

Si Efectivo muestra $2,999.24, es porque TRANSACCIONES fila 2
tiene ese valor.
""")

print("\n" + "="*80)
