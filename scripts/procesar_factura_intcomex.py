#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PROCESADOR DE FACTURA ELECTRÓNICA - INTCOMEX
Parsea XML de factura electrónica y registra en Excel
"""
import openpyxl
import xml.etree.ElementTree as ET
from datetime import datetime

EXCEL_FILE = "AlvaroVelasco_Finanzas_v2.0.xlsx"

# Factura XML (pegada directamente en el script)
XML_FACTURA = """<?xml version="1.0" encoding="utf-8"?>
<FacturaElectronica xmlns="https://cdn.comprobanteselectronicos.go.cr/xml-schemas/v4.4/facturaElectronica">
   <Clave>50610112500310127328900100002010000832067101832067</Clave>
   <NumeroConsecutivo>00100002010000832067</NumeroConsecutivo>
   <FechaEmision>2025-11-10T15:00:19</FechaEmision>
   <Emisor>
      <Nombre>Intcomex Costa Rica Mayorista en Equipo de Computo S.A.</Nombre>
   </Emisor>
   <Receptor>
      <Nombre>ALVARO VELASCONET SRL</Nombre>
   </Receptor>
   <CondicionVenta>02</CondicionVenta>
   <PlazoCredito>30</PlazoCredito>
   <DetalleServicio>
      <LineaDetalle>
         <NumeroLinea>1</NumeroLinea>
         <Detalle>HP 83A - CF283A - toner cartridge</Detalle>
         <MontoTotalLinea>87.24</MontoTotalLinea>
      </LineaDetalle>
      <LineaDetalle>
         <NumeroLinea>2</NumeroLinea>
         <Detalle>HP 230X Yellow High Yield Original LaserJet Toner Cartridge</Detalle>
         <MontoTotalLinea>235.66</MontoTotalLinea>
      </LineaDetalle>
      <LineaDetalle>
         <NumeroLinea>3</NumeroLinea>
         <Detalle>HP 230X Magenta High Yield Original LaserJet Toner Cartridge</Detalle>
         <MontoTotalLinea>235.66</MontoTotalLinea>
      </LineaDetalle>
      <LineaDetalle>
         <NumeroLinea>4</NumeroLinea>
         <Detalle>Flete por Envio</Detalle>
         <MontoTotalLinea>6.78</MontoTotalLinea>
      </LineaDetalle>
   </DetalleServicio>
   <ResumenFactura>
      <CodigoTipoMoneda>
         <CodigoMoneda>USD</CodigoMoneda>
         <TipoCambio>508.46000</TipoCambio>
      </CodigoTipoMoneda>
      <TotalComprobante>565.34</TotalComprobante>
   </ResumenFactura>
</FacturaElectronica>
"""

def parsear_factura_xml():
    """
    Extrae información clave del XML de factura electrónica
    """
    print("="*80)
    print("PROCESANDO FACTURA ELECTRÓNICA INTCOMEX")
    print("="*80)

    # Parsear XML
    root = ET.fromstring(XML_FACTURA)

    # Namespace de factura electrónica Costa Rica
    ns = {'fe': 'https://cdn.comprobanteselectronicos.go.cr/xml-schemas/v4.4/facturaElectronica'}

    # Extraer datos principales
    clave = root.find('fe:Clave', ns).text
    numero = root.find('fe:NumeroConsecutivo', ns).text
    fecha_str = root.find('fe:FechaEmision', ns).text
    fecha = datetime.strptime(fecha_str, '%Y-%m-%dT%H:%M:%S')

    emisor = root.find('fe:Emisor/fe:Nombre', ns).text
    receptor = root.find('fe:Receptor/fe:Nombre', ns).text

    plazo_credito = int(root.find('fe:PlazoCredito', ns).text)

    # Total
    total = float(root.find('fe:ResumenFactura/fe:TotalComprobante', ns).text)

    # Tipo de cambio
    tc = float(root.find('fe:ResumenFactura/fe:CodigoTipoMoneda/fe:TipoCambio', ns).text)

    # Detalle de productos
    lineas = root.findall('fe:DetalleServicio/fe:LineaDetalle', ns)
    productos = []
    for linea in lineas:
        detalle = linea.find('fe:Detalle', ns).text
        monto = float(linea.find('fe:MontoTotalLinea', ns).text)
        productos.append({'detalle': detalle, 'monto': monto})

    # Calcular fecha de vencimiento
    from datetime import timedelta
    fecha_vencimiento = fecha + timedelta(days=plazo_credito)

    # Mostrar resumen
    print(f"\n📄 DATOS DE LA FACTURA")
    print(f"   Número: {numero}")
    print(f"   Clave: {clave[:20]}...")
    print(f"   Fecha emisión: {fecha.strftime('%d/%m/%Y')}")
    print(f"   Fecha vencimiento: {fecha_vencimiento.strftime('%d/%m/%Y')} ({plazo_credito} días)")
    print(f"   Emisor: {emisor}")
    print(f"   Receptor: {receptor}")
    print(f"   Total: ${total:,.2f} USD")
    print(f"   Tipo cambio: ₡{tc:,.2f}")
    print(f"   Total CRC: ₡{total * tc:,.2f}")

    print(f"\n📦 PRODUCTOS ({len(productos)}):")
    for i, prod in enumerate(productos, 1):
        print(f"   {i}. {prod['detalle'][:60]}")
        print(f"      ${prod['monto']:,.2f}")

    return {
        'numero': numero,
        'clave': clave,
        'fecha': fecha,
        'fecha_vencimiento': fecha_vencimiento,
        'emisor': emisor,
        'receptor': receptor,
        'total_usd': total,
        'tipo_cambio': tc,
        'plazo_credito': plazo_credito,
        'productos': productos
    }


def registrar_en_excel(datos_factura):
    """
    Registra la factura en el Excel de finanzas
    """
    print(f"\n{'='*80}")
    print("REGISTRANDO EN EXCEL")
    print(f"{'='*80}")

    # Abrir Excel
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb['TRANSACCIONES']

    # Buscar próxima fila vacía
    next_row = ws.max_row + 1

    # Formato de fecha compacto (d/m/yy)
    fecha_formato = datos_factura['fecha'].strftime('%-d/%-m/%y')

    # Preparar concepto detallado
    concepto_items = [p['detalle'][:40] for p in datos_factura['productos'][:2]]
    if len(datos_factura['productos']) > 2:
        concepto_items.append(f"+ {len(datos_factura['productos']) - 2} items más")
    concepto = " | ".join(concepto_items)

    # Registrar movimiento
    # A: Fecha
    ws[f'A{next_row}'] = datos_factura['fecha']
    ws[f'A{next_row}'].number_format = 'd/m/yy'

    # B: Tipo
    ws[f'B{next_row}'] = 'Egreso'

    # C: Categoría
    ws[f'C{next_row}'] = 'Gastos Operativos'

    # D: Subcategoría
    ws[f'D{next_row}'] = 'Suministros de Oficina'

    # E: Cuenta
    ws[f'E{next_row}'] = 'Por Pagar'

    # F: Moneda
    ws[f'F{next_row}'] = 'USD'

    # G: Concepto
    ws[f'G{next_row}'] = f"Factura Intcomex #{datos_factura['numero'][-6:]} - {concepto}"

    # H: Referencia (clave de factura - primeros 20 caracteres)
    ws[f'H{next_row}'] = datos_factura['clave'][:20]

    # I: Monto USD
    ws[f'I{next_row}'] = datos_factura['total_usd']

    # J: Monto CRC
    ws[f'J{next_row}'] = datos_factura['total_usd'] * datos_factura['tipo_cambio']

    # K: Ingreso/Egreso
    ws[f'K{next_row}'] = 'Egreso'

    # L: Método de pago
    ws[f'L{next_row}'] = f'Crédito {datos_factura["plazo_credito"]} días'

    # M: Estado
    ws[f'M{next_row}'] = 'Pendiente'

    # N: Proveedor
    ws[f'N{next_row}'] = 'Intcomex Costa Rica'

    # O: Notas
    ws[f'O{next_row}'] = f"Vence: {datos_factura['fecha_vencimiento'].strftime('%d/%m/%Y')}"

    # Guardar
    wb.save(EXCEL_FILE)

    print(f"\n✅ FACTURA REGISTRADA EXITOSAMENTE")
    print(f"   Fila: {next_row}")
    print(f"   Fecha: {fecha_formato}")
    print(f"   Cuenta: Por Pagar")
    print(f"   Monto: ${datos_factura['total_usd']:,.2f}")
    print(f"   Vencimiento: {datos_factura['fecha_vencimiento'].strftime('%d/%m/%Y')}")
    print(f"   Estado: Pendiente de pago")

    print(f"\n📋 RESUMEN DEL REGISTRO:")
    print(f"   • Esta factura se agregó a 'Por Pagar'")
    print(f"   • Tienes {datos_factura['plazo_credito']} días para pagar (vence {datos_factura['fecha_vencimiento'].strftime('%d/%m')})")
    print(f"   • Cuando la pagues, registra el pago desde la cuenta bancaria")
    print(f"   • El sistema detectará que es pago de factura por la referencia")


def main():
    """
    Proceso principal
    """
    # Parsear factura
    datos = parsear_factura_xml()

    # Registrar en Excel
    registrar_en_excel(datos)

    print(f"\n{'='*80}")
    print("PROCESO COMPLETADO")
    print(f"{'='*80}")

    print(f"\n💡 PRÓXIMOS PASOS:")
    print(f"   1. Abre el Excel y verifica la fila {openpyxl.load_workbook(EXCEL_FILE)['TRANSACCIONES'].max_row}")
    print(f"   2. Cuando pagues la factura, registra:")
    print(f"      - Egreso desde cuenta bancaria (ej: Promerica USD)")
    print(f"      - Referencia: misma clave de factura")
    print(f"      - Concepto: Pago factura Intcomex #832067")
    print(f"   3. Cambia Estado de 'Pendiente' a 'Pagado'")


if __name__ == "__main__":
    main()
