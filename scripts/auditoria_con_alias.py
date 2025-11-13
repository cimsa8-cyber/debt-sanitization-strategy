#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AUDITORÍA GLOBAL CON SISTEMA DE ALIAS
Reconoce automáticamente todas las variaciones de nombres
y consolida saldos correctamente
"""
import openpyxl
from datetime import datetime
from collections import defaultdict
import sys
import os

# Importar sistema de alias
try:
    from alias_cuentas import (
        obtener_nombre_canonico,
        listar_cuentas,
        es_misma_cuenta,
        es_balance_inicial
    )
except ImportError:
    print("ERROR: No se pudo importar alias_cuentas.py")
    print("Asegúrese de que alias_cuentas.py esté en el mismo directorio")
    sys.exit(1)

EXCEL_FILE = "AlvaroVelasco_Finanzas_v2.0.xlsx"

print("="*80)
print("AUDITORIA GLOBAL CON SISTEMA DE ALIAS")
print("="*80)

if not os.path.exists(EXCEL_FILE):
    print(f"\nERROR: No se encontró {EXCEL_FILE}")
    sys.exit(1)

# Cargar Excel
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
ws_trans = wb['TRANSACCIONES']

# Leer transacciones y agrupar por cuenta canónica
print("\nLeyendo transacciones y aplicando alias...")

movimientos_por_cuenta = defaultdict(list)
nombres_originales_por_cuenta = defaultdict(set)
total_transacciones = 0
transacciones_no_reconocidas = []

for row in range(2, ws_trans.max_row + 1):
    cuenta_original = ws_trans[f'E{row}'].value

    if not cuenta_original or str(cuenta_original).strip() == 'Cuenta Bancaria':
        continue

    cuenta_original = str(cuenta_original).strip()

    # Obtener nombre canónico usando sistema de alias
    cuenta_canonica = obtener_nombre_canonico(cuenta_original)

    if not cuenta_canonica:
        # No se reconoció - agregar a lista de no reconocidos
        transacciones_no_reconocidas.append({
            'fila': row,
            'cuenta_original': cuenta_original
        })
        continue

    # Guardar nombre original usado
    nombres_originales_por_cuenta[cuenta_canonica].add(cuenta_original)

    # Leer datos del movimiento
    fecha = ws_trans[f'A{row}'].value
    concepto = ws_trans[f'G{row}'].value
    monto_usd = ws_trans[f'I{row}'].value
    monto_crc = ws_trans[f'J{row}'].value
    tipo_mov = ws_trans[f'K{row}'].value

    # Determinar monto y moneda
    try:
        if monto_usd and float(monto_usd) != 0:
            monto = float(monto_usd)
            moneda = 'USD'
        elif monto_crc and float(monto_crc) != 0:
            monto = float(monto_crc)
            moneda = 'CRC'
        else:
            continue
    except:
        continue

    # Aplicar signo
    if tipo_mov and 'Egreso' in str(tipo_mov):
        monto = -abs(monto)
    else:
        monto = abs(monto)

    movimientos_por_cuenta[cuenta_canonica].append({
        'fila': row,
        'fecha': fecha,
        'concepto': concepto,
        'monto': monto,
        'moneda': moneda,
        'tipo_mov': tipo_mov,
        'nombre_usado': cuenta_original
    })

    total_transacciones += 1

print(f"✓ Total transacciones analizadas: {total_transacciones}")
print(f"✓ Cuentas canónicas detectadas: {len(movimientos_por_cuenta)}")

if transacciones_no_reconocidas:
    print(f"⚠️ Transacciones NO reconocidas: {len(transacciones_no_reconocidas)}")
    print("\nPrimeras 10 no reconocidas:")
    for t in transacciones_no_reconocidas[:10]:
        print(f"   Fila {t['fila']}: '{t['cuenta_original']}'")

# Leer balances iniciales desde TRANSACCIONES
# (La hoja Efectivo tiene fórmulas que apuntan a TRANSACCIONES)
print("\nLeyendo balances iniciales desde TRANSACCIONES...")

balances_efectivo = {}

try:
    # Buscar en TRANSACCIONES los movimientos tipo "Apertura Inicial" o "Balance inicial"
    for row in range(2, ws_trans.max_row + 1):
        tipo = ws_trans[f'B{row}'].value
        cuenta = ws_trans[f'E{row}'].value
        monto_usd = ws_trans[f'I{row}'].value
        fecha = ws_trans[f'A{row}'].value

        # Verificar si es un balance/apertura inicial
        if not tipo or not cuenta:
            continue

        # Usar función del sistema de alias para reconocer todos los formatos
        if not es_balance_inicial(tipo):
            continue

        # Convertir nombre de cuenta a canónico
        cuenta_str = str(cuenta).strip()
        cuenta_canonica = obtener_nombre_canonico(cuenta_str)

        if cuenta_canonica and monto_usd:
            try:
                # Si ya existe una apertura para esta cuenta, usar el más reciente
                if cuenta_canonica in balances_efectivo:
                    # Comparar fechas y usar el más reciente
                    if fecha and balances_efectivo[cuenta_canonica]['fecha']:
                        if fecha > balances_efectivo[cuenta_canonica]['fecha']:
                            balances_efectivo[cuenta_canonica] = {
                                'balance': float(monto_usd),
                                'fecha': fecha,
                                'nombre_original': cuenta_str,
                                'fila': row
                            }
                else:
                    balances_efectivo[cuenta_canonica] = {
                        'balance': float(monto_usd),
                        'fecha': fecha,
                        'nombre_original': cuenta_str,
                        'fila': row
                    }
            except:
                pass

    print(f"✓ Balances iniciales encontrados: {len(balances_efectivo)}")
    if len(balances_efectivo) > 0:
        print(f"   Cuentas con balance inicial:")
        for cuenta_can in sorted(balances_efectivo.keys()):
            bal = balances_efectivo[cuenta_can]['balance']
            fila = balances_efectivo[cuenta_can]['fila']
            print(f"      - {cuenta_can}: ${bal:,.2f} (TRANSACCIONES fila {fila})")

except Exception as e:
    print(f"⚠️ Error al leer balances: {e}")

# Calcular saldos consolidados
print("\n" + "="*80)
print("SALDOS CONSOLIDADOS POR CUENTA")
print("="*80)

saldos_calculados = {}
problemas_detectados = []

for cuenta_canonica in sorted(movimientos_por_cuenta.keys()):
    movs = movimientos_por_cuenta[cuenta_canonica]

    # Calcular saldo
    saldo_total = sum([m['monto'] for m in movs])
    moneda = movs[0]['moneda'] if movs else 'USD'

    # Fechas
    movs_con_fecha = [m for m in movs if m['fecha']]
    if movs_con_fecha:
        movs_ordenados = sorted(movs_con_fecha, key=lambda x: x['fecha'])
        fecha_primer_mov = movs_ordenados[0]['fecha']
        fecha_ultimo_mov = movs_ordenados[-1]['fecha']
    else:
        fecha_primer_mov = None
        fecha_ultimo_mov = None

    saldos_calculados[cuenta_canonica] = {
        'saldo': saldo_total,
        'moneda': moneda,
        'num_movimientos': len(movs),
        'fecha_primer_mov': fecha_primer_mov,
        'fecha_ultimo_mov': fecha_ultimo_mov,
        'nombres_usados': nombres_originales_por_cuenta[cuenta_canonica]
    }

    # Mostrar
    print(f"\n📊 {cuenta_canonica}")
    print(f"   Movimientos: {len(movs)}")

    # Mostrar nombres usados (si hay variaciones)
    if len(nombres_originales_por_cuenta[cuenta_canonica]) > 1:
        print(f"   ⚠️ Nombres usados ({len(nombres_originales_por_cuenta[cuenta_canonica])}):")
        for nombre in sorted(nombres_originales_por_cuenta[cuenta_canonica]):
            count = sum(1 for m in movs if m['nombre_usado'] == nombre)
            print(f"      - '{nombre}' ({count} transacciones)")
    else:
        nombre_unico = list(nombres_originales_por_cuenta[cuenta_canonica])[0]
        print(f"   ✓ Nombre único: '{nombre_unico}'")

    # Saldo calculado
    simbolo_moneda = '$' if moneda == 'USD' else '₡'
    print(f"   Saldo calculado: {simbolo_moneda}{saldo_total:,.2f}")

    # Comparar con balance inicial (Apertura Inicial)
    if cuenta_canonica in balances_efectivo:
        balance_inicial = balances_efectivo[cuenta_canonica]['balance']
        nombre_inicial = balances_efectivo[cuenta_canonica]['nombre_original']
        fila_inicial = balances_efectivo[cuenta_canonica]['fila']

        diferencia = saldo_total - balance_inicial

        print(f"   Balance Inicial: ${balance_inicial:,.2f} (TRANSACCIONES:{fila_inicial})")
        if nombres_originales_por_cuenta[cuenta_canonica] != {nombre_inicial}:
            print(f"      (registrado como: '{nombre_inicial}')")

        if abs(diferencia) > 0.01:
            print(f"   ⚠️ DIFERENCIA: ${diferencia:,.2f}")

            # Diagnóstico
            if abs(diferencia) < abs(balance_inicial) * 0.1:
                porcentaje = abs(diferencia / balance_inicial) * 100
                print(f"      💡 Diferencia pequeña (~{porcentaje:.1f}%) - movimientos faltantes o duplicados")
                problemas_detectados.append({
                    'cuenta': cuenta_canonica,
                    'tipo': 'DIFERENCIA_MENOR',
                    'diferencia': diferencia
                })
            else:
                print(f"      💡 Discrepancia significativa - revisar extractos y categorización")
                problemas_detectados.append({
                    'cuenta': cuenta_canonica,
                    'tipo': 'DISCREPANCIA_GRANDE',
                    'diferencia': diferencia
                })
        else:
            print(f"   ✅ COINCIDE con Balance Inicial")
    else:
        print(f"   ⚠️ NO tiene Balance Inicial (Apertura Inicial)")
        problemas_detectados.append({
            'cuenta': cuenta_canonica,
            'tipo': 'SIN_BALANCE_INICIAL'
        })

# Resumen
print("\n" + "="*80)
print("RESUMEN DE AUDITORIA")
print("="*80)

total_cuentas = len(saldos_calculados)
cuentas_con_problemas = len(set([p['cuenta'] for p in problemas_detectados]))
cuentas_ok = total_cuentas - cuentas_con_problemas

print(f"\nTotal cuentas (canónicas): {total_cuentas}")
print(f"Cuentas correctas: {cuentas_ok}")
print(f"Cuentas con problemas: {cuentas_con_problemas}")

if cuentas_con_problemas > 0:
    print(f"\n⚠️ PROBLEMAS DETECTADOS:")

    for tipo_problema in ['DIFERENCIA_MENOR', 'DISCREPANCIA_GRANDE', 'SIN_BALANCE_INICIAL']:
        problemas_tipo = [p for p in problemas_detectados if p['tipo'] == tipo_problema]
        if problemas_tipo:
            print(f"\n{tipo_problema} ({len(problemas_tipo)} cuentas):")
            for prob in problemas_tipo:
                if tipo_problema in ['DIFERENCIA_MENOR', 'DISCREPANCIA_GRANDE']:
                    print(f"   - {prob['cuenta']}: Diferencia ${prob['diferencia']:,.2f}")
                else:
                    print(f"   - {prob['cuenta']}")

# Análisis de nombres inconsistentes
print("\n" + "="*80)
print("ANALISIS: NOMBRES INCONSISTENTES")
print("="*80)

cuentas_con_multiples_nombres = []

for cuenta in sorted(saldos_calculados.keys()):
    if len(nombres_originales_por_cuenta[cuenta]) > 1:
        cuentas_con_multiples_nombres.append(cuenta)

if cuentas_con_multiples_nombres:
    print(f"\n⚠️ {len(cuentas_con_multiples_nombres)} cuentas usan múltiples nombres:")
    for cuenta in cuentas_con_multiples_nombres:
        print(f"\n   {cuenta}:")
        for nombre in sorted(nombres_originales_por_cuenta[cuenta]):
            print(f"      - '{nombre}'")
    print(f"\n💡 RECOMENDACIÓN: Normalizar a un solo nombre por cuenta")
    print(f"   (Aunque el sistema de alias ya los reconoce correctamente)")
else:
    print("\n✅ Todas las cuentas usan un nombre único")

# Fiabilidad
print("\n" + "="*80)
print("ÍNDICE DE FIABILIDAD")
print("="*80)

if total_cuentas > 0:
    fiabilidad = (cuentas_ok / total_cuentas) * 100
else:
    fiabilidad = 0

print(f"\n{'='*40}")
print(f"FIABILIDAD: {fiabilidad:.1f}%")
print(f"{'='*40}")

if fiabilidad >= 90:
    print("\n✅ EXCELENTE: Excel muy confiable")
elif fiabilidad >= 75:
    print("\n✅ BUENO: Excel generalmente confiable")
elif fiabilidad >= 50:
    print("\n⚠️ REGULAR: Necesita correcciones")
else:
    print("\n❌ BAJO: Necesita auditoría completa")

print("\n" + "="*80)
print("SIGUIENTE PASOS")
print("="*80)

if cuentas_con_problemas > 0:
    print("\n1. CORREGIR problemas detectados (ver arriba)")
    print("2. AGREGAR saldos iniciales faltantes")
    print("3. VALIDAR con extractos bancarios")
else:
    print("\n✅ No se detectaron problemas significativos")

if cuentas_con_multiples_nombres:
    print(f"\n4. OPCIONAL: Normalizar {len(cuentas_con_multiples_nombres)} cuentas a un solo nombre")
    print("   (Mejora legibilidad, pero el sistema de alias ya funciona)")

print("\n" + "="*80)
