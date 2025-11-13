#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ANÁLISIS COMPLETO ESTRUCTURA EFECTIVO
Examina TODAS las hojas, fórmulas y dependencias de la fila 3 de Efectivo
"""
import openpyxl

EXCEL_FILE = "AlvaroVelasco_Finanzas_v2.0.xlsx"

def analizar():
    print("=" * 80)
    print("ANÁLISIS COMPLETO ESTRUCTURA EXCEL")
    print("=" * 80)
    print()

    wb = openpyxl.load_workbook(EXCEL_FILE)
    wb_data = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

    # =========================================================================
    # PASO 1: LISTAR TODAS LAS HOJAS
    # =========================================================================
    print("📋 HOJAS EN EL EXCEL:")
    print()

    for idx, hoja in enumerate(wb.sheetnames, 1):
        print(f"{idx}. {hoja}")

    print()

    # =========================================================================
    # PASO 2: ANALIZAR HOJA EFECTIVO COMPLETA
    # =========================================================================
    if 'Efectivo' not in wb.sheetnames:
        print("❌ No existe hoja 'Efectivo'")
        return

    ws_efectivo = wb['Efectivo']
    ws_efectivo_data = wb_data['Efectivo']

    print("=" * 80)
    print("📊 ANÁLISIS DETALLADO HOJA 'EFECTIVO' - FILA 3")
    print("=" * 80)
    print()

    # Analizar TODAS las columnas de fila 3
    print("🔍 TODAS LAS CELDAS FILA 3:")
    print()

    for col in range(1, ws_efectivo.max_column + 1):
        celda_formula = ws_efectivo.cell(3, col)
        celda_valor = ws_efectivo_data.cell(3, col)

        letra_col = openpyxl.utils.get_column_letter(col)

        contenido = celda_formula.value
        valor_evaluado = celda_valor.value

        es_formula = isinstance(contenido, str) and contenido.startswith('=')

        print(f"{letra_col}3:")

        if es_formula:
            print(f"   FÓRMULA: {contenido}")
            print(f"   VALOR: {valor_evaluado}")
        else:
            print(f"   VALOR: {contenido}")

        print()

    # =========================================================================
    # PASO 3: ANALIZAR FILA 1 (ENCABEZADOS)
    # =========================================================================
    print("=" * 80)
    print("📋 ENCABEZADOS (Fila 1):")
    print("=" * 80)
    print()

    for col in range(1, ws_efectivo.max_column + 1):
        letra_col = openpyxl.utils.get_column_letter(col)
        valor = ws_efectivo.cell(1, col).value
        print(f"{letra_col}1: {valor}")

    print()

    # =========================================================================
    # PASO 4: VERIFICAR OTRAS HOJAS
    # =========================================================================
    print("=" * 80)
    print("🔍 VERIFICANDO OTRAS HOJAS QUE PUEDAN ALIMENTAR EFECTIVO")
    print("=" * 80)
    print()

    for nombre_hoja in wb.sheetnames:
        if nombre_hoja in ['Efectivo', 'TRANSACCIONES', 'CUENTAS_ALIAS']:
            continue

        ws_otra = wb[nombre_hoja]

        print(f"📄 Hoja: {nombre_hoja}")
        print(f"   Dimensiones: {ws_otra.max_row} filas x {ws_otra.max_column} columnas")

        # Mostrar encabezados si existen
        encabezados = []
        for col in range(1, min(ws_otra.max_column + 1, 20)):
            valor = ws_otra.cell(1, col).value
            if valor:
                encabezados.append(str(valor))

        if encabezados:
            print(f"   Encabezados: {', '.join(encabezados[:10])}" + (" ..." if len(encabezados) > 10 else ""))

        print()

    # =========================================================================
    # PASO 5: BUSCAR REFERENCIAS A "Promerica" EN OTRAS HOJAS
    # =========================================================================
    print("=" * 80)
    print("🔍 BUSCANDO REFERENCIAS A 'PROMERICA' EN OTRAS HOJAS")
    print("=" * 80)
    print()

    for nombre_hoja in wb.sheetnames:
        if nombre_hoja == 'TRANSACCIONES':
            continue

        ws_buscar = wb[nombre_hoja]

        referencias_encontradas = []

        for row in range(1, min(ws_buscar.max_row + 1, 100)):
            for col in range(1, min(ws_buscar.max_column + 1, 20)):
                celda = ws_buscar.cell(row, col)
                valor = celda.value

                if valor and 'Promerica' in str(valor):
                    letra_col = openpyxl.utils.get_column_letter(col)
                    referencias_encontradas.append({
                        'celda': f"{letra_col}{row}",
                        'valor': str(valor)[:100]
                    })

        if referencias_encontradas:
            print(f"📄 Hoja '{nombre_hoja}': {len(referencias_encontradas)} referencias")
            for ref in referencias_encontradas[:5]:
                print(f"   {ref['celda']}: {ref['valor']}")
            if len(referencias_encontradas) > 5:
                print(f"   ... y {len(referencias_encontradas) - 5} más")
            print()

    # =========================================================================
    # PASO 6: INFORMACIÓN DE PROTECCIÓN
    # =========================================================================
    print("=" * 80)
    print("🔒 INFORMACIÓN DE PROTECCIÓN DE HOJAS")
    print("=" * 80)
    print()

    for nombre_hoja in wb.sheetnames:
        ws = wb[nombre_hoja]

        if ws.protection.sheet:
            print(f"🔒 Hoja '{nombre_hoja}': PROTEGIDA")
        else:
            print(f"🔓 Hoja '{nombre_hoja}': Sin protección")

    print()

    # =========================================================================
    # RESUMEN
    # =========================================================================
    print("=" * 80)
    print("📊 RESUMEN")
    print("=" * 80)
    print()

    print(f"Total de hojas: {len(wb.sheetnames)}")
    print(f"Hoja Efectivo: {ws_efectivo.max_row} filas x {ws_efectivo.max_column} columnas")
    print()
    print("🔍 PRÓXIMO PASO:")
    print("   Revisar si hay hojas adicionales que alimentan Efectivo")
    print("   y si las fórmulas en D3, E3, F3 son las únicas o hay dependencias ocultas")

    print()
    print("=" * 80)
    print("✅ ANÁLISIS COMPLETADO")
    print("=" * 80)

if __name__ == "__main__":
    try:
        analizar()
    except Exception as e:
        print(f"❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
