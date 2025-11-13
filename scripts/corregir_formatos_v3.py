#!/usr/bin/env python3
"""
Corrige formatos de celdas en v3.0
- Fechas: DD/MM/YY
- Montos CRC: ₡#,##0.00
- Montos USD: $#,##0.00
"""

import openpyxl

V3_FILE = "AlvaroVelasco_Finanzas_v3.0.xlsx"

print("Corrigiendo formatos...")

wb = openpyxl.load_workbook(V3_FILE)

# HOJA TRANSACCIONES
ws = wb["TRANSACCIONES"]
print(f"  TRANSACCIONES: {ws.max_row - 1} filas")

for row in range(2, ws.max_row + 1):
    ws.cell(row, 1).number_format = 'DD/MM/YY'  # Fecha
    ws.cell(row, 8).number_format = '₡#,##0.00'  # CRC
    ws.cell(row, 9).number_format = '$#,##0.00'  # USD
    ws.cell(row, 10).number_format = '₡#,##0.00'  # TC
    ws.cell(row, 17).number_format = 'DD/MM/YY'  # Fecha Creación (col Q)

# HOJA EFECTIVO
ws = wb["EFECTIVO"]
print("  EFECTIVO: Aplicando formatos...")

# Cuentas bancarias (filas 5-13 aprox)
for row in range(5, 20):
    cell_moneda = ws.cell(row, 4).value  # Columna D: Moneda
    if cell_moneda == "CRC":
        ws.cell(row, 5).number_format = '₡#,##0.00'
        ws.cell(row, 8).number_format = '₡#,##0.00'
    elif cell_moneda == "USD":
        ws.cell(row, 5).number_format = '$#,##0.00'
        ws.cell(row, 8).number_format = '$#,##0.00'

# HOJA DASHBOARD
ws = wb["DASHBOARD"]
print("  DASHBOARD: Aplicando formatos...")

# Columnas con USD
for row in range(6, 20):
    if ws.cell(row, 3).value is not None:
        ws.cell(row, 3).number_format = '$#,##0.00'

wb.save(V3_FILE)
wb.close()

print("✅ Formatos corregidos con símbolos de moneda")
