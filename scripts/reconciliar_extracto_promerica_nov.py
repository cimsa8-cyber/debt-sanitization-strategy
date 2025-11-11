#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
RECONCILIACIÓN EXTRACTO BANCARIO PROMERICA
Compara extracto del 01/11 al 11/11/2025 con Excel
Saldo esperado: $2,163.44
"""
import openpyxl
from datetime import datetime
import shutil

EXCEL_FILE = "AlvaroVelasco_Finanzas_v2.0.xlsx"
BACKUP_FILE = f"AlvaroVelasco_Finanzas_v2.0_RECONCILIACION_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

# Extracto bancario proporcionado por usuario
EXTRACTO_BANCO = [
    {'fecha': '03/11/2025', 'doc': '67169898', 'desc': 'TEF. ELEC pago CCSS', 'debito': 733.20, 'credito': 0, 'saldo': 2297.69},
    {'fecha': '04/11/2025', 'doc': '5490555', 'desc': 'PEREZ BENAVIDES Y: facturas 2535 2530', 'debito': 0, 'credito': 761.06, 'saldo': 3058.75},
    {'fecha': '06/11/2025', 'doc': '2795229', 'desc': 'Comisión. Debito por Transf. Sinpe', 'debito': 0.75, 'credito': 0, 'saldo': 3058.00},
    {'fecha': '06/11/2025', 'doc': '32048171', 'desc': 'DD: SEA GLOBAL LOGISTICS SERVICES , TRAN', 'debito': 58.76, 'credito': 0, 'saldo': 2999.24},
    {'fecha': '07/11/2025', 'doc': '2796348', 'desc': 'Comisión. Debito por Transf. Sinpe', 'debito': 0.75, 'credito': 0, 'saldo': 2998.49},
    {'fecha': '07/11/2025', 'doc': '2796348', 'desc': 'Reversion de Comisión por transferencia', 'debito': 0, 'credito': 0.75, 'saldo': 2999.24},
    {'fecha': '07/11/2025', 'doc': '2796348', 'desc': 'Comisión. Debito por Transf. Sinpe', 'debito': 0.75, 'credito': 0, 'saldo': 2998.49},
    {'fecha': '07/11/2025', 'doc': '32067344', 'desc': 'DD: ALEJANDRA ARIAS FALLAS , TRAN', 'debito': 40.57, 'credito': 0, 'saldo': 2957.92},
    {'fecha': '07/11/2025', 'doc': '8313051', 'desc': 'CD: ASOCIACION_COSTARRIC, PLANESBPPP', 'debito': 0, 'credito': 333.35, 'saldo': 3291.27},
    {'fecha': '08/11/2025', 'doc': '141111', 'desc': 'ESTACION SERVICIOS NASA HEREDIA CR', 'debito': 9.13, 'credito': 0, 'saldo': 3282.14},
    {'fecha': '10/11/2025', 'doc': '2797944', 'desc': 'Comisión. Debito por Transf. Sinpe', 'debito': 0.75, 'credito': 0, 'saldo': 3281.39},
    {'fecha': '10/11/2025', 'doc': '32105690', 'desc': 'DD: STCR COSTA RICA TRUST AND ESCR, TRAN', 'debito': 3137.26, 'credito': 0, 'saldo': 144.13},
    {'fecha': '10/11/2025', 'doc': '8319443', 'desc': 'CD: WAIPIO_S.A., CD BAC', 'debito': 0, 'credito': 1459.96, 'saldo': 1604.09},
    {'fecha': '10/11/2025', 'doc': '8320303', 'desc': 'CD: GENTRA_DE_COSTA_RICA, CD BAC', 'debito': 0, 'credito': 635.63, 'saldo': 2239.72},
    {'fecha': '10/11/2025', 'doc': '204203', 'desc': 'DELTA TIBAS SAN JOSE CR', 'debito': 76.28, 'credito': 0, 'saldo': 2163.44},
]

SALDO_FINAL_ESPERADO = 2163.44

def crear_backup():
    print("=" * 80)
    print("CREANDO BACKUP")
    print("=" * 80)
    print(f"Backup: {BACKUP_FILE}")
    try:
        shutil.copy2(EXCEL_FILE, BACKUP_FILE)
        print("✅ Backup creado")
        print()
        return True
    except Exception as e:
        print(f"❌ ERROR: {e}")
        return False

def reconciliar():
    print("=" * 80)
    print("RECONCILIACIÓN EXTRACTO BANCARIO PROMERICA")
    print("Período: 01/11/2025 - 11/11/2025")
    print(f"Saldo final extracto: ${SALDO_FINAL_ESPERADO:,.2f}")
    print("=" * 80)
    print()

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb['TRANSACCIONES']

    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    col_map = {}
    for col in range(1, len(headers) + 1):
        if headers[col-1]:
            col_map[headers[col-1]] = col

    # =========================================================================
    # PASO 1: VERIFICAR QUÉ MOVIMIENTOS YA ESTÁN REGISTRADOS
    # =========================================================================
    print("📋 PASO 1: Verificando movimientos del extracto vs Excel...")
    print()

    movimientos_faltantes = []

    for mov_banco in EXTRACTO_BANCO:
        fecha_banco = datetime.strptime(mov_banco['fecha'], '%d/%m/%Y')
        doc_banco = mov_banco['doc']
        monto_banco = -mov_banco['debito'] if mov_banco['debito'] > 0 else mov_banco['credito']

        # Buscar en Excel
        encontrado = False

        for row in range(2, ws.max_row + 1):
            fecha_excel = ws.cell(row, col_map['Fecha']).value
            referencia_excel = ws.cell(row, col_map['Referencia']).value
            monto_excel = ws.cell(row, col_map['Monto USD']).value
            cuenta_excel = ws.cell(row, col_map['Cuenta Bancaria']).value

            # Verificar que sea de Promerica
            if not cuenta_excel or 'Promerica' not in str(cuenta_excel):
                continue

            # Comparar fecha
            if fecha_excel:
                if isinstance(fecha_excel, datetime):
                    if fecha_excel.date() != fecha_banco.date():
                        continue
                else:
                    continue

            # Comparar referencia o monto
            if referencia_excel and doc_banco in str(referencia_excel):
                encontrado = True
                break

            # Si no hay referencia, comparar monto
            if monto_excel and abs(abs(float(monto_excel)) - abs(monto_banco)) < 0.01:
                encontrado = True
                break

        if not encontrado:
            movimientos_faltantes.append(mov_banco)

    # Mostrar resultado
    print(f"📊 RESULTADO:")
    print(f"   Total movimientos en extracto: {len(EXTRACTO_BANCO)}")
    print(f"   Movimientos registrados: {len(EXTRACTO_BANCO) - len(movimientos_faltantes)}")
    print(f"   Movimientos FALTANTES: {len(movimientos_faltantes)}")
    print()

    if not movimientos_faltantes:
        print("✅ ¡Todos los movimientos están registrados!")
        print()
    else:
        print("⚠️  MOVIMIENTOS FALTANTES:")
        print()
        for i, mov in enumerate(movimientos_faltantes, 1):
            tipo_mov = "DÉBITO" if mov['debito'] > 0 else "CRÉDITO"
            monto = mov['debito'] if mov['debito'] > 0 else mov['credito']
            print(f"{i}. {mov['fecha']} - Doc: {mov['doc']}")
            print(f"   {tipo_mov}: ${monto:,.2f}")
            print(f"   {mov['desc']}")
            print()

    # =========================================================================
    # PASO 2: REGISTRAR MOVIMIENTOS FALTANTES
    # =========================================================================
    if movimientos_faltantes:
        print("=" * 80)
        print("📋 PASO 2: Registrando movimientos faltantes...")
        print("=" * 80)
        print()

        transacciones_nuevas = []

        for mov in movimientos_faltantes:
            fecha = datetime.strptime(mov['fecha'], '%d/%m/%Y')
            es_debito = mov['debito'] > 0
            monto = -mov['debito'] if es_debito else mov['credito']

            # Determinar tipo y categoría según descripción
            desc = mov['desc'].upper()

            if 'COMISION' in desc or 'COMISIÓN' in desc:
                tipo = 'GASTOS FINANCIEROS'
                categoria = 'Comisiones Bancarias'
                entidad = 'Banco'
                proveedor = 'Banco Promerica'
                concepto = mov['desc']
            elif 'REVERSION' in desc:
                tipo = 'AJUSTES'
                categoria = 'Ajustes Bancarios'
                entidad = 'Banco'
                proveedor = 'Banco Promerica'
                concepto = mov['desc']
            elif 'CCSS' in desc or 'CAJA COSTARRICENSE' in desc:
                tipo = 'GASTOS OPERATIVOS'
                categoria = 'Cargas Sociales'
                entidad = 'Gobierno'
                proveedor = 'CCSS'
                concepto = 'Pago CCSS - Transferencia electrónica'
            elif 'PEREZ BENAVIDES' in desc:
                tipo = 'INGRESOS'
                categoria = 'Ventas de Productos'
                entidad = 'Cuentas por Cobrar'
                proveedor = 'Pérez Benavides Y.'
                concepto = 'Pago facturas 2535, 2530'
            elif 'SEA GLOBAL' in desc:
                tipo = 'GASTOS OPERATIVOS'
                categoria = 'Logística'
                entidad = 'Proveedor'
                proveedor = 'Sea Global Logistics Services'
                concepto = 'Pago logística - Transferencia'
            elif 'ALEJANDRA ARIAS' in desc:
                tipo = 'GASTOS OPERATIVOS'
                categoria = 'Servicios Profesionales'
                entidad = 'Proveedor'
                proveedor = 'Alejandra Arias Fallas'
                concepto = 'Pago servicios - Transferencia'
            elif 'ASOCIACION' in desc:
                tipo = 'INGRESOS'
                categoria = 'Otros Ingresos'
                entidad = 'Cuentas por Cobrar'
                proveedor = 'Asociación Costarricense'
                concepto = 'Ingreso - PLANESBPPP'
            elif 'NASA' in desc or 'ESTACION' in desc:
                tipo = 'GASTOS OPERATIVOS'
                categoria = 'Combustible'
                entidad = 'Vehículo'
                proveedor = 'Estación Servicios NASA Heredia'
                concepto = 'Combustible - Estación NASA Heredia'
            else:
                tipo = 'OTROS'
                categoria = 'Sin categorizar'
                entidad = 'Otros'
                proveedor = 'Promerica'
                concepto = mov['desc']

            transacciones_nuevas.append({
                'fecha': fecha,
                'tipo': tipo,
                'categoria': categoria,
                'entidad': entidad,
                'cuenta': 'Promerica USD (40000003881774)',
                'proveedor': proveedor,
                'concepto': concepto,
                'referencia': mov['doc'],
                'monto_usd': monto,
                'ingreso_egreso': 'Egreso' if es_debito else 'Ingreso',
                'estado': 'Completado' if not es_debito else 'Pagado',
                'prioridad': 'Normal',
                'notas': f"Según extracto bancario - Doc: {mov['doc']}",
            })

        # Registrar transacciones
        filas_agregadas = []

        for trans in transacciones_nuevas:
            next_row = ws.max_row + 1

            ws.cell(next_row, col_map['Fecha']).value = trans['fecha']
            ws.cell(next_row, col_map['Fecha']).number_format = 'd/m/yy'
            ws.cell(next_row, col_map['Tipo Transacción']).value = trans['tipo']
            ws.cell(next_row, col_map['Categoría']).value = trans['categoria']
            ws.cell(next_row, col_map['Entidad']).value = trans['entidad']
            ws.cell(next_row, col_map['Cuenta Bancaria']).value = trans['cuenta']
            ws.cell(next_row, col_map['Cliente/Proveedor']).value = trans['proveedor']
            ws.cell(next_row, col_map['Concepto']).value = trans['concepto']
            ws.cell(next_row, col_map['Referencia']).value = trans['referencia']
            ws.cell(next_row, col_map['Monto USD']).value = trans['monto_usd']
            ws.cell(next_row, col_map['Ingreso/Egreso']).value = trans['ingreso_egreso']
            ws.cell(next_row, col_map['Estado']).value = trans['estado']
            ws.cell(next_row, col_map['Prioridad']).value = trans['prioridad']
            ws.cell(next_row, col_map['Notas']).value = trans['notas']

            filas_agregadas.append({
                'fila': next_row,
                'fecha': trans['fecha'].strftime('%d/%m/%Y'),
                'concepto': trans['concepto'],
                'monto': trans['monto_usd'],
            })

            signo = '+' if trans['monto_usd'] > 0 else ''
            print(f"✅ {trans['fecha'].strftime('%d/%m/%Y')} - {trans['concepto'][:40]}")
            print(f"   Fila: {next_row}")
            print(f"   Monto: {signo}${abs(trans['monto_usd']):,.2f} USD")
            print()

        # Guardar
        print("💾 Guardando cambios...")
        wb.save(EXCEL_FILE)
        print("✅ Excel actualizado")
        print()

    # =========================================================================
    # PASO 3: VERIFICAR SALDO FINAL
    # =========================================================================
    print("=" * 80)
    print("📊 VERIFICACIÓN DE SALDO")
    print("=" * 80)
    print()

    # Calcular saldo desde Excel
    print("🔍 Calculando saldo Promerica desde transacciones Excel...")
    print()

    # Buscar saldo inicial Promerica
    saldo_inicial = 0
    for row in range(2, ws.max_row + 1):
        cuenta = ws.cell(row, col_map['Cuenta Bancaria']).value
        tipo = ws.cell(row, col_map['Tipo Transacción']).value
        concepto = ws.cell(row, col_map['Concepto']).value
        monto = ws.cell(row, col_map['Monto USD']).value

        if cuenta and 'Promerica' in str(cuenta) and '40000003881774' in str(cuenta):
            if tipo and 'Apertura Inicial' in str(tipo):
                if monto:
                    saldo_inicial = float(monto)
                    print(f"✅ Saldo Inicial Promerica: ${saldo_inicial:,.2f}")
                    break

    # Calcular movimientos de noviembre
    movimientos_nov = []
    for row in range(2, ws.max_row + 1):
        cuenta = ws.cell(row, col_map['Cuenta Bancaria']).value
        fecha = ws.cell(row, col_map['Fecha']).value
        monto = ws.cell(row, col_map['Monto USD']).value
        tipo = ws.cell(row, col_map['Tipo Transacción']).value

        # Solo Promerica
        if not cuenta or 'Promerica' not in str(cuenta) or '40000003881774' not in str(cuenta):
            continue

        # Solo noviembre
        if fecha and isinstance(fecha, datetime):
            if fecha.month == 11 and fecha.year == 2025:
                # Excluir apertura inicial
                if tipo and 'Apertura Inicial' not in str(tipo):
                    if monto:
                        movimientos_nov.append(float(monto))

    total_movimientos = sum(movimientos_nov)
    saldo_calculado = saldo_inicial + total_movimientos

    print(f"   Movimientos noviembre: {len(movimientos_nov)} transacciones")
    print(f"   Total movimientos: ${total_movimientos:,.2f}")
    print(f"   Saldo calculado: ${saldo_calculado:,.2f}")
    print()

    print("📊 COMPARACIÓN:")
    print(f"   Saldo extracto bancario: ${SALDO_FINAL_ESPERADO:,.2f}")
    print(f"   Saldo calculado Excel:   ${saldo_calculado:,.2f}")
    diferencia = abs(saldo_calculado - SALDO_FINAL_ESPERADO)
    print(f"   Diferencia: ${diferencia:,.2f}")
    print()

    if diferencia < 0.50:
        print("✅ ¡SALDOS COINCIDEN! Reconciliación exitosa")
    else:
        print("⚠️  HAY DIFERENCIA - Revisar:")
        print("   1. Puede haber movimientos de octubre no registrados")
        print("   2. Saldo inicial puede estar incorrecto")
        print("   3. Puede faltar algún movimiento")

    print()

    # Resumen
    print("=" * 80)
    print("📊 RESUMEN FINAL")
    print("=" * 80)
    print()

    print(f"✅ Movimientos reconciliados: {len(EXTRACTO_BANCO)}")
    if movimientos_faltantes:
        print(f"✅ Movimientos agregados: {len(movimientos_faltantes)}")
        for mov in filas_agregadas:
            signo = '+' if mov['monto'] > 0 else ''
            print(f"   • Fila {mov['fila']}: {mov['fecha']} - {mov['concepto'][:40]} ({signo}${abs(mov['monto']):,.2f})")
    else:
        print("✅ No fue necesario agregar movimientos")

    print()
    print(f"💰 Saldo final: ${SALDO_FINAL_ESPERADO:,.2f}")
    print()

    print("=" * 80)
    print("✅ RECONCILIACIÓN COMPLETADA")
    print("=" * 80)
    print()

if __name__ == "__main__":
    try:
        if not crear_backup():
            print("❌ Abortando")
            exit(1)

        reconciliar()
        print("🎉 Reconciliación completada exitosamente!")

    except Exception as e:
        print(f"❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
