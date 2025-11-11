#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SINCRONIZACIÓN MAESTRA - 10 NOVIEMBRE 2025
Script único que corrige y sincroniza todo:
1. Corrige fila 206 (Intcomex - columna Cliente/Proveedor)
2. Registra 3 transacciones pendientes del 10/11
"""
import openpyxl
from datetime import datetime
import shutil

EXCEL_FILE = "AlvaroVelasco_Finanzas_v2.0.xlsx"
BACKUP_FILE = f"AlvaroVelasco_Finanzas_v2.0_SYNC_BACKUP_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

def crear_backup():
    """Crea backup antes de modificar"""
    print("=" * 80)
    print("CREANDO BACKUP DE SEGURIDAD")
    print("=" * 80)
    print(f"Backup: {BACKUP_FILE}")

    try:
        shutil.copy2(EXCEL_FILE, BACKUP_FILE)
        print("✅ Backup creado exitosamente")
        print()
        return True
    except Exception as e:
        print(f"❌ ERROR: {e}")
        return False

def sincronizar():
    """Sincroniza todo: correcciones + nuevas transacciones"""

    print("=" * 80)
    print("SINCRONIZACIÓN MAESTRA - 10 NOVIEMBRE 2025")
    print("=" * 80)
    print()

    # Cargar Excel
    print("📂 Cargando Excel...")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb['TRANSACCIONES']
    print(f"✓ Hoja TRANSACCIONES cargada ({ws.max_row} filas)")
    print()

    # Identificar columnas
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]

    col_map = {}
    for col in range(1, len(headers) + 1):
        if headers[col-1]:
            col_map[headers[col-1]] = col

    print("=" * 80)
    print("PASO 1: CORRIGIENDO FILA 206 (Intcomex #832067)")
    print("=" * 80)
    print()

    # Corregir fila 206
    fila_206_antes = ws.cell(206, col_map['Cliente/Proveedor']).value
    ws.cell(206, col_map['Cliente/Proveedor']).value = "Intcomex Costa Rica"

    print(f"✅ Fila 206 corregida:")
    print(f"   Columna: Cliente/Proveedor")
    print(f"   Antes: '{fila_206_antes}'")
    print(f"   Después: 'Intcomex Costa Rica'")
    print()

    print("=" * 80)
    print("PASO 2: REGISTRANDO TRANSACCIONES DEL 10/11/2025")
    print("=" * 80)
    print()

    # Transacciones a registrar
    transacciones = [
        {
            'num': 1,
            'descripcion': 'Ingreso Gentra - Pago facturas',
            'fecha': datetime(2025, 11, 10, 17, 16, 59),
            'tipo': 'INGRESOS',
            'categoria': 'Ventas de Productos',
            'entidad': 'Cuentas por Cobrar',
            'cuenta': 'Promerica USD (40000003881774)',
            'cliente_proveedor': 'Gentra de Costa Rica',
            'concepto': 'Pago facturas electrónicas #02516, #02518, #02555',
            'referencia': '950456696',
            'monto_usd': 635.63,
            'monto_crc': None,
            'ingreso_egreso': 'Ingreso',
            'estado': 'Completado',
            'prioridad': 'Normal',
            'vencimiento': None,
            'notas': 'Transferencia SINPE - Aplicada 10pm',
        },
        {
            'num': 2,
            'descripcion': 'Pago Intcomex factura anterior #821720',
            'fecha': datetime(2025, 11, 10, 11, 31, 0),
            'tipo': 'COMPRAS PARA REVENTA',
            'categoria': 'Productos Tecnológicos',
            'entidad': 'Productos',
            'cuenta': 'Promerica USD (40000003881774)',
            'cliente_proveedor': 'Intcomex Costa Rica',
            'concepto': 'Pago Factura Intcomex #821720 - AlvaroVelascoNet',
            'referencia': '2025111011631000083186200',
            'monto_usd': -3137.26,
            'monto_crc': None,
            'ingreso_egreso': 'Egreso',
            'estado': 'Pagado',
            'prioridad': 'Normal',
            'vencimiento': None,
            'notas': 'SINPE Programado - Doc: 2797944',
        },
        {
            'num': 3,
            'descripcion': 'Diesel Petróleos Delta',
            'fecha': datetime(2025, 11, 10, 18, 33, 16),
            'tipo': 'GASTOS OPERATIVOS',
            'categoria': 'Combustible',
            'entidad': 'Vehículo',
            'cuenta': 'Efectivo',
            'cliente_proveedor': 'Petróleos Delta CR',
            'concepto': 'Diesel 67.759 litros - Estación Tibas',
            'referencia': '50610112500310102878',
            'monto_usd': None,
            'monto_crc': -37606.00,
            'ingreso_egreso': 'Egreso',
            'estado': 'Pagado',
            'prioridad': 'Normal',
            'vencimiento': None,
            'notas': 'Factura #8282 - Contado',
        },
    ]

    filas_agregadas = []

    for trans in transacciones:
        next_row = ws.max_row + 1

        # Escribir datos
        ws.cell(next_row, col_map['Fecha']).value = trans['fecha']
        ws.cell(next_row, col_map['Tipo Transacción']).value = trans['tipo']
        ws.cell(next_row, col_map['Categoría']).value = trans['categoria']
        ws.cell(next_row, col_map['Entidad']).value = trans['entidad']
        ws.cell(next_row, col_map['Cuenta Bancaria']).value = trans['cuenta']
        ws.cell(next_row, col_map['Cliente/Proveedor']).value = trans['cliente_proveedor']
        ws.cell(next_row, col_map['Concepto']).value = trans['concepto']
        ws.cell(next_row, col_map['Referencia']).value = trans['referencia']

        if trans['monto_usd']:
            ws.cell(next_row, col_map['Monto USD']).value = trans['monto_usd']
        if trans['monto_crc']:
            ws.cell(next_row, col_map['Monto CRC']).value = trans['monto_crc']

        ws.cell(next_row, col_map['Ingreso/Egreso']).value = trans['ingreso_egreso']
        ws.cell(next_row, col_map['Estado']).value = trans['estado']
        ws.cell(next_row, col_map['Prioridad']).value = trans['prioridad']

        if trans['vencimiento']:
            ws.cell(next_row, col_map['Vencimiento']).value = trans['vencimiento']

        ws.cell(next_row, col_map['Notas']).value = trans['notas']

        filas_agregadas.append({
            'num': trans['num'],
            'fila': next_row,
            'descripcion': trans['descripcion'],
            'tipo': trans['tipo'],
            'monto_usd': trans['monto_usd'],
            'monto_crc': trans['monto_crc'],
        })

        print(f"✅ Transacción {trans['num']}: {trans['descripcion']}")
        print(f"   Fila: {next_row}")
        print(f"   Tipo: {trans['tipo']}")
        if trans['monto_usd']:
            signo = '+' if trans['monto_usd'] > 0 else ''
            print(f"   Monto: {signo}${trans['monto_usd']:,.2f} USD")
        if trans['monto_crc']:
            signo = '+' if trans['monto_crc'] > 0 else ''
            print(f"   Monto: {signo}₡{trans['monto_crc']:,.2f} CRC")
        print()

    # Guardar
    print("💾 Guardando todos los cambios...")
    wb.save(EXCEL_FILE)
    print("✅ Excel actualizado exitosamente")
    print()

    # Resumen final
    print("=" * 80)
    print("📊 RESUMEN DE SINCRONIZACIÓN")
    print("=" * 80)
    print()

    print("✅ CORRECCIONES:")
    print(f"   • Fila 206: Cliente/Proveedor corregido")
    print()

    print("✅ TRANSACCIONES AGREGADAS:")
    for t in filas_agregadas:
        print(f"   • Fila {t['fila']}: {t['descripcion']}")
    print()

    # Balance del día
    print("💰 BALANCE DEL DÍA (10/11/2025):")
    total_usd = sum(t['monto_usd'] for t in filas_agregadas if t['monto_usd'])
    total_crc = sum(t['monto_crc'] for t in filas_agregadas if t['monto_crc'])

    print(f"   Ingresos USD:  +$635.63")
    print(f"   Egresos USD:   -$3,137.26")
    print(f"   ───────────────────────")
    print(f"   Balance USD:   ${total_usd:,.2f}")
    print()
    print(f"   Egresos CRC:   ₡{total_crc:,.2f}")
    print()

    # Estado final
    print("📈 ESTADO FINAL DEL EXCEL:")
    print(f"   Total filas: {ws.max_row}")
    print(f"   Transacciones: {ws.max_row - 1}")
    print()

    print("=" * 80)
    print("✅ SINCRONIZACIÓN COMPLETADA")
    print("=" * 80)
    print()

    print("📋 VERIFICACIÓN:")
    print("   1. Abre el Excel")
    print("   2. Ve a fila 206 y verifica columna F (Cliente/Proveedor)")
    print(f"   3. Ve a filas {filas_agregadas[0]['fila']}-{filas_agregadas[-1]['fila']} y verifica las 3 nuevas transacciones")
    print("   4. Todo debería estar sincronizado correctamente")
    print()

    return True

if __name__ == "__main__":
    try:
        # Backup
        if not crear_backup():
            print("❌ Abortando: No se pudo crear backup")
            exit(1)

        # Sincronizar
        if sincronizar():
            print("🎉 Sincronización exitosa!")
            print()
            print("💡 Si algo salió mal, puedes restaurar desde:")
            print(f"   {BACKUP_FILE}")
        else:
            print("❌ Error en sincronización")

    except FileNotFoundError:
        print(f"❌ ERROR: No se encontró el archivo {EXCEL_FILE}")
    except Exception as e:
        print(f"❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
