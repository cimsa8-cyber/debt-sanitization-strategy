#!/usr/bin/env python3
"""
CORRECTOR DE EXCEL - Agrega fórmulas faltantes
Corrige hojas A_R, A_P, Tarjetas_Credito, Efectivo
Versión: 1.0
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from datetime import datetime

COLOR_HEADER = "1F4E78"

def corregir_excel(archivo="AlvaroVelasco_Finanzas_v1.0.xlsx"):
    print("="*70)
    print("CORRECTOR DE EXCEL - AGREGANDO FÓRMULAS FALTANTES")
    print("="*70)
    print(f"Archivo: {archivo}")
    print()

    # Abrir workbook existente
    print("⏳ Abriendo archivo Excel...")
    wb = openpyxl.load_workbook(archivo)
    print("✅ Archivo abierto")
    print()

    # CORRECCIÓN 1: Hoja A_R (Cuentas por Cobrar)
    print("⏳ Corrigiendo hoja A_R (Cuentas por Cobrar)...")
    ws_ar = wb["A_R"]

    # Agregar headers si no existen
    if ws_ar["A2"].value != "Cliente":
        ws_ar["A2"] = "Cliente"
        ws_ar["B2"] = "Factura/Ref"
        ws_ar["C2"] = "Fecha"
        ws_ar["D2"] = "Monto USD"
        ws_ar["E2"] = "Estado"
        ws_ar["F2"] = "Prioridad"
        ws_ar["G2"] = "Días Mora"

        # Formatear headers
        for col in range(1, 8):
            cell = ws_ar.cell(2, col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")

    # Leer TRANSACCIONES para encontrar facturas clientes
    ws_trans = wb["TRANSACCIONES"]
    row_ar = 3

    for row in range(2, ws_trans.max_row + 1):
        tipo = ws_trans[f"B{row}"].value
        if tipo == "Factura Cliente":
            # Cliente
            ws_ar[f"A{row_ar}"] = f"=TRANSACCIONES!F{row}"
            # Referencia
            ws_ar[f"B{row_ar}"] = f"=TRANSACCIONES!H{row}"
            # Fecha
            ws_ar[f"C{row_ar}"] = f"=TRANSACCIONES!A{row}"
            # Monto
            ws_ar[f"D{row_ar}"] = f"=TRANSACCIONES!I{row}"
            ws_ar[f"D{row_ar}"].number_format = '"$"#,##0.00'
            # Estado
            ws_ar[f"E{row_ar}"] = f"=TRANSACCIONES!L{row}"
            # Prioridad
            ws_ar[f"F{row_ar}"] = f"=TRANSACCIONES!M{row}"
            # Días mora (solo si está pendiente)
            ws_ar[f"G{row_ar}"] = f'=IF(E{row_ar}="Pendiente",TODAY()-C{row_ar},"")'

            row_ar += 1

    print(f"✅ A_R corregida: {row_ar - 3} clientes agregados")
    print()

    # CORRECCIÓN 2: Hoja A_P (Cuentas por Pagar)
    print("⏳ Corrigiendo hoja A_P (Cuentas por Pagar)...")
    ws_ap = wb["A_P"]

    # Agregar headers
    ws_ap["A2"] = "Proveedor"
    ws_ap["B2"] = "Factura/Ref"
    ws_ap["C2"] = "Fecha"
    ws_ap["D2"] = "Monto USD"
    ws_ap["E2"] = "Vencimiento"
    ws_ap["F2"] = "Estado"
    ws_ap["G2"] = "Días Mora"

    # Formatear headers
    for col in range(1, 8):
        cell = ws_ap.cell(2, col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")

    # Agregar facturas proveedores
    row_ap = 3
    for row in range(2, ws_trans.max_row + 1):
        tipo = ws_trans[f"B{row}"].value
        if tipo == "Factura Proveedor":
            ws_ap[f"A{row_ap}"] = f"=TRANSACCIONES!F{row}"
            ws_ap[f"B{row_ap}"] = f"=TRANSACCIONES!H{row}"
            ws_ap[f"C{row_ap}"] = f"=TRANSACCIONES!A{row}"
            ws_ap[f"D{row_ap}"] = f"=TRANSACCIONES!I{row}"
            ws_ap[f"D{row_ap}"].number_format = '"$"#,##0.00'
            ws_ap[f"E{row_ap}"] = f"=TRANSACCIONES!N{row}"
            ws_ap[f"F{row_ap}"] = f"=TRANSACCIONES!L{row}"
            ws_ap[f"G{row_ap}"] = f'=IF(F{row_ap}="Pendiente",IF(E{row_ap}<>"",TODAY()-E{row_ap},""),"")'
            row_ap += 1

    print(f"✅ A_P corregida: {row_ap - 3} proveedores agregados")
    print()

    # CORRECCIÓN 3: Hoja Tarjetas_Credito
    print("⏳ Corrigiendo hoja Tarjetas_Credito...")
    ws_tc = wb["Tarjetas_Credito"]

    # Agregar headers
    ws_tc["A2"] = "Banco"
    ws_tc["B2"] = "Número"
    ws_tc["C2"] = "Saldo USD"
    ws_tc["D2"] = "Estado"
    ws_tc["E2"] = "Fecha Venc."

    # Formatear headers
    for col in range(1, 6):
        cell = ws_tc.cell(2, col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")

    # Agregar tarjetas desde TRANSACCIONES (Apertura Inicial + Categoría TC)
    row_tc = 3
    tarjetas_agregadas = set()

    for row in range(2, ws_trans.max_row + 1):
        tipo = ws_trans[f"B{row}"].value
        categoria = ws_trans[f"C{row}"].value
        cuenta = ws_trans[f"E{row}"].value

        if tipo == "Apertura Inicial" and categoria == "Tarjeta Crédito":
            # Extraer info de la cuenta
            if cuenta and "TC " in str(cuenta):
                if cuenta not in tarjetas_agregadas:
                    # Banco (extraer de cuenta)
                    if "BNCR" in str(cuenta):
                        ws_tc[f"A{row_tc}"] = "BNCR"
                    elif "BAC" in str(cuenta):
                        ws_tc[f"A{row_tc}"] = "BAC"

                    # Número (extraer de cuenta)
                    ws_tc[f"B{row_tc}"] = f"=TRANSACCIONES!E{row}"

                    # Saldo
                    ws_tc[f"C{row_tc}"] = f"=TRANSACCIONES!I{row}"
                    ws_tc[f"C{row_tc}"].number_format = '"$"#,##0.00'

                    # Estado
                    ws_tc[f"D{row_tc}"] = f"=TRANSACCIONES!L{row}"

                    # Fecha vencimiento
                    ws_tc[f"E{row_tc}"] = f"=TRANSACCIONES!N{row}"

                    tarjetas_agregadas.add(cuenta)
                    row_tc += 1

    print(f"✅ Tarjetas_Credito corregida: {row_tc - 3} tarjetas agregadas")
    print()

    # CORRECCIÓN 4: Hoja Efectivo (agregar todas las transacciones)
    print("⏳ Corrigiendo hoja Efectivo...")
    ws_efec = wb["Efectivo"]

    # Copiar todas las transacciones de efectivo/ahorro
    row_efec = 3
    balance = 0

    for row in range(2, ws_trans.max_row + 1):
        categoria = ws_trans[f"C{row}"].value
        estado = ws_trans[f"L{row}"].value

        if categoria in ["Efectivo", "Ahorro"] and estado == "Cobrado":
            ws_efec[f"A{row_efec}"] = f"=TRANSACCIONES!A{row}"
            ws_efec[f"B{row_efec}"] = f"=TRANSACCIONES!G{row}"
            ws_efec[f"C{row_efec}"] = f"=TRANSACCIONES!E{row}"
            ws_efec[f"D{row_efec}"] = f'=IF(TRANSACCIONES!K{row}="Ingreso",TRANSACCIONES!I{row},"")'
            ws_efec[f"E{row_efec}"] = f'=IF(TRANSACCIONES!K{row}="Egreso",TRANSACCIONES!I{row},"")'

            # Balance running
            if row_efec == 3:
                ws_efec[f"F{row_efec}"] = f"=D{row_efec}-E{row_efec}"
            else:
                ws_efec[f"F{row_efec}"] = f"=F{row_efec-1}+D{row_efec}-E{row_efec}"

            ws_efec[f"F{row_efec}"].number_format = '"$"#,##0.00'
            row_efec += 1

    print(f"✅ Efectivo corregida: {row_efec - 3} movimientos agregados")
    print()

    # Guardar archivo corregido
    print("⏳ Guardando archivo corregido...")
    output_file = "AlvaroVelasco_Finanzas_v1.0_CORREGIDO.xlsx"
    wb.save(output_file)
    print(f"✅ Archivo guardado: {output_file}")
    print()

    print("="*70)
    print("🎉 CORRECCIÓN COMPLETADA EXITOSAMENTE")
    print("="*70)
    print()
    print(f"📂 Archivo original: AlvaroVelasco_Finanzas_v1.0.xlsx (sin cambios)")
    print(f"📂 Archivo corregido: {output_file}")
    print()
    print("CORRECCIONES APLICADAS:")
    print(f"  ✅ Hoja A_R: {row_ar - 3} clientes con saldos pendientes")
    print(f"  ✅ Hoja A_P: {row_ap - 3} facturas proveedores")
    print(f"  ✅ Hoja Tarjetas_Credito: {row_tc - 3} tarjetas de crédito")
    print(f"  ✅ Hoja Efectivo: {row_efec - 3} movimientos con balance running")
    print()
    print("PRÓXIMO PASO:")
    print(f"  Abrir: {output_file}")
    print("  Verificar hoja A_R muestra todos tus clientes")
    print()

    return True

if __name__ == "__main__":
    try:
        corregir_excel()
    except Exception as e:
        print(f"❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
