#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ANÁLISIS DE MOVIMIENTOS UBER
Busca todos los movimientos de Uber/transporte para identificar faltantes
"""
import openpyxl
from datetime import datetime, timedelta

EXCEL_FILE = "AlvaroVelasco_Finanzas_v2.0.xlsx"

def analizar_uber():
    print("=" * 80)
    print("ANÁLISIS DE MOVIMIENTOS UBER/TRANSPORTE")
    print("=" * 80)
    print()

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb['TRANSACCIONES']

    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    col_map = {}
    for col in range(1, len(headers) + 1):
        if headers[col-1]:
            col_map[headers[col-1]] = col

    print("🔍 Buscando todos los movimientos de Uber/Transporte...")
    print()

    # Buscar movimientos Uber
    uber_movimientos = []

    for row in range(2, ws.max_row + 1):
        concepto = ws.cell(row, col_map['Concepto']).value
        categoria = ws.cell(row, col_map['Categoría']).value
        cliente_prov = ws.cell(row, col_map['Cliente/Proveedor']).value
        fecha = ws.cell(row, col_map['Fecha']).value
        monto_usd = ws.cell(row, col_map['Monto USD']).value
        monto_crc = ws.cell(row, col_map['Monto CRC']).value
        tipo = ws.cell(row, col_map['Tipo Transacción']).value
        cuenta = ws.cell(row, col_map['Cuenta Bancaria']).value

        # Buscar por Uber o Transporte
        es_uber = False
        texto_buscar = []

        if concepto:
            texto_buscar.append(str(concepto).upper())
        if categoria:
            texto_buscar.append(str(categoria).upper())
        if cliente_prov:
            texto_buscar.append(str(cliente_prov).upper())

        texto_completo = " ".join(texto_buscar)

        if 'UBER' in texto_completo or 'TRANSPORTE' in texto_completo:
            es_uber = True

        if es_uber:
            fecha_obj = None
            if fecha:
                if isinstance(fecha, datetime):
                    fecha_obj = fecha
                else:
                    try:
                        fecha_obj = datetime.strptime(str(fecha), '%Y-%m-%d %H:%M:%S')
                    except:
                        pass

            uber_movimientos.append({
                'fila': row,
                'fecha': fecha_obj,
                'tipo': tipo,
                'categoria': categoria,
                'cuenta': cuenta,
                'cliente_prov': cliente_prov,
                'concepto': concepto,
                'monto_usd': float(monto_usd) if monto_usd else None,
                'monto_crc': float(monto_crc) if monto_crc else None,
            })

    if not uber_movimientos:
        print("⚠️  No se encontraron movimientos de Uber/Transporte")
        print()
        return

    # Ordenar por fecha
    uber_movimientos.sort(key=lambda x: x['fecha'] if x['fecha'] else datetime.min)

    print(f"📋 MOVIMIENTOS ENCONTRADOS: {len(uber_movimientos)}")
    print()

    # Mostrar todos
    for i, mov in enumerate(uber_movimientos, 1):
        fecha_str = mov['fecha'].strftime('%d/%m/%Y %H:%M') if mov['fecha'] else 'Sin fecha'

        print(f"{i}. Fila {mov['fila']} - {fecha_str}")
        print(f"   Tipo: {mov['tipo']}")
        print(f"   Categoría: {mov['categoria']}")
        print(f"   Cliente/Proveedor: {mov['cliente_prov']}")
        print(f"   Cuenta: {mov['cuenta']}")
        print(f"   Concepto: {mov['concepto'][:60] if mov['concepto'] else 'N/A'}...")

        if mov['monto_usd']:
            print(f"   Monto USD: ${abs(mov['monto_usd']):,.2f}")
        if mov['monto_crc']:
            print(f"   Monto CRC: ₡{abs(mov['monto_crc']):,.2f}")
        print()

    # Análisis por fecha
    print("=" * 80)
    print("📊 ANÁLISIS POR FECHAS RECIENTES")
    print("=" * 80)
    print()

    # Últimos 7 días
    hoy = datetime.now()
    hace_7_dias = hoy - timedelta(days=7)

    recientes = [m for m in uber_movimientos if m['fecha'] and m['fecha'] >= hace_7_dias]

    if recientes:
        print(f"🕐 ÚLTIMOS 7 DÍAS ({len(recientes)} movimientos):")
        print()

        for mov in recientes:
            fecha_str = mov['fecha'].strftime('%d/%m/%Y')
            monto = mov['monto_usd'] if mov['monto_usd'] else mov['monto_crc']
            moneda = 'USD' if mov['monto_usd'] else 'CRC'

            print(f"   • {fecha_str}: ${abs(monto):,.2f} {moneda} - {mov['concepto'][:50] if mov['concepto'] else 'Sin concepto'}")
    else:
        print("⚠️  No hay movimientos Uber en los últimos 7 días")

    print()

    # Verificación de ayer
    ayer = hoy - timedelta(days=1)
    movimientos_ayer = [m for m in uber_movimientos
                        if m['fecha'] and m['fecha'].date() == ayer.date()]

    print("=" * 80)
    print(f"🔍 VERIFICACIÓN DÍA {ayer.strftime('%d/%m/%Y')}")
    print("=" * 80)
    print()

    if movimientos_ayer:
        print(f"📋 Movimientos encontrados: {len(movimientos_ayer)}")
        print()
        for mov in movimientos_ayer:
            print(f"   Fila {mov['fila']}: {mov['concepto'][:50] if mov['concepto'] else 'Sin concepto'}")
            monto = mov['monto_usd'] if mov['monto_usd'] else mov['monto_crc']
            moneda = 'USD' if mov['monto_usd'] else 'CRC'
            print(f"   Monto: ${abs(monto):,.2f} {moneda}")
            print()
    else:
        print(f"⚠️  No se encontraron movimientos Uber del {ayer.strftime('%d/%m/%Y')}")
        print()

    # Resumen
    print("=" * 80)
    print("📊 RESUMEN")
    print("=" * 80)
    print()

    total_movimientos = len(uber_movimientos)
    print(f"Total de movimientos Uber/Transporte registrados: {total_movimientos}")
    print(f"Movimientos últimos 7 días: {len(recientes)}")
    print(f"Movimientos de ayer ({ayer.strftime('%d/%m/%Y')}): {len(movimientos_ayer)}")
    print()

    if len(movimientos_ayer) == 1:
        print("⚠️  DISCREPANCIA DETECTADA:")
        print(f"   Usuario reportó: 3 movimientos ayer")
        print(f"   Sistema encontró: {len(movimientos_ayer)} movimiento")
        print(f"   Faltantes: 2 movimientos")
        print()
        print("📋 ACCIÓN REQUERIDA:")
        print("   1. Verificar datos proporcionados por usuario")
        print("   2. Buscar en otros registros o extractos")
        print("   3. Registrar los 2 movimientos faltantes")
    elif len(movimientos_ayer) == 3:
        print("✅ Sistema correcto: 3 movimientos registrados")
    else:
        print(f"⚠️  Situación inusual: {len(movimientos_ayer)} movimientos encontrados")

    print()
    print("=" * 80)
    print("✅ ANÁLISIS COMPLETADO")
    print("=" * 80)
    print()

if __name__ == "__main__":
    try:
        analizar_uber()
        print("🎉 Análisis completado!")
    except Exception as e:
        print(f"❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
