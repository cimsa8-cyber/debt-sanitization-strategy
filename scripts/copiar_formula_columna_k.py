#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
COPIAR FÓRMULA COLUMNA K A FILAS FALTANTES
Las filas 211-221 no tienen fórmula en columna K (Ingreso/Egreso)
"""
import openpyxl
from datetime import datetime
import shutil

EXCEL_FILE = "AlvaroVelasco_Finanzas_v2.0.xlsx"
BACKUP_FILE = f"AlvaroVelasco_Finanzas_v2.0_COPIAR_FORMULA_K_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

def crear_backup():
    print("=" * 80)
    print("CREANDO BACKUP")
    print("=" * 80)
    print(f"Backup: {BACKUP_FILE}")
    try:
        shutil.copy2(EXCEL_FILE, BACKUP_FILE)
        print("✅ Backup creado")
        print()
        return True
    except Exception as e:
        print(f"❌ ERROR: {e}")
        return False

def copiar_formula():
    print("=" * 80)
    print("COPIANDO FÓRMULA COLUMNA K (Ingreso/Egreso)")
    print("=" * 80)
    print()

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb['TRANSACCIONES']

    # Encontrar columna K
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]

    try:
        col_k = headers.index('Ingreso/Egreso') + 1
        col_b = headers.index('Tipo Transacción') + 1
    except ValueError as e:
        print(f"❌ ERROR: No se encontró columna necesaria: {e}")
        return False

    print(f"✅ Columna K (Ingreso/Egreso) = columna {col_k}")
    print(f"✅ Columna B (Tipo Transacción) = columna {col_b}")
    print()

    # Obtener fórmula de K2 (fila con Balance Inicial)
    formula_k2 = ws.cell(2, col_k).value

    print(f"📋 Fórmula original en K2:")
    print(f"   {formula_k2}")
    print()

    if not formula_k2 or not isinstance(formula_k2, str) or not formula_k2.startswith('='):
        print("❌ ERROR: K2 no tiene una fórmula válida")
        return False

    # =========================================================================
    # PASO 1: IDENTIFICAR FILAS SIN VALOR EN COLUMNA K
    # =========================================================================
    print("=" * 80)
    print("📋 PASO 1: Identificando filas sin valor en columna K...")
    print("=" * 80)
    print()

    filas_sin_valor = []

    for row in range(2, ws.max_row + 1):
        celda_k = ws.cell(row, col_k)

        if not celda_k.value:
            # Verificar si la fila tiene datos (no está vacía)
            tipo = ws.cell(row, col_b).value

            if tipo:  # Si tiene tipo de transacción, debería tener Ingreso/Egreso
                filas_sin_valor.append(row)

    print(f"⚠️  Filas sin valor en columna K: {len(filas_sin_valor)}")
    print()

    if len(filas_sin_valor) > 20:
        print(f"📋 Primeras 10: {filas_sin_valor[:10]}")
        print(f"📋 Últimas 10: {filas_sin_valor[-10:]}")
    else:
        print(f"📋 Filas: {filas_sin_valor}")

    print()

    # =========================================================================
    # PASO 2: COPIAR FÓRMULA A FILAS FALTANTES
    # =========================================================================
    print("=" * 80)
    print("✏️  PASO 2: Copiando fórmula a filas faltantes...")
    print("=" * 80)
    print()

    for row in filas_sin_valor:
        # Adaptar la fórmula para esta fila
        # Cambiar B2 por B{row}
        formula_adaptada = formula_k2.replace('B2', f'B{row}')

        ws.cell(row, col_k).value = formula_adaptada

        tipo = ws.cell(row, col_b).value
        concepto = ws.cell(row, headers.index('Concepto') + 1).value if 'Concepto' in headers else None

        print(f"✅ Fila {row}: {concepto[:40] if concepto else 'N/A'}...")
        print(f"   Tipo: {tipo}")
        print(f"   Fórmula: {formula_adaptada}")
        print()

    print(f"📊 Total filas actualizadas: {len(filas_sin_valor)}")
    print()

    # =========================================================================
    # PASO 3: GUARDAR
    # =========================================================================
    print("=" * 80)
    print("💾 PASO 3: Guardando cambios...")
    print("=" * 80)
    print()

    wb.save(EXCEL_FILE)
    print("✅ Excel actualizado")
    print()

    # =========================================================================
    # PASO 4: VERIFICAR RESULTADO
    # =========================================================================
    print("=" * 80)
    print("📊 PASO 4: Verificación...")
    print("=" * 80)
    print()

    # Recargar con data_only para ver valores calculados
    wb_verif = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws_verif = wb_verif['TRANSACCIONES']

    print("🔍 Verificando filas que se actualizaron:")
    print()

    for row in filas_sin_valor[:5]:  # Solo las primeras 5 para no saturar
        celda_k = ws_verif.cell(row, col_k)
        tipo = ws_verif.cell(row, col_b).value
        concepto = ws_verif.cell(row, headers.index('Concepto') + 1).value if 'Concepto' in headers else None

        print(f"Fila {row}: {concepto[:40] if concepto else 'N/A'}...")
        print(f"   Tipo: {tipo}")
        print(f"   Ingreso/Egreso calculado: {celda_k.value}")
        print()

    # =========================================================================
    # RESUMEN
    # =========================================================================
    print("=" * 80)
    print("📊 RESUMEN")
    print("=" * 80)
    print()

    print(f"✅ Fórmula copiada a {len(filas_sin_valor)} filas")
    print()
    print("🔧 PRÓXIMOS PASOS:")
    print("   1. Cierra Excel (si está abierto)")
    print("   2. Vuelve a abrirlo")
    print("   3. Excel recalculará las fórmulas automáticamente")
    print("   4. Ve a hoja Efectivo, fila 3")
    print("   5. Verifica que el Balance ahora sea correcto")
    print()
    print("💡 Balance esperado:")
    print("   Ingreso: ~$11,951.71")
    print("   Egreso: ~$10,170.96")
    print("   Balance: ~$1,780.75")
    print()
    print("   (Nota: Este es el balance calculado antes, puede variar")
    print("    ligeramente dependiendo del saldo inicial correcto)")

    print()
    print("=" * 80)
    print("✅ PROCESO COMPLETADO")
    print("=" * 80)
    print()

    return True

if __name__ == "__main__":
    try:
        if not crear_backup():
            print("❌ Abortando")
            exit(1)

        if copiar_formula():
            print("🎉 Fórmulas copiadas exitosamente!")
            print()
            print("👉 Ahora cierra y vuelve a abrir el Excel para ver los cambios")
        else:
            print("❌ Proceso falló")

    except Exception as e:
        print(f"❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
