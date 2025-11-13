#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CREACIÓN DE HOJA CUENTAS_ALIAS
Crea una hoja de mapeo de alias para normalización de nombres de cuentas
"""
import openpyxl
from datetime import datetime
import shutil

EXCEL_FILE = "AlvaroVelasco_Finanzas_v2.0.xlsx"
BACKUP_FILE = f"AlvaroVelasco_Finanzas_v2.0_CREAR_ALIAS_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

def crear_backup():
    print("=" * 80)
    print("CREANDO BACKUP")
    print("=" * 80)
    print(f"Backup: {BACKUP_FILE}")
    try:
        shutil.copy2(EXCEL_FILE, BACKUP_FILE)
        print("✅ Backup creado")
        print()
        return True
    except Exception as e:
        print(f"❌ ERROR: {e}")
        return False

def crear_hoja_alias():
    print("=" * 80)
    print("CREACIÓN HOJA CUENTAS_ALIAS")
    print("=" * 80)
    print()

    wb = openpyxl.load_workbook(EXCEL_FILE)

    # Verificar si ya existe
    if 'CUENTAS_ALIAS' in wb.sheetnames:
        print("⚠️  La hoja 'CUENTAS_ALIAS' ya existe")
        respuesta = input("¿Deseas recrearla? (Se perderán los datos actuales) [s/N]: ")
        if respuesta.lower() != 's':
            print("❌ Operación cancelada")
            return False

        # Eliminar hoja existente
        del wb['CUENTAS_ALIAS']
        print("🗑️  Hoja existente eliminada")
        print()

    # Crear nueva hoja
    ws = wb.create_sheet('CUENTAS_ALIAS', 0)  # Posición 0 = primera hoja

    print("📋 Creando estructura de la hoja...")
    print()

    # =========================================================================
    # ENCABEZADOS
    # =========================================================================
    encabezados = [
        'Cuenta Estándar',
        'Alias 1',
        'Alias 2',
        'Alias 3',
        'Alias 4',
        'Alias 5',
        'Notas'
    ]

    for col, header in enumerate(encabezados, 1):
        celda = ws.cell(1, col)
        celda.value = header
        celda.font = openpyxl.styles.Font(bold=True, size=11)
        celda.fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        celda.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        celda.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    # =========================================================================
    # DATOS INICIALES (basados en las cuentas detectadas)
    # =========================================================================
    cuentas_iniciales = [
        {
            'estandar': 'Promerica USD (40000003881774)',
            'alias1': 'Promerica USD',
            'alias2': 'Promerica USD 1774',
            'alias3': 'Promerica USD (*1774)',
            'alias4': 'Promerica 1774',
            'alias5': '',
            'notas': 'Cuenta corriente dólares Promerica'
        },
        {
            'estandar': 'Promerica CRC (10000003881708)',
            'alias1': 'Promerica CRC',
            'alias2': 'Promerica Colones',
            'alias3': 'Promerica CRC 1708',
            'alias4': '',
            'alias5': '',
            'notas': 'Cuenta corriente colones Promerica'
        },
        {
            'estandar': 'BNCR USD (100-01-000-123456-7)',
            'alias1': 'BNCR USD',
            'alias2': 'Banco Nacional USD',
            'alias3': 'BNCR Dólares',
            'alias4': '',
            'alias5': '',
            'notas': 'Cuenta dólares BNCR'
        },
        {
            'estandar': 'BNCR CRC (100-01-000-654321-8)',
            'alias1': 'BNCR CRC',
            'alias2': 'Banco Nacional CRC',
            'alias3': 'BNCR Colones',
            'alias4': '',
            'alias5': '',
            'notas': 'Cuenta colones BNCR'
        },
        {
            'estandar': 'Tarjeta BNCR Visa 3519',
            'alias1': 'BNCR 3519',
            'alias2': 'Tarjeta BNCR',
            'alias3': 'BNCR Visa *3519',
            'alias4': 'Visa 3519',
            'alias5': 'BNCR ****3519',
            'notas': 'Tarjeta de crédito BNCR Visa terminación 3519'
        },
        {
            'estandar': 'Efectivo',
            'alias1': 'Caja',
            'alias2': 'Efectivo CRC',
            'alias3': 'Efectivo USD',
            'alias4': 'Cash',
            'alias5': '',
            'notas': 'Efectivo físico'
        },
        {
            'estandar': 'Cuentas por Cobrar',
            'alias1': 'CxC',
            'alias2': 'Por Cobrar',
            'alias3': 'Clientes',
            'alias4': '',
            'alias5': '',
            'notas': 'Cuentas por cobrar a clientes'
        },
        {
            'estandar': 'Pasivos',
            'alias1': 'Cuentas por Pagar',
            'alias2': 'CxP',
            'alias3': 'Por Pagar',
            'alias4': 'Proveedores',
            'alias5': '',
            'notas': 'Pasivos y cuentas por pagar'
        },
    ]

    # Agregar datos
    for row_idx, cuenta in enumerate(cuentas_iniciales, 2):
        ws.cell(row_idx, 1).value = cuenta['estandar']
        ws.cell(row_idx, 2).value = cuenta['alias1']
        ws.cell(row_idx, 3).value = cuenta['alias2']
        ws.cell(row_idx, 4).value = cuenta['alias3']
        ws.cell(row_idx, 5).value = cuenta['alias4']
        ws.cell(row_idx, 6).value = cuenta['alias5']
        ws.cell(row_idx, 7).value = cuenta['notas']

    # =========================================================================
    # FORMATO
    # =========================================================================
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 40

    # Congelar primera fila
    ws.freeze_panes = 'A2'

    # Agregar bordes
    thin_border = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style='thin'),
        right=openpyxl.styles.Side(style='thin'),
        top=openpyxl.styles.Side(style='thin'),
        bottom=openpyxl.styles.Side(style='thin')
    )

    for row in range(1, len(cuentas_iniciales) + 2):
        for col in range(1, 8):
            ws.cell(row, col).border = thin_border

    # =========================================================================
    # INSTRUCCIONES (en fila inferior)
    # =========================================================================
    instrucciones_row = len(cuentas_iniciales) + 4

    ws.cell(instrucciones_row, 1).value = "INSTRUCCIONES:"
    ws.cell(instrucciones_row, 1).font = openpyxl.styles.Font(bold=True, size=11)

    instrucciones = [
        "• Cuenta Estándar: Nombre oficial único de la cuenta (usado en reportes)",
        "• Alias 1-5: Variaciones del nombre que pueden aparecer en extractos, XML, PDFs",
        "• Para agregar nueva cuenta: Agrega fila con nombre estándar y sus alias conocidos",
        "• Para agregar nuevo alias: Agrégalo en cualquier columna Alias disponible",
        "• El script de normalización usará esta tabla para unificar nombres automáticamente",
        "• Deja celdas de alias vacías si no hay más variaciones",
        "• Notas: Descripción opcional de la cuenta"
    ]

    for idx, instruccion in enumerate(instrucciones):
        ws.cell(instrucciones_row + 1 + idx, 1).value = instruccion
        ws.merge_cells(f'A{instrucciones_row + 1 + idx}:G{instrucciones_row + 1 + idx}')
        ws.cell(instrucciones_row + 1 + idx, 1).alignment = openpyxl.styles.Alignment(wrap_text=True)

    print("✅ Estructura creada")
    print(f"✅ {len(cuentas_iniciales)} cuentas iniciales agregadas")
    print()

    # =========================================================================
    # GUARDAR
    # =========================================================================
    print("💾 Guardando cambios...")
    wb.save(EXCEL_FILE)
    print("✅ Excel actualizado")
    print()

    # =========================================================================
    # RESUMEN
    # =========================================================================
    print("=" * 80)
    print("📊 RESUMEN")
    print("=" * 80)
    print()

    print("✅ Hoja 'CUENTAS_ALIAS' creada exitosamente")
    print()
    print("📋 Cuentas configuradas:")
    for cuenta in cuentas_iniciales:
        alias_count = sum(1 for a in [cuenta['alias1'], cuenta['alias2'], cuenta['alias3'],
                                       cuenta['alias4'], cuenta['alias5']] if a)
        print(f"   • {cuenta['estandar']}")
        print(f"     Alias: {alias_count} variaciones")

    print()
    print("🔧 PRÓXIMOS PASOS:")
    print("   1. Abre el Excel y ve a la hoja 'CUENTAS_ALIAS'")
    print("   2. Revisa las cuentas y alias configurados")
    print("   3. Agrega más cuentas o alias según necesites")
    print("   4. Ejecuta el script de normalización universal")

    print()
    print("=" * 80)
    print("✅ HOJA CUENTAS_ALIAS CREADA")
    print("=" * 80)
    print()

    return True

if __name__ == "__main__":
    try:
        if not crear_backup():
            print("❌ Abortando")
            exit(1)

        if crear_hoja_alias():
            print("🎉 Proceso completado exitosamente!")
        else:
            print("❌ Proceso cancelado o falló")

    except Exception as e:
        print(f"❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
