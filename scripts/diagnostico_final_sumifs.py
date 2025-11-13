#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
DIAGNÓSTICO FINAL - QUÉ SUMA SUMIFS AHORA
Muestra exactamente qué transacciones está sumando después de corregir signos
"""
import openpyxl
from datetime import datetime

EXCEL_FILE = "AlvaroVelasco_Finanzas_v2.0.xlsx"

def diagnosticar():
    print("=" * 80)
    print("DIAGNÓSTICO FINAL - TRANSACCIONES QUE SUMA SUMIFS")
    print("=" * 80)
    print()

    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb['TRANSACCIONES']

    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    col_map = {}
    for col in range(1, len(headers) + 1):
        if headers[col-1]:
            col_map[headers[col-1]] = col

    # Simular SUMIFS para Promerica
    cuenta_buscar = "Promerica USD (40000003881774)"

    print(f"🔍 Buscando transacciones de: {cuenta_buscar}")
    print()

    ingresos = []
    egresos = []

    for row in range(2, ws.max_row + 1):
        cuenta = ws.cell(row, col_map['Cuenta Bancaria']).value
        monto = ws.cell(row, col_map['Monto USD']).value
        ing_egr = ws.cell(row, col_map['Ingreso/Egreso']).value
        fecha = ws.cell(row, col_map['Fecha']).value
        concepto = ws.cell(row, col_map['Concepto']).value
        tipo = ws.cell(row, col_map['Tipo Transacción']).value

        if cuenta and str(cuenta).strip() == cuenta_buscar:
            if monto and ing_egr:
                monto_val = float(monto)

                trans_info = {
                    'fila': row,
                    'fecha': fecha.strftime('%d/%m/%Y') if isinstance(fecha, datetime) else 'Sin fecha',
                    'tipo': tipo,
                    'concepto': concepto[:50] if concepto else 'N/A',
                    'monto': monto_val
                }

                if ing_egr == 'Ingreso':
                    ingresos.append(trans_info)
                elif ing_egr == 'Egreso':
                    egresos.append(trans_info)

    # Mostrar INGRESOS
    print("=" * 80)
    print(f"📈 INGRESOS: {len(ingresos)} transacciones")
    print("=" * 80)
    print()

    for i, trans in enumerate(ingresos, 1):
        print(f"{i}. Fila {trans['fila']}: {trans['fecha']} - +${trans['monto']:,.2f}")
        print(f"   Tipo: {trans['tipo']}")
        print(f"   {trans['concepto']}")
        print()

    total_ingresos = sum(t['monto'] for t in ingresos)
    print(f"💰 TOTAL INGRESOS: ${total_ingresos:,.2f}")
    print(f"📊 Valor en Excel D3: $14,983")
    print(f"⚖️  Diferencia: ${abs(total_ingresos - 14983):,.2f}")
    print()

    # Buscar duplicados en ingresos
    print("🔍 Buscando duplicados en INGRESOS...")
    print()

    duplicados_ing = {}
    for trans in ingresos:
        key = f"{trans['fecha']}_{trans['monto']}"
        if key not in duplicados_ing:
            duplicados_ing[key] = []
        duplicados_ing[key].append(trans)

    hay_duplicados_ing = False
    for key, trans_list in duplicados_ing.items():
        if len(trans_list) > 1:
            hay_duplicados_ing = True
            print(f"⚠️  DUPLICADO: {trans_list[0]['fecha']} - ${trans_list[0]['monto']:,.2f}")
            print(f"   {trans_list[0]['concepto']}")
            print(f"   Aparece {len(trans_list)} veces:")
            for t in trans_list:
                print(f"      • Fila {t['fila']}")
            print()

    if not hay_duplicados_ing:
        print("✅ No hay duplicados en ingresos")
        print()

    # Mostrar EGRESOS
    print("=" * 80)
    print(f"📉 EGRESOS: {len(egresos)} transacciones")
    print("=" * 80)
    print()

    for i, trans in enumerate(egresos, 1):
        print(f"{i}. Fila {trans['fila']}: {trans['fecha']} - ${abs(trans['monto']):,.2f}")
        print(f"   Tipo: {trans['tipo']}")
        print(f"   {trans['concepto']}")
        print()

    # Total egresos (valor absoluto)
    total_egresos = sum(abs(t['monto']) for t in egresos)
    print(f"💰 TOTAL EGRESOS (valor absoluto): ${total_egresos:,.2f}")
    print(f"📊 Valor en Excel E3: $13,057")
    print(f"⚖️  Diferencia: ${abs(total_egresos - 13057):,.2f}")
    print()

    # Buscar duplicados en egresos
    print("🔍 Buscando duplicados en EGRESOS...")
    print()

    duplicados_egr = {}
    for trans in egresos:
        key = f"{trans['fecha']}_{abs(trans['monto'])}"
        if key not in duplicados_egr:
            duplicados_egr[key] = []
        duplicados_egr[key].append(trans)

    hay_duplicados_egr = False
    for key, trans_list in duplicados_egr.items():
        if len(trans_list) > 1:
            hay_duplicados_egr = True
            print(f"⚠️  DUPLICADO: {trans_list[0]['fecha']} - ${abs(trans_list[0]['monto']):,.2f}")
            print(f"   {trans_list[0]['concepto']}")
            print(f"   Aparece {len(trans_list)} veces:")
            for t in trans_list:
                print(f"      • Fila {t['fila']}")
            print()

    if not hay_duplicados_egr:
        print("✅ No hay duplicados en egresos")
        print()

    # BALANCE
    print("=" * 80)
    print("💰 BALANCE CALCULADO")
    print("=" * 80)
    print()

    balance_calc = total_ingresos - total_egresos

    print(f"   Ingresos: ${total_ingresos:,.2f}")
    print(f"   Egresos: ${total_egresos:,.2f}")
    print(f"   Balance: ${balance_calc:,.2f}")
    print()
    print(f"📊 Balance en Excel F3: $1,925.63")
    print(f"🏦 Balance extracto bancario: $2,163.44")
    print()
    print(f"⚖️  Diferencia Excel vs Calculado: ${abs(balance_calc - 1925.63):,.2f}")
    print(f"⚖️  Diferencia Calculado vs Extracto: ${abs(balance_calc - 2163.44):,.2f}")
    print()

    # RESUMEN
    print("=" * 80)
    print("📊 RESUMEN")
    print("=" * 80)
    print()

    if hay_duplicados_ing or hay_duplicados_egr:
        print("🚨 HAY TRANSACCIONES DUPLICADAS")
        print("   Necesitas eliminar los duplicados")
    else:
        print("✅ No hay duplicados")
        print()
        print("💡 ANÁLISIS:")
        print(f"   Diferencia vs extracto: ${abs(balance_calc - 2163.44):,.2f}")
        print()
        if abs(balance_calc - 2163.44) < 500:
            print("   Posible causa: Saldo inicial incorrecto")
            print(f"   Saldo inicial actual: $3,030.89")
            print(f"   Ajuste necesario: ${2163.44 - balance_calc:,.2f}")
        else:
            print("   Hay transacciones faltantes o incorrectas")

    print()
    print("=" * 80)
    print("✅ DIAGNÓSTICO COMPLETADO")
    print("=" * 80)

if __name__ == "__main__":
    try:
        diagnosticar()
    except Exception as e:
        print(f"❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
