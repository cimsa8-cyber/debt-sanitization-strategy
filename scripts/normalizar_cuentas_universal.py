#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
NORMALIZACIÓN UNIVERSAL DE CUENTAS
Lee la hoja CUENTAS_ALIAS y normaliza todos los nombres de cuenta en TRANSACCIONES
"""
import openpyxl
from datetime import datetime
import shutil

EXCEL_FILE = "AlvaroVelasco_Finanzas_v2.0.xlsx"
BACKUP_FILE = f"AlvaroVelasco_Finanzas_v2.0_NORMALIZAR_UNIVERSAL_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

def crear_backup():
    print("=" * 80)
    print("CREANDO BACKUP")
    print("=" * 80)
    print(f"Backup: {BACKUP_FILE}")
    try:
        shutil.copy2(EXCEL_FILE, BACKUP_FILE)
        print("✅ Backup creado")
        print()
        return True
    except Exception as e:
        print(f"❌ ERROR: {e}")
        return False

def cargar_mapeo_alias(wb):
    """Carga el mapeo de alias desde la hoja CUENTAS_ALIAS"""
    print("=" * 80)
    print("📋 PASO 1: Cargando mapeo de alias...")
    print("=" * 80)
    print()

    if 'CUENTAS_ALIAS' not in wb.sheetnames:
        print("❌ ERROR: No existe la hoja 'CUENTAS_ALIAS'")
        print("   Ejecuta primero: python scripts/crear_hoja_alias_cuentas.py")
        return None

    ws_alias = wb['CUENTAS_ALIAS']

    # Construir diccionario: {alias: cuenta_estandar}
    mapeo = {}

    for row in range(2, ws_alias.max_row + 1):
        cuenta_estandar = ws_alias.cell(row, 1).value

        if not cuenta_estandar:
            continue

        # El nombre estándar también mapea a sí mismo
        mapeo[cuenta_estandar.strip()] = cuenta_estandar.strip()

        # Leer alias (columnas 2-6)
        for col in range(2, 7):
            alias = ws_alias.cell(row, col).value

            if alias and str(alias).strip():
                alias_clean = str(alias).strip()
                mapeo[alias_clean] = cuenta_estandar.strip()

    print(f"✅ Mapeo cargado: {len(mapeo)} alias/nombres configurados")
    print()

    # Mostrar estadísticas
    cuentas_estandar = set(mapeo.values())
    print(f"📊 Cuentas estándar definidas: {len(cuentas_estandar)}")
    print()

    for cuenta in sorted(cuentas_estandar):
        alias_list = [k for k, v in mapeo.items() if v == cuenta and k != cuenta]
        print(f"   • {cuenta}")
        print(f"     Alias: {len(alias_list)} variaciones")
        if alias_list:
            print(f"     → {', '.join(f'\"{a}\"' for a in alias_list[:3])}" +
                  (" ..." if len(alias_list) > 3 else ""))

    print()

    return mapeo

def normalizar_cuentas(wb, mapeo):
    """Normaliza todas las cuentas en TRANSACCIONES usando el mapeo"""
    print("=" * 80)
    print("🔧 PASO 2: Normalizando cuentas en TRANSACCIONES...")
    print("=" * 80)
    print()

    ws_trans = wb['TRANSACCIONES']

    # Encontrar columna "Cuenta Bancaria"
    headers = [ws_trans.cell(1, col).value for col in range(1, ws_trans.max_column + 1)]

    try:
        col_cuenta = headers.index('Cuenta Bancaria') + 1
    except ValueError:
        print("❌ ERROR: No se encontró la columna 'Cuenta Bancaria'")
        return False

    # Analizar todas las cuentas
    cuentas_encontradas = {}
    transacciones_por_cuenta = {}

    for row in range(2, ws_trans.max_row + 1):
        cuenta = ws_trans.cell(row, col_cuenta).value

        if cuenta:
            cuenta_str = str(cuenta).strip()

            if cuenta_str not in cuentas_encontradas:
                cuentas_encontradas[cuenta_str] = []

            cuentas_encontradas[cuenta_str].append(row)

    print(f"📊 Cuentas únicas encontradas: {len(cuentas_encontradas)}")
    print()

    # Clasificar cuentas
    cuentas_a_normalizar = {}
    cuentas_sin_alias = []
    cuentas_ya_estandar = []

    for cuenta, filas in cuentas_encontradas.items():
        if cuenta in mapeo:
            nombre_estandar = mapeo[cuenta]

            if cuenta == nombre_estandar:
                # Ya es nombre estándar
                cuentas_ya_estandar.append((cuenta, len(filas)))
            else:
                # Necesita normalización
                if nombre_estandar not in cuentas_a_normalizar:
                    cuentas_a_normalizar[nombre_estandar] = []

                cuentas_a_normalizar[nombre_estandar].append({
                    'nombre_viejo': cuenta,
                    'filas': filas,
                    'cantidad': len(filas)
                })
        else:
            # No está en el mapeo
            cuentas_sin_alias.append((cuenta, len(filas)))

    # Mostrar resumen
    print("✅ Cuentas con nombre estándar correcto:")
    for cuenta, cantidad in sorted(cuentas_ya_estandar):
        print(f"   • {cuenta}: {cantidad} transacciones")
    print()

    if cuentas_a_normalizar:
        print("🔧 Cuentas que serán normalizadas:")
        for nombre_estandar, variaciones in sorted(cuentas_a_normalizar.items()):
            total = sum(v['cantidad'] for v in variaciones)
            print(f"   • {nombre_estandar}: {total} transacciones")
            for var in variaciones:
                print(f"     - \"{var['nombre_viejo']}\" ({var['cantidad']} trans.)")
        print()
    else:
        print("✅ No hay cuentas que normalizar (todas ya están correctas)")
        print()

    if cuentas_sin_alias:
        print("⚠️  Cuentas SIN MAPEO (no se cambiarán):")
        for cuenta, cantidad in sorted(cuentas_sin_alias):
            print(f"   • {cuenta}: {cantidad} transacciones")
        print()
        print("💡 Para normalizar estas cuentas:")
        print("   1. Abre el Excel, ve a hoja 'CUENTAS_ALIAS'")
        print("   2. Agrega estas cuentas con su nombre estándar y alias")
        print("   3. Vuelve a ejecutar este script")
        print()

    # Aplicar cambios
    if cuentas_a_normalizar:
        print("=" * 80)
        print("✏️  PASO 3: Aplicando cambios...")
        print("=" * 80)
        print()

        total_cambios = 0

        for nombre_estandar, variaciones in cuentas_a_normalizar.items():
            for var in variaciones:
                for fila in var['filas']:
                    ws_trans.cell(fila, col_cuenta).value = nombre_estandar
                    total_cambios += 1

                print(f"✅ \"{var['nombre_viejo']}\" → \"{nombre_estandar}\"")
                print(f"   {var['cantidad']} transacciones actualizadas")

        print()
        print(f"📊 Total transacciones actualizadas: {total_cambios}")
        print()

        return True
    else:
        print("✅ No se requieren cambios")
        return False

def verificar_resultado(wb):
    """Verifica el resultado después de la normalización"""
    print("=" * 80)
    print("📊 PASO 4: Verificación final...")
    print("=" * 80)
    print()

    wb_verificar = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws_trans = wb_verificar['TRANSACCIONES']

    headers = [ws_trans.cell(1, col).value for col in range(1, ws_trans.max_column + 1)]
    col_cuenta = headers.index('Cuenta Bancaria') + 1

    # Contar cuentas después
    cuentas_despues = {}

    for row in range(2, ws_trans.max_row + 1):
        cuenta = ws_trans.cell(row, col_cuenta).value

        if cuenta:
            cuenta_str = str(cuenta).strip()

            if cuenta_str not in cuentas_despues:
                cuentas_despues[cuenta_str] = 0

            cuentas_despues[cuenta_str] += 1

    print(f"📋 Cuentas únicas después de normalizar: {len(cuentas_despues)}")
    print()

    for cuenta, cantidad in sorted(cuentas_despues.items()):
        print(f"   • {cuenta}: {cantidad} transacciones")

    print()

def main():
    print("=" * 80)
    print("NORMALIZACIÓN UNIVERSAL DE CUENTAS")
    print("=" * 80)
    print()

    wb = openpyxl.load_workbook(EXCEL_FILE)

    # Cargar mapeo de alias
    mapeo = cargar_mapeo_alias(wb)
    if not mapeo:
        return False

    # Normalizar cuentas
    cambios_aplicados = normalizar_cuentas(wb, mapeo)

    if cambios_aplicados:
        # Guardar
        print("=" * 80)
        print("💾 Guardando cambios...")
        print("=" * 80)
        print()

        wb.save(EXCEL_FILE)
        print("✅ Excel actualizado")
        print()

        # Verificar
        verificar_resultado(wb)

    # Resumen final
    print("=" * 80)
    print("📊 RESUMEN")
    print("=" * 80)
    print()

    if cambios_aplicados:
        print("✅ Normalización completada exitosamente")
        print()
        print("🔧 PRÓXIMOS PASOS:")
        print("   1. Cierra y vuelve a abrir el Excel")
        print("   2. Verifica la hoja TRANSACCIONES - columna 'Cuenta Bancaria'")
        print("   3. Ve a la hoja Efectivo y verifica los saldos")
        print("   4. Si aparecen nuevas variaciones de nombres en el futuro:")
        print("      - Agrégalas a la hoja CUENTAS_ALIAS")
        print("      - Vuelve a ejecutar este script")
    else:
        print("✅ No se requirieron cambios - todas las cuentas ya están normalizadas")
        print()
        print("💡 Este script se ejecuta automáticamente para:")
        print("   • Unificar nombres de cuentas según el mapeo en CUENTAS_ALIAS")
        print("   • Detectar variaciones nuevas que necesitan ser agregadas al mapeo")

    print()
    print("=" * 80)
    print("✅ PROCESO COMPLETADO")
    print("=" * 80)
    print()

    return True

if __name__ == "__main__":
    try:
        if not crear_backup():
            print("❌ Abortando")
            exit(1)

        if main():
            print("🎉 Normalización universal completada!")
        else:
            print("❌ Proceso falló o fue cancelado")

    except Exception as e:
        print(f"❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
