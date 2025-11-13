#!/usr/bin/env python3
"""
Script interactivo para agregar transacciones al Excel v3.0
Con validaciones automáticas y sincronización a IVA_CONTROL
"""

import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime
import sys

V3_FILE = "AlvaroVelasco_Finanzas_v3.0.xlsx"

COLOR_EDITABLE = "FFF2CC"
COLOR_CALCULATED = "E7E6E6"

# Tipos de transacción válidos
TIPOS_VALIDOS = [
    "INGRESO",
    "GASTO OPERATIVO",
    "GASTO PERSONAL",
    "GASTO FINANCIERO",
    "COMPRA PARA REVENTA",
    "TRANSFERENCIA",
    "PAGO TARJETA",
    "PAGO PRESTAMO",
    "AJUSTE"
]

# Clientes zona franca
ZONA_FRANCA = ["VWR International", "RSHughes", "VWR", "RS Hughes"]

def detectar_duplicado(ws, fecha, cuenta, monto):
    """Detecta si existe una transacción duplicada"""
    for row in range(2, ws.max_row + 1):
        f = ws.cell(row, 1).value
        c = ws.cell(row, 5).value
        m_crc = ws.cell(row, 8).value or 0
        m_usd = ws.cell(row, 9).value or 0

        if f and c and str(f).split()[0] == str(fecha).split()[0]:
            if str(c) == str(cuenta) and (abs(m_crc - monto) < 0.01 or abs(m_usd - monto) < 0.01):
                return True
    return False

def agregar_a_iva(wb, tipo, entidad, monto_usd, factura, fecha):
    """Agrega automáticamente a IVA_CONTROL si aplica"""
    ws_iva = wb['IVA_CONTROL']

    if tipo == "INGRESO" and monto_usd > 0:
        # Buscar primera fila vacía en VENTAS (6-20)
        row = 6
        while ws_iva.cell(row, 4).value and row < 21:
            row += 1

        if row < 21:
            # Detectar zona franca
            es_zona_franca = any(zf.lower() in str(entidad).lower() for zf in ZONA_FRANCA)

            ws_iva.cell(row, 1, fecha)
            ws_iva.cell(row, 1).number_format = 'DD/MM/YY'
            ws_iva.cell(row, 2, factura or "N/D")
            ws_iva.cell(row, 3, entidad)
            ws_iva.cell(row, 4, monto_usd)
            ws_iva.cell(row, 4).number_format = '$#,##0.00'
            ws_iva.cell(row, 4).fill = PatternFill(start_color=COLOR_EDITABLE, end_color=COLOR_EDITABLE, fill_type="solid")
            ws_iva.cell(row, 5, f'=D{row}*0.13')
            ws_iva.cell(row, 5).number_format = '$#,##0.00'
            ws_iva.cell(row, 5).fill = PatternFill(start_color=COLOR_CALCULATED, end_color=COLOR_CALCULATED, fill_type="solid")
            ws_iva.cell(row, 6, f'=D{row}+E{row}')
            ws_iva.cell(row, 6).number_format = '$#,##0.00'
            ws_iva.cell(row, 6).fill = PatternFill(start_color=COLOR_CALCULATED, end_color=COLOR_CALCULATED, fill_type="solid")
            ws_iva.cell(row, 7, "SI" if es_zona_franca else "NO")
            ws_iva.cell(row, 7).fill = PatternFill(start_color=COLOR_EDITABLE, end_color=COLOR_EDITABLE, fill_type="solid")
            ws_iva.cell(row, 8, f'=IF(G{row}="SI",0,D{row}*0.02)')
            ws_iva.cell(row, 8).number_format = '$#,##0.00'
            ws_iva.cell(row, 8).fill = PatternFill(start_color=COLOR_CALCULATED, end_color=COLOR_CALCULATED, fill_type="solid")
            ws_iva.cell(row, 9, f'=F{row}-H{row}')
            ws_iva.cell(row, 9).number_format = '$#,##0.00'
            ws_iva.cell(row, 9).fill = PatternFill(start_color=COLOR_CALCULATED, end_color=COLOR_CALCULATED, fill_type="solid")
            ws_iva.cell(row, 10, "EMITIDA")

            if es_zona_franca:
                ws_iva.cell(row, 12, "ZONA FRANCA - No aplica IVA")

            return True

    elif tipo in ["GASTO OPERATIVO", "COMPRA PARA REVENTA"] and monto_usd > 0:
        # Buscar primera fila vacía en COMPRAS (25-40)
        row = 25
        while ws_iva.cell(row, 4).value and row < 41:
            row += 1

        if row < 41:
            base_gravable = monto_usd / 1.13

            ws_iva.cell(row, 1, fecha)
            ws_iva.cell(row, 1).number_format = 'DD/MM/YY'
            ws_iva.cell(row, 2, factura or "N/D")
            ws_iva.cell(row, 3, entidad)
            ws_iva.cell(row, 4, base_gravable)
            ws_iva.cell(row, 4).number_format = '$#,##0.00'
            ws_iva.cell(row, 4).fill = PatternFill(start_color=COLOR_EDITABLE, end_color=COLOR_EDITABLE, fill_type="solid")
            ws_iva.cell(row, 5, f'=D{row}*0.13')
            ws_iva.cell(row, 5).number_format = '$#,##0.00'
            ws_iva.cell(row, 5).fill = PatternFill(start_color=COLOR_CALCULATED, end_color=COLOR_CALCULATED, fill_type="solid")
            ws_iva.cell(row, 6, f'=D{row}+E{row}')
            ws_iva.cell(row, 6).number_format = '$#,##0.00'
            ws_iva.cell(row, 6).fill = PatternFill(start_color=COLOR_CALCULATED, end_color=COLOR_CALCULATED, fill_type="solid")
            ws_iva.cell(row, 7, "SI")
            ws_iva.cell(row, 7).fill = PatternFill(start_color=COLOR_EDITABLE, end_color=COLOR_EDITABLE, fill_type="solid")
            ws_iva.cell(row, 8, f'=IF(G{row}="SI",E{row},0)')
            ws_iva.cell(row, 8).number_format = '$#,##0.00'
            ws_iva.cell(row, 8).fill = PatternFill(start_color=COLOR_CALCULATED, end_color=COLOR_CALCULATED, fill_type="solid")
            ws_iva.cell(row, 10, "PAGADA")

            return True

    return False

def main():
    print("\n" + "="*60)
    print("AGREGAR TRANSACCIÓN - Excel v3.0")
    print("="*60)

    # Abrir Excel
    try:
        wb = openpyxl.load_workbook(V3_FILE)
        ws = wb['TRANSACCIONES']
    except Exception as e:
        print(f"❌ Error abriendo Excel: {e}")
        return

    # Input de datos
    print("\n📝 Ingresá los datos de la transacción:")

    # Fecha
    fecha_str = input("Fecha (DD/MM/YYYY) [Enter=hoy]: ").strip()
    if not fecha_str:
        fecha = datetime.now()
    else:
        try:
            fecha = datetime.strptime(fecha_str, "%d/%m/%Y")
        except:
            print("❌ Formato de fecha inválido")
            return

    # Tipo
    print("\nTipos disponibles:")
    for i, tipo in enumerate(TIPOS_VALIDOS, 1):
        print(f"  {i}. {tipo}")
    tipo_idx = input("Tipo (número): ").strip()
    try:
        tipo = TIPOS_VALIDOS[int(tipo_idx) - 1]
    except:
        print("❌ Tipo inválido")
        return

    # Descripción
    descripcion = input("Descripción: ").strip()
    if not descripcion:
        print("❌ Descripción requerida")
        return

    # Cuenta
    cuenta = input("Cuenta (ej: BAC USD, SINPE): ").strip()
    if not cuenta:
        print("❌ Cuenta requerida")
        return

    # Entidad
    entidad = input("Entidad/Cliente/Proveedor: ").strip()

    # Factura
    factura = input("N° Factura [opcional]: ").strip()

    # Monto
    moneda = input("Moneda (1=CRC, 2=USD): ").strip()
    monto = input("Monto: ").strip()
    try:
        monto = float(monto)
    except:
        print("❌ Monto inválido")
        return

    # TC
    tc = 508  # Default
    if moneda == "1":  # CRC
        tc_input = input(f"Tipo Cambio [Enter={tc}]: ").strip()
        if tc_input:
            tc = float(tc_input)

    # Método pago
    metodo = input("Método (EFECTIVO/TRANSFERENCIA/TARJETA/SINPE/CHEQUE): ").strip().upper()
    if not metodo:
        metodo = "TRANSFERENCIA"

    # Validar duplicado
    if detectar_duplicado(ws, fecha, cuenta, monto):
        confirmacion = input("⚠️  POSIBLE DUPLICADO detectado. ¿Continuar? (s/n): ")
        if confirmacion.lower() != 's':
            print("❌ Cancelado")
            return

    # Buscar primera fila vacía
    row = 2
    while ws.cell(row, 1).value:
        row += 1

    # Agregar transacción
    ws.cell(row, 1, fecha)  # Fecha
    ws.cell(row, 1).number_format = 'DD/MM/YY'
    ws.cell(row, 2, tipo)  # Tipo
    ws.cell(row, 4, descripcion)  # Descripción
    ws.cell(row, 5, cuenta)  # Cuenta
    ws.cell(row, 6, entidad)  # Entidad
    ws.cell(row, 7, factura)  # Factura

    # Montos
    if moneda == "1":  # CRC
        ws.cell(row, 8, monto)  # CRC
        ws.cell(row, 8).number_format = '₡#,##0.00'
        ws.cell(row, 9, 0)  # USD
        ws.cell(row, 9).number_format = '$#,##0.00'
        monto_usd = monto / tc
    else:  # USD
        ws.cell(row, 8, 0)  # CRC
        ws.cell(row, 8).number_format = '₡#,##0.00'
        ws.cell(row, 9, monto)  # USD
        ws.cell(row, 9).number_format = '$#,##0.00'
        monto_usd = monto

    ws.cell(row, 10, tc)  # TC
    ws.cell(row, 10).number_format = '₡#,##0.00'
    ws.cell(row, 11, metodo)  # Método
    ws.cell(row, 12, "COMPLETADA")  # Estado
    ws.cell(row, 17, datetime.now())  # Fecha creación
    ws.cell(row, 17).number_format = 'DD/MM/YY'

    # Duplicado (fórmula)
    ws.cell(row, 19, f'=COUNTIFS($A$2:$A${row-1},$A{row},$E$2:$E${row-1},$E{row},$H$2:$H${row-1},$H{row})')

    print(f"\n✅ Transacción agregada en fila {row}")

    # Sincronizar con IVA_CONTROL
    if agregar_a_iva(wb, tipo, entidad, monto_usd, factura, fecha):
        print("✅ Sincronizado con IVA_CONTROL")

    # Guardar
    wb.save(V3_FILE)
    wb.close()

    print("\n" + "="*60)
    print("✅ TRANSACCIÓN GUARDADA")
    print("="*60)
    print(f"   Tipo: {tipo}")
    print(f"   Monto: {'₡' if moneda == '1' else '$'}{monto:,.2f}")
    print(f"   Cuenta: {cuenta}")
    print(f"   Fila: {row}")
    print("="*60)

if __name__ == "__main__":
    main()
