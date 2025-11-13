#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
DIAGNÓSTICO NOMBRES DE CUENTA PROMERICA
Identifica todas las variaciones del nombre de cuenta Promerica
"""
import openpyxl

EXCEL_FILE = "AlvaroVelasco_Finanzas_v2.0.xlsx"

def diagnosticar():
    print("=" * 80)
    print("DIAGNÓSTICO - VARIACIONES NOMBRE CUENTA PROMERICA")
    print("=" * 80)
    print()

    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb['TRANSACCIONES']

    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    col_map = {}
    for col in range(1, len(headers) + 1):
        if headers[col-1]:
            col_map[headers[col-1]] = col

    # Buscar todas las variaciones de Promerica
    variaciones = {}

    for row in range(2, ws.max_row + 1):
        cuenta = ws.cell(row, col_map['Cuenta Bancaria']).value

        if cuenta and 'Promerica' in str(cuenta):
            cuenta_str = str(cuenta).strip()

            if cuenta_str not in variaciones:
                variaciones[cuenta_str] = []

            variaciones[cuenta_str].append(row)

    print(f"🔍 Se encontraron {len(variaciones)} variaciones del nombre 'Promerica':")
    print()

    for i, (nombre, filas) in enumerate(sorted(variaciones.items()), 1):
        print(f"{i}. \"{nombre}\"")
        print(f"   Cantidad de transacciones: {len(filas)}")
        print(f"   Filas: {', '.join(map(str, filas[:10]))}" + ("..." if len(filas) > 10 else ""))
        print()

    # Mostrar totales por variación
    print("=" * 80)
    print("💰 TOTALES POR VARIACIÓN")
    print("=" * 80)
    print()

    for nombre, filas in sorted(variaciones.items()):
        total_ingreso = 0
        total_egreso = 0

        for fila in filas:
            monto = ws.cell(fila, col_map['Monto USD']).value
            ing_egr = ws.cell(fila, col_map['Ingreso/Egreso']).value

            if monto:
                monto_val = float(monto)
                if ing_egr == 'Ingreso':
                    total_ingreso += monto_val
                elif ing_egr == 'Egreso':
                    total_egreso += abs(monto_val)

        balance = total_ingreso - total_egreso

        print(f"\"{nombre}\"")
        print(f"   Transacciones: {len(filas)}")
        print(f"   Ingreso: ${total_ingreso:,.2f}")
        print(f"   Egreso: ${total_egreso:,.2f}")
        print(f"   Balance: ${balance:,.2f}")
        print()

    # Total general
    print("=" * 80)
    print("📊 TOTAL GENERAL PROMERICA (TODAS LAS VARIACIONES)")
    print("=" * 80)
    print()

    total_trans = sum(len(filas) for filas in variaciones.values())

    total_ingreso_general = 0
    total_egreso_general = 0

    for filas in variaciones.values():
        for fila in filas:
            monto = ws.cell(fila, col_map['Monto USD']).value
            ing_egr = ws.cell(fila, col_map['Ingreso/Egreso']).value

            if monto:
                monto_val = float(monto)
                if ing_egr == 'Ingreso':
                    total_ingreso_general += monto_val
                elif ing_egr == 'Egreso':
                    total_egreso_general += abs(monto_val)

    balance_general = total_ingreso_general - total_egreso_general

    print(f"Total transacciones: {total_trans}")
    print(f"Ingreso total: ${total_ingreso_general:,.2f}")
    print(f"Egreso total: ${total_egreso_general:,.2f}")
    print(f"Balance total: ${balance_general:,.2f}")
    print()

    print("🏦 Balance esperado según extracto: $2,163.44")
    diferencia = abs(balance_general - 2163.44)
    print(f"⚖️  Diferencia: ${diferencia:,.2f}")
    print()

    # Recomendación
    print("=" * 80)
    print("📋 RECOMENDACIÓN")
    print("=" * 80)
    print()

    if len(variaciones) > 1:
        print("⚠️  HAY MÚLTIPLES VARIACIONES DEL NOMBRE DE CUENTA")
        print()
        print("✅ SOLUCIÓN 1: Normalizar todos los nombres a uno solo")
        print("   Elegir UN nombre estándar y cambiar todas las transacciones a ese nombre")
        print()
        print("✅ SOLUCIÓN 2: Usar fórmula flexible en Efectivo")
        print("   Usar SUMIFS con comodín o SEARCH para capturar todas las variaciones")
        print()
        print("🎯 RECOMENDACIÓN: Normalizar todos a un solo nombre (más limpio)")
    else:
        print("✅ Solo hay una variación del nombre - no es necesario normalizar")

    print()
    print("=" * 80)
    print("✅ DIAGNÓSTICO COMPLETADO")
    print("=" * 80)

if __name__ == "__main__":
    try:
        diagnosticar()
    except Exception as e:
        print(f"❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
