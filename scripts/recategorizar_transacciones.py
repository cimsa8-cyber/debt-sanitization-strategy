#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
RECATEGORIZACIÓN MASIVA - SISTEMA FINANCIERO
Aplica nueva estructura de Tipo Transacción + Categoría a todas las transacciones
"""
import openpyxl
from datetime import datetime
import shutil

EXCEL_FILE = "AlvaroVelasco_Finanzas_v2.0.xlsx"
BACKUP_FILE = f"AlvaroVelasco_Finanzas_v2.0_BACKUP_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

# MAPEO: Categoría Actual → Tipo Transacción
MAPEO_TIPO = {
    # INGRESOS
    'Cuentas por Cobrar': 'INGRESOS',
    'Ingresos Clientes': 'INGRESOS',
    'Ingresos Varios': 'INGRESOS',
    'Salario': 'INGRESOS',
    'Reintegros': 'INGRESOS',

    # COMPRAS PARA REVENTA
    'Compras': 'COMPRAS PARA REVENTA',
    'Proveedores': 'COMPRAS PARA REVENTA',
    'Inventario': 'COMPRAS PARA REVENTA',
    'Tecnología': 'COMPRAS PARA REVENTA',  # Productos tecnológicos para reventa
    'Logística': 'COMPRAS PARA REVENTA',
    'Logistica': 'COMPRAS PARA REVENTA',
    'Gastos Operativos': 'COMPRAS PARA REVENTA',  # ¡Fila 206 Intcomex mal categorizada!

    # GASTOS OPERATIVOS
    'Servicios': 'GASTOS OPERATIVOS',
    'Comisiones': 'GASTOS OPERATIVOS',
    'Alimentación': 'GASTOS OPERATIVOS',
    'Supermercado': 'GASTOS OPERATIVOS',
    'Combustible': 'GASTOS OPERATIVOS',
    'Servicios Públicos': 'GASTOS OPERATIVOS',
    'Vivienda': 'GASTOS OPERATIVOS',
    'Personal': 'GASTOS OPERATIVOS',
    'Entretenimiento': 'GASTOS OPERATIVOS',
    'Capacitación': 'GASTOS OPERATIVOS',
    'Capacitacion': 'GASTOS OPERATIVOS',
    'Educación': 'GASTOS OPERATIVOS',
    'Vehiculo': 'GASTOS OPERATIVOS',
    'Transporte': 'GASTOS OPERATIVOS',
    'CCSS': 'GASTOS OPERATIVOS',
    'Hacienda': 'GASTOS OPERATIVOS',
    'Otros Gastos': 'GASTOS OPERATIVOS',
    'Servicios Administrativos': 'GASTOS OPERATIVOS',

    # GASTOS FINANCIEROS
    'Comisiones Bancarias': 'GASTOS FINANCIEROS',
    'Gastos Bancarios': 'GASTOS FINANCIEROS',
    'Tarjetas de Crédito': 'GASTOS FINANCIEROS',
    'Tarjeta Crédito': 'GASTOS FINANCIEROS',
    'Tarjetas de Credito': 'GASTOS FINANCIEROS',
    'Financiamiento Vehículo': 'GASTOS FINANCIEROS',
    'Deudas': 'GASTOS FINANCIEROS',

    # TRANSFERENCIAS INTERNAS
    'Efectivo': 'TRANSFERENCIAS',
    'Ahorro': 'TRANSFERENCIAS',
    'Ahorro Personal': 'TRANSFERENCIAS',
    'Transferencias': 'TRANSFERENCIAS',
    'Cambio de Moneda': 'TRANSFERENCIAS',
    'Ajustes': 'TRANSFERENCIAS',
    'Saldos Iniciales': 'TRANSFERENCIAS',
}

# RENOMBRADO: Categoría Antigua → Categoría Nueva (para consistencia)
RENOMBRAR_CATEGORIA = {
    # Unificar compras tecnológicas
    'Compras': 'Productos Tecnológicos',
    'Tecnología': 'Productos Tecnológicos',
    'Proveedores': 'Productos Tecnológicos',
    'Inventario': 'Productos Tecnológicos',

    # Unificar logística
    'Logística': 'Flete y Logística',
    'Logistica': 'Flete y Logística',

    # Unificar gastos financieros
    'Tarjetas de Crédito': 'Intereses Tarjetas Crédito',
    'Tarjeta Crédito': 'Intereses Tarjetas Crédito',
    'Tarjetas de Credito': 'Intereses Tarjetas Crédito',

    # Corregir Intcomex (fila 206)
    'Gastos Operativos': 'Productos Tecnológicos',  # Era mal categorizada

    # Unificar capacitación
    'Capacitacion': 'Capacitación',

    # Unificar ingresos
    'Ingresos Clientes': 'Ventas de Productos',
    'Cuentas por Cobrar': 'Ventas de Productos',

    # Mantener otros nombres
    'Servicios': 'Servicios',
    'Comisiones': 'Comisiones',
    'Alimentación': 'Alimentación',
    'Supermercado': 'Supermercado',
    'Combustible': 'Combustible',
}

def crear_backup():
    """Crea backup del Excel antes de modificar"""
    print("=" * 80)
    print("CREANDO BACKUP")
    print("=" * 80)
    print(f"Archivo original: {EXCEL_FILE}")
    print(f"Backup: {BACKUP_FILE}")

    try:
        shutil.copy2(EXCEL_FILE, BACKUP_FILE)
        print("✅ Backup creado exitosamente")
        print()
        return True
    except Exception as e:
        print(f"❌ ERROR creando backup: {e}")
        return False

def recategorizar():
    """Aplica nueva estructura de categorías"""

    print("=" * 80)
    print("RECATEGORIZACIÓN MASIVA - SISTEMA FINANCIERO")
    print("=" * 80)
    print()

    # Cargar Excel
    print("📂 Cargando Excel...")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb['TRANSACCIONES']
    print(f"✓ {ws.max_row - 1} transacciones encontradas")
    print()

    # Identificar columnas
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]

    try:
        idx_tipo = headers.index('Tipo Transacción') + 1
        idx_categoria = headers.index('Categoría') + 1
    except ValueError as e:
        print(f"❌ ERROR: No se encontró columna esperada: {e}")
        return False

    # Estadísticas
    stats = {
        'total': 0,
        'actualizadas': 0,
        'sin_mapeo': 0,
        'por_tipo': {},
        'renombradas': 0,
    }

    transacciones_sin_mapeo = []

    print("=" * 80)
    print("APLICANDO RECATEGORIZACIÓN")
    print("=" * 80)
    print()

    # Procesar cada fila
    for row in range(2, ws.max_row + 1):
        stats['total'] += 1

        categoria_actual = ws.cell(row, idx_categoria).value

        if not categoria_actual:
            continue

        categoria_str = str(categoria_actual).strip()

        # Buscar tipo transacción
        tipo_nuevo = MAPEO_TIPO.get(categoria_str)

        if tipo_nuevo:
            # Actualizar Tipo Transacción (Columna B)
            ws.cell(row, idx_tipo).value = tipo_nuevo
            stats['actualizadas'] += 1

            # Contabilizar por tipo
            if tipo_nuevo not in stats['por_tipo']:
                stats['por_tipo'][tipo_nuevo] = 0
            stats['por_tipo'][tipo_nuevo] += 1

            # Renombrar categoría si corresponde
            if categoria_str in RENOMBRAR_CATEGORIA:
                categoria_nueva = RENOMBRAR_CATEGORIA[categoria_str]
                ws.cell(row, idx_categoria).value = categoria_nueva
                stats['renombradas'] += 1

                # Logging especial para fila 206 (Intcomex)
                if row == 206:
                    print(f"✨ FILA 206 (Intcomex):")
                    print(f"   Tipo: {tipo_nuevo}")
                    print(f"   Categoría: {categoria_str} → {categoria_nueva}")
                    print()
        else:
            stats['sin_mapeo'] += 1
            transacciones_sin_mapeo.append({
                'fila': row,
                'categoria': categoria_str
            })

    # Guardar
    print("💾 Guardando cambios...")
    wb.save(EXCEL_FILE)
    print("✅ Excel actualizado exitosamente")
    print()

    # Reporte de resultados
    print("=" * 80)
    print("📊 RESULTADOS DE RECATEGORIZACIÓN")
    print("=" * 80)
    print()

    print(f"Total transacciones procesadas: {stats['total']}")
    print(f"✅ Actualizadas con Tipo: {stats['actualizadas']}")
    print(f"✅ Categorías renombradas: {stats['renombradas']}")
    print(f"⚠️  Sin mapeo (revisar manualmente): {stats['sin_mapeo']}")
    print()

    if stats['por_tipo']:
        print("📊 DISTRIBUCIÓN POR TIPO:")
        for tipo, count in sorted(stats['por_tipo'].items()):
            porcentaje = (count / stats['total']) * 100
            print(f"   • {tipo}: {count} ({porcentaje:.1f}%)")
        print()

    if transacciones_sin_mapeo:
        print("⚠️  TRANSACCIONES SIN MAPEO (requieren revisión manual):")
        for t in transacciones_sin_mapeo[:10]:  # Mostrar primeras 10
            print(f"   Fila {t['fila']}: {t['categoria']}")
        if len(transacciones_sin_mapeo) > 10:
            print(f"   ... y {len(transacciones_sin_mapeo) - 10} más")
        print()

    print("=" * 80)
    print("✅ RECATEGORIZACIÓN COMPLETADA")
    print("=" * 80)
    print()

    print("📋 PRÓXIMOS PASOS:")
    print("   1. Abre el Excel y verifica fila 206 (Intcomex)")
    print("   2. Revisa la columna 'Tipo Transacción' (columna B)")
    print("   3. Verifica que las categorías se actualizaron correctamente")
    print("   4. Si todo está correcto, podemos continuar con análisis de utilidades")
    print()

    return True

if __name__ == "__main__":
    try:
        # Crear backup
        if not crear_backup():
            print("❌ Abortando: No se pudo crear backup")
            exit(1)

        # Recategorizar
        if recategorizar():
            print("🎉 Proceso completado exitosamente!")
        else:
            print("❌ Proceso completado con errores")
            print(f"💡 Puedes restaurar desde: {BACKUP_FILE}")

    except FileNotFoundError:
        print(f"❌ ERROR: No se encontró el archivo {EXCEL_FILE}")
        print(f"   Asegúrate de ejecutar este script desde la carpeta del proyecto")
    except Exception as e:
        print(f"❌ ERROR INESPERADO: {e}")
        import traceback
        traceback.print_exc()
        print()
        print(f"💡 Puedes restaurar desde: {BACKUP_FILE}")
