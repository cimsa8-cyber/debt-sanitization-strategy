#!/usr/bin/env python3
"""
Actualiza DASHBOARD con KPIs en tiempo real
Calcula automáticamente desde TRANSACCIONES, EFECTIVO, CxC, CxP
"""

import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import datetime

V3_FILE = "AlvaroVelasco_Finanzas_v3.0.xlsx"

print("\n" + "="*60)
print("ACTUALIZAR DASHBOARD - KPIs Tiempo Real")
print("="*60)

wb = openpyxl.load_workbook(V3_FILE)

# ============================================================================
# CALCULAR KPIs
# ============================================================================

print("\n📊 Calculando KPIs...")

# --- EFECTIVO ---
ws_efectivo = wb['EFECTIVO']
total_bancos = 0
total_tarjetas = 0

for row in range(5, 14):  # Bancos (filas 5-13)
    saldo = ws_efectivo.cell(row, 5).value or 0
    if isinstance(saldo, (int, float)):
        total_bancos += saldo

for row in range(16, 21):  # Tarjetas (filas 16-20)
    saldo = ws_efectivo.cell(row, 5).value or 0
    if isinstance(saldo, (int, float)):
        total_tarjetas += abs(saldo)  # Tarjetas son negativas

efectivo_neto = total_bancos - total_tarjetas

# --- CxC ---
ws_cxc = wb['CxC']
total_cxc = 0
cxc_vencida = 0

for row in range(3, 25):
    saldo = ws_cxc.cell(row, 6).value or 0
    dias_vencido = ws_cxc.cell(row, 8).value or 0

    if isinstance(saldo, (int, float)) and saldo > 0:
        total_cxc += saldo
        if isinstance(dias_vencido, (int, float)) and dias_vencido > 0:
            cxc_vencida += saldo

# --- CxP ---
ws_cxp = wb['CxP']
total_cxp = 0
cxp_critica = 0

for row in range(3, 25):
    saldo = ws_cxp.cell(row, 6).value or 0
    prioridad = ws_cxp.cell(row, 8).value or ""

    if isinstance(saldo, (int, float)) and saldo > 0:
        total_cxp += saldo
        if "CRÍTICA" in str(prioridad).upper():
            cxp_critica += saldo

# --- TRANSACCIONES (Ingresos/Gastos del mes) ---
ws_trans = wb['TRANSACCIONES']
ingresos_mes = 0
gastos_mes = 0
mes_actual = datetime.now().month

for row in range(2, ws_trans.max_row + 1):
    fecha = ws_trans.cell(row, 1).value
    tipo = ws_trans.cell(row, 2).value
    monto_usd = ws_trans.cell(row, 9).value or 0

    if fecha and isinstance(fecha, datetime) and fecha.month == mes_actual:
        if tipo and "INGRESO" in str(tipo).upper():
            ingresos_mes += monto_usd
        elif tipo and ("GASTO" in str(tipo).upper() or "COMPRA" in str(tipo).upper()):
            gastos_mes += monto_usd

flujo_neto_mes = ingresos_mes - gastos_mes

# --- IVA ---
ws_iva = wb['IVA_CONTROL']
# Calcular IVA cobrado manualmente
iva_cobrado = 0
for row in range(6, 21):
    base = ws_iva.cell(row, 4).value or 0
    if isinstance(base, (int, float)) and base > 0:
        iva_cobrado += base * 0.13

# Calcular IVA acreditable manualmente
iva_acreditable = 0
for row in range(25, 41):
    base = ws_iva.cell(row, 4).value or 0
    deducible = ws_iva.cell(row, 7).value
    if isinstance(base, (int, float)) and base > 0 and deducible == "SI":
        iva_acreditable += base * 0.13

iva_neto = iva_cobrado - iva_acreditable

# --- Días de cobertura ---
# Calcular gasto promedio de TODOS los gastos (no solo del mes)
gastos_totales = 0
dias_periodo = 0

for row in range(2, ws_trans.max_row + 1):
    tipo = ws_trans.cell(row, 2).value
    monto_usd = ws_trans.cell(row, 9).value or 0

    if tipo and ("GASTO" in str(tipo).upper() or "COMPRA" in str(tipo).upper()):
        gastos_totales += monto_usd

# Usar 30 días como periodo default
dias_periodo = 30

if gastos_totales > 0:
    gasto_diario = gastos_totales / dias_periodo
    dias_cobertura = efectivo_neto / gasto_diario if gasto_diario > 0 else 0
else:
    dias_cobertura = 0

# ============================================================================
# ACTUALIZAR DASHBOARD
# ============================================================================

print("\n📈 Actualizando DASHBOARD...")

ws_dash = wb['DASHBOARD']

# Limpiar área de KPIs (filas 3-20)
for row in range(3, 21):
    for col in range(1, 6):
        ws_dash.cell(row, col).value = None
        ws_dash.cell(row, col).fill = PatternFill(fill_type=None)

# --- HEADER ---
ws_dash['A1'] = f"DASHBOARD FINANCIERO - {datetime.now().strftime('%d/%b/%Y %H:%M')}"
ws_dash['A1'].font = Font(size=14, bold=True)

row = 3

# --- EFECTIVO ---
ws_dash.cell(row, 1, "💰 EFECTIVO")
ws_dash.cell(row, 1).font = Font(size=12, bold=True)
ws_dash.cell(row, 1).fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
ws_dash.cell(row, 1).font = Font(size=12, bold=True, color='FFFFFF')
row += 1

ws_dash.cell(row, 1, "Bancos")
ws_dash.cell(row, 2, total_bancos)
ws_dash.cell(row, 2).number_format = '$#,##0.00'
row += 1

ws_dash.cell(row, 1, "Tarjetas (debe)")
ws_dash.cell(row, 2, -total_tarjetas)
ws_dash.cell(row, 2).number_format = '$#,##0.00'
ws_dash.cell(row, 2).font = Font(color='FF0000')
row += 1

ws_dash.cell(row, 1, "NETO")
ws_dash.cell(row, 2, efectivo_neto)
ws_dash.cell(row, 2).number_format = '$#,##0.00'
ws_dash.cell(row, 2).font = Font(bold=True)
if efectivo_neto < 0:
    ws_dash.cell(row, 2).fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
else:
    ws_dash.cell(row, 2).fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
row += 2

# --- FLUJO DEL MES ---
ws_dash.cell(row, 1, "📊 FLUJO NOVIEMBRE")
ws_dash.cell(row, 1).font = Font(size=12, bold=True, color='FFFFFF')
ws_dash.cell(row, 1).fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
row += 1

ws_dash.cell(row, 1, "Ingresos")
ws_dash.cell(row, 2, ingresos_mes)
ws_dash.cell(row, 2).number_format = '$#,##0.00'
ws_dash.cell(row, 2).fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
row += 1

ws_dash.cell(row, 1, "Gastos")
ws_dash.cell(row, 2, -gastos_mes)
ws_dash.cell(row, 2).number_format = '$#,##0.00'
ws_dash.cell(row, 2).fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
row += 1

ws_dash.cell(row, 1, "NETO")
ws_dash.cell(row, 2, flujo_neto_mes)
ws_dash.cell(row, 2).number_format = '$#,##0.00'
ws_dash.cell(row, 2).font = Font(bold=True)
row += 2

# --- CxC / CxP ---
ws_dash.cell(row, 1, "📋 CUENTAS")
ws_dash.cell(row, 1).font = Font(size=12, bold=True, color='FFFFFF')
ws_dash.cell(row, 1).fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
row += 1

ws_dash.cell(row, 1, "CxC Total")
ws_dash.cell(row, 2, total_cxc)
ws_dash.cell(row, 2).number_format = '$#,##0.00'
row += 1

ws_dash.cell(row, 1, "CxC Vencida")
ws_dash.cell(row, 2, cxc_vencida)
ws_dash.cell(row, 2).number_format = '$#,##0.00'
ws_dash.cell(row, 2).fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
row += 1

ws_dash.cell(row, 1, "CxP Total")
ws_dash.cell(row, 2, -total_cxp)
ws_dash.cell(row, 2).number_format = '$#,##0.00'
ws_dash.cell(row, 2).font = Font(color='FF0000')
row += 1

ws_dash.cell(row, 1, "CxP Crítica")
ws_dash.cell(row, 2, -cxp_critica)
ws_dash.cell(row, 2).number_format = '$#,##0.00'
ws_dash.cell(row, 2).fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
row += 2

# --- IVA ---
ws_dash.cell(row, 1, "🧾 IVA NOVIEMBRE")
ws_dash.cell(row, 1).font = Font(size=12, bold=True, color='FFFFFF')
ws_dash.cell(row, 1).fill = PatternFill(start_color='7030A0', end_color='7030A0', fill_type='solid')
row += 1

ws_dash.cell(row, 1, "IVA Cobrado")
ws_dash.cell(row, 2, iva_cobrado)
ws_dash.cell(row, 2).number_format = '$#,##0.00'
row += 1

ws_dash.cell(row, 1, "IVA Acreditable")
ws_dash.cell(row, 2, -iva_acreditable)
ws_dash.cell(row, 2).number_format = '$#,##0.00'
row += 1

ws_dash.cell(row, 1, "IVA Neto")
ws_dash.cell(row, 2, iva_neto)
ws_dash.cell(row, 2).number_format = '$#,##0.00'
ws_dash.cell(row, 2).font = Font(bold=True)
if iva_neto > 0:
    ws_dash.cell(row, 2).fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
else:
    ws_dash.cell(row, 2).fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
row += 2

# --- DÍAS DE COBERTURA ---
ws_dash.cell(row, 1, "⏰ DÍAS DE COBERTURA")
ws_dash.cell(row, 1).font = Font(size=12, bold=True, color='FFFFFF')
ws_dash.cell(row, 1).fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
row += 1

ws_dash.cell(row, 1, "Efectivo / Gasto Diario")
ws_dash.cell(row, 2, dias_cobertura)
ws_dash.cell(row, 2).number_format = '0.0'
ws_dash.cell(row, 2).font = Font(size=14, bold=True)
if dias_cobertura < 15:
    ws_dash.cell(row, 2).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    ws_dash.cell(row, 2).font = Font(size=14, bold=True, color='FFFFFF')
elif dias_cobertura < 30:
    ws_dash.cell(row, 2).fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
else:
    ws_dash.cell(row, 2).fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

# Anchos de columna
ws_dash.column_dimensions['A'].width = 25
ws_dash.column_dimensions['B'].width = 15

# ============================================================================
# GUARDAR
# ============================================================================

print(f"\n💾 Guardando {V3_FILE}...")
wb.save(V3_FILE)
wb.close()

# ============================================================================
# RESUMEN
# ============================================================================

print("\n" + "="*60)
print("✅ DASHBOARD ACTUALIZADO")
print("="*60)
print(f"   💰 Efectivo Neto: ${efectivo_neto:,.2f}")
print(f"   📊 Flujo Nov: ${flujo_neto_mes:,.2f}")
print(f"   📋 CxC: ${total_cxc:,.2f} | CxP: ${total_cxp:,.2f}")
print(f"   🧾 IVA Neto: ${iva_neto:,.2f}")
print(f"   ⏰ Días Cobertura: {dias_cobertura:.1f} días")
print("="*60)
