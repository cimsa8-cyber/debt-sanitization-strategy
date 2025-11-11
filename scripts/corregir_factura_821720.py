#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CORRECCIÓN FACTURA #821720
Corrige 2 cosas que faltaron:
1. Buscar factura #821720 original y marcarla como Pagado
2. Cambiar fila 208 (pago) de COMPRAS PARA REVENTA → TRANSFERENCIAS
"""
import openpyxl

EXCEL_FILE = "AlvaroVelasco_Finanzas_v2.0.xlsx"

def corregir():
    print("=" * 80)
    print("CORRECCIÓN FINAL - FACTURA #821720")
    print("=" * 80)
    print()

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb['TRANSACCIONES']

    # Mapeo de columnas
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    col_map = {}
    for col in range(1, len(headers) + 1):
        if headers[col-1]:
            col_map[headers[col-1]] = col

    print("📋 PASO 1: Buscando factura #821720 original...")
    print()

    # Buscar factura #821720 (debería estar antes de fila 206)
    factura_fila = None
    for row in range(2, 206):  # Buscar solo en registros anteriores
        referencia = ws.cell(row, col_map['Referencia']).value
        concepto = ws.cell(row, col_map['Concepto']).value
        fecha = ws.cell(row, col_map['Fecha']).value

        if referencia and '821720' in str(referencia):
            factura_fila = row
            break
        elif concepto and '821720' in str(concepto):
            factura_fila = row
            break

    if factura_fila:
        estado_antes = ws.cell(factura_fila, col_map['Estado']).value
        fecha_factura = ws.cell(factura_fila, col_map['Fecha']).value
        monto = ws.cell(factura_fila, col_map['Monto USD']).value

        ws.cell(factura_fila, col_map['Estado']).value = 'Pagado'

        print(f"✅ Factura #821720 encontrada:")
        print(f"   Fila: {factura_fila}")
        print(f"   Fecha: {fecha_factura}")
        print(f"   Monto: ${monto:,.2f} USD")
        print(f"   Estado: {estado_antes} → Pagado")
        print()
    else:
        print("⚠️  Factura #821720 no encontrada")
        print("   (Puede que no tenga el número en Referencia o Concepto)")
        print()

    print("=" * 80)
    print("📋 PASO 2: Corrigiendo fila 208 (pago)...")
    print()

    # Corregir fila 208 - cambiar tipo
    tipo_antes = ws.cell(208, col_map['Tipo Transacción']).value
    categoria_antes = ws.cell(208, col_map['Categoría']).value

    ws.cell(208, col_map['Tipo Transacción']).value = 'TRANSFERENCIAS'
    ws.cell(208, col_map['Categoría']).value = 'Transferencias'
    ws.cell(208, col_map['Entidad']).value = 'Pago a Proveedor'

    print(f"✅ Fila 208 corregida:")
    print(f"   Tipo: {tipo_antes} → TRANSFERENCIAS")
    print(f"   Categoría: {categoria_antes} → Transferencias")
    print()

    # Guardar
    print("💾 Guardando cambios...")
    wb.save(EXCEL_FILE)
    print("✅ Excel actualizado")
    print()

    print("=" * 80)
    print("📊 RESUMEN")
    print("=" * 80)
    print()

    if factura_fila:
        print(f"✅ Fila {factura_fila}: Factura #821720 marcada como Pagado")
    print(f"✅ Fila 208: Pago recategorizado como TRANSFERENCIAS")
    print()

    print("=" * 80)
    print("✅ CORRECCIÓN COMPLETADA")
    print("=" * 80)
    print()

    print("📋 VERIFICACIÓN:")
    if factura_fila:
        print(f"   1. Abre Excel y ve a fila {factura_fila}")
        print(f"      Verifica que Estado = 'Pagado'")
    print(f"   2. Ve a fila 208")
    print(f"      Verifica que Tipo = 'TRANSFERENCIAS'")
    print()

if __name__ == "__main__":
    try:
        corregir()
        print("🎉 Todo sincronizado correctamente!")
    except Exception as e:
        print(f"❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
