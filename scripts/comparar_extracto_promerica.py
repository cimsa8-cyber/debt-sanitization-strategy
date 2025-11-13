#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Comparación extracto bancario Promerica vs Excel
Identifica movimientos faltantes exactos
"""
import openpyxl
from datetime import datetime

EXCEL_FILE = "AlvaroVelasco_Finanzas_v2.0.xlsx"

# Movimientos del extracto bancario (31/10/2025) - TODOS
extracto = [
    # 16/10
    {"fecha": "16/10/2025", "ref": "585896", "monto": -3.00, "desc": "Comision TFT Salario"},
    {"fecha": "16/10/2025", "ref": "585896", "monto": -500.00, "desc": "TFT Salario Quincena"},
    # 17/10
    {"fecha": "17/10/2025", "ref": "8210835", "monto": 350.00, "desc": "VWR INTERNATIONAL LT"},
    # 21/10
    {"fecha": "21/10/2025", "ref": "595744", "monto": -3.00, "desc": "Comision TFT fact 199488"},
    {"fecha": "21/10/2025", "ref": "595744", "monto": -149.01, "desc": "TFT fact 199488"},
    {"fecha": "21/10/2025", "ref": "596247", "monto": -3.00, "desc": "Comision TFT Casa 10E"},
    {"fecha": "21/10/2025", "ref": "596247", "monto": -775.00, "desc": "TFT Casa 10E Alquiler"},
    # 22/10
    {"fecha": "22/10/2025", "ref": "597597", "monto": -3.00, "desc": "Comision TFT fact 2502060 INTCOMEX"},
    {"fecha": "22/10/2025", "ref": "597597", "monto": -410.09, "desc": "TFT fact 2502060 INTCOMEX"},
    # 23/10
    {"fecha": "23/10/2025", "ref": "8226004", "monto": 1186.50, "desc": "CORPORACION TIERRARE"},
    {"fecha": "23/10/2025", "ref": "547493", "monto": -73.61, "desc": "UNOPETROL BARREAL"},
    # 24/10
    {"fecha": "24/10/2025", "ref": "2785908", "monto": -0.75, "desc": "Comision Debito SINPE"},
    {"fecha": "24/10/2025", "ref": "31821661", "monto": -29.38, "desc": "DD SEA GLOBAL LOGISTICS"},
    # 27/10
    {"fecha": "27/10/2025", "ref": "5325194", "monto": -39.59, "desc": "DEBITO ESPH AGUA"},
    {"fecha": "27/10/2025", "ref": "5325195", "monto": -189.73, "desc": "DEBITO ESPH ELECT"},
    {"fecha": "27/10/2025", "ref": "5325197", "monto": -392.68, "desc": "ICE TELEFONICO"},
    # 28/10
    {"fecha": "28/10/2025", "ref": "8241248", "monto": 56.50, "desc": "CPF SERVICIOS RADIOL"},
    {"fecha": "28/10/2025", "ref": "8241249", "monto": 56.50, "desc": "ORTODEC SERVICIOS"},
    {"fecha": "28/10/2025", "ref": "8241251", "monto": 356.50, "desc": "ORTODONCIA DE LA CRU"},
    {"fecha": "28/10/2025", "ref": "8241539", "monto": 1237.35, "desc": "SMART WEB SERVICES"},
    # 29/10
    {"fecha": "29/10/2025", "ref": "611261", "monto": -3.00, "desc": "Comision TFT Pago tarjeta"},
    {"fecha": "29/10/2025", "ref": "611261", "monto": -305.50, "desc": "TFT Pago de tarjeta"},
    {"fecha": "29/10/2025", "ref": "611329", "monto": -3.00, "desc": "Comision TFT pago tarjeta"},
    {"fecha": "29/10/2025", "ref": "611329", "monto": -101.83, "desc": "TFT pago tarjeta"},
    {"fecha": "29/10/2025", "ref": "611345", "monto": -3.00, "desc": "Comision TFT Curso Pricing"},
    {"fecha": "29/10/2025", "ref": "611345", "monto": -101.83, "desc": "TFT Curso Pricing"},
    {"fecha": "29/10/2025", "ref": "2788366", "monto": -0.75, "desc": "Comision Debito SINPE"},
    {"fecha": "29/10/2025", "ref": "31889087", "monto": -800.00, "desc": "DD CARROFACIL"},
    # 30/10
    {"fecha": "30/10/2025", "ref": "93194651", "monto": 1171.18, "desc": "GRUPO PORCINAS"},
    {"fecha": "30/10/2025", "ref": "251030212028809", "monto": 282.50, "desc": "Transferencia ALVARO VELASCONET"},
    {"fecha": "30/10/2025", "ref": "66679628", "monto": 284.76, "desc": "TEF ELEC Volio Partners"},
    {"fecha": "30/10/2025", "ref": "8254868", "monto": 149.16, "desc": "SMART WEB SERVICES"},
    {"fecha": "30/10/2025", "ref": "8254872", "monto": 226.00, "desc": "GENTRA DE COSTA RICA"},
    {"fecha": "30/10/2025", "ref": "730298", "monto": -226.83, "desc": "DON FERNANDO HEREDIA"},
    {"fecha": "30/10/2025", "ref": "737072", "monto": -40.46, "desc": "FARMAVALUE HEREDIA"},
    {"fecha": "30/10/2025", "ref": "754410", "monto": -63.03, "desc": "UNOPETROL BARREAL"},
]

SALDO_FINAL_EXTRACTO = 3030.89

print("="*80)
print("COMPARACION: EXTRACTO BANCARIO VS EXCEL")
print("="*80)

# Cargar Excel
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
ws = wb['TRANSACCIONES']

# Leer referencias existentes en Excel para Promerica
referencias_excel = {}

for row in range(2, ws.max_row + 1):
    cuenta = ws[f'E{row}'].value
    if cuenta and 'Promerica USD 1774' in str(cuenta):
        fecha = ws[f'A{row}'].value
        referencia = ws[f'H{row}'].value
        monto = ws[f'I{row}'].value
        tipo_mov = ws[f'K{row}'].value

        if referencia and fecha:
            fecha_str = fecha.strftime('%d/%m/%Y') if hasattr(fecha, 'strftime') else str(fecha)
            clave = f"{fecha_str}_{referencia}"

            # Determinar signo
            try:
                monto_num = float(monto) if monto else 0
                if tipo_mov and 'Egreso' in str(tipo_mov):
                    monto_num = -abs(monto_num)
                else:
                    monto_num = abs(monto_num)
            except:
                monto_num = 0

            referencias_excel[clave] = {
                'fila': row,
                'monto': monto_num,
                'fecha': fecha_str
            }

print(f"\nMovimientos Promerica en Excel: {len(referencias_excel)}")
print(f"Movimientos en extracto bancario: {len(extracto)}")

# Comparar
faltantes = []
encontrados = []

for mov_extracto in extracto:
    clave = f"{mov_extracto['fecha']}_{mov_extracto['ref']}"

    if clave in referencias_excel:
        encontrados.append(mov_extracto)
    else:
        faltantes.append(mov_extracto)

print(f"\n✅ Movimientos encontrados: {len(encontrados)}")
print(f"❌ Movimientos FALTANTES: {len(faltantes)}")

if len(faltantes) > 0:
    print("\n" + "="*80)
    print("MOVIMIENTOS QUE FALTAN EN EL EXCEL")
    print("="*80)

    total_faltante = 0
    for mov in faltantes:
        signo = "+" if mov['monto'] > 0 else ""
        print(f"{mov['fecha']} | {mov['ref']:15} | {signo}${mov['monto']:>8.2f} | {mov['desc']}")
        total_faltante += mov['monto']

    print(f"\nImpacto total de movimientos faltantes: ${total_faltante:,.2f}")

# Calcular saldo según Excel
saldo_excel = sum([r['monto'] for r in referencias_excel.values()])

print("\n" + "="*80)
print("ANALISIS DE SALDOS")
print("="*80)

print(f"\nSaldo según extracto bancario (31/10): ${SALDO_FINAL_EXTRACTO:,.2f}")
print(f"Saldo según Excel (calculado): ${saldo_excel:,.2f}")
print(f"Diferencia: ${SALDO_FINAL_EXTRACTO - saldo_excel:,.2f}")

# Calcular saldo si agregáramos los faltantes
if len(faltantes) > 0:
    total_faltante = sum([m['monto'] for m in faltantes])
    saldo_con_faltantes = saldo_excel + total_faltante
    print(f"\nSi agregamos los {len(faltantes)} movimientos faltantes:")
    print(f"Nuevo saldo Excel: ${saldo_con_faltantes:,.2f}")
    print(f"VS extracto: ${SALDO_FINAL_EXTRACTO:,.2f}")
    print(f"Nueva diferencia: ${SALDO_FINAL_EXTRACTO - saldo_con_faltantes:,.2f}")

# Verificar balance inicial en hoja Efectivo
try:
    ws_efectivo = wb['Efectivo']
    for row in range(1, 20):
        concepto = ws_efectivo[f'B{row}'].value
        if concepto and 'Balance inicial Promerica USD' in str(concepto):
            balance_efectivo = ws_efectivo[f'F{row}'].value
            print(f"\n⚠️ Hoja Efectivo muestra balance inicial: ${balance_efectivo:,.2f}")
            print(f"   Debería ser (según extracto 31/10): ${SALDO_FINAL_EXTRACTO:,.2f}")
            print(f"   Diferencia: ${abs(balance_efectivo - SALDO_FINAL_EXTRACTO):,.2f}")
            break
except:
    pass

print("\n" + "="*80)
print("RECOMENDACIONES")
print("="*80)

if len(faltantes) > 0:
    print(f"\n1. AGREGAR {len(faltantes)} movimientos faltantes al Excel")
    print(f"2. CORREGIR balance inicial en hoja Efectivo a ${SALDO_FINAL_EXTRACTO:,.2f}")
    print(f"3. Después de esto, el saldo de Promerica quedará correcto")
else:
    print("\n✅ Todos los movimientos del extracto están en el Excel")
    print(f"⚠️ Solo falta corregir balance inicial a ${SALDO_FINAL_EXTRACTO:,.2f}")

print("\n" + "="*80)
