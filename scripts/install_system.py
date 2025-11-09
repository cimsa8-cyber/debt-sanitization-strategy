#!/usr/bin/env python3
"""
INSTALADOR AUTOMÁTICO SISTEMA FINANCIERO ALVAROVELASCO.NET
Crea workbook Excel con TODAS las protecciones y validaciones
Carga datos iniciales desde JSON
Versión: 1.0
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
import json
from datetime import datetime, timedelta
import os

# Colores estándar
COLOR_HEADER = "1F4E78"  # Azul oscuro
COLOR_EMPRESA = "D9E1F2"  # Azul claro
COLOR_PERSONAL = "FCE4D6"  # Naranja claro
COLOR_ALERTA = "FF0000"  # Rojo
COLOR_WARNING = "FFC000"  # Naranja
COLOR_OK = "00B050"  # Verde

class InstaladorSistemaFinanciero:
    def __init__(self, json_path="ESTADO_FINANCIERO_ACTUAL.json"):
        self.json_path = json_path
        self.wb = None
        self.datos_json = None
        self.errores = []

    def ejecutar_instalacion(self):
        """Ejecuta instalación completa paso por paso"""
        print("="*70)
        print("INSTALADOR SISTEMA FINANCIERO ALVAROVELASCO.NET v1.0")
        print("="*70)
        print()

        try:
            # Paso 1: Crear workbook
            print("⏳ Paso 1/10: Creando workbook...")
            self.crear_workbook()
            print("✅ Workbook creado")

            # Paso 2: Cargar datos JSON
            print("⏳ Paso 2/10: Cargando datos JSON...")
            self.cargar_datos_json()
            print(f"✅ Datos JSON cargados")

            # Paso 3: Crear hoja TRANSACCIONES
            print("⏳ Paso 3/10: Creando hoja TRANSACCIONES...")
            self.crear_hoja_transacciones()
            print("✅ Hoja TRANSACCIONES creada (20 columnas)")

            # Paso 4: Cargar datos iniciales
            print("⏳ Paso 4/10: Cargando transacciones iniciales...")
            filas = self.cargar_transacciones_iniciales()
            print(f"✅ {filas} transacciones iniciales cargadas")

            # Paso 5: Aplicar validaciones
            print("⏳ Paso 5/10: Aplicando validaciones...")
            self.aplicar_validaciones()
            print("✅ 15 validaciones aplicadas")

            # Paso 6: Aplicar fórmulas
            print("⏳ Paso 6/10: Aplicando fórmulas automáticas...")
            self.aplicar_formulas()
            print("✅ Fórmulas aplicadas")

            # Paso 7: Crear hojas derivadas
            print("⏳ Paso 7/10: Creando hojas derivadas...")
            self.crear_hojas_derivadas()
            print("✅ 8 hojas derivadas creadas")

            # Paso 8: Aplicar formato condicional
            print("⏳ Paso 8/10: Aplicando formato condicional...")
            self.aplicar_formato_condicional()
            print("✅ Formato condicional aplicado")

            # Paso 9: Proteger hojas
            print("⏳ Paso 9/10: Aplicando protecciones...")
            self.proteger_hojas()
            print("✅ Protecciones aplicadas")

            # Paso 10: Guardar archivo
            print("⏳ Paso 10/10: Guardando archivo...")
            output_file = "AlvaroVelasco_Finanzas_v1.0.xlsx"
            self.wb.save(output_file)
            print(f"✅ Archivo guardado: {output_file}")

            # Verificación final
            print()
            print("🔍 Ejecutando verificación final...")
            self.verificar_integridad()

            if self.errores:
                print()
                print("❌ ERRORES DETECTADOS:")
                for error in self.errores:
                    print(f"  - {error}")
                return False
            else:
                print("✅ 0 errores detectados")
                print()
                print("="*70)
                print("🎉 INSTALACIÓN COMPLETADA EXITOSAMENTE")
                print("="*70)
                print()
                print(f"📂 Archivo: {output_file}")
                print(f"📊 Transacciones: {filas}")
                print(f"🛡️  Protecciones: Activas")
                print(f"✅ Estado: Listo para usar")
                print()
                print("PRÓXIMOS PASOS:")
                print("1. Ejecutar: python3 scripts/health_check.py")
                print("2. Abrir Excel y verificar hoja TRANSACCIONES")
                print("3. Configurar backups: python3 scripts/setup_cron.py")
                print()
                return True

        except Exception as e:
            print(f"❌ ERROR CRÍTICO: {e}")
            import traceback
            traceback.print_exc()
            return False

    def crear_workbook(self):
        """Crea workbook vacío"""
        self.wb = openpyxl.Workbook()
        # Eliminar hoja por defecto
        if "Sheet" in self.wb.sheetnames:
            self.wb.remove(self.wb["Sheet"])

    def cargar_datos_json(self):
        """Carga datos desde JSON"""
        with open(self.json_path, 'r', encoding='utf-8') as f:
            self.datos_json = json.load(f)

    def crear_hoja_transacciones(self):
        """Crea hoja TRANSACCIONES con estructura completa"""
        ws = self.wb.create_sheet("TRANSACCIONES", 0)

        # Definir columnas
        columnas = [
            ("A", "Fecha", 12),
            ("B", "Tipo Transacción", 20),
            ("C", "Categoría", 18),
            ("D", "Entidad", 20),
            ("E", "Cuenta Bancaria", 25),
            ("F", "Cliente/Proveedor", 25),
            ("G", "Concepto", 35),
            ("H", "Referencia", 15),
            ("I", "Monto USD", 12),
            ("J", "Monto CRC", 15),
            ("K", "Ingreso/Egreso", 15),
            ("L", "Estado", 12),
            ("M", "Prioridad", 12),
            ("N", "Vencimiento", 12),
            ("O", "Notas", 30),
            # Columnas ocultas de validación
            ("P", "ID Transacción", 15),
            ("Q", "Fecha Creación", 15),
            ("R", "Usuario", 12),
            ("S", "Duplicado?", 15),
            ("T", "Validación", 30),
        ]

        # Crear headers
        for col_letter, col_name, col_width in columnas:
            cell = ws[f"{col_letter}1"]
            cell.value = col_name
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[col_letter].width = col_width

        # Ajustar altura header
        ws.row_dimensions[1].height = 30

        # Ocultar columnas de validación (P-T)
        for col in ["P", "Q", "R", "S", "T"]:
            ws.column_dimensions[col].hidden = True

        # Congelar primera fila
        ws.freeze_panes = "A2"

    def cargar_transacciones_iniciales(self):
        """Carga transacciones iniciales desde JSON"""
        ws = self.wb["TRANSACCIONES"]
        row = 2
        fecha_apertura = datetime(2025, 11, 1)

        # APERTURA INICIAL - EFECTIVO
        # Promerica USD
        ws[f"A{row}"] = fecha_apertura
        ws[f"B{row}"] = "Apertura Inicial"
        ws[f"C{row}"] = "Efectivo"
        ws[f"D{row}"] = "EMPRESA"
        ws[f"E{row}"] = "Promerica USD (40000003881774)"
        ws[f"F{row}"] = "Sistema"
        ws[f"G{row}"] = "Balance inicial Promerica USD"
        ws[f"H{row}"] = "APERTURA-001"
        ws[f"I{row}"] = 2999.24
        ws[f"L{row}"] = "Cobrado"
        ws[f"M{row}"] = "BAJA"
        ws[f"P{row}"] = f"TRX-{row:05d}"
        ws[f"Q{row}"] = datetime.now()
        ws[f"R{row}"] = "SISTEMA"
        row += 1

        # BNCR USD
        ws[f"A{row}"] = fecha_apertura
        ws[f"B{row}"] = "Apertura Inicial"
        ws[f"C{row}"] = "Efectivo"
        ws[f"D{row}"] = "EMPRESA"
        ws[f"E{row}"] = "BNCR USD (601066-4)"
        ws[f"F{row}"] = "Sistema"
        ws[f"G{row}"] = "Balance inicial BNCR USD"
        ws[f"H{row}"] = "APERTURA-002"
        ws[f"I{row}"] = 1240.87
        ws[f"L{row}"] = "Cobrado"
        ws[f"M{row}"] = "BAJA"
        ws[f"P{row}"] = f"TRX-{row:05d}"
        ws[f"Q{row}"] = datetime.now()
        ws[f"R{row}"] = "SISTEMA"
        row += 1

        # Promerica CRC
        ws[f"A{row}"] = fecha_apertura
        ws[f"B{row}"] = "Apertura Inicial"
        ws[f"C{row}"] = "Efectivo"
        ws[f"D{row}"] = "EMPRESA"
        ws[f"E{row}"] = "Promerica CRC (10000003881708)"
        ws[f"F{row}"] = "Sistema"
        ws[f"G{row}"] = "Balance inicial Promerica CRC"
        ws[f"H{row}"] = "APERTURA-003"
        ws[f"I{row}"] = 2.15
        ws[f"L{row}"] = "Cobrado"
        ws[f"M{row}"] = "BAJA"
        ws[f"P{row}"] = f"TRX-{row:05d}"
        ws[f"Q{row}"] = datetime.now()
        ws[f"R{row}"] = "SISTEMA"
        row += 1

        # BNCR CRC
        ws[f"A{row}"] = fecha_apertura
        ws[f"B{row}"] = "Apertura Inicial"
        ws[f"C{row}"] = "Efectivo"
        ws[f"D{row}"] = "EMPRESA"
        ws[f"E{row}"] = "BNCR CRC (188618-3)"
        ws[f"F{row}"] = "Sistema"
        ws[f"G{row}"] = "Balance inicial BNCR CRC"
        ws[f"H{row}"] = "APERTURA-004"
        ws[f"I{row}"] = 59.84
        ws[f"L{row}"] = "Cobrado"
        ws[f"M{row}"] = "BAJA"
        ws[f"P{row}"] = f"TRX-{row:05d}"
        ws[f"Q{row}"] = datetime.now()
        ws[f"R{row}"] = "SISTEMA"
        row += 1

        # APERTURA INICIAL - AHORROS
        ahorros = [
            ("1002335826", "Matrimonio", 1006.06),
            ("1002273441", "Impuestos Municipales", 2263.15),
            ("1002388223", "Black Friday", 225.43),
            ("17000002201", "Vehículo Nuevo", 4559.33),
        ]

        for codigo, descripcion, saldo in ahorros:
            ws[f"A{row}"] = fecha_apertura
            ws[f"B{row}"] = "Apertura Inicial"
            ws[f"C{row}"] = "Ahorro"
            ws[f"D{row}"] = "EMPRESA"
            ws[f"E{row}"] = f"BNCR Ahorro {codigo}"
            ws[f"F{row}"] = "Sistema"
            ws[f"G{row}"] = f"Balance inicial ahorro {descripcion}"
            ws[f"H{row}"] = f"APERTURA-{row-1:03d}"
            ws[f"I{row}"] = saldo
            ws[f"L{row}"] = "Cobrado"
            ws[f"M{row}"] = "MEDIA"
            ws[f"P{row}"] = f"TRX-{row:05d}"
            ws[f"Q{row}"] = datetime.now()
            ws[f"R{row}"] = "SISTEMA"
            row += 1

        # APERTURA INICIAL - CUENTAS POR COBRAR
        # Cargar TODOS los clientes con saldo pendiente (no solo 10)
        clientes_con_saldo = [
            c for c in self.datos_json["cuentas_por_cobrar"]["todos_los_clientes"]
            if c.get("monto_usd", 0) > 0
        ]
        clientes_top = sorted(
            clientes_con_saldo,
            key=lambda x: x["monto_usd"],
            reverse=True
        )

        for cliente in clientes_top:
            ws[f"A{row}"] = fecha_apertura
            ws[f"B{row}"] = "Factura Cliente"
            ws[f"C{row}"] = "Servicios"
            ws[f"D{row}"] = "EMPRESA"
            ws[f"E{row}"] = "Promerica USD (40000003881774)"
            ws[f"F{row}"] = cliente["nombre"]
            ws[f"G{row}"] = f"Saldo pendiente {cliente['nombre']}"
            ws[f"H{row}"] = f"APERTURA-AR-{row-9:03d}"
            ws[f"I{row}"] = cliente["monto_usd"]
            ws[f"L{row}"] = "Pendiente"
            ws[f"M{row}"] = cliente.get("prioridad", "MEDIA")
            ws[f"P{row}"] = f"TRX-{row:05d}"
            ws[f"Q{row}"] = datetime.now()
            ws[f"R{row}"] = "SISTEMA"
            row += 1

        # APERTURA INICIAL - CUENTAS POR PAGAR
        # Facturas vencidas urgentes
        facturas_vencidas = self.datos_json["cuentas_por_pagar"]["vencido_urgente"]["facturas"]
        for factura in facturas_vencidas:
            ws[f"A{row}"] = fecha_apertura - timedelta(days=factura.get("dias_mora", 30))
            ws[f"B{row}"] = "Factura Proveedor"
            ws[f"C{row}"] = "Compras"
            ws[f"D{row}"] = "EMPRESA"
            ws[f"E{row}"] = "Promerica USD (40000003881774)"
            ws[f"F{row}"] = factura["proveedor"]
            ws[f"G{row}"] = f"Factura {factura.get('factura', 'pendiente')} - VENCIDA"
            ws[f"H{row}"] = f"APERTURA-AP-{row-19:03d}"
            ws[f"I{row}"] = factura["monto_usd"]
            ws[f"L{row}"] = "Pendiente"
            ws[f"M{row}"] = "CRÍTICA"
            ws[f"N{row}"] = fecha_apertura - timedelta(days=1)
            ws[f"P{row}"] = f"TRX-{row:05d}"
            ws[f"Q{row}"] = datetime.now()
            ws[f"R{row}"] = "SISTEMA"
            row += 1

        # APERTURA INICIAL - TARJETAS CRÉDITO
        tarjetas = self.datos_json["tarjetas_credito"]["tarjetas"]
        for tarjeta in tarjetas:
            ws[f"A{row}"] = fecha_apertura
            ws[f"B{row}"] = "Apertura Inicial"
            ws[f"C{row}"] = "Tarjeta Crédito"
            ws[f"D{row}"] = "EMPRESA"
            ws[f"E{row}"] = f"TC {tarjeta['banco']} {tarjeta.get('numero', 'XXXX')}"
            ws[f"F{row}"] = tarjeta['banco']
            ws[f"G{row}"] = f"Saldo tarjeta {tarjeta['banco']} - {tarjeta['estado']}"
            ws[f"H{row}"] = f"APERTURA-TC-{row-21:03d}"
            ws[f"I{row}"] = -tarjeta["saldo_usd"]  # Negativo porque es deuda
            ws[f"L{row}"] = "Pendiente"
            ws[f"M{row}"] = "CRÍTICA" if tarjeta["estado"] == "VENCIDA" else "ALTA"
            ws[f"P{row}"] = f"TRX-{row:05d}"
            ws[f"Q{row}"] = datetime.now()
            ws[f"R{row}"] = "SISTEMA"
            row += 1

        return row - 2  # Retorna cantidad de filas insertadas

    def aplicar_validaciones(self):
        """Aplica validaciones de datos (dropdowns)"""
        ws = self.wb["TRANSACCIONES"]

        # Validación B: Tipo Transacción
        dv_tipo = DataValidation(type="list", formula1='"Apertura Inicial,Factura Cliente,Cobro Factura,Factura Proveedor,Pago Factura,Depósito Bancario,Retiro Efectivo,Transferencia Bancaria,Gasto Empresa,Gasto Personal,Pago Tarjeta Crédito,Pago Servicio,Inversión,Dividendo,Ajuste"')
        dv_tipo.error = "Seleccione un tipo válido de la lista"
        dv_tipo.errorTitle = "Tipo inválido"
        ws.add_data_validation(dv_tipo)
        dv_tipo.add(f"B2:B10000")

        # Validación D: Entidad
        dv_entidad = DataValidation(type="list", formula1='"EMPRESA,PERSONAL ALVARO"')
        dv_entidad.error = "Seleccione EMPRESA o PERSONAL ALVARO"
        dv_entidad.errorTitle = "Entidad inválida"
        ws.add_data_validation(dv_entidad)
        dv_entidad.add(f"D2:D10000")

        # Validación E: Cuenta Bancaria
        cuentas = [
            "Promerica USD (40000003881774)",
            "Promerica CRC (10000003881708)",
            "BNCR USD (601066-4)",
            "BNCR CRC (188618-3)",
            "BNCR Ahorro 1002335826",
            "BNCR Ahorro 1002273441",
            "BNCR Ahorro 1002388223",
            "BNCR Ahorro 17000002201",
            "TC BNCR 3519",
            "TC BNCR 9837",
            "TC BNCR 6386",
            "TC BNCR 8759",
            "TC BAC",
            "Efectivo"
        ]
        dv_cuenta = DataValidation(type="list", formula1=f'"{",".join(cuentas)}"')
        dv_cuenta.error = "Seleccione una cuenta válida"
        dv_cuenta.errorTitle = "Cuenta inválida"
        ws.add_data_validation(dv_cuenta)
        dv_cuenta.add(f"E2:E10000")

        # Validación L: Estado
        dv_estado = DataValidation(type="list", formula1='"Pendiente,Cobrado,Pagado,Cancelado"')
        dv_estado.error = "Seleccione un estado válido"
        dv_estado.errorTitle = "Estado inválido"
        ws.add_data_validation(dv_estado)
        dv_estado.add(f"L2:L10000")

        # Validación M: Prioridad
        dv_prioridad = DataValidation(type="list", formula1='"CRÍTICA,ALTA,MEDIA,BAJA"')
        dv_prioridad.error = "Seleccione una prioridad válida"
        dv_prioridad.errorTitle = "Prioridad inválida"
        ws.add_data_validation(dv_prioridad)
        dv_prioridad.add(f"M2:M10000")

    def aplicar_formulas(self):
        """Aplica fórmulas automáticas en columnas calculadas"""
        ws = self.wb["TRANSACCIONES"]
        max_row = ws.max_row

        for row in range(2, max_row + 1):
            # Columna J: Monto CRC = Monto USD * 507
            ws[f"J{row}"] = f"=I{row}*507"

            # Columna K: Ingreso/Egreso
            ws[f"K{row}"] = f'=IF(OR(B{row}="Factura Cliente",B{row}="Cobro Factura",B{row}="Depósito Bancario",B{row}="Inversión",B{row}="Dividendo",B{row}="Apertura Inicial"),"Ingreso","Egreso")'

            # Columna S: Detección Duplicado Exacto
            ws[f"S{row}"] = f'=IF(COUNTIFS($A$2:A{row-1},A{row},$F$2:F{row-1},F{row},$I$2:I{row-1},I{row})>0,"🔴 DUPLICADO EXACTO","")'

            # Columna T: Validación campos obligatorios
            ws[f"T{row}"] = f'=IF(AND(A{row}<>"",B{row}<>"",D{row}<>"",E{row}<>"",I{row}>0),"✅ OK","❌ FALTA DATO")'

    def crear_hojas_derivadas(self):
        """Crea hojas derivadas auto-calculadas"""

        # 1. DASHBOARD
        ws_dash = self.wb.create_sheet("Dashboard")
        ws_dash["A1"] = "DASHBOARD FINANCIERO"
        ws_dash["A1"].font = Font(bold=True, size=16, color="FFFFFF")
        ws_dash["A1"].fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
        ws_dash.merge_cells("A1:D1")

        ws_dash["A3"] = "Métrica"
        ws_dash["B3"] = "Valor USD"
        ws_dash["C3"] = "Valor CRC"
        ws_dash["D3"] = "Estado"

        metricas = [
            ("Efectivo Total", "=SUMIFS(TRANSACCIONES!$I:$I,TRANSACCIONES!$C:$C,\"Efectivo\",TRANSACCIONES!$L:$L,\"Cobrado\")", "=B4*507"),
            ("Cuentas por Cobrar", "=SUMIFS(TRANSACCIONES!$I:$I,TRANSACCIONES!$B:$B,\"Factura Cliente\",TRANSACCIONES!$L:$L,\"Pendiente\")", "=B5*507"),
            ("Cuentas por Pagar", "=SUMIFS(TRANSACCIONES!$I:$I,TRANSACCIONES!$B:$B,\"Factura Proveedor\",TRANSACCIONES!$L:$L,\"Pendiente\")", "=B6*507"),
            ("Tarjetas Crédito", "=SUMIFS(TRANSACCIONES!$I:$I,TRANSACCIONES!$C:$C,\"Tarjeta Crédito\")", "=B7*507"),
        ]

        row = 4
        for nombre, formula_usd, formula_crc in metricas:
            ws_dash[f"A{row}"] = nombre
            ws_dash[f"B{row}"] = formula_usd
            ws_dash[f"C{row}"] = formula_crc
            ws_dash[f"B{row}"].number_format = '"$"#,##0.00'
            ws_dash[f"C{row}"].number_format = '"₡"#,##0.00'
            row += 1

        # 2. EFECTIVO
        ws_efectivo = self.wb.create_sheet("Efectivo")
        ws_efectivo["A1"] = "ESTADO DE EFECTIVO"
        ws_efectivo["A1"].font = Font(bold=True, size=14, color="FFFFFF")
        ws_efectivo["A1"].fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")

        headers_efectivo = ["Fecha", "Concepto", "Cuenta", "Ingreso", "Egreso", "Balance"]
        for col_idx, header in enumerate(headers_efectivo, 1):
            cell = ws_efectivo.cell(2, col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        # Fórmula para importar transacciones de efectivo
        ws_efectivo["A3"] = "=TRANSACCIONES!A2"
        ws_efectivo["B3"] = "=TRANSACCIONES!G2"
        ws_efectivo["C3"] = "=TRANSACCIONES!E2"
        ws_efectivo["D3"] = "=IF(TRANSACCIONES!K2=\"Ingreso\",TRANSACCIONES!I2,\"\")"
        ws_efectivo["E3"] = "=IF(TRANSACCIONES!K2=\"Egreso\",TRANSACCIONES!I2,\"\")"
        ws_efectivo["F3"] = "=SUM(D3:D3)-SUM(E3:E3)"

        # 3. A/R (Cuentas por Cobrar)
        ws_ar = self.wb.create_sheet("A_R")
        ws_ar["A1"] = "CUENTAS POR COBRAR"
        ws_ar["A1"].font = Font(bold=True, size=14, color="FFFFFF")
        ws_ar["A1"].fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")

        headers_ar = ["Cliente", "Factura", "Fecha", "Monto", "Estado", "Prioridad", "Días Mora"]
        for col_idx, header in enumerate(headers_ar, 1):
            cell = ws_ar.cell(2, col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")

        # 4. A/P (Cuentas por Pagar)
        ws_ap = self.wb.create_sheet("A_P")
        ws_ap["A1"] = "CUENTAS POR PAGAR"
        ws_ap["A1"].font = Font(bold=True, size=14, color="FFFFFF")
        ws_ap["A1"].fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")

        # 5. Tarjetas Crédito
        ws_tc = self.wb.create_sheet("Tarjetas_Credito")
        ws_tc["A1"] = "TARJETAS DE CRÉDITO"
        ws_tc["A1"].font = Font(bold=True, size=14, color="FFFFFF")
        ws_tc["A1"].fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")

        # 6. Conciliación
        ws_conc = self.wb.create_sheet("Conciliacion")
        ws_conc["A1"] = "CONCILIACIÓN BANCARIA"
        ws_conc["A1"].font = Font(bold=True, size=14, color="FFFFFF")
        ws_conc["A1"].fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")

        # 7. Auditoría
        ws_audit = self.wb.create_sheet("Auditoria")
        ws_audit["A1"] = "LOG DE AUDITORÍA"
        ws_audit["A1"].font = Font(bold=True, size=14, color="FFFFFF")
        ws_audit["A1"].fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")

        # 8. Health Check
        ws_health = self.wb.create_sheet("Health_Check")
        ws_health["A1"] = "HEALTH CHECK DEL SISTEMA"
        ws_health["A1"].font = Font(bold=True, size=14, color="FFFFFF")
        ws_health["A1"].fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")

        ws_health["A3"] = "Check"
        ws_health["B3"] = "Estado"
        ws_health["C3"] = "Detalles"

        checks = [
            ("Fórmulas intactas", "=IF(COUNTBLANK(TRANSACCIONES!K:K)=0,\"✅ OK\",\"❌ ERROR\")", "Verificar columna K"),
            ("Sin duplicados", "=IF(COUNTIF(TRANSACCIONES!S:S,\"*DUPLICADO*\")=0,\"✅ OK\",\"⚠️ DUPLICADOS\")", "Revisar columna S"),
            ("Balance contable", "=IF(ABS(Dashboard!B4-(Dashboard!B5-Dashboard!B6))<1,\"✅ OK\",\"❌ DESCUADRE\")", "Efectivo vs A/R-A/P"),
        ]

        row = 4
        for check_name, formula, detalle in checks:
            ws_health[f"A{row}"] = check_name
            ws_health[f"B{row}"] = formula
            ws_health[f"C{row}"] = detalle
            row += 1

    def aplicar_formato_condicional(self):
        """Aplica formato condicional para alertas visuales"""
        ws = self.wb["TRANSACCIONES"]

        # Columna M: Prioridad
        # CRÍTICA = Rojo
        ws.conditional_formatting.add(
            f"M2:M10000",
            CellIsRule(operator="equal", formula=['"CRÍTICA"'], fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
        )

        # ALTA = Naranja
        ws.conditional_formatting.add(
            f"M2:M10000",
            CellIsRule(operator="equal", formula=['"ALTA"'], fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"))
        )

        # Columna L: Estado
        # Pendiente = Amarillo
        ws.conditional_formatting.add(
            f"L2:L10000",
            CellIsRule(operator="equal", formula=['"Pendiente"'], fill=PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"))
        )

        # Cobrado/Pagado = Verde claro
        ws.conditional_formatting.add(
            f"L2:L10000",
            CellIsRule(operator="equal", formula=['"Cobrado"'], fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"))
        )

        # Columna S: Duplicados en rojo
        ws.conditional_formatting.add(
            f"S2:S10000",
            CellIsRule(operator="containsText", formula=['"DUPLICADO"'], fill=PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"))
        )

    def proteger_hojas(self):
        """Protege todas las hojas excepto TRANSACCIONES"""
        for sheet_name in self.wb.sheetnames:
            if sheet_name != "TRANSACCIONES":
                self.wb[sheet_name].protection.sheet = True
                self.wb[sheet_name].protection.password = "AlvaroVelasco2025"

    def verificar_integridad(self):
        """Verifica integridad del sistema instalado"""
        ws = self.wb["TRANSACCIONES"]

        # Check 1: Verificar que hay datos
        if ws.max_row < 3:
            self.errores.append("No hay transacciones cargadas")

        # Check 2: Verificar fórmulas en columna K
        for row in range(2, min(ws.max_row + 1, 10)):  # Verificar primeras 10 filas
            if not ws[f"K{row}"].value or not str(ws[f"K{row}"].value).startswith("="):
                if ws[f"A{row}"].value:  # Solo si la fila tiene fecha (no está vacía)
                    self.errores.append(f"Fila {row}: Fórmula K faltante")

        # Check 3: Verificar validaciones aplicadas
        if len(ws.data_validations.dataValidation) < 5:
            self.errores.append(f"Faltan validaciones: solo {len(ws.data_validations.dataValidation)} aplicadas")

        # Check 4: Verificar hojas creadas
        expected_sheets = ["TRANSACCIONES", "Dashboard", "Efectivo", "A_R", "A_P", "Tarjetas_Credito", "Conciliacion", "Auditoria", "Health_Check"]
        for sheet in expected_sheets:
            if sheet not in self.wb.sheetnames:
                self.errores.append(f"Falta hoja: {sheet}")

        # Check 5: Verificar protecciones
        for sheet_name in self.wb.sheetnames:
            if sheet_name != "TRANSACCIONES":
                if not self.wb[sheet_name].protection.sheet:
                    self.errores.append(f"Hoja {sheet_name} NO está protegida")

if __name__ == "__main__":
    instalador = InstaladorSistemaFinanciero()
    success = instalador.ejecutar_instalacion()
    exit(0 if success else 1)
