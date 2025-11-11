#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CORRECCIÓN FINAL COMPLETA
1. Busca factura #821720 de manera más flexible y marca como Pagado
2. Corrige fórmulas en hoja Efectivo ("" → 0)
"""
import openpyxl

EXCEL_FILE = "AlvaroVelasco_Finanzas_v2.0.xlsx"

def corregir_todo():
    print("=" * 80)
    print("CORRECCIÓN FINAL COMPLETA")
    print("=" * 80)
    print()

    wb = openpyxl.load_workbook(EXCEL_FILE)

    # =========================================================================
    # PARTE 1: BUSCAR Y ACTUALIZAR FACTURA #821720
    # =========================================================================

    print("📋 PARTE 1: Buscando factura #821720...")
    print()

    ws_trans = wb['TRANSACCIONES']
    headers = [ws_trans.cell(1, col).value for col in range(1, ws_trans.max_column + 1)]

    col_map = {}
    for col in range(1, len(headers) + 1):
        if headers[col-1]:
            col_map[headers[col-1]] = col

    # Buscar en TODAS las columnas
    factura_fila = None
    for row in range(2, 207):  # Antes de las nuevas transacciones
        for col in range(1, ws_trans.max_column + 1):
            valor = ws_trans.cell(row, col).value
            if valor and '821720' in str(valor):
                factura_fila = row
                break
        if factura_fila:
            break

    if factura_fila:
        fecha = ws_trans.cell(factura_fila, col_map['Fecha']).value
        estado_antes = ws_trans.cell(factura_fila, col_map['Estado']).value
        monto = ws_trans.cell(factura_fila, col_map['Monto USD']).value
        concepto = ws_trans.cell(factura_fila, col_map['Concepto']).value

        ws_trans.cell(factura_fila, col_map['Estado']).value = 'Pagado'

        print(f"✅ Factura #821720 encontrada y actualizada:")
        print(f"   Fila: {factura_fila}")
        print(f"   Fecha: {fecha}")
        print(f"   Concepto: {concepto[:50] if concepto else 'N/A'}")
        print(f"   Monto: ${abs(float(monto)):,.2f} USD" if monto else "   Monto: N/A")
        print(f"   Estado: {estado_antes} → Pagado")
        print()
    else:
        print("⚠️  Factura #821720 NO encontrada en filas 2-206")
        print("   Puede que ya esté marcada como Pagado o use otro formato")
        print()

    # =========================================================================
    # PARTE 2: CORREGIR FÓRMULAS EN HOJA EFECTIVO
    # =========================================================================

    print("=" * 80)
    print("📋 PARTE 2: Corrigiendo fórmulas en hoja Efectivo...")
    print()

    if 'Efectivo' not in wb.sheetnames:
        print("⚠️  Hoja 'Efectivo' no encontrada")
        print()
    else:
        ws_efectivo = wb['Efectivo']

        formulas_corregidas = 0

        # Revisar todas las celdas con fórmulas
        for row in range(1, ws_efectivo.max_row + 1):
            for col in range(1, ws_efectivo.max_column + 1):
                celda = ws_efectivo.cell(row, col)

                if celda.value and isinstance(celda.value, str) and celda.value.startswith('='):
                    formula_original = celda.value

                    # Reemplazar ;"") con ;0)
                    if '""' in formula_original or ';""' in formula_original:
                        formula_nueva = formula_original.replace(';"")', ';0)')
                        formula_nueva = formula_nueva.replace('""', '0')

                        if formula_nueva != formula_original:
                            celda.value = formula_nueva
                            formulas_corregidas += 1

                            col_letter = openpyxl.utils.get_column_letter(col)
                            print(f"   ✓ {col_letter}{row}: Fórmula corregida")

        if formulas_corregidas > 0:
            print()
            print(f"✅ {formulas_corregidas} fórmulas corregidas en hoja Efectivo")
        else:
            print("✅ No se encontraron fórmulas con \"\" para corregir")
        print()

    # =========================================================================
    # GUARDAR CAMBIOS
    # =========================================================================

    print("=" * 80)
    print("💾 Guardando todos los cambios...")
    wb.save(EXCEL_FILE)
    print("✅ Excel actualizado")
    print()

    # =========================================================================
    # RESUMEN
    # =========================================================================

    print("=" * 80)
    print("📊 RESUMEN FINAL")
    print("=" * 80)
    print()

    if factura_fila:
        print(f"✅ Fila {factura_fila}: Factura #821720 marcada como Pagado")
    else:
        print("⚠️  Factura #821720 no encontrada (revisar manualmente)")

    if formulas_corregidas > 0:
        print(f"✅ Hoja Efectivo: {formulas_corregidas} fórmulas corregidas")

    print()
    print("=" * 80)
    print("✅ CORRECCIÓN COMPLETADA")
    print("=" * 80)
    print()

    print("📋 VERIFICACIÓN FINAL:")
    if factura_fila:
        print(f"   1. Abre Excel y ve a TRANSACCIONES fila {factura_fila}")
        print(f"      Verifica Estado = 'Pagado'")
    print(f"   2. Abre hoja Efectivo")
    print(f"      Verifica que no haya errores #VALUE!")
    print(f"   3. Todo debería estar sincronizado")
    print()

if __name__ == "__main__":
    try:
        corregir_todo()
        print("🎉 Sistema completamente sincronizado!")
    except Exception as e:
        print(f"❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
