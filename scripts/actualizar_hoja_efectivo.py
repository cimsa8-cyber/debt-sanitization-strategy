#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ACTUALIZAR HOJA EFECTIVO
Corrige el balance inicial de Promerica USD en la hoja Efectivo
y elimina el balance duplicado del 01/11 en TRANSACCIONES
"""
import openpyxl
from datetime import datetime
import sys
import os

EXCEL_FILE = "AlvaroVelasco_Finanzas_v2.0.xlsx"

print("="*80)
print("ACTUALIZACIÓN: HOJA EFECTIVO Y TRANSACCIONES")
print("="*80)

if not os.path.exists(EXCEL_FILE):
    print(f"\n❌ ERROR: No se encontró {EXCEL_FILE}")
    print("   Verifique que está en el directorio correcto")
    sys.exit(1)

print(f"\n✓ Archivo encontrado: {EXCEL_FILE}")

# Cargar Excel
print("\nCargando Excel...")
wb = openpyxl.load_workbook(EXCEL_FILE)

# Verificar hojas
hojas_disponibles = wb.sheetnames
print(f"\n📊 Hojas disponibles: {hojas_disponibles}")

if 'Efectivo' not in hojas_disponibles:
    print("\n❌ ERROR: No se encontró la hoja 'Efectivo'")
    sys.exit(1)

if 'TRANSACCIONES' not in hojas_disponibles:
    print("\n❌ ERROR: No se encontró la hoja 'TRANSACCIONES'")
    sys.exit(1)

# ============================================================================
# PARTE 1: ACTUALIZAR HOJA EFECTIVO
# ============================================================================

print("\n" + "="*80)
print("PARTE 1: ACTUALIZAR HOJA EFECTIVO")
print("="*80)

ws_efectivo = wb['Efectivo']

# Buscar la fila de Promerica USD
print("\nBuscando balance de Promerica USD en hoja Efectivo...")

fila_promerica = None
for row in range(1, 30):
    concepto = ws_efectivo[f'B{row}'].value
    cuenta = ws_efectivo[f'C{row}'].value

    if concepto and 'Balance inicial Promerica USD' in str(concepto):
        if cuenta and 'Promerica USD' in str(cuenta):
            fila_promerica = row
            print(f"\n✓ Encontrado en fila {row}:")
            print(f"   Concepto: {concepto}")
            print(f"   Cuenta: {cuenta}")
            print(f"   Balance actual: ${ws_efectivo[f'F{row}'].value}")
            break

if fila_promerica:
    # Actualizar balance
    saldo_correcto = 3030.89  # Según extracto bancario 31/10

    print(f"\n📝 Actualizando balance:")
    print(f"   Valor anterior: ${ws_efectivo[f'F{fila_promerica}'].value}")
    print(f"   Valor nuevo: ${saldo_correcto}")

    # Actualizar columna D (Ingreso) y F (Balance)
    ws_efectivo[f'D{fila_promerica}'] = saldo_correcto
    ws_efectivo[f'F{fila_promerica}'] = saldo_correcto

    print(f"   ✅ Balance actualizado")
else:
    print("\n⚠️ No se encontró el balance de Promerica USD en hoja Efectivo")

# ============================================================================
# PARTE 2: ELIMINAR BALANCE DUPLICADO EN TRANSACCIONES
# ============================================================================

print("\n" + "="*80)
print("PARTE 2: ELIMINAR BALANCE DUPLICADO EN TRANSACCIONES")
print("="*80)

ws_trans = wb['TRANSACCIONES']

print("\nBuscando balance duplicado del 01/11/2025...")

filas_a_eliminar = []

for row in range(2, ws_trans.max_row + 1):
    cuenta = ws_trans[f'E{row}'].value
    tipo = ws_trans[f'B{row}'].value
    fecha = ws_trans[f'A{row}'].value
    monto = ws_trans[f'I{row}'].value

    if not cuenta or not tipo:
        continue

    # Buscar balance inicial de Promerica del 01/11
    if 'Promerica USD' in str(cuenta):
        if 'Balance inicial' in str(tipo) or 'Saldo Inicial' in str(tipo):
            if fecha and fecha.month == 11 and fecha.day == 1:
                print(f"\n⚠️ Balance duplicado encontrado (Fila {row}):")
                print(f"   Fecha: {fecha.strftime('%d/%m/%Y')}")
                print(f"   Tipo: {tipo}")
                print(f"   Monto: ${monto}")
                print(f"   Cuenta: {cuenta}")
                filas_a_eliminar.append(row)

if filas_a_eliminar:
    print(f"\n📝 Eliminando {len(filas_a_eliminar)} fila(s) duplicada(s)...")

    # Eliminar de abajo hacia arriba para no afectar índices
    for row in sorted(filas_a_eliminar, reverse=True):
        print(f"   Eliminando fila {row}...")
        ws_trans.delete_rows(row, 1)

    print(f"   ✅ Fila(s) eliminada(s)")
else:
    print("\n✓ No se encontraron balances duplicados para eliminar")

# ============================================================================
# GUARDAR CAMBIOS
# ============================================================================

print("\n" + "="*80)
print("GUARDANDO CAMBIOS")
print("="*80)

if fila_promerica or filas_a_eliminar:
    print(f"\nGuardando {EXCEL_FILE}...")
    wb.save(EXCEL_FILE)
    print("✅ Cambios guardados exitosamente")

    print("\n" + "="*80)
    print("RESUMEN DE CAMBIOS")
    print("="*80)

    if fila_promerica:
        print(f"\n✅ Hoja Efectivo:")
        print(f"   Balance Promerica actualizado a: ${saldo_correcto:,.2f}")

    if filas_a_eliminar:
        print(f"\n✅ Hoja TRANSACCIONES:")
        print(f"   {len(filas_a_eliminar)} balance(s) duplicado(s) eliminado(s)")

    print("\n" + "="*80)
    print("SIGUIENTE PASO")
    print("="*80)

    print("""
1. Abrir Excel y verificar:
   - Hoja Efectivo: Balance Promerica = $3,030.89
   - Hoja TRANSACCIONES: Solo debe haber 1 saldo inicial (15/10)

2. Ejecutar auditoría:
   python scripts/auditoria_con_alias.py

3. El saldo de Promerica ahora debe coincidir mejor con el extracto
""")

else:
    print("\n⚠️ No se realizaron cambios")

print("\n" + "="*80)
