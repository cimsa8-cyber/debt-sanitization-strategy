#!/usr/bin/env python3
"""
Limpia hojas duplicadas CxC/CxP
Mantiene la versión correcta (CxP1 con Nissan y Hacienda corregidos)
"""

import openpyxl

V3_FILE = "AlvaroVelasco_Finanzas_v3.0.xlsx"

print("🧹 Limpiando hojas duplicadas...")

wb = openpyxl.load_workbook(V3_FILE)

# Eliminar hojas antiguas (sin "1")
if "CxC" in wb.sheetnames:
    del wb["CxC"]
    print("   ❌ Eliminada hoja CxC (antigua)")

if "CxP" in wb.sheetnames:
    del wb["CxP"]
    print("   ❌ Eliminada hoja CxP (antigua, sin Nissan)")

# Renombrar hojas nuevas (quitar el "1")
if "CxC1" in wb.sheetnames:
    wb["CxC1"].title = "CxC"
    print("   ✅ CxC1 → CxC")

if "CxP1" in wb.sheetnames:
    wb["CxP1"].title = "CxP"
    print("   ✅ CxP1 → CxP (con Nissan $800 y Hacienda BAJA)")

wb.save(V3_FILE)
wb.close()

print("✅ Limpieza completada")
