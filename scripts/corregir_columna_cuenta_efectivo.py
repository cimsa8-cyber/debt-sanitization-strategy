#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CORRECCIÓN COLUMNA CUENTA EN EFECTIVO
Cambia las celdas C3-C10 de fórmulas a texto fijo con nombres estándar
"""
import openpyxl
from datetime import datetime
import shutil

EXCEL_FILE = "AlvaroVelasco_Finanzas_v2.0.xlsx"
BACKUP_FILE = f"AlvaroVelasco_Finanzas_v2.0_CORREGIR_CUENTA_EFECTIVO_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

# Mapeo manual de cuentas estándar según lo que hemos identificado
CUENTAS_ESTANDAR = {
    3: "Promerica USD (40000003881774)",
    4: "BNCR USD (100-01-000-123456-7)",
    5: "Promerica CRC (10000003881708)",
    6: "BNCR CRC (100-01-000-654321-8)",
    7: "Ahorro Matrimonio",
    8: "Ahorro Impuestos Municipales",
    9: "Ahorro Black Friday",
    10: "Ahorro Vehículo Nuevo",
}

def crear_backup():
    print("=" * 80)
    print("CREANDO BACKUP")
    print("=" * 80)
    print(f"Backup: {BACKUP_FILE}")
    try:
        shutil.copy2(EXCEL_FILE, BACKUP_FILE)
        print("✅ Backup creado")
        print()
        return True
    except Exception as e:
        print(f"❌ ERROR: {e}")
        return False

def corregir():
    print("=" * 80)
    print("CORRECCIÓN COLUMNA CUENTA EN EFECTIVO")
    print("=" * 80)
    print()

    # Primero leer con data_only para ver valores actuales
    wb_data = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws_efectivo_data = wb_data['Efectivo']
    ws_trans_data = wb_data['TRANSACCIONES']

    # Leer con fórmulas para ver qué tienen las celdas
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws_efectivo = wb['Efectivo']
    ws_trans = wb['TRANSACCIONES']

    print("📋 ESTADO ACTUAL DE COLUMNA C (Cuenta) EN EFECTIVO:")
    print()

    for fila in range(3, 11):
        celda_c = ws_efectivo.cell(fila, 3)  # Columna C
        celda_b = ws_efectivo.cell(fila, 2)  # Columna B (nombre descriptivo)

        contenido_c = celda_c.value
        nombre_b = celda_b.value

        # Mostrar qué tiene actualmente
        es_formula = isinstance(contenido_c, str) and contenido_c.startswith('=')

        print(f"Fila {fila}:")
        print(f"   B{fila}: {nombre_b}")
        print(f"   C{fila}: {contenido_c}")

        if es_formula:
            # Intentar evaluar la fórmula
            valor_evaluado = ws_efectivo_data.cell(fila, 3).value
            print(f"   Evaluado: {valor_evaluado}")
            print(f"   ⚠️  ES UNA FÓRMULA - necesita corrección")
        else:
            print(f"   ✅ Es texto fijo")

        print()

    # =========================================================================
    # APLICAR CORRECCIONES
    # =========================================================================
    print("=" * 80)
    print("✏️  APLICANDO CORRECCIONES...")
    print("=" * 80)
    print()

    # Estrategia: Leer el Balance Inicial de cada cuenta en TRANSACCIONES
    # para determinar el nombre estándar actual

    # Buscar columnas en TRANSACCIONES
    headers_trans = [ws_trans.cell(1, col).value for col in range(1, ws_trans.max_column + 1)]
    col_map = {}
    for col in range(1, len(headers_trans) + 1):
        if headers_trans[col-1]:
            col_map[headers_trans[col-1]] = col

    # Leer las primeras filas de TRANSACCIONES para obtener cuentas
    print("🔍 Detectando cuentas desde TRANSACCIONES (Balance Inicial)...")
    print()

    cuentas_detectadas = {}

    # Filas 2-9 de TRANSACCIONES contienen los balances iniciales
    for row_trans in range(2, 10):
        tipo = ws_trans.cell(row_trans, col_map['Tipo Transacción']).value
        cuenta = ws_trans.cell(row_trans, col_map['Cuenta Bancaria']).value
        concepto = ws_trans.cell(row_trans, col_map['Concepto']).value

        if tipo and 'Apertura Inicial' in str(tipo):
            # Esta es una cuenta
            fila_efectivo = row_trans + 1  # Fila 2 TRANS → Fila 3 EFECTIVO

            if fila_efectivo >= 3 and fila_efectivo <= 10:
                cuentas_detectadas[fila_efectivo] = str(cuenta).strip() if cuenta else None
                print(f"   Fila Efectivo {fila_efectivo}: {cuenta}")

    print()

    # Si no detectamos todas, usar el mapeo manual
    print("📋 Usando mapeo de cuentas...")
    print()

    cambios_aplicados = 0

    for fila in range(3, 11):
        celda_c = ws_efectivo.cell(fila, 3)
        contenido_actual = celda_c.value

        es_formula = isinstance(contenido_actual, str) and contenido_actual.startswith('=')

        # Determinar cuenta estándar
        if fila in cuentas_detectadas and cuentas_detectadas[fila]:
            cuenta_estandar = cuentas_detectadas[fila]
        elif fila in CUENTAS_ESTANDAR:
            cuenta_estandar = CUENTAS_ESTANDAR[fila]
        else:
            print(f"⚠️  Fila {fila}: No se pudo determinar cuenta estándar - SALTANDO")
            continue

        # Si es fórmula o está vacío, cambiar a texto fijo
        if es_formula or not contenido_actual:
            print(f"✅ Fila {fila}: Cambiando a \"{cuenta_estandar}\"")
            celda_c.value = cuenta_estandar
            cambios_aplicados += 1
        else:
            # Verificar si el texto actual coincide con el estándar
            if str(contenido_actual).strip() != cuenta_estandar:
                print(f"⚠️  Fila {fila}: Actualizando \"{contenido_actual}\" → \"{cuenta_estandar}\"")
                celda_c.value = cuenta_estandar
                cambios_aplicados += 1
            else:
                print(f"✅ Fila {fila}: Ya tiene el valor correcto")

    print()
    print(f"📊 Cambios aplicados: {cambios_aplicados}")
    print()

    # =========================================================================
    # GUARDAR
    # =========================================================================
    print("=" * 80)
    print("💾 Guardando cambios...")
    print("=" * 80)
    print()

    wb.save(EXCEL_FILE)
    print("✅ Excel actualizado")
    print()

    # =========================================================================
    # VERIFICACIÓN
    # =========================================================================
    print("=" * 80)
    print("📊 VERIFICACIÓN FINAL")
    print("=" * 80)
    print()

    # Recargar y verificar
    wb_verif = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws_efectivo_verif = wb_verif['Efectivo']

    print("✅ Columna C después de corrección:")
    print()

    for fila in range(3, 11):
        nombre = ws_efectivo_verif.cell(fila, 2).value
        cuenta = ws_efectivo_verif.cell(fila, 3).value
        ingreso = ws_efectivo_verif.cell(fila, 4).value
        egreso = ws_efectivo_verif.cell(fila, 5).value
        balance = ws_efectivo_verif.cell(fila, 6).value

        print(f"Fila {fila}: {nombre}")
        print(f"   Cuenta: {cuenta}")
        print(f"   Ingreso: ${float(ingreso):,.2f}" if ingreso else "   Ingreso: $0.00")
        print(f"   Egreso: ${float(egreso):,.2f}" if egreso else "   Egreso: $0.00")
        print(f"   Balance: ${float(balance):,.2f}" if balance else "   Balance: $0.00")
        print()

    # Mostrar balance Promerica
    balance_promerica = ws_efectivo_verif.cell(3, 6).value

    print("=" * 80)
    print("💰 BALANCE PROMERICA (Fila 3)")
    print("=" * 80)
    print()

    if balance_promerica:
        balance_val = float(balance_promerica)
        print(f"   Balance: ${balance_val:,.2f}")
        print(f"   Esperado: $2,163.44")
        diferencia = abs(balance_val - 2163.44)
        print(f"   Diferencia: ${diferencia:,.2f}")
        print()

        if diferencia < 1.00:
            print("✅ ¡BALANCE CORRECTO!")
        else:
            print("⚠️  Balance aún no coincide")
            print()
            print("📋 POSIBLES CAUSAS:")
            print("   1. Excel necesita recalcular (cierra y abre el archivo)")
            print("   2. Hay transacciones incorrectas o duplicadas")
            print("   3. Saldo inicial incorrecto")
    else:
        print("⚠️  Balance no disponible (Excel necesita recalcular)")
        print()
        print("🔧 ACCIÓN REQUERIDA:")
        print("   1. Cierra completamente el Excel")
        print("   2. Vuelve a abrirlo")
        print("   3. Excel recalculará automáticamente las fórmulas")
        print("   4. Verifica que el balance Promerica sea $2,163.44")

    print()

    # =========================================================================
    # RESUMEN
    # =========================================================================
    print("=" * 80)
    print("📊 RESUMEN")
    print("=" * 80)
    print()

    print(f"✅ Celdas corregidas: {cambios_aplicados}")
    print()
    print("🔧 PRÓXIMOS PASOS:")
    print("   1. Cierra completamente el Excel")
    print("   2. Vuelve a abrirlo (para recalcular fórmulas)")
    print("   3. Ve a hoja Efectivo, fila 3")
    print("   4. Verifica que el Balance sea $2,163.44")
    print()
    print("   Si el balance sigue incorrecto, ejecuta:")
    print("   python scripts/diagnosticar_formula_efectivo_promerica.py")

    print()
    print("=" * 80)
    print("✅ CORRECCIÓN COMPLETADA")
    print("=" * 80)
    print()

if __name__ == "__main__":
    try:
        if not crear_backup():
            print("❌ Abortando")
            exit(1)

        corregir()
        print("🎉 Proceso completado exitosamente!")

    except Exception as e:
        print(f"❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
