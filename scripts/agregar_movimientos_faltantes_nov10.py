#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Agregar movimientos faltantes al Excel V.20
- 16 movimientos Promerica USD 1774 (30/10 al 08/11)
- 1 movimiento TC BNCR MC 8759 (10/11)
Total: 17 movimientos
"""
import openpyxl
from datetime import datetime
import sys
import os

# Configuración
EXCEL_FILE = "AlvaroVelasco_Finanzas_v2.0.xlsx"

print("="*80)
print("AGREGANDO MOVIMIENTOS FALTANTES - NOVIEMBRE 10, 2025")
print("="*80)

# Verificar que el archivo existe
if not os.path.exists(EXCEL_FILE):
    print(f"\n ERROR: No se encontro el archivo {EXCEL_FILE}")
    sys.exit(1)

# Movimientos a agregar
movimientos = []

# ============================================================================
# PROMERICA USD 1774 - MOVIMIENTOS FALTANTES
# ============================================================================

# Comisiones bancarias no registradas (12 movimientos)
movimientos.extend([
    # 21/10/2025
    {"fecha": "21/10/2025", "tipo": "Comision Bancaria", "categoria": "Comisiones", "entidad": "Banco Promerica", "cuenta": "Promerica USD 1774", "proveedor": "Banco Promerica", "concepto": "Comision TFT fact 199488", "referencia": "595744-COM", "monto_usd": 3.00, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "BAJA", "notas": "Comision transferencia"},
    {"fecha": "21/10/2025", "tipo": "Comision Bancaria", "categoria": "Comisiones", "entidad": "Banco Promerica", "cuenta": "Promerica USD 1774", "proveedor": "Banco Promerica", "concepto": "Comision TFT alquiler Casa 10E", "referencia": "596247-COM", "monto_usd": 3.00, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "BAJA", "notas": "Comision transferencia"},

    # 22/10/2025
    {"fecha": "22/10/2025", "tipo": "Comision Bancaria", "categoria": "Comisiones", "entidad": "Banco Promerica", "cuenta": "Promerica USD 1774", "proveedor": "Banco Promerica", "concepto": "Comision TFT fact 2502060 INTCOMEX", "referencia": "597597-COM", "monto_usd": 3.00, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "BAJA", "notas": "Comision transferencia"},

    # 24/10/2025
    {"fecha": "24/10/2025", "tipo": "Comision Bancaria", "categoria": "Comisiones", "entidad": "Banco Promerica", "cuenta": "Promerica USD 1774", "proveedor": "Banco Promerica", "concepto": "Comision debito SINPE SEA Global", "referencia": "2785908", "monto_usd": 0.75, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "BAJA", "notas": "Comision debito directo"},

    # 29/10/2025
    {"fecha": "29/10/2025", "tipo": "Comision Bancaria", "categoria": "Comisiones", "entidad": "Banco Promerica", "cuenta": "Promerica USD 1774", "proveedor": "Banco Promerica", "concepto": "Comision TFT pago tarjeta", "referencia": "611261-COM", "monto_usd": 3.00, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "BAJA", "notas": "Comision transferencia"},
    {"fecha": "29/10/2025", "tipo": "Comision Bancaria", "categoria": "Comisiones", "entidad": "Banco Promerica", "cuenta": "Promerica USD 1774", "proveedor": "Banco Promerica", "concepto": "Comision TFT pago tarjeta", "referencia": "611329-COM", "monto_usd": 3.00, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "BAJA", "notas": "Comision transferencia"},
    {"fecha": "29/10/2025", "tipo": "Comision Bancaria", "categoria": "Comisiones", "entidad": "Banco Promerica", "cuenta": "Promerica USD 1774", "proveedor": "Banco Promerica", "concepto": "Comision TFT Curso Pricing", "referencia": "611345-COM", "monto_usd": 3.00, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "BAJA", "notas": "Comision transferencia"},
    {"fecha": "29/10/2025", "tipo": "Comision Bancaria", "categoria": "Comisiones", "entidad": "Banco Promerica", "cuenta": "Promerica USD 1774", "proveedor": "Banco Promerica", "concepto": "Comision debito SINPE CarroFacil", "referencia": "2788366", "monto_usd": 0.75, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "BAJA", "notas": "Comision debito directo"},

    # 06/11/2025
    {"fecha": "06/11/2025", "tipo": "Comision Bancaria", "categoria": "Comisiones", "entidad": "Banco Promerica", "cuenta": "Promerica USD 1774", "proveedor": "Banco Promerica", "concepto": "Comision debito SINPE SEA Global", "referencia": "2795229", "monto_usd": 0.75, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "BAJA", "notas": "Comision debito directo"},

    # 07/11/2025
    {"fecha": "07/11/2025", "tipo": "Comision Bancaria", "categoria": "Comisiones", "entidad": "Banco Promerica", "cuenta": "Promerica USD 1774", "proveedor": "Banco Promerica", "concepto": "Comision debito SINPE", "referencia": "2796348-1", "monto_usd": 0.75, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "BAJA", "notas": "Comision debito directo"},
    {"fecha": "07/11/2025", "tipo": "Ingreso", "categoria": "Ajustes", "entidad": "Banco Promerica", "cuenta": "Promerica USD 1774", "proveedor": "Banco Promerica", "concepto": "Reversion de comision", "referencia": "2796348-REV", "monto_usd": 0.75, "monto_crc": 0, "tipo_mov": "Ingreso", "estado": "Completado", "prioridad": "BAJA", "notas": "Reversion comision"},
    {"fecha": "07/11/2025", "tipo": "Comision Bancaria", "categoria": "Comisiones", "entidad": "Banco Promerica", "cuenta": "Promerica USD 1774", "proveedor": "Banco Promerica", "concepto": "Comision debito SINPE Alejandra", "referencia": "2796348-2", "monto_usd": 0.75, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "BAJA", "notas": "Comision debito directo"},
])

# Otros movimientos operativos no registrados (4 movimientos)
movimientos.extend([
    # 30/10/2025
    {"fecha": "30/10/2025", "tipo": "Gasto", "categoria": "Combustible", "entidad": "Unopetrol", "cuenta": "Promerica USD 1774", "proveedor": "Unopetrol Barreal", "concepto": "Combustible vehiculo empresa", "referencia": "754410", "monto_usd": 63.03, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "MEDIA", "notas": "Gasto operativo combustible"},

    # 06/11/2025
    {"fecha": "06/11/2025", "tipo": "Pago Proveedor", "categoria": "Logistica", "entidad": "SEA Global", "cuenta": "Promerica USD 1774", "proveedor": "SEA Global Logistics", "concepto": "Pago SEA Global servicios", "referencia": "32048171", "monto_usd": 58.76, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "ALTA", "notas": "DD debito directo SEA Global"},

    # 07/11/2025
    {"fecha": "07/11/2025", "tipo": "Pago", "categoria": "Servicios Administrativos", "entidad": "Alejandra Arias", "cuenta": "Promerica USD 1774", "proveedor": "Alejandra Arias Fallas", "concepto": "Pago semanal Alejandra Arias", "referencia": "32067344", "monto_usd": 40.57, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "ALTA", "notas": "DD debito directo pago semanal"},

    # 08/11/2025
    {"fecha": "08/11/2025", "tipo": "Gasto", "categoria": "Combustible", "entidad": "NASA Heredia", "cuenta": "Promerica USD 1774", "proveedor": "Estacion Servicios NASA", "concepto": "Combustible vehiculo empresa", "referencia": "141111", "monto_usd": 9.13, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "MEDIA", "notas": "Gasto operativo combustible"},
])

# ============================================================================
# TC BNCR MC 8759 - MOVIMIENTO NUEVO 10/11/2025
# ============================================================================

movimientos.extend([
    {"fecha": "10/11/2025", "tipo": "Compra", "categoria": "Inventario", "entidad": "Amazon", "cuenta": "TC BNCR MC 8759", "proveedor": "Amazon Marketplace", "concepto": "Compra inventario para reventa Amazon", "referencia": "111023129834", "monto_usd": 112.66, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "ALTA", "notas": "Compra mercancia reventa - Aut: 212015"},
])

print(f"\nTotal movimientos preparados: {len(movimientos)}")
print(f"  - Promerica comisiones: 12 movimientos")
print(f"  - Promerica operativos: 4 movimientos")
print(f"  - TC MC 8759 Amazon: 1 movimiento")

# Cargar el Excel
print("\n Cargando Excel V.20...")
wb = openpyxl.load_workbook(EXCEL_FILE)
ws = wb['TRANSACCIONES']

# Obtener último ID y última fila
last_id = 0
last_row = 1
print("\n Analizando Excel existente...")

for row in range(2, ws.max_row + 1):
    id_val = ws[f'P{row}'].value
    if id_val:
        try:
            last_id = max(last_id, int(id_val))
            last_row = row
        except:
            pass

print(f"   Ultimo ID encontrado: {last_id}")
print(f"   Ultima fila con datos: {last_row}")

# Crear conjunto de referencias existentes para evitar duplicados
referencias_existentes = set()
for row in range(2, ws.max_row + 1):
    ref = ws[f'H{row}'].value
    fecha = ws[f'A{row}'].value
    if ref and fecha:
        try:
            fecha_str = fecha.strftime('%Y-%m-%d') if hasattr(fecha, 'strftime') else str(fecha)
            referencias_existentes.add(f"{fecha_str}_{ref}")
        except:
            pass

print(f"   Referencias unicas existentes: {len(referencias_existentes)}")

# Filtrar movimientos que no existen
movimientos_nuevos = []
movimientos_duplicados = []

print("\n Verificando duplicados...")
for mov in movimientos:
    try:
        fecha_obj = datetime.strptime(mov['fecha'], '%d/%m/%Y')
        fecha_str = fecha_obj.strftime('%Y-%m-%d')
        clave = f"{fecha_str}_{mov['referencia']}"

        if clave in referencias_existentes:
            movimientos_duplicados.append(mov)
        else:
            movimientos_nuevos.append(mov)
    except:
        # Si hay error, agregar de todos modos
        movimientos_nuevos.append(mov)

print(f"   Movimientos nuevos a agregar: {len(movimientos_nuevos)}")
print(f"   Movimientos duplicados (omitidos): {len(movimientos_duplicados)}")

if len(movimientos_duplicados) > 0:
    print("\n   DUPLICADOS ENCONTRADOS:")
    for mov in movimientos_duplicados:
        print(f"      - {mov['fecha']} | {mov['referencia']} | {mov['concepto'][:40]}")

if len(movimientos_nuevos) == 0:
    print("\n No hay movimientos nuevos que agregar.")
    print("   Todos los movimientos ya existen en el Excel.")
    sys.exit(0)

# Agregar movimientos nuevos
print(f"\n Agregando {len(movimientos_nuevos)} movimientos nuevos...")

next_id = last_id + 1
next_row = last_row + 1

movimientos_agregados = []

for mov in movimientos_nuevos:
    try:
        # Convertir fecha
        fecha_obj = datetime.strptime(mov['fecha'], '%d/%m/%Y')

        # Agregar en la siguiente fila
        ws[f'A{next_row}'] = fecha_obj
        ws[f'B{next_row}'] = mov['tipo']
        ws[f'C{next_row}'] = mov['categoria']
        ws[f'D{next_row}'] = mov['entidad']
        ws[f'E{next_row}'] = mov['cuenta']
        ws[f'F{next_row}'] = mov['proveedor']
        ws[f'G{next_row}'] = mov['concepto']
        ws[f'H{next_row}'] = mov['referencia']
        ws[f'I{next_row}'] = mov['monto_usd']
        ws[f'J{next_row}'] = mov['monto_crc']
        ws[f'K{next_row}'] = mov['tipo_mov']
        ws[f'L{next_row}'] = mov['estado']
        ws[f'M{next_row}'] = mov.get('prioridad', 'MEDIA')
        ws[f'N{next_row}'] = ''  # Vencimiento vacío
        ws[f'O{next_row}'] = mov.get('notas', '')
        ws[f'P{next_row}'] = next_id
        ws[f'Q{next_row}'] = datetime.now()
        ws[f'R{next_row}'] = 'Alvaro Velasco'
        ws[f'S{next_row}'] = False  # Duplicado
        ws[f'T{next_row}'] = 'OK'  # Validación

        movimientos_agregados.append({
            'id': next_id,
            'fila': next_row,
            'fecha': mov['fecha'],
            'cuenta': mov['cuenta'],
            'concepto': mov['concepto'],
            'monto_usd': mov['monto_usd']
        })

        next_row += 1
        next_id += 1

    except Exception as e:
        print(f"   Advertencia: Error al agregar movimiento: {mov.get('concepto', 'desconocido')} - {e}")

# Guardar archivo
print(f"\n Guardando archivo {EXCEL_FILE}...")
wb.save(EXCEL_FILE)

print("\n" + "="*80)
print("MOVIMIENTOS AGREGADOS EXITOSAMENTE")
print("="*80)

print(f"\n RESUMEN:")
print(f"   Total movimientos procesados: {len(movimientos)}")
print(f"   Movimientos nuevos agregados: {len(movimientos_agregados)}")
print(f"   Movimientos duplicados omitidos: {len(movimientos_duplicados)}")

# Calcular totales
total_ingresos = sum([m['monto_usd'] for m in movimientos_agregados if 'Ingreso' in movimientos_nuevos[movimientos_agregados.index(m)].get('tipo_mov', '')])
total_egresos = sum([m['monto_usd'] for m in movimientos_agregados if 'Egreso' in movimientos_nuevos[movimientos_agregados.index(m)].get('tipo_mov', '')])

print(f"\n IMPACTO FINANCIERO:")
print(f"   Total ingresos agregados: ${total_ingresos:,.2f}")
print(f"   Total egresos agregados: ${total_egresos:,.2f}")
print(f"   Movimiento neto: ${total_ingresos - total_egresos:,.2f}")

# Agrupar por cuenta
por_cuenta = {}
for mov in movimientos_agregados:
    cuenta = mov['cuenta']
    if cuenta not in por_cuenta:
        por_cuenta[cuenta] = []
    por_cuenta[cuenta].append(mov)

print(f"\n MOVIMIENTOS POR CUENTA:")
for cuenta, movs in por_cuenta.items():
    total_cuenta = sum([m['monto_usd'] for m in movs])
    print(f"\n   {cuenta}: {len(movs)} movimientos")
    for mov in movs[:10]:  # Primeros 10
        signo = "+" if mov['monto_usd'] > 0 else "-"
        print(f"      Fila {mov['fila']} | {mov['fecha']} | {signo}${abs(mov['monto_usd']):>7.2f} | {mov['concepto'][:45]}")
    if len(movs) > 10:
        print(f"      ... y {len(movs) - 10} movimientos mas")

print("\n" + "="*80)
print("PROCESO COMPLETADO")
print("="*80)
print(f"\n Siguiente paso: Abrir {EXCEL_FILE} y verificar los movimientos")
print(" Los saldos de Promerica y TC MC 8759 fueron actualizados")
print("\n" + "="*80)
