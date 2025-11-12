#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ANÁLISIS COMPLETO DE MÉTRICAS - Excel v2.0
Genera reporte agregado sin exponer transacciones individuales
"""
import openpyxl
from datetime import datetime
from collections import defaultdict
import statistics
import os

# Buscar el archivo Excel en múltiples ubicaciones
EXCEL_NAMES = [
    "AlvaroVelasco_Finanzas_v2.0.xlsx",
    "AlvaroVelasco_Finanzas_v2.0.xls",
    "../AlvaroVelasco_Finanzas_v2.0.xlsx",
    "../AlvaroVelasco_Finanzas_v2.0.xls"
]

EXCEL_FILE = None
for name in EXCEL_NAMES:
    if os.path.exists(name):
        EXCEL_FILE = name
        break

if not EXCEL_FILE:
    print("❌ ERROR: No se encontró el archivo Excel")
    print("Buscando:")
    for name in EXCEL_NAMES:
        print(f"  - {name}")
    print()
    print("Coloca el archivo Excel en la misma carpeta que este script")
    print("o en la carpeta padre")
    exit(1)

def analizar_metricas():
    print("=" * 80)
    print("📊 ANÁLISIS DE MÉTRICAS FINANCIERAS - Excel v2.0")
    print("=" * 80)
    print()

    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        ws = wb['TRANSACCIONES']

        # Encontrar columnas
        headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
        col_map = {}
        for col in range(1, len(headers) + 1):
            if headers[col-1]:
                col_map[headers[col-1]] = col

        print(f"✅ Excel cargado: {ws.max_row - 1} transacciones encontradas")
        print()
        print(f"📋 Columnas detectadas: {', '.join(headers[:10])}...")
        print()

        # ==================================================================
        # ESTRUCTURAS DE DATOS
        # ==================================================================
        ingresos_por_mes = defaultdict(float)
        egresos_por_mes = defaultdict(float)
        ingresos_por_cliente = defaultdict(float)
        compras_por_proveedor = defaultdict(float)
        transacciones_por_categoria = defaultdict(int)
        ingresos_por_categoria = defaultdict(float)
        egresos_por_categoria = defaultdict(float)
        saldos_por_cuenta = defaultdict(float)
        efectivo_por_fecha = []

        # ==================================================================
        # LECTURA DE TRANSACCIONES
        # ==================================================================
        print("=" * 80)
        print("🔍 ANALIZANDO TRANSACCIONES...")
        print("=" * 80)
        print()

        for row in range(2, ws.max_row + 1):
            # Extraer datos
            fecha = ws.cell(row, col_map.get('Fecha', 1)).value
            tipo = ws.cell(row, col_map.get('Tipo Transacción', 2)).value
            categoria = ws.cell(row, col_map.get('Categoría', 3)).value
            cuenta = ws.cell(row, col_map.get('Cuenta Bancaria', 4)).value
            entidad = ws.cell(row, col_map.get('Cliente/Proveedor', 5)).value
            monto_usd = ws.cell(row, col_map.get('Monto USD', 6)).value
            ing_egr = ws.cell(row, col_map.get('Ingreso/Egreso', 7)).value

            # Validar datos
            if not monto_usd or not fecha:
                continue

            try:
                monto = float(monto_usd)
            except:
                continue

            # Convertir fecha a mes
            if isinstance(fecha, datetime):
                mes = fecha.strftime('%Y-%m')
            else:
                try:
                    if '/' in str(fecha):
                        partes = str(fecha).split('/')
                        if len(partes) == 3:
                            mes = f"20{partes[2]}-{partes[1].zfill(2)}" if len(partes[2]) == 2 else f"{partes[2]}-{partes[1].zfill(2)}"
                        else:
                            mes = "2024-00"
                    else:
                        mes = "2024-00"
                except:
                    mes = "2024-00"

            # INGRESOS POR MES
            if ing_egr == 'Ingreso' and monto > 0:
                ingresos_por_mes[mes] += monto

            # EGRESOS POR MES
            elif ing_egr == 'Egreso' and monto < 0:
                egresos_por_mes[mes] += abs(monto)

            # INGRESOS POR CLIENTE
            if ing_egr == 'Ingreso' and monto > 0 and entidad:
                cliente = str(entidad).strip()
                ingresos_por_cliente[cliente] += monto

            # COMPRAS POR PROVEEDOR
            if ing_egr == 'Egreso' and monto < 0 and entidad:
                proveedor = str(entidad).strip()
                compras_por_proveedor[proveedor] += abs(monto)

            # POR CATEGORÍA
            if categoria:
                cat = str(categoria).strip()
                transacciones_por_categoria[cat] += 1
                if ing_egr == 'Ingreso' and monto > 0:
                    ingresos_por_categoria[cat] += monto
                elif ing_egr == 'Egreso' and monto < 0:
                    egresos_por_categoria[cat] += abs(monto)

            # SALDOS POR CUENTA
            if cuenta:
                cta = str(cuenta).strip()
                saldos_por_cuenta[cta] += monto

        # ==================================================================
        # 1. ESTACIONALIDAD
        # ==================================================================
        print("=" * 80)
        print("📅 1. ESTACIONALIDAD - INGRESOS Y EGRESOS POR MES")
        print("=" * 80)
        print()

        meses_ordenados = sorted([m for m in ingresos_por_mes.keys() if m != "2024-00"])

        print(f"{'MES':<12} {'INGRESOS':>12} {'EGRESOS':>12} {'NETO':>12}")
        print("-" * 50)

        for mes in meses_ordenados:
            ing = ingresos_por_mes.get(mes, 0)
            egr = egresos_por_mes.get(mes, 0)
            neto = ing - egr
            print(f"{mes:<12} ${ing:>10,.2f} ${egr:>10,.2f} ${neto:>10,.2f}")

        print()

        if len(meses_ordenados) > 0:
            mejor_mes = max(meses_ordenados, key=lambda m: ingresos_por_mes[m])
            peor_mes = min(meses_ordenados, key=lambda m: ingresos_por_mes[m])

            print(f"🏆 MEJOR MES: {mejor_mes} - ${ingresos_por_mes[mejor_mes]:,.2f}")
            print(f"📉 PEOR MES: {peor_mes} - ${ingresos_por_mes[peor_mes]:,.2f}")

            promedio_ingresos = statistics.mean([ingresos_por_mes[m] for m in meses_ordenados])
            promedio_egresos = statistics.mean([egresos_por_mes[m] for m in meses_ordenados])

            print(f"📊 PROMEDIO MENSUAL INGRESOS: ${promedio_ingresos:,.2f}")
            print(f"📊 PROMEDIO MENSUAL EGRESOS: ${promedio_egresos:,.2f}")
            print(f"📊 PROMEDIO MENSUAL NETO: ${promedio_ingresos - promedio_egresos:,.2f}")

        print()

        # ==================================================================
        # 2. TOP 10 CLIENTES
        # ==================================================================
        print("=" * 80)
        print("👥 2. TOP 10 CLIENTES POR INGRESOS")
        print("=" * 80)
        print()

        top_clientes = sorted(ingresos_por_cliente.items(), key=lambda x: x[1], reverse=True)[:10]
        total_ingresos = sum(ingresos_por_cliente.values())

        print(f"{'#':<4} {'CLIENTE':<40} {'INGRESOS':>12} {'%':>8}")
        print("-" * 68)

        for idx, (cliente, monto) in enumerate(top_clientes, 1):
            porcentaje = (monto / total_ingresos * 100) if total_ingresos > 0 else 0
            cliente_corto = cliente[:37] + "..." if len(cliente) > 40 else cliente
            print(f"{idx:<4} {cliente_corto:<40} ${monto:>10,.2f} {porcentaje:>6.1f}%")

        print()
        print(f"💰 TOTAL INGRESOS: ${total_ingresos:,.2f}")

        if len(top_clientes) >= 3:
            top3_total = sum([monto for _, monto in top_clientes[:3]])
            top3_pct = (top3_total / total_ingresos * 100) if total_ingresos > 0 else 0
            print(f"⚠️  CONCENTRACIÓN TOP 3: {top3_pct:.1f}% del total")

        print()

        # ==================================================================
        # 3. TOP 10 PROVEEDORES
        # ==================================================================
        print("=" * 80)
        print("🏭 3. TOP 10 PROVEEDORES POR COMPRAS")
        print("=" * 80)
        print()

        top_proveedores = sorted(compras_por_proveedor.items(), key=lambda x: x[1], reverse=True)[:10]
        total_compras = sum(compras_por_proveedor.values())

        print(f"{'#':<4} {'PROVEEDOR':<40} {'COMPRAS':>12} {'%':>8}")
        print("-" * 68)

        for idx, (proveedor, monto) in enumerate(top_proveedores, 1):
            porcentaje = (monto / total_compras * 100) if total_compras > 0 else 0
            proveedor_corto = proveedor[:37] + "..." if len(proveedor) > 40 else proveedor
            print(f"{idx:<4} {proveedor_corto:<40} ${monto:>10,.2f} {porcentaje:>6.1f}%")

        print()
        print(f"💰 TOTAL COMPRAS: ${total_compras:,.2f}")
        print()

        # ==================================================================
        # 4. DISTRIBUCIÓN POR CATEGORÍA
        # ==================================================================
        print("=" * 80)
        print("📂 4. DISTRIBUCIÓN POR CATEGORÍA")
        print("=" * 80)
        print()

        print("INGRESOS POR CATEGORÍA:")
        print(f"{'CATEGORÍA':<40} {'MONTO':>12} {'%':>8}")
        print("-" * 62)

        top_cat_ing = sorted(ingresos_por_categoria.items(), key=lambda x: x[1], reverse=True)
        total_ing_cat = sum(ingresos_por_categoria.values())

        for cat, monto in top_cat_ing:
            porcentaje = (monto / total_ing_cat * 100) if total_ing_cat > 0 else 0
            cat_corto = cat[:37] + "..." if len(cat) > 40 else cat
            print(f"{cat_corto:<40} ${monto:>10,.2f} {porcentaje:>6.1f}%")

        print()
        print("EGRESOS POR CATEGORÍA:")
        print(f"{'CATEGORÍA':<40} {'MONTO':>12} {'%':>8}")
        print("-" * 62)

        top_cat_egr = sorted(egresos_por_categoria.items(), key=lambda x: x[1], reverse=True)
        total_egr_cat = sum(egresos_por_categoria.values())

        for cat, monto in top_cat_egr:
            porcentaje = (monto / total_egr_cat * 100) if total_egr_cat > 0 else 0
            cat_corto = cat[:37] + "..." if len(cat) > 40 else cat
            print(f"{cat_corto:<40} ${monto:>10,.2f} {porcentaje:>6.1f}%")

        print()

        # ==================================================================
        # 5. SALDOS POR CUENTA
        # ==================================================================
        print("=" * 80)
        print("🏦 5. SALDOS POR CUENTA BANCARIA")
        print("=" * 80)
        print()

        print(f"{'CUENTA':<50} {'BALANCE':>12}")
        print("-" * 64)

        for cuenta, saldo in sorted(saldos_por_cuenta.items(), key=lambda x: x[1], reverse=True):
            cuenta_corta = cuenta[:47] + "..." if len(cuenta) > 50 else cuenta
            print(f"{cuenta_corta:<50} ${saldo:>10,.2f}")

        print()
        total_efectivo = sum(saldos_por_cuenta.values())
        print(f"💰 TOTAL EFECTIVO CALCULADO: ${total_efectivo:,.2f}")
        print()

        # ==================================================================
        # 6. ANÁLISIS DE CUENTAS POR COBRAR (si existe la columna)
        # ==================================================================
        if 'Estado' in col_map or 'Status' in col_map:
            print("=" * 80)
            print("📋 6. CUENTAS POR COBRAR")
            print("=" * 80)
            print()

            cxc_pendiente = 0
            cxc_count = 0

            for row in range(2, ws.max_row + 1):
                estado = ws.cell(row, col_map.get('Estado', col_map.get('Status', 999))).value
                ing_egr = ws.cell(row, col_map.get('Ingreso/Egreso', 7)).value
                monto = ws.cell(row, col_map.get('Monto USD', 6)).value

                if estado and str(estado).strip().lower() == 'pendiente' and ing_egr == 'Ingreso':
                    try:
                        monto_val = float(monto)
                        if monto_val > 0:
                            cxc_pendiente += monto_val
                            cxc_count += 1
                    except:
                        pass

            print(f"📊 Total CxC Pendiente: ${cxc_pendiente:,.2f}")
            print(f"📊 Facturas pendientes: {cxc_count}")
            print()

        # ==================================================================
        # 7. RESUMEN EJECUTIVO
        # ==================================================================
        print("=" * 80)
        print("📊 7. RESUMEN EJECUTIVO")
        print("=" * 80)
        print()

        print(f"Total Transacciones: {ws.max_row - 1}")
        print(f"Total Ingresos: ${total_ingresos:,.2f}")
        print(f"Total Egresos: ${total_compras:,.2f}")
        print(f"Utilidad Neta: ${total_ingresos - total_compras:,.2f}")
        print()

        if total_compras > 0:
            margen_bruto = ((total_ingresos - total_compras) / total_ingresos * 100)
            print(f"Margen Bruto Promedio: {margen_bruto:.1f}%")

        print()
        print(f"Clientes únicos: {len(ingresos_por_cliente)}")
        print(f"Proveedores únicos: {len(compras_por_proveedor)}")
        print(f"Categorías de ingreso: {len(ingresos_por_categoria)}")
        print(f"Categorías de egreso: {len(egresos_por_categoria)}")
        print()

        # ==================================================================
        # FINALIZACIÓN
        # ==================================================================
        print("=" * 80)
        print("✅ ANÁLISIS COMPLETADO")
        print("=" * 80)
        print()
        print("👉 Copia TODO este reporte y envíaselo a Claude")
        print("   Esto le permitirá reducir significativamente las preguntas")
        print()

    except FileNotFoundError:
        print(f"❌ ERROR: No se encontró el archivo '{EXCEL_FILE}'")
        print()
        print("Asegúrate de que:")
        print("1. El archivo existe en el directorio actual")
        print("2. El nombre es exactamente 'AlvaroVelasco_Finanzas_v2.0.xlsx'")
        print()
        print("O edita la variable EXCEL_FILE en la línea 10 del script")

    except Exception as e:
        print(f"❌ ERROR INESPERADO: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    analizar_metricas()
