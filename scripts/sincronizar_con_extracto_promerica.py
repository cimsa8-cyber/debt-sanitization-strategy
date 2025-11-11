#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CORRECCIÓN TRANSACCIONES SEGÚN EXTRACTO PROMERICA
Compara con extracto bancario 10/11/2025 y corrige:
1. Elimina Delta en CRC (fila 209 - error)
2. Registra Delta en USD ($76.28 desde Promerica)
3. Registra ingreso Waipio ($1,459.96)
4. Registra comisión SINPE ($0.75)
"""
import openpyxl
from datetime import datetime
import shutil

EXCEL_FILE = "AlvaroVelasco_Finanzas_v2.0.xlsx"
BACKUP_FILE = f"AlvaroVelasco_Finanzas_v2.0_EXTRACTO_BACKUP_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

def crear_backup():
    print("=" * 80)
    print("CREANDO BACKUP")
    print("=" * 80)
    print(f"Backup: {BACKUP_FILE}")
    try:
        shutil.copy2(EXCEL_FILE, BACKUP_FILE)
        print("✅ Backup creado")
        print()
        return True
    except Exception as e:
        print(f"❌ ERROR: {e}")
        return False

def corregir_segun_extracto():
    print("=" * 80)
    print("CORRECCIÓN SEGÚN EXTRACTO BANCARIO PROMERICA")
    print("Fecha: 10/11/2025")
    print("=" * 80)
    print()

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb['TRANSACCIONES']

    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    col_map = {}
    for col in range(1, len(headers) + 1):
        if headers[col-1]:
            col_map[headers[col-1]] = col

    # =========================================================================
    # PASO 1: ELIMINAR DELTA EN CRC (FILA 209 - ERROR)
    # =========================================================================
    print("📋 PASO 1: Eliminando registro incorrecto de Delta en CRC...")
    print()

    # Buscar Delta en CRC
    delta_crc_fila = None
    for row in range(200, ws.max_row + 1):  # Buscar en las últimas filas
        proveedor = ws.cell(row, col_map['Cliente/Proveedor']).value
        monto_crc = ws.cell(row, col_map['Monto CRC']).value

        if proveedor and 'Delta' in str(proveedor) and monto_crc:
            try:
                if abs(float(monto_crc)) > 30000:  # ₡37,606
                    delta_crc_fila = row
                    break
            except:
                pass

    if delta_crc_fila:
        concepto = ws.cell(delta_crc_fila, col_map['Concepto']).value
        monto = ws.cell(delta_crc_fila, col_map['Monto CRC']).value

        print(f"✅ Registro erróneo encontrado:")
        print(f"   Fila: {delta_crc_fila}")
        print(f"   Concepto: {concepto}")
        print(f"   Monto: ₡{abs(float(monto)):,.2f} CRC")
        print(f"   Motivo: Delta se pagó en USD desde Promerica, no efectivo CRC")
        print()

        ws.delete_rows(delta_crc_fila, 1)
        print("✅ Fila eliminada")
        print()
    else:
        print("⚠️  No se encontró registro Delta en CRC")
        print()

    # =========================================================================
    # PASO 2: REGISTRAR TRANSACCIONES FALTANTES
    # =========================================================================
    print("=" * 80)
    print("📋 PASO 2: Registrando transacciones faltantes...")
    print()

    transacciones = [
        {
            'num': 1,
            'descripcion': 'Ingreso Waipio S.A.',
            'fecha': datetime(2025, 11, 10, 0, 0, 0),
            'tipo': 'INGRESOS',
            'categoria': 'Ventas de Productos',
            'entidad': 'Cuentas por Cobrar',
            'cuenta': 'Promerica USD (40000003881774)',
            'cliente_proveedor': 'Waipio S.A.',
            'concepto': 'Pago de facturas - Transferencia BAC',
            'referencia': '8319443',
            'monto_usd': 1459.96,
            'monto_crc': None,
            'ingreso_egreso': 'Ingreso',
            'estado': 'Completado',
            'prioridad': 'Normal',
            'vencimiento': None,
            'notas': 'CD BAC - Doc: 8319443',
        },
        {
            'num': 2,
            'descripcion': 'Diesel Petróleos Delta (CORRECTO)',
            'fecha': datetime(2025, 11, 10, 18, 33, 16),
            'tipo': 'GASTOS OPERATIVOS',
            'categoria': 'Combustible',
            'entidad': 'Vehículo',
            'cuenta': 'Promerica USD (40000003881774)',
            'cliente_proveedor': 'Petróleos Delta CR',
            'concepto': 'Diesel - Estación Tibas - Factura #8282',
            'referencia': '204203',
            'monto_usd': -76.28,
            'monto_crc': None,
            'ingreso_egreso': 'Egreso',
            'estado': 'Pagado',
            'prioridad': 'Normal',
            'vencimiento': None,
            'notas': 'Pagado en USD desde Promerica - Doc: 204203',
        },
        {
            'num': 3,
            'descripcion': 'Comisión SINPE',
            'fecha': datetime(2025, 11, 10, 0, 0, 0),
            'tipo': 'GASTOS FINANCIEROS',
            'categoria': 'Comisiones Bancarias',
            'entidad': 'Banco',
            'cuenta': 'Promerica USD (40000003881774)',
            'cliente_proveedor': 'Banco Promerica',
            'concepto': 'Comisión Transferencia SINPE',
            'referencia': '2797944',
            'monto_usd': -0.75,
            'monto_crc': None,
            'ingreso_egreso': 'Egreso',
            'estado': 'Pagado',
            'prioridad': 'Normal',
            'vencimiento': None,
            'notas': 'Comisión débito por transferencia',
        },
    ]

    filas_agregadas = []

    for trans in transacciones:
        next_row = ws.max_row + 1

        ws.cell(next_row, col_map['Fecha']).value = trans['fecha']
        ws.cell(next_row, col_map['Tipo Transacción']).value = trans['tipo']
        ws.cell(next_row, col_map['Categoría']).value = trans['categoria']
        ws.cell(next_row, col_map['Entidad']).value = trans['entidad']
        ws.cell(next_row, col_map['Cuenta Bancaria']).value = trans['cuenta']
        ws.cell(next_row, col_map['Cliente/Proveedor']).value = trans['cliente_proveedor']
        ws.cell(next_row, col_map['Concepto']).value = trans['concepto']
        ws.cell(next_row, col_map['Referencia']).value = trans['referencia']

        if trans['monto_usd']:
            ws.cell(next_row, col_map['Monto USD']).value = trans['monto_usd']
        if trans['monto_crc']:
            ws.cell(next_row, col_map['Monto CRC']).value = trans['monto_crc']

        ws.cell(next_row, col_map['Ingreso/Egreso']).value = trans['ingreso_egreso']
        ws.cell(next_row, col_map['Estado']).value = trans['estado']
        ws.cell(next_row, col_map['Prioridad']).value = trans['prioridad']

        if trans['vencimiento']:
            ws.cell(next_row, col_map['Vencimiento']).value = trans['vencimiento']

        ws.cell(next_row, col_map['Notas']).value = trans['notas']

        filas_agregadas.append({
            'num': trans['num'],
            'fila': next_row,
            'descripcion': trans['descripcion'],
            'monto': trans['monto_usd'],
        })

        signo = '+' if trans['monto_usd'] > 0 else ''
        print(f"✅ Transacción {trans['num']}: {trans['descripcion']}")
        print(f"   Fila: {next_row}")
        print(f"   Monto: {signo}${abs(trans['monto_usd']):,.2f} USD")
        print()

    # Guardar
    print("💾 Guardando cambios...")
    wb.save(EXCEL_FILE)
    print("✅ Excel actualizado")
    print()

    # Resumen
    print("=" * 80)
    print("📊 RESUMEN DE CORRECCIONES")
    print("=" * 80)
    print()

    if delta_crc_fila:
        print(f"🗑️  ELIMINADO:")
        print(f"   • Fila {delta_crc_fila}: Delta en CRC (error)")
        print()

    print(f"✅ AGREGADO:")
    for t in filas_agregadas:
        signo = '+' if t['monto'] > 0 else ''
        print(f"   • Fila {t['fila']}: {t['descripcion']} ({signo}${abs(t['monto']):,.2f})")
    print()

    # Balance
    print("💰 BALANCE DEL DÍA SEGÚN EXTRACTO (10/11/2025):")
    ingresos = sum(t['monto'] for t in filas_agregadas if t['monto'] > 0)
    egresos = sum(t['monto'] for t in filas_agregadas if t['monto'] < 0)

    print(f"   Ingresos:  +${ingresos:,.2f}")
    print(f"   Egresos:   ${egresos:,.2f}")
    print(f"   ───────────────────────")
    print(f"   Neto:      ${ingresos + egresos:,.2f}")
    print()

    print("📊 SALDO PROMERICA SEGÚN EXTRACTO:")
    print(f"   Saldo final 10/11: $2,163.44")
    print()

    print("=" * 80)
    print("✅ CORRECCIÓN COMPLETADA")
    print("=" * 80)
    print()

    print("📋 VERIFICACIÓN:")
    print("   1. Abre Excel y verifica las nuevas filas")
    print("   2. Ejecuta auditoría para verificar saldo Promerica")
    print("   3. Debería coincidir con $2,163.44")
    print()

if __name__ == "__main__":
    try:
        if not crear_backup():
            print("❌ Abortando")
            exit(1)

        corregir_segun_extracto()
        print("🎉 Sincronización con extracto bancario completada!")

    except Exception as e:
        print(f"❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
