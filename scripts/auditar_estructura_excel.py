#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AUDITORÍA DE ESTRUCTURA EXCEL
Analiza categorías existentes, subcategorías, y estructura de TRANSACCIONES
"""
import openpyxl
from collections import Counter

EXCEL_FILE = "AlvaroVelasco_Finanzas_v2.0.xlsx"

def auditar_estructura():
    """Analiza estructura actual del Excel"""

    print("=" * 80)
    print("AUDITORÍA DE ESTRUCTURA - SISTEMA FINANCIERO")
    print("=" * 80)
    print()

    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

    # Verificar hojas existentes
    print("📊 HOJAS EXISTENTES:")
    for i, sheet_name in enumerate(wb.sheetnames, 1):
        ws = wb[sheet_name]
        print(f"   {i}. {sheet_name} ({ws.max_row} filas, {ws.max_column} columnas)")
    print()

    # Analizar hoja TRANSACCIONES
    if 'TRANSACCIONES' not in wb.sheetnames:
        print("❌ ERROR: No existe hoja TRANSACCIONES")
        return

    ws = wb['TRANSACCIONES']

    print("=" * 80)
    print("ANÁLISIS DE TRANSACCIONES")
    print("=" * 80)
    print()

    # Leer encabezados
    print("📋 COLUMNAS (Encabezados):")
    headers = []
    for col in range(1, ws.max_column + 1):
        header = ws.cell(1, col).value
        headers.append(header)
        col_letter = openpyxl.utils.get_column_letter(col)
        print(f"   {col_letter}. {header}")
    print()

    # Encontrar índices de columnas importantes
    try:
        idx_tipo = headers.index('Tipo') + 1 if 'Tipo' in headers else None
        idx_categoria = headers.index('Categoría') + 1 if 'Categoría' in headers else None
        idx_subcategoria = headers.index('Subcategoría') + 1 if 'Subcategoría' in headers else None
        idx_cuenta = headers.index('Cuenta') + 1 if 'Cuenta' in headers else None
        idx_concepto = headers.index('Concepto') + 1 if 'Concepto' in headers else None
        idx_ingreso_egreso = headers.index('Ingreso/Egreso') + 1 if 'Ingreso/Egreso' in headers else None
    except Exception as e:
        print(f"❌ Error identificando columnas: {e}")
        return

    # Contadores
    tipos = Counter()
    categorias = Counter()
    subcategorias = Counter()
    cuentas = Counter()

    # Analizar todas las filas
    print("🔍 ANALIZANDO TRANSACCIONES...")
    for row in range(2, ws.max_row + 1):
        if idx_tipo:
            tipo = ws.cell(row, idx_tipo).value
            if tipo:
                tipos[str(tipo).strip()] += 1

        if idx_categoria:
            cat = ws.cell(row, idx_categoria).value
            if cat:
                categorias[str(cat).strip()] += 1

        if idx_subcategoria:
            subcat = ws.cell(row, idx_subcategoria).value
            if subcat:
                subcategorias[str(subcat).strip()] += 1

        if idx_cuenta:
            cuenta = ws.cell(row, idx_cuenta).value
            if cuenta:
                cuentas[str(cuenta).strip()] += 1

    print(f"✓ {ws.max_row - 1} transacciones analizadas")
    print()

    # Reportes
    print("=" * 80)
    print("📊 TIPOS EXISTENTES")
    print("=" * 80)
    if tipos:
        for tipo, count in tipos.most_common():
            print(f"   • {tipo}: {count} transacciones")
    else:
        print("   (No se encontraron tipos)")
    print()

    print("=" * 80)
    print("📂 CATEGORÍAS EXISTENTES")
    print("=" * 80)
    if categorias:
        for cat, count in categorias.most_common():
            print(f"   • {cat}: {count} transacciones")
    else:
        print("   (No se encontraron categorías)")
    print()

    print("=" * 80)
    print("📑 SUBCATEGORÍAS EXISTENTES")
    print("=" * 80)
    if subcategorias:
        for subcat, count in subcategorias.most_common():
            print(f"   • {subcat}: {count} transacciones")
    else:
        print("   (No se encontraron subcategorías)")
    print()

    print("=" * 80)
    print("🏦 CUENTAS EXISTENTES (Top 20)")
    print("=" * 80)
    if cuentas:
        for i, (cuenta, count) in enumerate(cuentas.most_common(20), 1):
            print(f"   {i:2}. {cuenta}: {count} transacciones")
    else:
        print("   (No se encontraron cuentas)")
    print()

    # Buscar ejemplos de compras a proveedores
    print("=" * 80)
    print("🔎 EJEMPLOS DE COMPRAS A PROVEEDORES (últimas 10)")
    print("=" * 80)

    ejemplos_compras = []
    for row in range(ws.max_row, 1, -1):  # De abajo hacia arriba
        if idx_tipo and idx_categoria and idx_concepto:
            tipo = ws.cell(row, idx_tipo).value
            cat = ws.cell(row, idx_categoria).value
            concepto = ws.cell(row, idx_concepto).value

            # Buscar egresos que podrían ser compras a proveedores
            if tipo and 'egreso' in str(tipo).lower():
                if concepto:
                    concepto_lower = str(concepto).lower()
                    # Palabras clave de compras
                    if any(kw in concepto_lower for kw in ['intcomex', 'factura', 'compra', 'proveedor', 'producto']):
                        ejemplos_compras.append({
                            'fila': row,
                            'tipo': tipo,
                            'categoria': cat,
                            'concepto': str(concepto)[:60]
                        })
                        if len(ejemplos_compras) >= 10:
                            break

    if ejemplos_compras:
        for ej in ejemplos_compras:
            print(f"   Fila {ej['fila']:3}: [{ej['categoria']}] {ej['concepto']}")
    else:
        print("   (No se encontraron ejemplos obvios)")
    print()

    # Resumen y recomendaciones
    print("=" * 80)
    print("💡 RESUMEN Y RECOMENDACIONES")
    print("=" * 80)
    print()

    total_transacciones = ws.max_row - 1
    transacciones_sin_categoria = total_transacciones - sum(categorias.values())
    transacciones_sin_subcategoria = total_transacciones - sum(subcategorias.values())

    print(f"📊 Estadísticas:")
    print(f"   • Total transacciones: {total_transacciones}")
    print(f"   • Con categoría: {sum(categorias.values())} ({sum(categorias.values())/total_transacciones*100:.1f}%)")
    print(f"   • Sin categoría: {transacciones_sin_categoria} ({transacciones_sin_categoria/total_transacciones*100:.1f}%)")
    print(f"   • Con subcategoría: {sum(subcategorias.values())} ({sum(subcategorias.values())/total_transacciones*100:.1f}%)")
    print(f"   • Sin subcategoría: {transacciones_sin_subcategoria} ({transacciones_sin_subcategoria/total_transacciones*100:.1f}%)")
    print()

    print("🎯 Próximos pasos recomendados:")
    print()
    print("   1. Revisar categorías existentes vs estructura propuesta")
    print("   2. Identificar qué categoría se usa actualmente para compras a proveedores")
    print("   3. Crear categoría 'Costo de Ventas' o 'COGS' si no existe")
    print("   4. Recategorizar compras a proveedores: Gastos → COGS")
    print("   5. Estandarizar subcategorías para análisis detallado")
    print()

    print("=" * 80)
    print("AUDITORÍA COMPLETADA")
    print("=" * 80)

if __name__ == "__main__":
    try:
        auditar_estructura()
    except FileNotFoundError:
        print(f"❌ ERROR: No se encontró el archivo {EXCEL_FILE}")
        print(f"   Asegúrate de ejecutar este script desde la carpeta del proyecto")
    except Exception as e:
        print(f"❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
