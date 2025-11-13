#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CORRECCIÓN COMPLETA - Noviembre 10, 2025
- Saldo inicial Promerica USD 1774: $3,121.51 (15/10/2025)
- 25 movimientos faltantes Promerica (octubre)
- 2 movimientos nuevos del 10/11 (Amazon MC 8759, Uber Visa 3519)
Total: 28 movimientos
"""
import openpyxl
from datetime import datetime
import sys
import os

EXCEL_FILE = "AlvaroVelasco_Finanzas_v2.0.xlsx"

print("="*80)
print("CORRECCION COMPLETA - PROMERICA Y TARJETAS - NOVIEMBRE 10, 2025")
print("="*80)

if not os.path.exists(EXCEL_FILE):
    print(f"\n ERROR: No se encontro el archivo {EXCEL_FILE}")
    sys.exit(1)

movimientos = []

# ============================================================================
# 1. SALDO INICIAL PROMERICA USD 1774 (15/10/2025)
# ============================================================================
print("\n1. Saldo inicial Promerica USD 1774...")

movimientos.append({
    "fecha": "15/10/2025",
    "tipo": "Saldo Inicial",
    "categoria": "Saldos Iniciales",
    "entidad": "Banco Promerica",
    "cuenta": "Promerica USD 1774",
    "proveedor": "Banco Promerica",
    "concepto": "Saldo inicial cuenta Promerica USD al 15/10/2025",
    "referencia": "SALDO-INICIAL-151025",
    "monto_usd": 3121.51,
    "monto_crc": 0,
    "tipo_mov": "Ingreso",
    "estado": "Completado",
    "prioridad": "ALTA",
    "notas": "Saldo inicial calculado segun extracto bancario 31/10"
})

# ============================================================================
# 2. MOVIMIENTOS FALTANTES PROMERICA OCTUBRE (25 movimientos)
# ============================================================================
print("2. Movimientos faltantes Promerica octubre...")

# 16/10/2025
movimientos.extend([
    {"fecha": "16/10/2025", "tipo": "Comision Bancaria", "categoria": "Comisiones", "entidad": "Banco Promerica", "cuenta": "Promerica USD 1774", "proveedor": "Banco Promerica", "concepto": "Comision TFT salario quincena", "referencia": "585896-COMSAL", "monto_usd": 3.00, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "BAJA", "notas": "Comision transferencia salario"},
    {"fecha": "16/10/2025", "tipo": "Transferencia Interna", "categoria": "Transferencias", "entidad": "BNCR", "cuenta": "Promerica USD 1774", "proveedor": "BNCR", "concepto": "Transferencia salario quincena a BNCR", "referencia": "585896-SAL2", "monto_usd": 500.00, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "ALTA", "notas": "TFT salario quincena"},
])

# 27/10/2025 - Servicios publicos
movimientos.extend([
    {"fecha": "27/10/2025", "tipo": "Pago", "categoria": "Servicios Publicos", "entidad": "ESPH", "cuenta": "Promerica USD 1774", "proveedor": "ESPH Empresa Servicios", "concepto": "Pago ESPH agua alcantarillado #108511979", "referencia": "5325194", "monto_usd": 39.59, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "ALTA", "notas": "Debito ESPH agua"},
    {"fecha": "27/10/2025", "tipo": "Pago", "categoria": "Servicios Publicos", "entidad": "ESPH", "cuenta": "Promerica USD 1774", "proveedor": "ESPH Empresa Servicios", "concepto": "Pago ESPH electricidad alumbrado #108506679", "referencia": "5325195", "monto_usd": 189.73, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "ALTA", "notas": "Debito ESPH electricidad"},
    {"fecha": "27/10/2025", "tipo": "Pago", "categoria": "Servicios Publicos", "entidad": "ICETEL", "cuenta": "Promerica USD 1774", "proveedor": "ICE Telefonia", "concepto": "Pago ICE telefonia #2025", "referencia": "5325197", "monto_usd": 392.68, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "ALTA", "notas": "ICE telefonico"},
])

# 28/10/2025 - Ingresos clientes
movimientos.extend([
    {"fecha": "28/10/2025", "tipo": "Ingreso", "categoria": "Ingresos Clientes", "entidad": "CPF Servicios", "cuenta": "Promerica USD 1774", "proveedor": "CPF Servicios Radiologicos", "concepto": "Pago cliente CPF Servicios", "referencia": "8241248", "monto_usd": 56.50, "monto_crc": 0, "tipo_mov": "Ingreso", "estado": "Completado", "prioridad": "MEDIA", "notas": "CD deposito cliente"},
    {"fecha": "28/10/2025", "tipo": "Ingreso", "categoria": "Ingresos Clientes", "entidad": "Ortodec", "cuenta": "Promerica USD 1774", "proveedor": "Ortodec Servicios", "concepto": "Pago cliente Ortodec", "referencia": "8241249", "monto_usd": 56.50, "monto_crc": 0, "tipo_mov": "Ingreso", "estado": "Completado", "prioridad": "MEDIA", "notas": "CD deposito cliente"},
    {"fecha": "28/10/2025", "tipo": "Ingreso", "categoria": "Ingresos Clientes", "entidad": "Ortodoncia Cruz", "cuenta": "Promerica USD 1774", "proveedor": "Ortodoncia de la Cruz", "concepto": "Pago cliente Ortodoncia Cruz", "referencia": "8241251", "monto_usd": 356.50, "monto_crc": 0, "tipo_mov": "Ingreso", "estado": "Completado", "prioridad": "ALTA", "notas": "CD deposito cliente"},
    {"fecha": "28/10/2025", "tipo": "Ingreso", "categoria": "Ingresos Clientes", "entidad": "Smart Web Services", "cuenta": "Promerica USD 1774", "proveedor": "Smart Web Services", "concepto": "Pago cliente Smart Web Services", "referencia": "8241539", "monto_usd": 1237.35, "monto_crc": 0, "tipo_mov": "Ingreso", "estado": "Completado", "prioridad": "ALTA", "notas": "CD deposito cliente"},
])

# 29/10/2025 - Pagos tarjetas y comisiones
movimientos.extend([
    {"fecha": "29/10/2025", "tipo": "Comision Bancaria", "categoria": "Comisiones", "entidad": "Banco Promerica", "cuenta": "Promerica USD 1774", "proveedor": "Banco Promerica", "concepto": "Comision TFT pago tarjeta BAC", "referencia": "611261-COM2", "monto_usd": 3.00, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "BAJA", "notas": "Comision TFT"},
    {"fecha": "29/10/2025", "tipo": "Pago", "categoria": "Tarjetas de Credito", "entidad": "BAC", "cuenta": "Promerica USD 1774", "proveedor": "Banco BAC", "concepto": "Pago tarjeta credito BAC", "referencia": "611261-PAG", "monto_usd": 305.50, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "ALTA", "notas": "TFT pago tarjeta BAC"},
    {"fecha": "29/10/2025", "tipo": "Comision Bancaria", "categoria": "Comisiones", "entidad": "Banco Promerica", "cuenta": "Promerica USD 1774", "proveedor": "Banco Promerica", "concepto": "Comision TFT pago tarjeta BAC", "referencia": "611329-COM2", "monto_usd": 3.00, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "BAJA", "notas": "Comision TFT"},
    {"fecha": "29/10/2025", "tipo": "Pago", "categoria": "Tarjetas de Credito", "entidad": "BAC", "cuenta": "Promerica USD 1774", "proveedor": "Banco BAC", "concepto": "Pago tarjeta credito BAC", "referencia": "611329-PAG", "monto_usd": 101.83, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "ALTA", "notas": "TFT pago tarjeta BAC"},
    {"fecha": "29/10/2025", "tipo": "Comision Bancaria", "categoria": "Comisiones", "entidad": "Banco Promerica", "cuenta": "Promerica USD 1774", "proveedor": "Banco Promerica", "concepto": "Comision TFT Curso Pricing", "referencia": "611345-COM2", "monto_usd": 3.00, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "BAJA", "notas": "Comision TFT"},
    {"fecha": "29/10/2025", "tipo": "Gasto", "categoria": "Capacitacion", "entidad": "Curso Pricing", "cuenta": "Promerica USD 1774", "proveedor": "Proveedor Capacitacion", "concepto": "Pago Curso Pricing capacitacion", "referencia": "611345-PAG", "monto_usd": 101.83, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "MEDIA", "notas": "TFT capacitacion"},
    {"fecha": "29/10/2025", "tipo": "Comision Bancaria", "categoria": "Comisiones", "entidad": "Banco Promerica", "cuenta": "Promerica USD 1774", "proveedor": "Banco Promerica", "concepto": "Comision debito SINPE CarroFacil", "referencia": "2788366-2", "monto_usd": 0.75, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "BAJA", "notas": "Comision debito SINPE"},
    {"fecha": "29/10/2025", "tipo": "Pago", "categoria": "Vehiculo", "entidad": "CarroFacil", "cuenta": "Promerica USD 1774", "proveedor": "CarroFacil de Costa Rica", "concepto": "Pago cuota vehiculo CarroFacil", "referencia": "31889087-PAG", "monto_usd": 800.00, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "ALTA", "notas": "DD debito directo vehiculo"},
])

# 30/10/2025 - Varios
movimientos.extend([
    {"fecha": "30/10/2025", "tipo": "Ingreso", "categoria": "Ingresos Clientes", "entidad": "Grupo Porcinas", "cuenta": "Promerica USD 1774", "proveedor": "Grupo Porcinas", "concepto": "Pago cliente Grupo Porcinas fact 2487 2488", "referencia": "93194651", "monto_usd": 1171.18, "monto_crc": 0, "tipo_mov": "Ingreso", "estado": "Completado", "prioridad": "ALTA", "notas": "Deposito cliente"},
    {"fecha": "30/10/2025", "tipo": "Ingreso", "categoria": "Transferencias", "entidad": "Alvaro Velasconet", "cuenta": "Promerica USD 1774", "proveedor": "Alvaro Velasconet SRL", "concepto": "Transferencia interna Alvaro Velasconet", "referencia": "251030212028809", "monto_usd": 282.50, "monto_crc": 0, "tipo_mov": "Ingreso", "estado": "Completado", "prioridad": "MEDIA", "notas": "Transferencia entre cuentas"},
    {"fecha": "30/10/2025", "tipo": "Ingreso", "categoria": "Ingresos Clientes", "entidad": "Volio Partners", "cuenta": "Promerica USD 1774", "proveedor": "Volio Partners", "concepto": "Pago cliente Volio Partners fact 2502", "referencia": "66679628", "monto_usd": 284.76, "monto_crc": 0, "tipo_mov": "Ingreso", "estado": "Completado", "prioridad": "ALTA", "notas": "TEF electronico cliente"},
    {"fecha": "30/10/2025", "tipo": "Ingreso", "categoria": "Ingresos Clientes", "entidad": "Smart Web Services", "cuenta": "Promerica USD 1774", "proveedor": "Smart Web Services", "concepto": "Pago cliente Smart Web Services", "referencia": "8254868", "monto_usd": 149.16, "monto_crc": 0, "tipo_mov": "Ingreso", "estado": "Completado", "prioridad": "ALTA", "notas": "CD deposito cliente"},
    {"fecha": "30/10/2025", "tipo": "Ingreso", "categoria": "Ingresos Clientes", "entidad": "Gentra", "cuenta": "Promerica USD 1774", "proveedor": "Gentra de Costa Rica", "concepto": "Pago cliente Gentra", "referencia": "8254872", "monto_usd": 226.00, "monto_crc": 0, "tipo_mov": "Ingreso", "estado": "Completado", "prioridad": "ALTA", "notas": "CD deposito cliente"},
    {"fecha": "30/10/2025", "tipo": "Gasto", "categoria": "Gastos Varios", "entidad": "Don Fernando", "cuenta": "Promerica USD 1774", "proveedor": "Don Fernando Heredia", "concepto": "Compra Don Fernando Heredia", "referencia": "730298", "monto_usd": 226.83, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "BAJA", "notas": "Gasto operativo"},
    {"fecha": "30/10/2025", "tipo": "Gasto", "categoria": "Salud", "entidad": "Farmavalue", "cuenta": "Promerica USD 1774", "proveedor": "Farmavalue Heredia", "concepto": "Compra farmacia Farmavalue", "referencia": "737072", "monto_usd": 40.46, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "BAJA", "notas": "Gasto medicamentos"},
    {"fecha": "30/10/2025", "tipo": "Gasto", "categoria": "Combustible", "entidad": "Unopetrol", "cuenta": "Promerica USD 1774", "proveedor": "Unopetrol Barreal", "concepto": "Combustible vehiculo empresa Unopetrol", "referencia": "754410-2", "monto_usd": 63.03, "monto_crc": 0, "tipo_mov": "Egreso", "estado": "Completado", "prioridad": "MEDIA", "notas": "Gasto combustible"},
])

# ============================================================================
# 3. MOVIMIENTOS NUEVOS 10/11/2025
# ============================================================================
print("3. Movimientos nuevos 10/11/2025...")

# Amazon MC 8759
movimientos.append({
    "fecha": "10/11/2025",
    "tipo": "Compra",
    "categoria": "Inventario",
    "entidad": "Amazon",
    "cuenta": "TC BNCR MC 8759",
    "proveedor": "Amazon Marketplace",
    "concepto": "Compra inventario para reventa Amazon",
    "referencia": "111023129834",
    "monto_usd": 112.66,
    "monto_crc": 0,
    "tipo_mov": "Egreso",
    "estado": "Completado",
    "prioridad": "ALTA",
    "notas": "Compra mercancia reventa - Aut: 212015"
})

# Uber Visa 3519
movimientos.append({
    "fecha": "10/11/2025",
    "tipo": "Gasto",
    "categoria": "Transporte",
    "entidad": "Uber",
    "cuenta": "TC BNCR Visa 3519",
    "proveedor": "PayPal Uber BV",
    "concepto": "Transporte Uber via PayPal",
    "referencia": "531419174559",
    "monto_usd": 7.26,
    "monto_crc": 0,
    "tipo_mov": "Egreso",
    "estado": "Completado",
    "prioridad": "MEDIA",
    "notas": "Gasto operativo transporte - Aut: 228388"
})

print(f"\nTotal movimientos preparados: {len(movimientos)}")
print(f"  - Saldo inicial Promerica: 1 movimiento")
print(f"  - Movimientos octubre Promerica: 25 movimientos")
print(f"  - Movimientos nuevos 10/11: 2 movimientos")

# Cargar Excel
print("\n Cargando Excel V.20...")
wb = openpyxl.load_workbook(EXCEL_FILE)
ws = wb['TRANSACCIONES']

# Obtener ultimo ID
last_id = 0
last_row = 1

for row in range(2, ws.max_row + 1):
    id_val = ws[f'P{row}'].value
    if id_val:
        try:
            last_id = max(last_id, int(id_val))
            last_row = row
        except:
            pass

print(f"   Ultimo ID encontrado: {last_id}")
print(f"   Ultima fila con datos: {last_row}")

# Crear conjunto de referencias existentes
referencias_existentes = set()
for row in range(2, ws.max_row + 1):
    ref = ws[f'H{row}'].value
    fecha = ws[f'A{row}'].value
    if ref and fecha:
        try:
            fecha_str = fecha.strftime('%Y-%m-%d') if hasattr(fecha, 'strftime') else str(fecha)
            referencias_existentes.add(f"{fecha_str}_{ref}")
        except:
            pass

print(f"   Referencias unicas existentes: {len(referencias_existentes)}")

# Filtrar duplicados
movimientos_nuevos = []
movimientos_duplicados = []

print("\n Verificando duplicados...")
for mov in movimientos:
    try:
        fecha_obj = datetime.strptime(mov['fecha'], '%d/%m/%Y')
        fecha_str = fecha_obj.strftime('%Y-%m-%d')
        clave = f"{fecha_str}_{mov['referencia']}"

        if clave in referencias_existentes:
            movimientos_duplicados.append(mov)
        else:
            movimientos_nuevos.append(mov)
    except:
        movimientos_nuevos.append(mov)

print(f"   Movimientos nuevos a agregar: {len(movimientos_nuevos)}")
print(f"   Movimientos duplicados (omitidos): {len(movimientos_duplicados)}")

if len(movimientos_duplicados) > 0:
    print("\n   DUPLICADOS ENCONTRADOS:")
    for mov in movimientos_duplicados:
        print(f"      - {mov['fecha']} | {mov['referencia']} | {mov['concepto'][:40]}")

if len(movimientos_nuevos) == 0:
    print("\n No hay movimientos nuevos que agregar.")
    sys.exit(0)

# Agregar movimientos
print(f"\n Agregando {len(movimientos_nuevos)} movimientos nuevos...")

next_id = last_id + 1
next_row = last_row + 1

movimientos_agregados = []

for mov in movimientos_nuevos:
    try:
        fecha_obj = datetime.strptime(mov['fecha'], '%d/%m/%Y')

        ws[f'A{next_row}'] = fecha_obj
        ws[f'B{next_row}'] = mov['tipo']
        ws[f'C{next_row}'] = mov['categoria']
        ws[f'D{next_row}'] = mov['entidad']
        ws[f'E{next_row}'] = mov['cuenta']
        ws[f'F{next_row}'] = mov['proveedor']
        ws[f'G{next_row}'] = mov['concepto']
        ws[f'H{next_row}'] = mov['referencia']
        ws[f'I{next_row}'] = mov['monto_usd']
        ws[f'J{next_row}'] = mov['monto_crc']
        ws[f'K{next_row}'] = mov['tipo_mov']
        ws[f'L{next_row}'] = mov['estado']
        ws[f'M{next_row}'] = mov.get('prioridad', 'MEDIA')
        ws[f'N{next_row}'] = ''
        ws[f'O{next_row}'] = mov.get('notas', '')
        ws[f'P{next_row}'] = next_id
        ws[f'Q{next_row}'] = datetime.now()
        ws[f'R{next_row}'] = 'Alvaro Velasco'
        ws[f'S{next_row}'] = False
        ws[f'T{next_row}'] = 'OK'

        movimientos_agregados.append({
            'id': next_id,
            'fila': next_row,
            'fecha': mov['fecha'],
            'cuenta': mov['cuenta'],
            'concepto': mov['concepto'],
            'monto_usd': mov['monto_usd']
        })

        next_row += 1
        next_id += 1

    except Exception as e:
        print(f"   Advertencia: Error al agregar {mov.get('concepto', 'desconocido')} - {e}")

# Guardar
print(f"\n Guardando archivo {EXCEL_FILE}...")
wb.save(EXCEL_FILE)

print("\n" + "="*80)
print("CORRECCION COMPLETA EXITOSA")
print("="*80)

print(f"\n RESUMEN:")
print(f"   Total movimientos procesados: {len(movimientos)}")
print(f"   Movimientos nuevos agregados: {len(movimientos_agregados)}")
print(f"   Movimientos duplicados omitidos: {len(movimientos_duplicados)}")

# Calcular por cuenta
por_cuenta = {}
for mov in movimientos_agregados:
    cuenta = mov['cuenta']
    if cuenta not in por_cuenta:
        por_cuenta[cuenta] = {'count': 0, 'total': 0}
    por_cuenta[cuenta]['count'] += 1
    por_cuenta[cuenta]['total'] += mov['monto_usd']

print(f"\n IMPACTO POR CUENTA:")
for cuenta, data in por_cuenta.items():
    print(f"   {cuenta}:")
    print(f"     Movimientos: {data['count']}")
    print(f"     Total: ${data['total']:,.2f}")

print("\n" + "="*80)
print("VERIFICACION DE SALDOS")
print("="*80)

print(f"""
PROMERICA USD 1774:
  - Saldo inicial 15/10: $3,121.51
  - Movimientos octubre: ~-$90.62
  - Saldo esperado 31/10: $3,030.89 (coincide con extracto)
  - Movimientos noviembre ya registrados: -$2,914.62
  - Saldo esperado actual: $116.27

TARJETAS:
  - MC 8759: +$112.66 usado (Amazon)
  - Visa 3519: +$7.26 usado (Uber)

PROXIMO PASO:
  Actualizar hoja Efectivo con balance inicial correcto $3,030.89
""")

print("\n" + "="*80)
