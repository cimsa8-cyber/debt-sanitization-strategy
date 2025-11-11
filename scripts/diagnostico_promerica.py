#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
DIAGNÓSTICO DETALLADO PROMERICA
Muestra TODAS las transacciones de Promerica en noviembre para entender la diferencia
"""
import openpyxl
from datetime import datetime

EXCEL_FILE = "AlvaroVelasco_Finanzas_v2.0.xlsx"

def diagnostico():
    print("=" * 80)
    print("DIAGNÓSTICO DETALLADO - CUENTA PROMERICA")
    print("=" * 80)
    print()

    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb['TRANSACCIONES']

    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    col_map = {}
    for col in range(1, len(headers) + 1):
        if headers[col-1]:
            col_map[headers[col-1]] = col

    print("🔍 Buscando TODAS las transacciones de Promerica...")
    print()

    # Buscar TODAS las transacciones Promerica
    transacciones_promerica = []

    for row in range(2, ws.max_row + 1):
        cuenta = ws.cell(row, col_map['Cuenta Bancaria']).value

        if cuenta and 'Promerica' in str(cuenta) and '40000003881774' in str(cuenta):
            fecha = ws.cell(row, col_map['Fecha']).value
            tipo = ws.cell(row, col_map['Tipo Transacción']).value
            categoria = ws.cell(row, col_map['Categoría']).value
            concepto = ws.cell(row, col_map['Concepto']).value
            monto_usd = ws.cell(row, col_map['Monto USD']).value
            referencia = ws.cell(row, col_map['Referencia']).value
            ing_egr = ws.cell(row, col_map['Ingreso/Egreso']).value

            transacciones_promerica.append({
                'fila': row,
                'fecha': fecha,
                'tipo': tipo,
                'categoria': categoria,
                'concepto': concepto,
                'monto_usd': float(monto_usd) if monto_usd else 0,
                'referencia': referencia,
                'ing_egr': ing_egr,
            })

    # Mostrar todas
    print(f"📋 TOTAL TRANSACCIONES PROMERICA: {len(transacciones_promerica)}")
    print()

    # Separar por mes
    print("=" * 80)
    print("📅 TRANSACCIONES POR MES")
    print("=" * 80)
    print()

    # Octubre
    octubre = [t for t in transacciones_promerica
               if t['fecha'] and isinstance(t['fecha'], datetime)
               and t['fecha'].month == 10 and t['fecha'].year == 2025]

    print(f"🗓️  OCTUBRE 2025: {len(octubre)} transacciones")
    if octubre:
        total_octubre = sum(t['monto_usd'] for t in octubre)
        print(f"   Total impacto: ${total_octubre:,.2f}")
        for t in octubre:
            fecha_str = t['fecha'].strftime('%d/%m/%Y')
            signo = '+' if t['monto_usd'] > 0 else ''
            print(f"   • Fila {t['fila']}: {fecha_str} - {signo}${t['monto_usd']:,.2f} - {t['concepto'][:50] if t['concepto'] else 'Sin concepto'}")
    print()

    # Noviembre
    noviembre = [t for t in transacciones_promerica
                 if t['fecha'] and isinstance(t['fecha'], datetime)
                 and t['fecha'].month == 11 and t['fecha'].year == 2025]

    print(f"🗓️  NOVIEMBRE 2025: {len(noviembre)} transacciones")
    if noviembre:
        # Excluir apertura inicial
        nov_sin_apertura = [t for t in noviembre
                           if not (t['tipo'] and 'Apertura Inicial' in str(t['tipo']))]

        print(f"   (Sin contar apertura inicial: {len(nov_sin_apertura)} transacciones)")
        print()

        total_noviembre = sum(t['monto_usd'] for t in nov_sin_apertura)
        print(f"   Total impacto noviembre: ${total_noviembre:,.2f}")
        print()

        # Mostrar cada transacción
        for t in sorted(noviembre, key=lambda x: x['fecha'] if x['fecha'] else datetime.min):
            fecha_str = t['fecha'].strftime('%d/%m/%Y')
            signo = '+' if t['monto_usd'] > 0 else ''
            es_apertura = t['tipo'] and 'Apertura Inicial' in str(t['tipo'])
            marca = "🔷 APERTURA" if es_apertura else "  "

            print(f"{marca} Fila {t['fila']}: {fecha_str}")
            print(f"   Tipo: {t['tipo']}")
            print(f"   Monto: {signo}${t['monto_usd']:,.2f}")
            print(f"   Ing/Egr: {t['ing_egr']}")
            print(f"   Concepto: {t['concepto'][:60] if t['concepto'] else 'Sin concepto'}")
            print(f"   Ref: {t['referencia']}")
            print()

    # Otros meses
    otros = [t for t in transacciones_promerica
             if t['fecha'] and isinstance(t['fecha'], datetime)
             and (t['fecha'].month not in [10, 11] or t['fecha'].year != 2025)]

    if otros:
        print(f"🗓️  OTROS MESES: {len(otros)} transacciones")
        for t in otros:
            fecha_str = t['fecha'].strftime('%d/%m/%Y')
            print(f"   • Fila {t['fila']}: {fecha_str} - ${t['monto_usd']:,.2f}")
        print()

    # Sin fecha
    sin_fecha = [t for t in transacciones_promerica if not t['fecha']]
    if sin_fecha:
        print(f"⚠️  SIN FECHA: {len(sin_fecha)} transacciones")
        for t in sin_fecha:
            print(f"   • Fila {t['fila']}: ${t['monto_usd']:,.2f} - {t['concepto'][:50] if t['concepto'] else 'Sin concepto'}")
        print()

    # Calcular saldo
    print("=" * 80)
    print("💰 CÁLCULO DE SALDO")
    print("=" * 80)
    print()

    # Buscar saldo inicial
    saldo_inicial = 0
    apertura = [t for t in noviembre
                if t['tipo'] and 'Apertura Inicial' in str(t['tipo'])]

    if apertura:
        saldo_inicial = apertura[0]['monto_usd']
        print(f"✅ Saldo Inicial (01/11/2025): ${saldo_inicial:,.2f}")
    else:
        print("⚠️  No se encontró apertura inicial")

    # Sumar movimientos noviembre
    nov_sin_apertura = [t for t in noviembre
                       if not (t['tipo'] and 'Apertura Inicial' in str(t['tipo']))]

    total_movimientos = sum(t['monto_usd'] for t in nov_sin_apertura)
    saldo_calculado = saldo_inicial + total_movimientos

    print(f"📊 Movimientos noviembre: {len(nov_sin_apertura)} transacciones")
    print(f"📊 Total movimientos: ${total_movimientos:,.2f}")
    print(f"💰 Saldo calculado: ${saldo_calculado:,.2f}")
    print()

    print(f"🏦 Saldo extracto bancario: $2,163.44")
    print(f"📊 Diferencia: ${abs(saldo_calculado - 2163.44):,.2f}")
    print()

    # Análisis de débitos y créditos
    print("=" * 80)
    print("📊 ANÁLISIS DÉBITOS/CRÉDITOS NOVIEMBRE")
    print("=" * 80)
    print()

    debitos = [t for t in nov_sin_apertura if t['monto_usd'] < 0]
    creditos = [t for t in nov_sin_apertura if t['monto_usd'] > 0]

    total_debitos = sum(abs(t['monto_usd']) for t in debitos)
    total_creditos = sum(t['monto_usd'] for t in creditos)

    print(f"📉 DÉBITOS: {len(debitos)} transacciones")
    print(f"   Total: ${total_debitos:,.2f}")
    for t in debitos:
        fecha_str = t['fecha'].strftime('%d/%m/%Y') if t['fecha'] else 'Sin fecha'
        print(f"   • {fecha_str}: -${abs(t['monto_usd']):,.2f} - {t['concepto'][:40] if t['concepto'] else 'N/A'}")
    print()

    print(f"📈 CRÉDITOS: {len(creditos)} transacciones")
    print(f"   Total: ${total_creditos:,.2f}")
    for t in creditos:
        fecha_str = t['fecha'].strftime('%d/%m/%Y') if t['fecha'] else 'Sin fecha'
        print(f"   • {fecha_str}: +${t['monto_usd']:,.2f} - {t['concepto'][:40] if t['concepto'] else 'N/A'}")
    print()

    print(f"💰 IMPACTO NETO: ${total_creditos - total_debitos:,.2f}")
    print()

    # Comparar con extracto
    print("=" * 80)
    print("📊 COMPARACIÓN CON EXTRACTO BANCARIO")
    print("=" * 80)
    print()

    print("🏦 SEGÚN EXTRACTO:")
    print("   Débitos totales: $4,058.20")
    print("   Créditos totales: $3,190.75")
    print("   Impacto neto: -$867.45")
    print()

    print("📊 SEGÚN EXCEL:")
    print(f"   Débitos totales: ${total_debitos:,.2f}")
    print(f"   Créditos totales: ${total_creditos:,.2f}")
    print(f"   Impacto neto: ${total_creditos - total_debitos:,.2f}")
    print()

    dif_debitos = abs(total_debitos - 4058.20)
    dif_creditos = abs(total_creditos - 3190.75)

    print("⚖️  DIFERENCIAS:")
    print(f"   Débitos: ${dif_debitos:,.2f}")
    print(f"   Créditos: ${dif_creditos:,.2f}")
    print()

    if dif_debitos < 1 and dif_creditos < 1:
        print("✅ Los totales coinciden con el extracto")
    else:
        print("⚠️  Hay diferencias que investigar")

    print()
    print("=" * 80)
    print("✅ DIAGNÓSTICO COMPLETADO")
    print("=" * 80)

if __name__ == "__main__":
    try:
        diagnostico()
    except Exception as e:
        print(f"❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
